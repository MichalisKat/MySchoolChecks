# -*- coding: utf-8 -*-
"""
Οργάνωση τοποθετήσεων/συμπληρώσεων ανά ειδικότητα (ΠΕ06, ΠΕ05, ΠΕ07, ΠΕ08,
ΠΕ11, ΠΕ79, ΠΕ86, ΠΕ91) σε ένα ενιαίο βιβλίο εργασίας, ένα φύλλο ανά
ειδικότητα, με:
  • στήλη «Υποχρεωτικό MySchool» (από τα στατιστικά 4.1 / 4.2 — πρώτα 4.1,
    μετά 4.2 αν δεν βρεθεί εκεί),
  • στήλη «Υπόλοιπο για Τοποθέτηση» (πόσες ώρες λείπουν μέχρι το
    υποχρεωτικό, ώστε να μπουν στο σχολείο τοποθέτησης),
  • «Ενδεικτικές» ώρες προστιθέμενες κανονικά στο σύνολο.

Δύο δομές πρωτοτύπου εντοπίστηκαν στο αρχείο "Ομαδοποίηση ΕΙΔΙΚΟΤΗΤΩΝ":

  FAMILY A ("ΟΜΑΔΟΠΟΙΗΣΗ" — ΓΑΛΛ ΠΕ05, ΓΕΡΜ ΠΕ07): μία γραμμή ανά σχολείο
    (κύριο + συμπληρώσεις), ομάδες χωρισμένες με κενή γραμμή.
  FAMILY B ("ΘΕΣΕΙΣ" — ΠΕ06, ΠΕ08, ΠΕ11, ΠΕ79, ΠΕ86, ΠΕ91): μία γραμμή ανά
    εκπαιδευτικό, στήλες τοποθέτησης/συμπληρώσεων· η διάταξη στηλών
    διαφέρει ανά φύλλο, γι' αυτό ανιχνεύεται από τις επικεφαλίδες.

Αυτό το module εκτελεί όλη τη λογική δυναμικά, με είσοδο:
  • το αρχείο .xls "Ομαδοποίηση ΕΙΔΙΚΟΤΗΤΩΝ" (κύρια πηγή δεδομένων),
  • προαιρετικά ένα 2ο .xls "ΣΥΜΠΛΗΡΩΣΕΙΣ ΕΙΔΙΚΟΤΗΤΩΝ" για διασταύρωση,
  • τα στατιστικά 4.1 και 4.2 (CSV, cp1253) που κατεβαίνουν από το MySchool.
"""

import re
import csv
from collections import defaultdict

import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Ειδικότητες που υποστηρίζονται ──────────────────────────────────────────
# (sheet_name στο πρωτότυπο, κωδικός ειδικότητας MySchool, τίτλος εξόδου)
FAMILY_A_SPECIALTIES = [
    ('ΓΑΛΛ ΠΕ05', 'ΠΕ05', 'ΠΕ05 Γαλλικής'),
    ('ΓΕΡΜ ΠΕ07', 'ΠΕ07', 'ΠΕ07 Γερμανικής'),
]
FAMILY_B_SPECIALTIES = [
    ('ΠΕ06', 'ΠΕ06', 'ΠΕ06 Αγγλικής'),
    ('ΠΕ08', 'ΠΕ08', 'ΠΕ08 Καλλιτεχνικών'),
    ('ΠΕ11', 'ΠΕ11', 'ΠΕ11 Φυσικής Αγωγής'),
    ('ΠΕ79', 'ΠΕ79', 'ΠΕ79 Μουσικής'),
    ('ΠΕ86', 'ΠΕ86', 'ΠΕ86 Πληροφορικής'),
    ('ΠΕ91', 'ΠΕ91', 'ΠΕ91 Θεατρικής Αγωγής'),
]


# ── MySchool 4.1/4.2 lookup (γενικό, ανά κωδικό ειδικότητας) ────────────────

def _load_csv_rows(path):
    with open(path, encoding='cp1253') as f:
        r = csv.reader(f, delimiter=';')
        next(r, None)
        return list(r)


def _build_index(rows_all, code_prefix):
    idx = defaultdict(list)
    for row in rows_all:
        if len(row) > 25 and row[14].strip().startswith(code_prefix):
            idx[row[3].strip().upper()].append((row[4].strip().upper(), row[25].strip()))
    return idx


def _parse_name(raw):
    raw = raw.strip()
    raw_nonum = re.sub(r'\s*\d+\s*$', '', raw).strip()
    parts = raw_nonum.split()
    surname = parts[0].upper() if parts else ''
    fname_hint = ' '.join(parts[1:]).upper() if len(parts) > 1 else ''
    return surname, fname_hint


def _make_lookup(rows41_all, rows42_all, code_prefix):
    idx41 = _build_index(rows41_all, code_prefix)
    idx42 = _build_index(rows42_all, code_prefix)

    def lookup(surname, fname_hint):
        for idx, src in ((idx41, '4.1'), (idx42, '4.2')):
            cands = idx.get(surname)
            if not cands:
                continue
            if len(cands) == 1:
                return cands[0][1], src, 'unique'
            if fname_hint:
                m = [c for c in cands
                     if c[0].startswith(fname_hint)
                     or fname_hint.startswith(c[0].split()[0] if c[0] else '')]
                if len(m) == 1:
                    return m[0][1], src, 'disambiguated'
            return None, src, f'ambiguous x{len(cands)}'
        return None, None, 'δεν βρέθηκε'
    return lookup


def _num(x):
    return x if isinstance(x, (int, float)) and x != '' else None


def _s(x):
    return str(x).strip() if x is not None else ''


def _name_mandatory(raw_name):
    m = re.search(r'(\d+)\s*$', raw_name)
    return int(m.group(1)) if m else None


# ── Στυλ ─────────────────────────────────────────────────────────────────────
_HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
_HDR_FONT = Font(bold=True, color='FFFFFF')
_REMAIN_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
_MISMATCH_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
_GROUP_ALT_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')


def _style_sheet(ws, headers, remain_col, ms_col, name_mand_col, n_datarows, group_alt_rows=None):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.row_dimensions[1].height = 34
    for r_idx in range(2, n_datarows + 2):
        if group_alt_rows and r_idx in group_alt_rows:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = _GROUP_ALT_FILL
        if remain_col:
            rv = ws.cell(row=r_idx, column=remain_col).value
            if isinstance(rv, (int, float)) and rv > 0:
                cell = ws.cell(row=r_idx, column=remain_col)
                cell.fill = _REMAIN_FILL
                cell.font = Font(bold=True, color='9C0006')
        if ms_col and name_mand_col:
            msv = ws.cell(row=r_idx, column=ms_col).value
            nmv = ws.cell(row=r_idx, column=name_mand_col).value
            if msv is not None and nmv is not None and msv != nmv:
                for c_idx in (ms_col, name_mand_col):
                    ws.cell(row=r_idx, column=c_idx).fill = _MISMATCH_FILL
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{n_datarows + 1}'


# ══════════════════════════════════════════════════════════════════════════
# FAMILY B: ΘΕΣΕΙΣ / τοποθέτηση+συμπληρώσεις σε στήλες
# (γενικός ταξινομητής στηλών βάσει επικεφαλίδων — δουλεύει και για ΠΕ06)
# ══════════════════════════════════════════════════════════════════════════

def _classify_family_b(hdr):
    theseis = aa = name = apousies = paratiriseis = None
    blocks = []
    cur = None
    for i, h in enumerate(hdr):
        H = h.strip()
        if H == 'ΘΕΣΕΙΣ':
            theseis = i; continue
        if H.lower() == 'α/α':
            aa = i; continue
        if H == 'ΕΚΠΑΙΔΕΥΤΙΚΟΣ':
            name = i; continue
        if H == 'ΑΠΟΥΣΙΕΣ':
            apousies = i; continue
        if H == 'ΠΑΡΑΤΗΡΗΣΕΙΣ':
            paratiriseis = i; continue
        if not H:
            continue
        is_school_col = (('ΤΟΠΟΘΕΤΗΣΗΣ' in H or 'ΣΥΜΠΛΗΡΩΣΗΣ' in H)
                          and 'ΩΡΕΣ' not in H.upper() and 'ώρες' not in H
                          and 'ΕΝΔΕΙΚΤΙΚΕΣ' not in H and 'αναθέσεις' not in H)
        if is_school_col:
            kind = 'topothetisi' if 'ΤΟΠΟΘΕΤΗΣΗΣ' in H else 'symplirosi'
            cur = {'kind': kind, 'school': i, 'hours': None, 'indicative': []}
            blocks.append(cur)
            continue
        if cur is None:
            continue
        if 'ΕΝΔΕΙΚΤΙΚΕΣ' in H or 'αναθέσεις' in H:
            cur['indicative'].append(i)
        elif 'ΩΡΕΣ' in H.upper() or 'ώρες' in H:
            if cur['hours'] is None:
                cur['hours'] = i
    return dict(theseis=theseis, aa=aa, name=name, apousies=apousies,
                paratiriseis=paratiriseis, blocks=blocks)


def _build_family_b_sheet(wb, wb1, lookup_factory, sheet_name, code_prefix, out_title):
    sh = wb1.sheet_by_name(sheet_name)
    hdr = [_s(sh.cell_value(0, c)) for c in range(sh.ncols)]
    layout = _classify_family_b(hdr)
    lookup = lookup_factory(code_prefix)

    n_topo = 1
    n_symp = sum(1 for b in layout['blocks'] if b['kind'] == 'symplirosi')

    headers = ['Α/Α', 'Είδος Θέσης', 'Εκπαιδευτικός',
               'Υποχρεωτικό MySchool', 'Πηγή', 'Υποχρεωτικό (από όνομα)']
    for b in layout['blocks']:
        idx_in_symp = sum(1 for x in layout['blocks'][:layout['blocks'].index(b) + 1]
                           if x['kind'] == 'symplirosi')
        label = 'Τοποθέτησης' if b['kind'] == 'topothetisi' else f'Συμπλ.{idx_in_symp}'
        headers += [f'Σχολείο {label}', f'Ώρες {label}', f'Ενδεικτικές {label}']
    headers += ['Σύνολο Ωρών', 'Υπόλοιπο για Τοποθέτηση', 'Απουσίες', 'Παρατηρήσεις']

    out_rows = []
    unresolved, mismatches = [], []

    for r in range(1, sh.nrows):
        theseis_v = _s(sh.cell_value(r, layout['theseis'])) if layout['theseis'] is not None else ''
        raw_name = _s(sh.cell_value(r, layout['name'])) if layout['name'] is not None else ''
        if not raw_name and not theseis_v:
            continue
        aa_v = sh.cell_value(r, layout['aa']) if layout['aa'] is not None else ''

        block_vals = []
        all_hours = []
        for b in layout['blocks']:
            school = _s(sh.cell_value(r, b['school']))
            hrs = _num(sh.cell_value(r, b['hours'])) if b['hours'] is not None else None
            indic = sum(v for v in (_num(sh.cell_value(r, ic)) for ic in b['indicative']) if v is not None) or None
            block_vals += [school, hrs, indic]
            if hrs is not None:
                all_hours.append(hrs)
            if indic is not None:
                all_hours.append(indic)

        total = sum(all_hours) if all_hours else None
        apousies_v = _s(sh.cell_value(r, layout['apousies'])) if layout['apousies'] is not None else ''
        paratiriseis_v = _s(sh.cell_value(r, layout['paratiriseis'])) if layout['paratiriseis'] is not None else ''

        mandatory_name = _name_mandatory(raw_name)
        mandatory_ms, ms_src = None, ''
        if raw_name and not raw_name.startswith('#'):
            surname, fhint = _parse_name(raw_name)
            val, src, status = lookup(surname, fhint)
            if val:
                try:
                    mandatory_ms = int(val); ms_src = src
                except ValueError:
                    pass
            if mandatory_ms is None:
                unresolved.append((raw_name, status))
            elif mandatory_name is not None and mandatory_ms != mandatory_name:
                mismatches.append((raw_name, mandatory_name, mandatory_ms, ms_src))

        mandatory_final = mandatory_ms if mandatory_ms is not None else mandatory_name
        remaining = None
        if mandatory_final is not None:
            remaining = max(0, mandatory_final - total) if total is not None else mandatory_final

        out_rows.append(
            [int(aa_v) if isinstance(aa_v, (int, float)) and aa_v else aa_v,
             theseis_v, raw_name, mandatory_ms, ms_src, mandatory_name]
            + block_vals
            + [total, remaining, apousies_v, paratiriseis_v]
        )

    ws = wb.create_sheet(sheet_name.replace(' ', '_')[:31])
    ws.append(headers)
    for row in out_rows:
        ws.append(row)

    ms_col = headers.index('Υποχρεωτικό MySchool') + 1
    name_mand_col = headers.index('Υποχρεωτικό (από όνομα)') + 1
    remain_col = headers.index('Υπόλοιπο για Τοποθέτηση') + 1
    _style_sheet(ws, headers, remain_col, ms_col, name_mand_col, len(out_rows))

    widths = [6, 20, 22, 11, 7, 11] + [26, 8, 9] * len(layout['blocks']) + [10, 11, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    real_teachers = [rrow for rrow in out_rows if rrow[2] and not rrow[2].startswith('#')]
    covered = sum(1 for rrow in real_teachers if rrow[3] is not None)
    notes = [
        (f'{out_title} — Οργάνωση (φύλλο "{sheet_name}")', True, 12),
        ('', False, 10),
        (f'Δομή: {n_topo} σχολείο τοποθέτησης + {n_symp} σχολεία συμπλήρωσης ανά εκπαιδευτικό '
         '(ανιχνεύτηκε από τις επικεφαλίδες του πρωτότυπου αρχείου).', False, 10),
        ('Οι "Ενδεικτικές" ώρες κάθε σχολείου προστίθενται κανονικά στο "Σύνολο Ωρών".', False, 10),
        (f'Υποχρεωτικό MySchool: από 4.1 (πρώτα) → 4.2 (αν δεν βρεθεί), κωδικός ειδικότητας "{code_prefix}".', False, 10),
        (f'Κάλυψη: {covered}/{len(real_teachers)} πραγματικών εκπαιδευτικών.', False, 10),
        ('', False, 10),
    ]
    if mismatches:
        notes.append((f'Αναντιστοιχίες υποχρεωτικού (MySchool vs όνομα) — {len(mismatches)}:', True, 11))
        for name_, mn, ms, src in mismatches:
            notes.append((f'   - {name_}: όνομα={mn}, MySchool={ms} (πηγή {src})', False, 9))
        notes.append(('', False, 10))
    if unresolved:
        notes.append((f'Δεν βρέθηκαν στα 4.1/4.2 — {len(unresolved)}:', True, 11))
        for name_, status in unresolved:
            notes.append((f'   - {name_}  ({status})', False, 9))
    return notes


# ══════════════════════════════════════════════════════════════════════════
# FAMILY A: ΟΜΑΔΟΠΟΙΗΣΗ (ΓΑΛΛ ΠΕ05, ΓΕΡΜ ΠΕ07) — group ανά γραμμές
# ══════════════════════════════════════════════════════════════════════════

def _classify_family_a(hdr):
    aa = group_lbl = school = tm_e = tm_st = hours = total_hrs = other = name = apousies = paratiriseis = None
    for i, h in enumerate(hdr):
        H = h.strip()
        if H.lower() == 'α/α':
            aa = i
        elif 'ΟΜΑΔΟΠΟΙΗΣΗ' in H:
            group_lbl = i
        elif 'ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ' in H:
            school = i
        elif 'Ε΄' in H or 'E΄' in H:
            tm_e = i
        elif 'Στ΄' in H:
            tm_st = i
        elif 'ΣΥΝΟΛΟ' in H:
            total_hrs = i
        elif 'ΩΡΕΣ' in H.upper() and 'ΠΡΩΙΝ' in H.upper():
            hours = i
        elif 'άλλες αναθέσεις' in H:
            other = i
        elif H == 'ΕΚΠΑΙΔΕΥΤΙΚΟΣ':
            name = i
        elif H == 'ΑΠΟΥΣΙΕΣ':
            apousies = i
        elif H == 'ΠΑΡΑΤΗΡΗΣΕΙΣ':
            paratiriseis = i
    return dict(aa=aa, group_lbl=group_lbl, school=school, tm_e=tm_e, tm_st=tm_st,
                hours=hours, total_hrs=total_hrs, other=other, name=name,
                apousies=apousies, paratiriseis=paratiriseis)


def _build_family_a_sheet(wb, wb1, lookup_factory, sheet_name, code_prefix, out_title, max_symp=4):
    sh = wb1.sheet_by_name(sheet_name)
    hdr = [_s(sh.cell_value(0, c)) for c in range(sh.ncols)]
    L = _classify_family_a(hdr)
    lookup = lookup_factory(code_prefix)

    groups = []
    cur = []
    for r in range(1, sh.nrows):
        rowvals = [sh.cell_value(r, c) for c in range(sh.ncols)]
        if all(v == '' for v in rowvals):
            if cur:
                groups.append(cur); cur = []
            continue
        cur.append(rowvals)
    if cur:
        groups.append(cur)

    headers = ['Α/Α', 'Εκπαιδευτικός', 'Υποχρεωτικό MySchool', 'Πηγή', 'Υποχρεωτικό (από όνομα)',
               'Σχολείο Τοποθέτησης', 'Ώρες Τοπ.', 'Άλλες αναθ. Τοπ.']
    for k in range(1, max_symp + 1):
        headers += [f'Σχολείο Συμπλ.{k}', f'Ώρες Συμπλ.{k}']
    headers += ['Σύνολο Ωρών (MySchool αρχείο)', 'Σύνολο Ωρών (υπολογισμένο)',
                'Υπόλοιπο για Τοποθέτηση', 'Απουσίες', 'Παρατηρήσεις']

    out_rows = []
    unresolved, mismatches, total_mismatches = [], [], []
    group_alt_rows = set()

    for gi, g in enumerate(groups):
        main = g[0]
        symp_rows = g[1:]
        aa_v = main[L['aa']] if L['aa'] is not None else ''
        raw_name = _s(main[L['name']]) if L['name'] is not None else ''
        main_school = _s(main[L['school']]) if L['school'] is not None else ''
        main_hours = _num(main[L['hours']]) if L['hours'] is not None else None
        main_other = _num(main[L['other']]) if L['other'] is not None else None
        total_file = _num(main[L['total_hrs']]) if L['total_hrs'] is not None else None
        apousies_v = _s(main[L['apousies']]) if L['apousies'] is not None else ''
        paratiriseis_v = _s(main[L['paratiriseis']]) if L['paratiriseis'] is not None else ''

        symp_cols = []
        computed_total = (main_hours or 0) + (main_other or 0)
        for k in range(max_symp):
            if k < len(symp_rows):
                srow = symp_rows[k]
                ssch = _s(srow[L['school']]) if L['school'] is not None else ''
                shrs = _num(srow[L['hours']]) if L['hours'] is not None else None
                symp_cols += [ssch, shrs]
                if shrs is not None:
                    computed_total += shrs
            else:
                symp_cols += ['', None]

        mandatory_name = _name_mandatory(raw_name)
        mandatory_ms, ms_src = None, ''
        if raw_name and not raw_name.startswith('#'):
            surname, fhint = _parse_name(raw_name)
            val, src, status = lookup(surname, fhint)
            if val:
                try:
                    mandatory_ms = int(val); ms_src = src
                except ValueError:
                    pass
            if mandatory_ms is None:
                unresolved.append((raw_name, status))
            elif mandatory_name is not None and mandatory_ms != mandatory_name:
                mismatches.append((raw_name, mandatory_name, mandatory_ms, ms_src))

        if total_file is not None and computed_total != total_file:
            total_mismatches.append((raw_name, total_file, computed_total))

        mandatory_final = mandatory_ms if mandatory_ms is not None else mandatory_name
        total_best = total_file if total_file is not None else computed_total
        remaining = None
        if mandatory_final is not None:
            remaining = max(0, mandatory_final - total_best) if total_best is not None else mandatory_final

        row_out = ([int(aa_v) if isinstance(aa_v, (int, float)) and aa_v else aa_v,
                    raw_name, mandatory_ms, ms_src, mandatory_name,
                    main_school, main_hours, main_other]
                   + symp_cols
                   + [total_file, computed_total, remaining, apousies_v, paratiriseis_v])
        out_rows.append(row_out)
        if gi % 2 == 1:
            group_alt_rows.add(len(out_rows) + 1)

    ws = wb.create_sheet(sheet_name.replace(' ', '_')[:31])
    ws.append(headers)
    for row in out_rows:
        ws.append(row)

    ms_col = headers.index('Υποχρεωτικό MySchool') + 1
    name_mand_col = headers.index('Υποχρεωτικό (από όνομα)') + 1
    remain_col = headers.index('Υπόλοιπο για Τοποθέτηση') + 1
    _style_sheet(ws, headers, remain_col, ms_col, name_mand_col, len(out_rows), group_alt_rows)

    widths = [6, 22, 11, 7, 11, 26, 8, 10] + [26, 8] * max_symp + [12, 12, 12, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    real_teachers = [rrow for rrow in out_rows if rrow[1] and not rrow[1].startswith('#')]
    covered = sum(1 for rrow in real_teachers if rrow[2] is not None)
    notes = [
        (f'{out_title} — Οργάνωση (φύλλο "{sheet_name}")', True, 12),
        ('', False, 10),
        ('Δομή διαφορετική από ΠΕ06/08/κλπ: το πρωτότυπο αρχείο έχει ΜΙΑ ΓΡΑΜΜΗ ΑΝΑ ΣΧΟΛΕΙΟ (κύριο + '
         'συμπληρώσεις) αντί για στήλες — εδώ συμπτύχθηκαν σε μία γραμμή ανά εκπαιδευτικό.', False, 10),
        ('"Σύνολο Ωρών (MySchool αρχείο)" = η τιμή που είχε ήδη το πρωτότυπο αρχείο. '
         '"Σύνολο Ωρών (υπολογισμένο)" = άθροισμα όλων των ωρών/αναθέσεων — κρατήθηκαν και τα δύο '
         'για διασταύρωση.', False, 10),
        (f'Υποχρεωτικό MySchool: από 4.1 (πρώτα) → 4.2 (αν δεν βρεθεί), κωδικός ειδικότητας "{code_prefix}".', False, 10),
        (f'Κάλυψη: {covered}/{len(real_teachers)} πραγματικών εκπαιδευτικών.', False, 10),
        ('', False, 10),
    ]
    if total_mismatches:
        notes.append((f'Αναντιστοιχίες Συνόλου Ωρών (αρχείο vs υπολογισμένο) — {len(total_mismatches)}:', True, 11))
        for name_, tf, ct in total_mismatches:
            notes.append((f'   - {name_}: αρχείο={tf}, υπολογισμένο={ct}', False, 9))
        notes.append(('', False, 10))
    if mismatches:
        notes.append((f'Αναντιστοιχίες υποχρεωτικού (MySchool vs όνομα) — {len(mismatches)}:', True, 11))
        for name_, mn, ms, src in mismatches:
            notes.append((f'   - {name_}: όνομα={mn}, MySchool={ms} (πηγή {src})', False, 9))
        notes.append(('', False, 10))
    if unresolved:
        notes.append((f'Δεν βρέθηκαν στα 4.1/4.2 — {len(unresolved)}:', True, 11))
        for name_, status in unresolved:
            notes.append((f'   - {name_}  ({status})', False, 9))
    return notes


# ══════════════════════════════════════════════════════════════════════════
# Διασταύρωση με 2ο αρχείο (προαιρετικό)
# ══════════════════════════════════════════════════════════════════════════

def _cross_check_second_file(wb1, wb2_path):
    """Απλός έλεγχος: για κάθε κοινό φύλλο ανάμεσα στο κύριο και το 2ο αρχείο,
    συγκρίνει τιμές κελιών (raw) και μετράει διαφορές. Δεν χρησιμοποιείται
    για τον υπολογισμό — μόνο ενημερωτικά, σαν έλεγχος συνέπειας."""
    try:
        wb2 = xlrd.open_workbook(wb2_path)
    except Exception as e:
        return [f'⚠  Δεν ήταν δυνατό το άνοιγμα του 2ου αρχείου: {e}']

    lines = []
    common = [n for n in wb2.sheet_names() if n in wb1.sheet_names()]
    if not common:
        return ['ℹ  Το 2ο αρχείο δεν έχει κανένα κοινό φύλλο με το κύριο — δεν έγινε διασταύρωση.']
    for name in common:
        sh1 = wb1.sheet_by_name(name)
        sh2 = wb2.sheet_by_name(name)
        rows = min(sh1.nrows, sh2.nrows)
        cols = min(sh1.ncols, sh2.ncols)
        diffs = 0
        for r in range(rows):
            for c in range(cols):
                if sh1.cell_value(r, c) != sh2.cell_value(r, c):
                    diffs += 1
        extra_rows = abs(sh1.nrows - sh2.nrows)
        if diffs == 0 and extra_rows == 0:
            lines.append(f'✓  «{name}»: πανομοιότυπο με το κύριο αρχείο.')
        else:
            lines.append(f'⚠  «{name}»: {diffs} διαφορετικά κελιά, {extra_rows} επιπλέον/λιγότερες γραμμές '
                          '(ελέγχθηκε μόνο ως πληροφορία — χρησιμοποιήθηκε το κύριο αρχείο).')
    return lines


# ══════════════════════════════════════════════════════════════════════════
# MAIN entry point
# ══════════════════════════════════════════════════════════════════════════

def build_workbook(xls1_path, csv41_path, csv42_path, out_path, xls2_path=None,
                    progress_cb=None):
    """Χτίζει το βιβλίο εργασίας με όλες τις ειδικότητες και το αποθηκεύει
    στο `out_path`. Επιστρέφει dict με summary στατιστικά ανά ειδικότητα
    (για εμφάνιση στο UI μετά την εκτέλεση).

    progress_cb(text): προαιρετικό callback για ενημέρωση προόδου στο UI.
    """
    def _prog(msg):
        if progress_cb:
            progress_cb(msg)

    wb1 = xlrd.open_workbook(xls1_path)
    rows41_all = _load_csv_rows(csv41_path)
    rows42_all = _load_csv_rows(csv42_path)
    lookup_factory = lambda code_prefix: _make_lookup(rows41_all, rows42_all, code_prefix)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_notes = []
    summary = []  # [(title, ok, msg)]

    # ΠΕ06 πρώτο (family B), μετά ΠΕ05/ΠΕ07 (family A), μετά τα υπόλοιπα family B
    plan = [('B', *FAMILY_B_SPECIALTIES[0])] + \
           [('A', *sp) for sp in FAMILY_A_SPECIALTIES] + \
           [('B', *sp) for sp in FAMILY_B_SPECIALTIES[1:]]

    for family, sheet_name, code_prefix, title in plan:
        _prog(f'Επεξεργασία {title}…')
        if sheet_name not in wb1.sheet_names():
            summary.append((title, False, f'Το φύλλο «{sheet_name}» δεν βρέθηκε στο αρχείο — παραλείφθηκε.'))
            continue
        try:
            if family == 'A':
                notes = _build_family_a_sheet(wb, wb1, lookup_factory, sheet_name, code_prefix, title)
            else:
                notes = _build_family_b_sheet(wb, wb1, lookup_factory, sheet_name, code_prefix, title)
            all_notes.append((title, notes))
            summary.append((title, True, ''))
        except Exception as e:
            summary.append((title, False, f'Σφάλμα επεξεργασίας: {e}'))

    # Συγκεντρωτικό φύλλο Οδηγιών
    ws_idx = wb.create_sheet('Οδηγίες & Ευρήματα', 0)
    ws_idx.column_dimensions['A'].width = 110
    r = 1
    c = ws_idx.cell(row=r, column=1,
                     value='Οργάνωση ειδικοτήτων — Οδηγίες & Ευρήματα '
                           '(ΠΕ06, ΠΕ05, ΠΕ07, ΠΕ08, ΠΕ11, ΠΕ79, ΠΕ86, ΠΕ91)')
    c.font = Font(bold=True, size=14, color='1F4E79')
    r += 2

    if xls2_path:
        _prog('Διασταύρωση με 2ο αρχείο…')
        cross_lines = _cross_check_second_file(wb1, xls2_path)
        cc = ws_idx.cell(row=r, column=1, value='Διασταύρωση με 2ο αρχείο ομαδοποιήσεων:')
        cc.font = Font(bold=True, size=11, color='1F4E79')
        r += 1
        for line in cross_lines:
            ws_idx.cell(row=r, column=1, value=line).alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
        r += 1

    for title, notes in all_notes:
        for text, bold, size in notes:
            cell = ws_idx.cell(row=r, column=1, value=text)
            cell.font = Font(bold=bold, size=size, color='1F4E79' if bold else '000000')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
        r += 1

    _prog('Αποθήκευση αρχείου…')
    wb.save(out_path)
    return summary
