"""
smeae/compare.py
════════════════
Σύγκριση στατιστικών ΣΜΕΑΕ, διαχωρισμός ανά σχολείο, αποστολή email.
Προσαρμογή του xlsx.py από το smeae_stats_10.
"""
import os, re, ssl, logging, smtplib
from datetime import datetime
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.utils import formataddr, formatdate
from email import encoders

import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# Αντιστοιχίες ονομάτων στηλών (ίδιο με αρχικό xlsx.py)
COLUMN_NAMES = {
    "Συγκεντρωτικά Στοιχεία μαθητών-μαθητριών με ΕΕΑ":
        "Συγκεντρωτικά Στοιχεία μαθητών/μαθητριών με ΕΕΑ",
    "Συγκεντρωτικά Στοιχεία μαθητών-μαθητριών με ΕΕΑ που υποστηρίζονται από τον-την εκπαιδευτικό της τάξης":
        "Συγκεντρωτικά Στοιχεία μαθητών/μαθητριών με ΕΕΑ που υποστηρίζονται από τον/την εκπαιδευτικό της τάξης",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών σε Τμήματα Ένταξης με κοινό και εξειδικευμένο πρόγραμμα":
        "Σε Τμήματα Ένταξης με κοινό και εξειδικευμένο πρόγραμμα",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών σε Τμήματα Ένταξης διευρυμένου ωραρίου":
        "Σε Τμήματα Ένταξης διευρυμένου ωραρίου",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών σε Παράλληλη Στήριξη":
        "Με Παράλληλη Στήριξη",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών με Ειδικό Βοηθητικό Προσωπικό":
        "Με Ειδικό Βοηθητικό Προσωπικό",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών με Σχολικό Νοσηλευτή":
        "Με Σχολικό Νοσηλευτή",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών με ειδικό βοηθό (που διαθέτει η οικογένεια)":
        "Με ειδικό βοηθό (που διαθέτει η οικογένεια)",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών Μαθητών που υποστηρίζονται Κατ' οίκον":
        "Κατ' οίκον",
}

# Σύντομα ονόματα φύλλων Excel (max 31 chars συμπεριλαμβανομένου "N. ")
SHEET_NAMES = {
    "Συγκεντρωτικά Στοιχεία μαθητών-μαθητριών με ΕΕΑ που υποστηρίζονται από τον-την εκπαιδευτικό της τάξης":
        "Εκπαιδευτικός τάξης",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών σε Τμήματα Ένταξης με κοινό και εξειδικευμένο πρόγραμμα":
        "ΤΕ κοινό & εξειδικευμένο",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών σε Τμήματα Ένταξης διευρυμένου ωραρίου":
        "ΤΕ διευρυμένου ωραρίου",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών σε Παράλληλη Στήριξη":
        "Παράλληλη Στήριξη",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών με Ειδικό Βοηθητικό Προσωπικό":
        "Ειδικό Βοηθητικό Προσωπικό",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών με Σχολικό Νοσηλευτή":
        "Σχολικός Νοσηλευτής",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών με ειδικό βοηθό (που διαθέτει η οικογένεια)":
        "Ειδικός βοηθός οικογένειας",
    "Συγκεντρωτικά Στοιχεία Ειδικών Εκπαιδευτικών Αναγκών Μαθητών που υποστηρίζονται Κατ' οίκον":
        "Κατ' οίκον",
}

_IGNORE = '-------- Αγνόησε τη στήλη --------'


def sanitize_filename(filename):
    return re.sub(r'[^\w\s]', '_', filename)


def open_xlsx(file_path) -> DataFrame:
    return pd.read_excel(file_path, header=[0, 1])


def match_columns(dfm: DataFrame, dfs: DataFrame, column_matches: dict = None):
    """
    Αντιστοιχεί στήλες master↔slave.
    column_matches: dict από JSON (match_indices_l1) ή None για διαδραστική λειτουργία.
    Επιστρέφει (col_match_l1, col_match_l2).
    """
    mc_l1 = dfm.columns.get_level_values(0)
    sc_l1 = dfs.columns.get_level_values(0)

    slave_only  = sc_l1.difference(mc_l1, sort=False)
    master_only = mc_l1.difference(sc_l1, sort=False)
    master_only = master_only.insert(0, _IGNORE)

    col_match_l1 = {}
    for i, sc in enumerate(slave_only):
        if column_matches is None:
            idx = int(input(
                f'Slave col {i}: "{sc}"\n'
                f'Αντίστοιχη master: {list(master_only)}\nIndex: '
            ))
        else:
            idx = column_matches.get('match_indices_l1', {}).get(f'slave_col{i + 1}', 0)
            print(f'Auto: slave_col{i + 1} "{sc}" → master[{idx}] "{master_only[idx]}"')
        col_match_l1[sc] = master_only[idx]

    l2_master = pd.Index([
        item
        for sc in col_match_l1 if col_match_l1[sc] != _IGNORE
        for item in dfm[col_match_l1[sc]].columns.get_level_values(0).unique()
    ])
    l2_slave = pd.Index([
        item
        for sc in col_match_l1 if col_match_l1[sc] != _IGNORE
        for item in dfs[sc].columns.get_level_values(0).unique()
    ])
    diff_l2 = l2_slave.difference(l2_master, sort=False)
    col_match_l2 = l2_slave if diff_l2.size == 0 else pd.Index([])

    return col_match_l1, col_match_l2


def compare_xlsx(master_path, slave_path, dfm, dfs, col_l1, col_l2):
    """Συγκρίνει master και slave DataFrame. Επιστρέφει dict αποκλίσεων."""
    print(f'Εγγραφές: {len(dfm)} master, {len(dfs)} slave.')
    common = dfm.columns.intersection(dfs.columns)
    dfm = dfm.sort_values(by=list(common)).reset_index(drop=True)
    dfs = dfs.sort_values(by=list(common)).reset_index(drop=True)

    master_name = os.path.splitext(os.path.basename(master_path))[0].split('. ', 2)[-1]
    slave_name  = os.path.splitext(os.path.basename(slave_path))[0].split('. ', 2)[-1]
    return _compare_rows(master_name, slave_name, dfm, dfs, common, col_l1, col_l2)


def _first(s):
    """Ασφαλής πρόσβαση στο πρώτο στοιχείο Series/MultiIndex — αποφεύγει KeyError(0)."""
    try:
        return s.iloc[0]
    except Exception:
        return s


def _compare_rows(master_name, slave_name, dfm, dfs, common, col_l1, col_l2):
    s2m = {sc: mc for sc, mc in col_l1.items() if mc != _IGNORE}
    differences = {}
    count = 0

    for si, sr in dfs.iterrows():
        master_idx = dfm.index[dfm[common].eq(sr[common]).all(axis=1)]
        if not master_idx.empty:
            mr = dfm.loc[master_idx[0]]
            for sc, mc in s2m.items():
                for c2 in col_l2:
                    if c2 == 'Σ':
                        continue
                    if c2 in sr[sc] and c2 in mr[mc]:
                        if sr[sc][c2] != mr[mc][c2]:
                            differences[count] = {
                                'school'      : _first(sr['Ονομασία Μονάδας']),
                                'schCode'     : _first(sr['Κωδικός Υπουργείου']),
                                'class'       : _first(sr['Τάξη']),
                                'col_l1'      : sc,
                                'col_l2'      : c2,
                                'slave_value' : sr[sc][c2],
                                'master_value': mr[mc][c2],
                            }
                            count += 1
                    else:
                        print(f'Στήλη {c2} δεν βρέθηκε στη γραμμή {si}')
        else:
            # Γραμμή απούσα από master — καταγραφή μόνο αν μη-μηδενική
            non_zero = any(
                sr[sc][c2] != 0
                for sc in s2m
                for c2 in col_l2
                if c2 != 'Σ' and c2 in sr.get(sc, {})
            )
            if not non_zero:
                continue
            differences[count] = {
                'school'      : _first(sr['Ονομασία Μονάδας']),
                'schCode'     : _first(sr['Κωδικός Υπουργείου']),
                'class'       : _first(sr['Τάξη']),
                'col_l1'      : '',
                'col_l2'      : '',
                'slave_value' : f"Υπάρχει στο '{COLUMN_NAMES.get(slave_name, slave_name)}'",
                'master_value': f"Δεν υπάρχει στο '{COLUMN_NAMES.get(master_name, master_name)}'",
            }
            count += 1

    for mi, mr in dfm.iterrows():
        slave_idx = dfs.index[dfs[common].eq(mr[common]).all(axis=1)]
        if slave_idx.empty:
            for sc, mc in s2m.items():
                if any(mr[mc][c2] != 0 for c2 in col_l2 if c2 != 'Σ' and c2 in mr[mc]):
                    differences[count] = {
                        'school'      : _first(mr['Ονομασία Μονάδας']),
                        'schCode'     : _first(mr['Κωδικός Υπουργείου']),
                        'class'       : _first(mr['Τάξη']),
                        'col_l1'      : '',
                        'col_l2'      : '',
                        'slave_value' : f"Δεν υπάρχει στο '{slave_name}'",
                        'master_value': f"Υπάρχει στο '{master_name}'",
                    }
                    count += 1
                    break

    print(f'Βρέθηκαν {count} αποκλίσεις.')
    return differences


def write_to_excel(differences, sheet_name, output_dir, school_year):
    """Γράφει τις αποκλίσεις σε Excel. Επιστρέφει path αρχείου."""
    rows = [
        {
            'Α/Α'                      : key,
            'Σχολείο'                  : item['school'],
            'Τάξη'                     : item['class'],
            'Κωδικός Υπουργείου'       : item['schCode'],
            'Όνομα γενικής στήλης'     : item['col_l1'],
            'Όνομα στήλης μαθητών'     : item['col_l2'],
            'Επιμέρους στατιστικό'     : item['slave_value'],
            'Συγκεντρωτικό στατιστικό' : item['master_value'],
        }
        for key, item in differences.items()
    ]
    df_out = pd.DataFrame(rows)

    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime('%Y%m%d')
    filename = os.path.join(output_dir, f'differences_{school_year}_{date_str}.xlsx')

    _bn       = os.path.splitext(os.path.basename(sheet_name))[0]
    _num      = _bn.split('.')[0].strip()          # "2", "3", ... "9"
    raw_sname = _bn.split('. ', 1)[-1] if '. ' in _bn else _bn
    _short    = SHEET_NAMES.get(raw_sname,
                    sanitize_filename(COLUMN_NAMES.get(raw_sname, raw_sname)))
    # Prefix με αριθμό αρχείου: "2. Εκπαιδευτικός Τάξης"
    sname     = (f'{_num}. {_short}' if _num.isdigit() else _short)[:31]

    mode = 'a' if os.path.exists(filename) else 'w'
    kw   = {'if_sheet_exists': 'replace'} if mode == 'a' else {}
    with pd.ExcelWriter(filename, mode=mode, engine='openpyxl', **kw) as writer:
        df_out.to_excel(writer, sheet_name=sname, index=False)

    wb = load_workbook(filename)
    ws = wb[sname]
    for col in ws.columns:
        max_len   = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[col_letter].width = max_len + 5
    wb.save(filename)

    print(f'Αποθηκεύτηκε: {filename}')
    return filename


def split_xlsx(file_path, output_dir, school_year):
    """Διαχωρισμός αρχείου αποτελεσμάτων σε ξεχωριστά αρχεία ανά σχολείο."""
    wb = load_workbook(file_path)
    os.makedirs(output_dir, exist_ok=True)

    school_data = {}
    col_dims    = {}
    col_aligns  = {}

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        data = ws.values
        try:
            columns = next(data)[0:]
        except StopIteration:
            print(f"Φύλλο '{sheet_name}' χωρίς δεδομένα — παράλειψη.")
            continue
        df      = pd.DataFrame(data, columns=columns)
        grouped = df.groupby(['Κωδικός Υπουργείου', 'Σχολείο'])

        for (mcode, sname), group in grouped:
            school_data.setdefault(mcode, {})[sheet_name] = group

        col_dims[sheet_name]   = {
            get_column_letter(i): ws.column_dimensions[get_column_letter(i)].width
            for i in range(1, ws.max_column + 1)
        }
        col_aligns[sheet_name] = {
            get_column_letter(i): ws.cell(1, i).alignment
            for i in range(1, ws.max_column + 1)
        }

    for mcode, sheets in school_data.items():
        sname    = sheets[list(sheets.keys())[0]]['Σχολείο'].iloc[0]
        fname    = f'{mcode}_{sanitize_filename(sname)}_{school_year}.xlsx'
        fpath    = os.path.join(output_dir, fname)

        with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
            for sh, grp in sheets.items():
                grp = grp.copy()
                if 'Α/Α' in grp.columns:
                    grp.drop(columns=['Α/Α'], inplace=True)
                grp.reset_index(drop=True, inplace=True)
                grp.insert(0, 'Α/Α', range(1, 1 + len(grp)))
                grp.to_excel(writer, index=False, sheet_name=sh)

            new_wb = writer.book
            for sh in sheets:
                new_ws = new_wb[sh]
                for col_l, w in col_dims[sh].items():
                    new_ws.column_dimensions[col_l].width = w
                for col_l, aln in col_aligns[sh].items():
                    for cell in new_ws[col_l]:
                        cell.alignment = Alignment(
                            horizontal=aln.horizontal, vertical=aln.vertical)

    print(f'Διαχωρισμός ολοκληρώθηκε → {output_dir}')


# ── Προεπιλεγμένα πρότυπα email ─────────────────────────────────────────────

DEFAULT_SMEAE_SUBJECT = (
    'Αποτελέσματα ελέγχου για το σχολικό έτος {school_year}, '
    'των Στατιστικών Στοιχείων Ειδικών Εκπαιδευτικών Αναγκών '
    'στο Πληροφοριακό Σύστημα Myschool'
)

DEFAULT_SMEAE_BODY = """\
Προς τη Διεύθυνση του Σχολείου,

Σας αποστέλλουμε με συνημμένο αρχείο τα αποτελέσματα του ελέγχου των \
στατιστικών στοιχείων των Ειδικών Εκπαιδευτικών Αναγκών (ΕΕΑ) στο \
Πληροφοριακό Σύστημα Myschool για το σχολικό έτος {school_year}.

Ο έλεγχος αφορά τη σύγκριση μεταξύ του συγκεντρωτικού στατιστικού και \
των επιμέρους στατιστικών στοιχείων των μαθητών/μαθητριών με ΕΕΑ. Συγκεκριμένα:
- Εκπαιδευτικός τάξης
- Τμήματα Ένταξης (κοινό/εξειδικευμένο & διευρυμένο)
- Παράλληλη Στήριξη
- Ειδικό Βοηθητικό Προσωπικό
- Σχολικός Νοσηλευτής
- Ειδικός βοηθός οικογένειας
- Κατ' οίκον

Παρακαλούμε βρείτε στο συνημμένο αρχείο τα αποτελέσματα και προβείτε \
στις απαραίτητες διορθώσεις στο Myschool. Ιδιαίτερη προσοχή στις στήλες \
«Όνομα γενικής στήλης» και «Όνομα στήλης μαθητών».

Σημειώνεται ότι όταν ένας μαθητής/μαθήτρια έχει περισσότερες από μία ΕΕΑ, \
μια μικρή απόκλιση θεωρείται αναμενόμενη."""


def _plain_to_html(text):
    """Μετατρέπει απλό κείμενο σε HTML για αποστολή email."""
    import html as _html
    paragraphs = text.split('\n\n')
    html_parts = []
    for para in paragraphs:
        lines = para.split('\n')
        escaped_lines = []
        for line in lines:
            esc = _html.escape(line)
            if esc.startswith('- '):
                esc = '&bull; ' + esc[2:]
            escaped_lines.append(esc)
        html_parts.append('<br>'.join(escaped_lines))
    return '<html><body><p>' + '</p><p>'.join(html_parts) + '</p></body></html>'


def send_emails(output_dir, school_year, email_from, username, password,
                smtp_host, school_dir_path, dry_run=False,
                send_only_one=False, callback=None,
                custom_subject=None, custom_body_text=None):
    """Αποστολή email με συνημμένο αρχείο σε κάθε σχολείο."""
    log = callback or print

    school_df  = pd.read_excel(school_dir_path)
    school_map = {}
    for _, row in school_df.iterrows():
        school_map[row['Κωδικός Υπουργείου']] = (row['Ονομασία'], row['e-mail'])

    first = True
    sent_schools = []   # λίστα (ονομασία, email) που εστάλησαν επιτυχώς
    failed_schools = [] # λίστα (ονομασία, email) που απέτυχαν

    for fname in sorted(os.listdir(output_dir)):
        if not fname.endswith('.xlsx'):
            continue
        if not first and send_only_one:
            break

        mcode = fname.split('_')[0]
        try:
            info = school_map.get(int(mcode))
        except (ValueError, TypeError):
            continue
        if not info:
            log(f'  Σχολείο {mcode} δεν βρέθηκε στον κατάλογο.')
            continue

        school_name, email_to = info

        # Ανίχνευση επιτυχίας μέσω callback wrapper
        _success = [True]
        _orig_log = log
        def _tracked_log(msg, _name=school_name, _email=email_to):
            _orig_log(msg)
            if '✗' in msg:
                _success[0] = False

        send_email_with_attachment(
            receiver_email   = email_to,
            attachment_path  = os.path.join(output_dir, fname),
            dry_run          = dry_run,
            sender_email     = email_from,
            username         = username,
            password         = password,
            smtp_host        = smtp_host,
            first_email      = first,
            school_name      = school_name,
            school_year      = school_year,
            callback         = _tracked_log,
            custom_subject   = custom_subject,
            custom_body_text = custom_body_text,
        )

        if _success[0]:
            sent_schools.append((school_name, email_to))
        else:
            failed_schools.append((school_name, email_to))

        first = False

    # ── Επιβεβαιωτικό summary email στο FROM_EMAIL ───────────────────────────
    if not dry_run and sent_schools:
        _send_smeae_summary(
            email_from   = email_from,
            username     = username,
            password     = password,
            smtp_host    = smtp_host,
            school_year  = school_year,
            sent         = sent_schools,
            failed       = failed_schools,
            callback     = log,
        )


def _send_smeae_summary(email_from, username, password, smtp_host,
                        school_year, sent, failed, callback=None):
    """Αποστολή επιβεβαιωτικού summary email μετά την ολοκλήρωση της αποστολής."""
    log = callback or print
    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')

    sent_lines   = '\n'.join(f'  ✓  {name}  ({email})' for name, email in sorted(sent))
    failed_lines = ('\n'.join(f'  ✗  {name}  ({email})' for name, email in sorted(failed))
                    if failed else '  —')

    body_plain = (
        f'Αποστολή email ΕΕΑ — {now_str}\n'
        f'Σχολικό έτος: {school_year}\n\n'
        f'Εστάλησαν επιτυχώς: {len(sent)} σχολεία\n'
        f'{sent_lines}\n\n'
        + (f'Αποτυχίες ({len(failed)}):\n{failed_lines}\n' if failed else '')
    )
    html_body = _plain_to_html(body_plain)

    subject = (
        f'[ΣΜΕΑΕ] Αποστολή {school_year} ολοκληρώθηκε — '
        f'{len(sent)} σχολεία ✓'
        + (f', {len(failed)} αποτυχίες ✗' if failed else '')
    )

    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From']    = formataddr((
        str(Header('Πρόγραμμα Ελέγχου ΕΕΑ', 'utf-8')), email_from))
    msg['To']      = email_from
    msg['Date']    = formatdate(localtime=True)
    msg.attach(MIMEText(html_body, 'html'))

    log(f'\n  ✉ Αποστολή επιβεβαιωτικού summary → {email_from}')
    try:
        context = ssl.create_default_context()
        server  = smtplib.SMTP(smtp_host, 587)
        server.starttls(context=context)
        result = server.login(username, password)
        if result[0] == 235:
            server.sendmail(email_from, email_from, msg.as_string())
            log(f'  ✓ Επιβεβαιωτικό εστάλη.')
        else:
            log(f'  ✗ Επιβεβαιωτικό: login απέτυχε.')
        server.quit()
    except Exception as e:
        log(f'  ✗ Επιβεβαιωτικό: {e}')


def send_email_with_attachment(receiver_email, attachment_path, dry_run,
                                sender_email, username, password, smtp_host,
                                first_email, school_name, school_year,
                                callback=None,
                                custom_subject=None, custom_body_text=None):
    log = callback or print
    _log_dir = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')),
                            'MySchoolChecks')
    os.makedirs(_log_dir, exist_ok=True)
    logging.basicConfig(
        filename=os.path.join(_log_dir, 'email_sender.log'), level=logging.ERROR,
        format='%(asctime)s - %(levelname)s - %(message)s')

    # Subject
    if custom_subject:
        _subject = custom_subject.replace('{school_year}', school_year)
    else:
        _subject = DEFAULT_SMEAE_SUBJECT.replace('{school_year}', school_year)

    # Body
    if custom_body_text:
        html_body = _plain_to_html(custom_body_text.replace('{school_year}', school_year))
    else:
        html_body = _plain_to_html(DEFAULT_SMEAE_BODY.replace('{school_year}', school_year))

    msg = MIMEMultipart()
    msg['Subject'] = _subject
    msg['From']  = formataddr((
        str(Header(
            'Πρόγραμμα Ελέγχου Στατιστικών Στοιχείων ΕΕΑ — ΔΙΠΕ Ανατολικής Θεσ/νίκης',
            'utf-8')),
        sender_email
    ))
    msg['To']   = formataddr((str(Header(school_name, 'utf-8')), receiver_email))
    msg['Date'] = formatdate(localtime=True)
    msg.attach(MIMEText(html_body, 'html'))

    with open(attachment_path, 'rb') as fh:
        ext = os.path.splitext(attachment_path)[1].lower()
        if ext == '.xls':
            part = MIMEBase('application', 'vnd.ms-excel')
        elif ext == '.xlsx':
            part = MIMEBase('application',
                            'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            part = MIMEBase('application', 'octet-stream')
        part.set_payload(fh.read())
        encoders.encode_base64(part)
        fname_enc = Header(os.path.basename(attachment_path), 'utf-8').encode()
        part.add_header('Content-Disposition', 'attachment', filename=fname_enc)
        msg.attach(part)

    if first_email:
        log(f'Email → {receiver_email}')
        log(f'Θέμα: {msg["Subject"][:70]}...')
        log(f'Συνημμένο: {os.path.basename(attachment_path)}')

    if not dry_run:
        context = ssl.create_default_context()
        server  = None
        try:
            server = smtplib.SMTP(smtp_host, 587)
            server.starttls(context=context)
            result = server.login(username, password)
            if result[0] == 235:
                server.sendmail(sender_email, receiver_email, msg.as_string())
                log(f'  ✓ Εστάλη → {receiver_email}')
            else:
                log(f'  ✗ Login απέτυχε: {result}')
        except Exception as e:
            log(f'  ✗ Σφάλμα: {e}')
            logging.error(f'send_email failed: {e}')
        finally:
            if server:
                try:
                    server.quit()
                except Exception:
                    pass
    else:
        log(f'  (dry-run) Θα εστέλνετο → {receiver_email} — {os.path.basename(attachment_path)}')
