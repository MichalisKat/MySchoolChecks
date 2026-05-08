#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
placements.py
=============
Αυτόματη καταχώρηση τοποθετήσεων εκπαιδευτικών στο MySchool.
Διαβάζει δεδομένα από Excel και συμπληρώνει τη φόρμα τοποθέτησης.
"""

import os
import re
import sys
import time
from datetime import datetime

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

CHECK_TITLE       = 'Τοποθετήσεις - Αυτόματη Καταχώρηση'
CHECK_DESCRIPTION = 'Αυτόματη καταχώρηση τοποθετήσεων εκπαιδευτικών από Excel'
RESULTS_FOLDER    = 'placements_results'
HAS_EMAIL         = False
CUSTOM_RUN        = True

BASE_URL      = 'https://app.myschool.sch.gr'
LIST_URL      = BASE_URL + '/Management.list.Act.Placement.aspx'
ADD_BTN_ID    = 'ctl00_ContentData_gridActsTemp_header0_new'
FORM_URL_PART = 'Management.Act.Placement.Massive.aspx'

STRIKE_INTERVAL = 0.3
TIME_TO_WAIT    = 30

EXCEL_COLUMNS = [
    'ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ', 'ΑΜ', 'Α.Φ.Μ.', 'ΕΠΙΘΕΤΟ', 'ΟΝΟΜΑ',
    'ΚΩΔ. ΣΧΟΛΕΙΟΥ', 'ΣΧΟΛΕΙΟ', 'ΩΡΕΣ', 'ΑΠΟ', 'ΕΩΣ', 'OK', 'ΣΧΟΛΙΟ',
]


# ── Βοηθητικές ────────────────────────────────────────────────────────────────

def _send_keys_slow(element, text, delay=STRIKE_INTERVAL):
    """Πληκτρολογεί έναν-έναν χαρακτήρες με καθυστέρηση (απαραίτητο για MySchool)."""
    for char in str(text):
        element.send_keys(char)
        time.sleep(delay)


def _parse_date(value):
    """Μετατρέπει οποιαδήποτε μορφή ημερομηνίας σε DD/MM/YYYY string."""
    if value is None:
        return None
    try:
        if isinstance(value, float) and pd.isna(value):
            return None
    except Exception:
        pass

    # pandas Timestamp ή datetime object
    if hasattr(value, 'strftime'):
        return value.strftime('%d/%m/%Y')

    s = str(value).strip()
    if not s or s.lower() in ('nat', 'nan', 'none', ''):
        return None

    # YYYY-MM-DD (Excel string ή ISO)
    if len(s) >= 10 and s[4] == '-':
        try:
            return datetime.strptime(s[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
        except ValueError:
            pass

    # DD/MM/YYYY
    if '/' in s:
        try:
            return datetime.strptime(s[:10], '%d/%m/%Y').strftime('%d/%m/%Y')
        except ValueError:
            pass

    return None


# ── Selenium — φόρμα τοποθέτησης ─────────────────────────────────────────────

def _select_teacher(driver, afm, wait, log):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.keys import Keys
    try:
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, 'ctl00_ContentData_cmbWorker_B0')))
        driver.find_element(By.ID, 'ctl00_ContentData_cmbWorker_B0').click()

        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, 'ctl00_ContentData_cmbWorker_I')))
        field = driver.find_element(By.ID, 'ctl00_ContentData_cmbWorker_I')
        field.clear()
        time.sleep(2)
        WebDriverWait(driver, wait).until(lambda d: field.get_attribute('value') == '')

        # Στέλνω το ΑΦΜ σε δύο δόσεις (όπως το πρωτότυπο)
        _send_keys_slow(field, afm[:1])
        time.sleep(2)
        _send_keys_slow(field, afm[1:])
        time.sleep(2)

        # Αναμονή για dropdown αποτέλεσμα
        for _ in range(20):
            try:
                table = WebDriverWait(driver, 1).until(
                    EC.presence_of_element_located(
                        (By.ID, 'ctl00_ContentData_cmbWorker_DDD_gv_DXMainTable')))
                if table.text:
                    row0 = WebDriverWait(driver, 1).until(
                        EC.presence_of_element_located(
                            (By.ID, 'ctl00_ContentData_cmbWorker_DDD_gv_DXDataRow0')))
                    if row0.text and afm in row0.text:
                        break
            except Exception:
                pass
            time.sleep(2)
        else:
            log('  ⚠ Δεν βρέθηκε ο εκπαιδευτικός στο dropdown')
            return False

        field.send_keys(Keys.TAB)
        return True
    except Exception as e:
        log(f'  ❌ selectTeacher: {e}')
        return False


def _verify_teacher(driver, afm, wait):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, 'ctl00_ContentData_cmbWorker_I')))
        val = driver.find_element(By.ID, 'ctl00_ContentData_cmbWorker_I').get_attribute('value')
        return afm in val
    except Exception:
        return False


def _select_school(driver, sch_code, wait, log):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, 'ctl00_ContentData_cmbUnit_B0')))
        driver.find_element(By.ID, 'ctl00_ContentData_cmbUnit_B0').click()

        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, 'ctl00_ContentData_cmbUnit_I')))
        field = driver.find_element(By.ID, 'ctl00_ContentData_cmbUnit_I')
        field.clear()
        WebDriverWait(driver, wait).until(lambda d: field.get_attribute('value') == '')

        _send_keys_slow(field, sch_code, STRIKE_INTERVAL / 2)
        field.click()
    except Exception as e:
        log(f'  ❌ selectSchool: {e}')


def _verify_school(driver, sch_code, wait):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, 'ctl00_ContentData_cmbUnit_I')))
        val = driver.find_element(By.ID, 'ctl00_ContentData_cmbUnit_I').get_attribute('value')
        return sch_code in val
    except Exception:
        return False


def _select_placement_type(driver, type_text, wait, log):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located(
                (By.ID, 'ctl00_ContentData_cbpEmploymentType_lbEmploymentType_LBT')))
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.XPATH, f"//td[text()='{type_text}']")))
        driver.find_element(By.XPATH, f"//td[text()='{type_text}']").click()
    except Exception as e:
        log(f'  ❌ selectPlacementType ("{type_text}"): {e}')


def _set_date(driver, field_id, date_str, wait, log):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, field_id)))
        field = driver.find_element(By.ID, field_id)
        field.clear()
        WebDriverWait(driver, wait).until(lambda d: field.get_attribute('value') == '')
        _send_keys_slow(field, date_str, STRIKE_INTERVAL / 2)
    except Exception as e:
        log(f'  ❌ setDate({field_id}): {e}')


def _verify_date(driver, field_id, date_str, wait):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, field_id)))
        val = driver.find_element(By.ID, field_id).get_attribute('value')
        if not val:
            return False
        d1 = datetime.strptime(date_str, '%d/%m/%Y')
        d2 = datetime.strptime(val,      '%d/%m/%Y')
        return d1 == d2
    except Exception:
        return False


def _select_hours(driver, hours, wait, log):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        if hours > -1:
            sid = f'ctl00_ContentData_lbEditHours_RB{hours - 1}_I_D'
            WebDriverWait(driver, wait).until(
                EC.presence_of_element_located((By.ID, sid)))
            driver.find_element(By.ID, sid).click()
    except Exception as e:
        log(f'  ❌ selectHours({hours}): {e}')


def _click_all_days(driver, all_days, log):
    try:
        if all_days:
            driver.execute_script(
                "$('input[name=\"ctl00$ContentData$btnAllDays\"]').click()")
        else:
            driver.execute_script(
                "$('input[name=\"ctl00$ContentData$btnUnknown\"]').click()")
    except Exception as e:
        log(f'  ❌ clickAllDays: {e}')


def _save_placement(driver, wait, log):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, 'ctl00_ContentData_savePlacement')))
        driver.find_element(By.ID, 'ctl00_ContentData_savePlacement').click()

        WebDriverWait(driver, wait).until(
            EC.presence_of_element_located((By.ID, 'ctl00_ContentData_info')))
        info_text = driver.find_element(By.ID, 'ctl00_ContentData_info').text
        success   = 'Η αποθήκευση δεν ολοκληρώθηκε' not in info_text
        return [success, info_text]
    except Exception as e:
        log(f'  ❌ savePlacement: {e}')
        return [False, f'Σφάλμα κατά αποθήκευση: {e}']


def _fill_form(driver, afm, sch_code, date_from, date_to, hours, type_text, wait, log):
    if not _select_teacher(driver, afm, wait, log):
        return [False, 'Απέτυχε καταχώριση ΑΦΜ']
    time.sleep(1)
    if not _verify_teacher(driver, afm, wait):
        return [False, 'Απέτυχε επαλήθευση ΑΦΜ']

    time.sleep(1)
    _select_school(driver, sch_code, wait, log)
    _set_date(driver, 'ctl00_ContentData_dateFrom_I', date_from, wait, log)
    _set_date(driver, 'ctl00_ContentData_dateTo_I',   date_to,   wait, log)
    _select_hours(driver, hours, wait, log)
    # Ο κατάλογος τοποθέτησης φορτώνει αργά — τον θέτω μετά τις ώρες
    _select_placement_type(driver, type_text, wait, log)
    _click_all_days(driver, hours == -1, log)

    if not _verify_school(driver, sch_code, wait):
        return [False, 'Απέτυχε επαλήθευση κωδικού σχολείου']
    if not _verify_date(driver, 'ctl00_ContentData_dateFrom_I', date_from, wait):
        return [False, 'Απέτυχε επαλήθευση ΑΠΟ']
    if not _verify_date(driver, 'ctl00_ContentData_dateTo_I', date_to, wait):
        return [False, 'Απέτυχε επαλήθευση ΕΩΣ']

    time.sleep(1)
    return [True, 'Επιτυχής πληκτρολόγηση']


# ── Ανάγνωση γραμμής Excel ────────────────────────────────────────────────────

def _read_row(row):
    try:
        str_cols = ['ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ', 'ΕΠΙΘΕΤΟ', 'ΟΝΟΜΑ',
                    'ΚΩΔ. ΣΧΟΛΕΙΟΥ', 'ΣΧΟΛΕΙΟ', 'ΑΠΟ', 'ΕΩΣ', 'OK', 'ΣΧΟΛΙΟ']
        for col in str_cols:
            if isinstance(row[col], (float, int)):
                row[col] = ''

        # ΑΦΜ/ΑΜ: Excel αποθηκεύει αριθμούς ως float — μετατροπή σε string
        def _to_str(val):
            if val is None:
                return ''
            s = str(val).strip()
            if s in ('', 'nan', 'None', 'NaN'):
                return ''
            try:
                return str(int(float(s)))
            except Exception:
                return s

        afm = _to_str(row['Α.Φ.Μ.'])
        am  = _to_str(row['ΑΜ'])

        if not afm and am:
            return [False, 'ΑΦΜ κενό — καταχωρήστε ΑΦΜ απευθείας στο Excel']
        if not afm:
            return [False, 'Κενό ΑΦΜ']

        while len(afm) < 9:
            afm = '0' + afm

        sch_code = str(row['ΚΩΔ. ΣΧΟΛΕΙΟΥ']).strip()
        if not sch_code:
            return [False, 'Κενός κωδικός σχολείου']

        date_from = _parse_date(row['ΑΠΟ'])
        if not date_from:
            return [False, 'Μη έγκυρη ημερομηνία ΑΠΟ']

        date_to = _parse_date(row['ΕΩΣ'])
        if not date_to:
            return [False, 'Μη έγκυρη ημερομηνία ΕΩΣ']

        hours = row['ΩΡΕΣ']
        try:
            hours = int(hours)
        except Exception:
            return [False, 'Μη έγκυρος αριθμός ωρών (χρησιμοποιήστε -1 για πλήρες ωράριο)']

        type_text = str(row['ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ']).strip()
        if not type_text:
            return [False, 'Κενό είδος τοποθέτησης']

        teacher_name = f"{row['ΕΠΙΘΕΤΟ']} {row['ΟΝΟΜΑ']}".strip()
        sch_name     = str(row['ΣΧΟΛΕΙΟ']).strip()

        ok_val = str(row['OK']).strip()
        skip   = ok_val in ('ΕΠΙΤΥΧΙΑ', 'ΑΛΛΗ ΤΟΠΟΘΕΤΗΣΗ')

        return [True, type_text, afm, date_from, date_to,
                teacher_name, sch_code, sch_name, hours, skip]
    except Exception as e:
        return [False, f'Σφάλμα ανάγνωσης γραμμής: {e}']


# ── Μετατροπή αρχείου ΠΔΕ/Υπουργείου ────────────────────────────────────────

_ORDINAL_ONLY = re.compile(r'^\d+\S*$')   # π.χ. "3ο", "12η"

def _split_schools(text):
    """
    "3ο & 1ο ΔΣ ΤΡΙΑΝΔΡΙΑΣ"  →  ["3ο ΔΣ ΤΡΙΑΝΔΡΙΑΣ", "1ο ΔΣ ΤΡΙΑΝΔΡΙΑΣ"]
    "59ο ΔΣ ΘΕΣΣ. & 61ο ΔΣ ΘΕΣΣ."  →  αμετάβλητο
    """
    parts = [s.strip() for s in re.split(r'\s*&\s*', text) if s.strip()]
    if len(parts) <= 1:
        return parts
    # Βρες το suffix από το πρώτο πλήρες μέρος
    suffix = None
    for p in parts:
        if not _ORDINAL_ONLY.match(p):
            m = re.match(r'^\S+\s+(.+)$', p)
            if m:
                suffix = m.group(1)
            break
    result = []
    for p in parts:
        result.append(f'{p} {suffix}' if (_ORDINAL_ONLY.match(p) and suffix) else p)
    return result


def convert_raw_file(src_path, dest_path=None):
    """
    Μετατρέπει αρχείο τοποθετήσεων από μορφή ΠΔΕ/Υπουργείου
    στη μορφή που χρειάζεται το placements.py.

    Επιστρέφει (dest_path, n_rows) ή raise Exception σε αποτυχία.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Χρώματα ──────────────────────────────────────────────────────────────
    HDR_FG   = 'FFFFFF'
    AUTO_BG  = 'E2EFDA'   # πράσινο — συμπληρώνεται αυτόματα
    WARN_BG  = 'FCE4D6'   # πορτοκαλί — χρειάζεται συμπλήρωση
    APP_BG   = 'F2F2F2'   # γκρι — εφαρμογή
    INPUT_BG = 'D9E1F2'   # μπλε — χειροκίνητο (ΩΡΕΣ για multi)
    MULTI_BG = 'FFF2CC'   # κίτρινο — ΩΡΕΣ σε πολλαπλά σχολεία

    HDR_COLORS = {
        'ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ': 'E74C3C', 'ΑΜ': '2E7D32', 'Α.Φ.Μ.': 'E74C3C',
        'ΕΠΙΘΕΤΟ': '2E7D32', 'ΟΝΟΜΑ': '2E7D32', 'ΚΩΔ. ΣΧΟΛΕΙΟΥ': 'E74C3C',
        'ΣΧΟΛΕΙΟ': '2E7D32', 'ΩΡΕΣ': 'E74C3C', 'ΑΠΟ': '2E7D32',
        'ΕΩΣ': 'E74C3C', 'OK': '616161', 'ΣΧΟΛΙΟ': '616161',
    }
    COL_BGS = {
        'ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ': WARN_BG, 'ΑΜ': AUTO_BG, 'Α.Φ.Μ.': WARN_BG,
        'ΕΠΙΘΕΤΟ': AUTO_BG, 'ΟΝΟΜΑ': AUTO_BG, 'ΚΩΔ. ΣΧΟΛΕΙΟΥ': WARN_BG,
        'ΣΧΟΛΕΙΟ': AUTO_BG, 'ΩΡΕΣ': INPUT_BG, 'ΑΠΟ': AUTO_BG,
        'ΕΩΣ': WARN_BG, 'OK': APP_BG, 'ΣΧΟΛΙΟ': APP_BG,
    }
    thin   = Side(style='thin', color='BFBFBF')
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Διάβασμα ──────────────────────────────────────────────────────────────
    raw = pd.read_excel(src_path, sheet_name=0, header=1)
    raw.columns = [str(c).strip() for c in raw.columns]
    raw = raw[raw['ΕΠΩΝΥΜΟ'].notna() &
              (raw['ΕΠΩΝΥΜΟ'].astype(str).str.strip() != '')].copy()
    raw.reset_index(drop=True, inplace=True)

    if raw.empty:
        raise ValueError('Το αρχείο δεν περιέχει αναγνωρίσιμα δεδομένα '
                         '(αναμένεται στήλη ΕΠΩΝΥΜΟ).')

    # ── Μετατροπή + split ─────────────────────────────────────────────────────
    rows = []
    for _, r in raw.iterrows():
        am_raw = r.get('A.M.') or r.get('ΑΜ') or r.get('Α.Μ.')
        am     = str(int(am_raw)) if pd.notna(am_raw) and str(am_raw) not in ('nan', '') else ''
        apo_raw = r.get('ΗΜΕΡΟΜΗΝΙΑ') or r.get('ΑΠΟ') or ''
        apo     = pd.to_datetime(apo_raw).strftime('%d/%m/%Y') if pd.notna(apo_raw) and apo_raw != '' else ''

        epan    = str(r.get('ΕΠΑΝΑΤΟΠΟΘΕΤΗΣΗ') or r.get('ΤΟΠΟΘΕΤΗΣΗ') or '').strip()
        schools = _split_schools(epan)
        is_multi = len(schools) > 1

        for sch in schools:
            rows.append({
                'ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ': '',
                'ΑΜ':  am,
                'Α.Φ.Μ.': '',
                'ΕΠΙΘΕΤΟ': str(r.get('ΕΠΩΝΥΜΟ', '')).strip(),
                'ΟΝΟΜΑ':   str(r.get('ΟΝΟΜΑ', '')).strip(),
                'ΚΩΔ. ΣΧΟΛΕΙΟΥ': '',
                'ΣΧΟΛΕΙΟ': sch,
                'ΩΡΕΣ':  None if is_multi else -1,
                'ΑΠΟ':   apo,
                'ΕΩΣ':   '',
                'OK':    '',
                'ΣΧΟΛΙΟ': '',
                '_multi': is_multi,
            })

    # ── Δημιουργία dest_path ──────────────────────────────────────────────────
    if dest_path is None:
        src_dir  = os.path.dirname(src_path)
        fname    = os.path.basename(src_path)
        m_date   = re.search(r'\d{2}-\d{2}-\d{4}', fname)
        tag      = m_date.group(0) if m_date else datetime.now().strftime('%d-%m-%Y')
        dest_path = os.path.join(src_dir, f'αρχικο_τοποθ_{tag}.xlsx')

    # ── Δημιουργία xlsx ───────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'ΤΟΠΟΘΕΤΗΣΕΙΣ'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value = (u'Συμπλήρωσε τα '
               u'πορτοκαλί κελιά: '
               u'EIDOS TOPOTHETHSHS, A.F.M., KWD. SXOLEIOU, EWS | '
               u'Κίτρινο WRES = πολλαπλά σχολεία')
    c.font      = Font(bold=True, color='7B3F00', name='Arial', size=10, italic=True)
    c.fill      = PatternFill('solid', fgColor=WARN_BG)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border    = brd
    ws.row_dimensions[1].height = 22

    col_greek = [
        'ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ',
        'ΑΜ',
        'Α.Φ.Μ.',
        'ΕΠΙΘΕΤΟ',
        'ΟΝΟΜΑ',
        'ΚΩΔ. ΣΧΟΛΕΙΟΥ',
        'ΣΧΟΛΕΙΟ',
        'ΩΡΕΣ',
        'ΑΠΟ',
        'ΕΩΣ',
        'OK',
        'ΣΧΟΛΙΟ',
    ]
    for ci, col in enumerate(col_greek, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font      = Font(bold=True, color=HDR_FG, name='Arial', size=10)
        cell.fill      = PatternFill('solid', fgColor=HDR_COLORS[col] if col in HDR_COLORS else '616161')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = brd
    ws.row_dimensions[2].height = 28

    left_cols = {
        'ΕΠΙΘΕΤΟ',
        'ΟΝΟΜΑ',
        'ΣΧΟΛΕΙΟ',
        'ΣΧΟΛΙΟ',
    }
    wres_col  = 'ΩΡΕΣ'
    apo_col   = 'ΑΠΟ'
    ews_col   = 'ΕΩΣ'

    for ri, row in enumerate(rows):
        er       = ri + 3
        is_multi = row['_multi']
        for ci, col in enumerate(col_greek, 1):
            val  = row.get(col, '')
            cell = ws.cell(row=er, column=ci)
            if col == wres_col:
                cell.value = int(val) if val is not None and val != '' else None
            else:
                cell.value = val if val not in ('', None) else None
            bg = MULTI_BG if (col == wres_col and is_multi) else COL_BGS.get(col, APP_BG)
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.font      = Font(name='Arial', size=10)
            cell.border    = brd
            cell.alignment = Alignment(
                horizontal='left' if col in left_cols else 'center',
                vertical='center')
            if col in (apo_col, ews_col):
                cell.number_format = 'DD/MM/YYYY'
        ws.row_dimensions[er].height = 18

    widths = [22, 10, 13, 18, 12, 14, 34, 10, 12, 12, 18, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(dest_path)
    return dest_path, len(rows)


# ── Σύνδεση & πλοήγηση ────────────────────────────────────────────

def connect(log=print):
    """
    Ανοίγει Chrome, κάνει SSO login, πλοηγείται στη λίστα τοποθετήσεων
    και κάνει κλικ στο κουμπί Προσθήκης.
    Επιστρέφει το driver ή None σε αποτυχία.
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
    except ImportError:
        log('pip install selenium')
        return None

    log('Chrome...')
    options = webdriver.ChromeOptions()
    options.add_argument('--window-size=1400,900')
    options.add_argument('--no-sandbox')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    try:
        from selenium.webdriver.chrome.service import Service as ChromeService
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            driver = webdriver.Chrome(
                service=ChromeService(ChromeDriverManager().install()),
                options=options)
        except Exception:
            driver = webdriver.Chrome(options=options)
    except Exception as e:
        log(f'Chrome: {e}')
        return None

    try:
        import config as _cfg
        log('MySchool...')
        driver.get(BASE_URL)
        time.sleep(2)

        if 'sso.sch.gr' in driver.current_url or 'login' in driver.current_url.lower():
            user_f = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,
                    '#username, input[name="username"], input[type="text"]')))
            user_f.clear()
            user_f.send_keys(_cfg.MYSCHOOL_USER)

            pass_f = driver.find_element(By.CSS_SELECTOR,
                '#password, input[name="password"], input[type="password"]')
            pass_f.clear()
            pass_f.send_keys(_cfg.MYSCHOOL_PASS)

            driver.find_element(By.CSS_SELECTOR,
                'button[type="submit"], input[type="submit"]').click()
            time.sleep(3)

        driver.get(LIST_URL)
        time.sleep(3)

        add_btn = WebDriverWait(driver, TIME_TO_WAIT).until(
            EC.element_to_be_clickable((By.ID, ADD_BTN_ID)))
        driver.execute_script('arguments[0].click();', add_btn)
        time.sleep(3)

        if FORM_URL_PART not in driver.current_url:
            log(f'page: {driver.current_url}')
            driver.quit()
            return None

        return driver

    except Exception as e:
        log(f'{e}')
        try:
            driver.quit()
        except Exception:
            pass
        return None


# ── Κύριος βρόχος ───────────────────────────────────────────────

def run(ctx, driver, callback=None):
    """
    Κύριος βρόχος επεξεργασίας.
    Απαιτεί driver ήδη συνδεδεμένο (από connect()).
    """
    log = callback or print

    excel_path = ctx.get('excel_path')

    try:
        col_types = {c: 'str' for c in EXCEL_COLUMNS if c != 'ΩΡΕΣ'}
        col_types['ΩΡΕΣ'] = 'Int64'
        df = pd.read_excel(excel_path, dtype=col_types)
        for col in EXCEL_COLUMNS:
            if col not in df.columns:
                df[col] = ''
        df = df[EXCEL_COLUMNS]
    except Exception as e:
        log(f'{e}')
        return

    if df.empty:
        return

    if FORM_URL_PART not in driver.current_url:
        return

    total = len(df)

    for idx, row in df.iterrows():
        result = _read_row(row)

        if not result[0]:
            df.at[idx, 'OK']     = 'ΑΠΟΤΥΧΙΑ'
            df.at[idx, 'ΣΧΟΛΙΟ'] = result[1]
            df.to_excel(excel_path, index=False)
            continue

        _, type_text, afm, date_from, date_to, teacher_name, sch_code, sch_name, hours, skip = result

        if skip:
            continue

        fill_ok = _fill_form(driver, afm, sch_code, date_from, date_to,
                             hours, type_text, TIME_TO_WAIT, log)
        if not fill_ok[0]:
            df.at[idx, 'OK']     = 'ΑΠΟΤΥΧΙΑ'
            df.at[idx, 'ΣΧΟΛΙΟ'] = fill_ok[1]
            df.to_excel(excel_path, index=False)
            continue

        save_ok = _save_placement(driver, TIME_TO_WAIT, log)
        if save_ok[0]:
            df.at[idx, 'OK']     = 'ΕΠΙΤΥΧΙΑ'
            df.at[idx, 'ΣΧΟΛΙΟ'] = save_ok[1]
        else:
            msg = save_ok[1]
            df.at[idx, 'OK']     = ('ΑΛΛΗ ΤΟΠΟΘΕΤΗΣΗ'
                                    if 'Υπάρχει και άλλη ενεργή τοποθέτηση' in msg
                                    else 'ΑΠΟΤΥΧΙΑ')
            df.at[idx, 'ΣΧΟΛΙΟ'] = msg

        df.to_excel(excel_path, index=False)
        time.sleep(1)
