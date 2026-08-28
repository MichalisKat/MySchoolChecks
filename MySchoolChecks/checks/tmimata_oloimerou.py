"""
checks/tmimata_oloimerou.py
════════════════════════════
Έλεγχος Τμημάτων Ολοημέρου (2026-2027).

Πηγές:
  5.3 (stat5_3) — Αποτύπωση νηπιαγωγείων.
  5.4 (stat5_4) — Αποτύπωση δημοτικών.
  2.2 (stat2_2) — προαιρετικό, μόνο για τη στήλη «Email Σχολείου».
Και τα δύο στατιστικά 5.3/5.4 κατεβαίνουν ήδη για το σχολικό έτος 2026-2027
από προεπιλογή (core/downloader.py) — δεν χρειάζεται override έτους, σε
αντίθεση με το tmimata_genikis (που χρειάζεται override για το 3.1).

Η λήψη γίνεται ΜΟΝΟ μέσω του tab «⬇ Λήψη» (CUSTOM_DOWNLOAD_INFO /
CUSTOM_DOWNLOAD παρακάτω) — το run() (tab «▶ Εκτέλεση») ΔΕΝ κατεβάζει ποτέ
τίποτα, μόνο ψάχνει για ήδη κατεβασμένα αρχεία (_find_downloaded_inputs).
Ίδιο μοτίβο με το checks/tmimata_genikis.py.

Λογική:
  Νηπιαγωγεία (5.3): σχολεία ΧΩΡΙΣ «Αναστολή λειτουργίας Ολοημέρου»
    (τιμή «Όχι») ΚΑΙ με «Μαθητές σε ολοήμερα» < 3.
  Δημοτικά (5.4): σχολεία ΧΩΡΙΣ «Αναστολή λειτουργίας Ολοημέρου»
    (τιμή «Όχι») ΚΑΙ με «Διακριτοί Μαθητές σε Ολοήμερα» < 3.

Αποτέλεσμα: 1 αρχείο Excel με 2 φύλλα (Νηπιαγωγεία, Δημοτικά) — μόνο τα
σχολεία που πληρούν τα παραπάνω κριτήρια (άρα ΟΛΑ όσα εμφανίζονται στο
αρχείο χρειάζονται ενημέρωση — δεν υπάρχει έννοια «απόκλισης» εδώ, σε
αντίθεση με το tmimata_genikis).

Tabs «✂ Διαχωρισμός» / «✉ Αποστολή»: ίδιο μοτίβο με το tmimata_genikis —
το συγκεντρωτικό αρχείο χωρίζεται σε ένα Excel ανά σχολείο (ΟΛΑ τα σχολεία
εδώ, χωρίς φίλτρο απόκλισης, αφού όλα ήδη πληρούν το κριτήριο), και
στέλνεται email ανά σχολείο (ή ένα συγκεντρωτικό test email).
"""

import glob as _glob
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Μεταδεδομένα ────────────────────────────────────────────────────────────
CHECK_TITLE       = 'Έλεγχος Τμημάτων Ολοημέρου (2026-2027)'
CHECK_DESCRIPTION = ('Νηπιαγωγεία (5.3) και Δημοτικά (5.4) χωρίς Αναστολή λειτουργίας '
                     'Ολοημέρου, με λιγότερους από 3 μαθητές σε Ολοήμερο.')
RESULTS_FOLDER    = 'tmimata_oloimerou'
HAS_EMAIL         = False
CUSTOM_RUN        = True
SCHOOL_YEAR       = '2026-2027'
THRESHOLD         = 3   # < 3 μαθητές σε ολοήμερο
REQUIRED_REPORTS  = [
    '5.3 — Αποτύπωση νηπιαγωγείων',
    '5.4 — Αποτύπωση δημοτικών',
]

# Αποτέλεσμα της τελευταίας επιτυχούς Εκτέλεσης (df_nip/df_ds/today/out_path)
# — το χρησιμοποιεί το tab «✉ Αποστολή» του CheckRunDialog (βλ. build_send_tab).
_LAST_RESULT = None

DEFAULT_EMAIL_SUBJECT = (
    'Τμήματα Ολοημέρου {school_year} — έλεγχος βιωσιμότητας'
)
DEFAULT_EMAIL_BODY = (
    'Καλημέρα,\n\n'
    'Κατά τον έλεγχο των στοιχείων Ολοημέρου για το σχολικό έτος '
    '{school_year} στο MySchool, το τμήμα Ολοημέρου του σχολείου σας '
    'δεν εμφανίζεται σε Αναστολή λειτουργίας, ενώ έχει λιγότερους από 3 '
    'μαθητές.\n\n'
    'Παρακαλούμε ελέγξτε τα στοιχεία και ενημερώστε μας για την τρέχουσα '
    'κατάσταση του τμήματος (λειτουργία / αναστολή).\n\n'
    'Παρακαλούμε για τις ενέργειές σας.'
)

COL_ANASTOLI = 'Αναστολή λειτουργίας Ολοημέρου'
COL_MO_NIP   = 'Μαθητές σε ολοήμερα'
COL_MO_DS    = 'Διακριτοί Μαθητές σε Ολοήμερα'

COLOR_HDR = '1F4E79'
COLOR_SUB = 'D6E4F0'
COLOR_ALT = 'EBF3FB'
DEFAULT_PERIFEREIA = 'ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ'

_LEFT_ALIGN_COLS = {'Τύπος Σχολείου', 'Ονομασία Σχολ. Μονάδας', 'Επώνυμο Διευθυντή',
                    'Email Σχολείου'}


# ── Στήλες εξόδου ────────────────────────────────────────────────────────────
def columns_nip():
    return [
        ('Τύπος Σχολείου',           40),
        ('Κωδ. Υπουργείου Σχολείου', 14),
        ('Ονομασία Σχολ. Μονάδας',   40),
        ('Email Σχολείου',           28),
        ('Λειτουργικότητα',          12),
        ('Οργανικότητα',             12),
        (COL_ANASTOLI,               14),
        ('Επώνυμο Διευθυντή',        18),
        (COL_MO_NIP,                 16),
        ('Μαθητές',                  10),
    ]

def columns_ds():
    return [
        ('Τύπος Σχολείου',           40),
        ('Κωδ. Υπουργείου Σχολείου', 14),
        ('Ονομασία Σχολ. Μονάδας',   40),
        ('Email Σχολείου',           28),
        ('Λειτουργικότητα',          12),
        ('Οργανικότητα',             12),
        (COL_ANASTOLI,               14),
        ('Επώνυμο Διευθυντή',        18),
        (COL_MO_DS,                  16),
        ('Μαθητές',                  10),
    ]


# ── Email lookup (2.2, προαιρετικό) ─────────────────────────────────────────
def _load_email_lookup(path_22):
    """
    Διαβάζει stat2_2 (xlsx ή CSV/zip, cp1253, ';'-delimited) και επιστρέφει
    {code_int: email}. col11 = Κωδ. ΥΠΠΘ, col18 = e-mail σχολείου.
    Αντίγραφο ίδιας λογικής με tmimata_genikis._load_email_lookup.
    """
    import re
    lower = path_22.lower()
    if lower.endswith('.xlsx') or lower.endswith('.xls'):
        df = pd.read_excel(path_22, dtype=str)
    else:
        import zipfile, io
        if lower.endswith('.zip'):
            with zipfile.ZipFile(path_22) as z:
                raw = z.read(z.namelist()[0])
        else:
            with open(path_22, 'rb') as f:
                raw = f.read()
        text = raw.decode('cp1253')
        df = pd.read_csv(io.StringIO(text), sep=';', dtype=str, header=0)

    if len(df.columns) <= 18:
        return {}

    c_code  = df.columns[11]
    c_email = df.columns[18]

    def _clean_code(val):
        s = str(val).strip().strip('"').lstrip('=').strip('"').strip()
        s = re.sub(r'\.0$', '', s)
        return s.lstrip('0') or s

    lookup = {}
    for _, row in df.iterrows():
        raw_code = _clean_code(row[c_code])
        email    = str(row[c_email]).strip() if pd.notna(row[c_email]) else ''
        if email.lower() in ('nan', 'none', ''):
            email = ''
        if raw_code.isdigit():
            lookup[int(raw_code)] = email
    return lookup


# ── Λογική ───────────────────────────────────────────────────────────────────
def _filter_schools(df, mo_col, threshold, email_lookup=None):
    """Επιστρέφει το subset του df με Αναστολή='Όχι' ΚΑΙ mo_col < threshold,
    με προστιθέμενη στήλη 'Email Σχολείου' (από email_lookup, αν δόθηκε)."""
    if email_lookup is None:
        email_lookup = {}
    if df.empty or COL_ANASTOLI not in df.columns or mo_col not in df.columns:
        return df.iloc[0:0]
    mo_num = pd.to_numeric(df[mo_col], errors='coerce')
    mask = (df[COL_ANASTOLI].astype(str).str.strip() == 'Όχι') & (mo_num < threshold)
    out = df[mask].copy()
    if not out.empty:
        code_num = pd.to_numeric(out['Κωδ. Υπουργείου Σχολείου'], errors='coerce')

        def _email(c):
            try:
                return email_lookup.get(int(c), '')
            except Exception:
                return ''

        out['Email Σχολείου'] = code_num.apply(_email)
        out = out.sort_values('Ονομασία Σχολ. Μονάδας').reset_index(drop=True)
    return out


def process(path_53, path_54, path_22=None, threshold=THRESHOLD):
    """Επιστρέφει (df_nip, df_ds, periferia_name)."""
    df53 = pd.read_excel(path_53)
    df54 = pd.read_excel(path_54)

    email_lookup = _load_email_lookup(path_22) if path_22 else {}

    df_nip = _filter_schools(df53, COL_MO_NIP, threshold, email_lookup)
    df_ds  = _filter_schools(df54, COL_MO_DS,  threshold, email_lookup)

    perif = _periferia_name(df54) if not df54.empty else _periferia_name(df53)
    return df_nip, df_ds, perif


def _periferia_name(df_src):
    try:
        raw = str(df_src['Περιφέρεια'].dropna().iloc[0])
        if 'ΕΚΠ/ΣΗΣ ' in raw:
            return raw.split('ΕΚΠ/ΣΗΣ ')[-1].strip()
    except Exception:
        pass
    return DEFAULT_PERIFEREIA


# ── Excel ─────────────────────────────────────────────────────────────────
def _brd():
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)


def _write_sheet(ws, df, title, columns, today, subtitle_extra=''):
    brd = _brd()
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    ncols = len(columns)

    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws['A1'] = f'{title}  —  {today.strftime("%d/%m/%Y")}{subtitle_extra}'
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', start_color=COLOR_HDR)
    ws['A1'].alignment = ctr
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
    ws['A2'] = f'Σύνολο σχολείων: {len(df)}'
    ws['A2'].font      = Font(name='Arial', italic=True, size=9)
    ws['A2'].fill      = PatternFill('solid', start_color=COLOR_SUB)
    ws['A2'].alignment = ctr
    ws.row_dimensions[2].height = 16

    for ci, (name, width) in enumerate(columns, 1):
        c = ws.cell(row=3, column=ci, value=name)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill      = PatternFill('solid', start_color=COLOR_HDR)
        c.alignment = ctr
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 28

    col_keys = [c[0] for c in columns]

    for ri, (_, row) in enumerate(df.iterrows(), start=4):
        base_fill = PatternFill('solid', start_color=COLOR_ALT) if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(col_keys, 1):
            val = row.get(key, '')
            if hasattr(val, 'item'):
                val = val.item()
            c = ws.cell(row=ri, column=ci, value=val)
            c.border    = brd
            c.font      = Font(name='Arial', size=9)
            c.fill      = base_fill
            c.alignment = lft if key in _LEFT_ALIGN_COLS else ctr
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}3'


def build_workbook(df_nip, df_ds, perif, today, out_path):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = f'ΝΗΠ-{perif}'[:31]
    _write_sheet(ws1, df_nip, CHECK_TITLE, columns_nip(), today,
                 subtitle_extra=f'  |  Νηπιαγωγεία (χωρίς αναστολή, < {THRESHOLD} μαθητές)')

    ws2 = wb.create_sheet(f'ΔΣ-{perif}'[:31])
    _write_sheet(ws2, df_ds, CHECK_TITLE, columns_ds(), today,
                 subtitle_extra=f'  |  Δημοτικά (χωρίς αναστολή, < {THRESHOLD} μαθητές)')

    wb.save(out_path)


# ── Email — Βοηθητικά ─────────────────────────────────────────────────────────
def _app_base():
    import sys
    if getattr(sys, 'frozen', False):
        lad = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'MySchoolChecks')
        return lad if os.path.isdir(lad) else os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _load_oloimerou_email_template():
    import json
    spath = os.path.join(_app_base(), 'data', 'local_settings.json')
    try:
        with open(spath, encoding='utf-8') as f:
            s = json.load(f)
        t = s.get('tmimata_oloimerou_email', {})
        return t.get('subject', DEFAULT_EMAIL_SUBJECT), t.get('body', DEFAULT_EMAIL_BODY)
    except Exception:
        return DEFAULT_EMAIL_SUBJECT, DEFAULT_EMAIL_BODY


def _save_oloimerou_email_template(subject, body):
    import json
    spath = os.path.join(_app_base(), 'data', 'local_settings.json')
    os.makedirs(os.path.dirname(spath), exist_ok=True)
    try:
        with open(spath, encoding='utf-8') as f:
            s = json.load(f)
    except Exception:
        s = {}
    s['tmimata_oloimerou_email'] = {'subject': subject, 'body': body}
    with open(spath, 'w', encoding='utf-8') as f:
        json.dump(s, f, ensure_ascii=False, indent=2)


def _build_mini_excel(row_df, col_defs, label, today):
    """Mini Excel (bytes) για ένα σχολείο."""
    import io as _io
    buf = _io.BytesIO()
    wb  = Workbook()
    ws  = wb.active
    ws.title = label[:31]
    _write_sheet(ws, row_df, CHECK_TITLE, col_defs, today,
                 subtitle_extra=f'  |  {label}')
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _split_dir_for(out_path):
    """Ο φάκελος «split» που αντιστοιχεί σε ένα συγκεντρωτικό αρχείο αποτελεσμάτων."""
    return os.path.join(os.path.dirname(out_path), 'split')


def _schools_from_split_dir(split_dir):
    """
    Διαβάζει κάθε αρχείο Excel μέσα στον φάκελο «split» (βλ.
    split_oloimerou_workbook) και επιστρέφει list of (code, name, email,
    file_path) — ένα ανά σχολείο. Αφού ΟΛΑ τα σχολεία σε αυτόν τον έλεγχο
    ήδη πληρούν το κριτήριο (δεν υπάρχει φίλτρο απόκλισης εδώ), ο φάκελος
    split περιέχει ένα αρχείο ανά σχολείο του συγκεντρωτικού αρχείου.
    """
    result = []
    if not split_dir or not os.path.isdir(split_dir):
        return result
    for fn in sorted(os.listdir(split_dir)):
        if not fn.lower().endswith('.xlsx'):
            continue
        fp = os.path.join(split_dir, fn)
        name, email, code = fn, '', ''
        try:
            df = pd.read_excel(fp, skiprows=2)
            if not df.empty:
                row   = df.iloc[0]
                name  = str(row.get('Ονομασία Σχολ. Μονάδας', '')).strip() or fn
                email = str(row.get('Email Σχολείου', '')).strip()
                if email.lower() in ('nan', 'none'):
                    email = ''
                code = row.get('Κωδ. Υπουργείου Σχολείου', '')
                try:
                    code = int(code)
                except Exception:
                    pass
        except Exception:
            pass
        result.append((code, name, email, fp))
    return result


def _smtp_send(config, to_addr, msg_str):
    """
    Αποστολή email — 3 προσπάθειες όπως ακριβώς στο framework.py:
    1) SMTP_SSL 465 με πλήρες cert
    2) SMTP_SSL 465 χωρίς cert check
    3) SMTP STARTTLS 587 χωρίς cert check
    """
    import ssl, smtplib

    def _lenient():
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode    = ssl.CERT_NONE
        return ctx

    ef = config.FROM_EMAIL
    pw = config.FROM_PASSWORD
    sh = config.SMTP_HOST

    sent = False
    try:
        with smtplib.SMTP_SSL(sh, 465, context=ssl.create_default_context()) as s:
            s.login(ef, pw)
            s.sendmail(ef, to_addr, msg_str)
            sent = True
    except Exception:
        pass
    if not sent:
        try:
            with smtplib.SMTP_SSL(sh, 465, context=_lenient()) as s:
                s.login(ef, pw)
                s.sendmail(ef, to_addr, msg_str)
                sent = True
        except Exception:
            pass
    if not sent:
        with smtplib.SMTP(sh, 587) as s:
            s.starttls(context=_lenient())
            s.login(ef, pw)
            s.sendmail(ef, to_addr, msg_str)


def _send_summary_oloimerou(config, sent, failed, today, log):
    """Summary email στο FROM_EMAIL μετά την ολοκλήρωση."""
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.utils import formataddr, formatdate
    from email.header import Header

    ef = getattr(config, 'FROM_EMAIL',    '')
    pw = getattr(config, 'FROM_PASSWORD', '')
    if not ef or not pw:
        return

    sent_lines   = '\n'.join(f'  ✓  {n}  ({e})' for n, e in sorted(sent))
    failed_lines = ('\n'.join(f'  ✗  {n}  ({e})' for n, e in sorted(failed))
                    if failed else '  —')
    body = (
        f'Αποστολή email Τμημάτων Ολοημέρου — {today.strftime("%d/%m/%Y")}\n\n'
        f'Εστάλησαν: {len(sent)}\n{sent_lines}\n\n'
        + (f'Αποτυχίες ({len(failed)}):\n{failed_lines}' if failed else '')
    )
    subject = (
        f'[Ολοήμερο {SCHOOL_YEAR}] Αποστολή ολοκληρώθηκε — {len(sent)} ✓'
        + (f', {len(failed)} ✗' if failed else '')
    )
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From']    = formataddr((str(Header('MySchool Checks', 'utf-8')), ef))
    msg['To']      = ef
    msg['Date']    = formatdate(localtime=True)
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    log(f'\n  ✉ Summary → {ef}')
    try:
        _smtp_send(config, ef, msg.as_string())
        log('  ✓ Summary εστάλη.')
    except Exception as e:
        log(f'  ✗ Summary: {e}')


def _do_send_emails(dry_run, subj_tpl, body_tpl, config, today, log,
                    out_path=None, split_dir=None):
    """
    Αποστολή email. Επιστρέφει (sent, failed).

    - Κανονική αποστολή (dry_run=False): ένα email ανά σχολείο, ΜΟΝΟ για όσα
      έχουν ατομικό αρχείο μέσα στον φάκελο «split» (βλ. tab «✂ Διαχωρισμός» /
      _schools_from_split_dir) — συνημμένο ακριβώς το ίδιο αρχείο.
    - Test mode (dry_run=True): ΕΝΑ συγκεντρωτικό email προς το FROM_EMAIL,
      με ολόκληρο το ΣΥΝΟΛΙΚΟ Excel (out_path) συνημμένο.
    """
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email.utils import formataddr, formatdate
    from email.header import Header
    from email import encoders

    ef = getattr(config, 'FROM_EMAIL',    '')
    pw = getattr(config, 'FROM_PASSWORD', '')
    if not ef:
        log('  ✗ FROM_EMAIL δεν έχει οριστεί στις Ρυθμίσεις.')
        return [], []
    if not pw:
        log('  ✗ Κωδικός email δεν έχει οριστεί στις Ρυθμίσεις.')
        return [], []

    schools = _schools_from_split_dir(split_dir)
    if not schools:
        log('  ℹ Δεν βρέθηκαν αρχεία στο φάκελο «split» — τρέξε πρώτα το tab '
            '«✂ Διαχωρισμός» (χωρίζει το συγκεντρωτικό αρχείο σε ένα Excel '
            'ανά σχολείο).')
        return [], []

    # ── Test mode: ένα συγκεντρωτικό email με το συνολικό Excel ────────────
    if dry_run:
        if not out_path or not os.path.exists(out_path):
            log('  ✗ Δεν βρέθηκε το συνολικό αρχείο Excel για το test email.')
            return [], []

        names = [name for _, name, *_ in schools]
        log(f'  → Test mode: 1 συγκεντρωτικό email ({len(schools)} σχολεία) → {ef}')

        subj = (f'[TEST] {subj_tpl}'.replace('{school_year}', SCHOOL_YEAR)
                .replace('{school_name}', f'Συγκεντρωτικό ({len(schools)} σχολεία)'))
        body_txt = (
            f'Δοκιμαστική αποστολή — συνολικό αρχείο Excel (όλα τα σχολεία).\n\n'
            f'Σχολεία: {len(schools)}\n'
            + '\n'.join(f'  • {n}' for n in sorted(names))
        )

        with open(out_path, 'rb') as f:
            xls_bytes = f.read()

        msg = MIMEMultipart()
        msg['Subject'] = subj
        msg['From']    = formataddr((str(Header('ΔΙ.Π.Ε. Αν. Θεσ/νίκης', 'utf-8')), ef))
        msg['To']      = ef
        msg['Date']    = formatdate(localtime=True)
        msg.attach(MIMEText(body_txt, 'plain', 'utf-8'))
        part = MIMEBase('application',
                        'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(xls_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment',
                        filename=Header(os.path.basename(out_path), 'utf-8').encode())
        msg.attach(part)

        try:
            _smtp_send(config, ef, msg.as_string())
            log(f'  ✓ → {ef}  (συγκεντρωτικό, {len(schools)} σχολεία)')
            return [(n, ef) for n in names], []
        except Exception as e:
            log(f'  ✗ {e}')
            return [], [(n, ef) for n in names]

    # ── Κανονική αποστολή: ένα email ανά σχολείο (μόνο όσα έχουν αρχείο split) ─
    log(f'  → {len(schools)} σχολεία για αποστολή (βρέθηκαν στο φάκελο split)')
    sent, failed = [], []

    for code, name, email_to, file_path in schools:
        if not email_to:
            log(f'  ⚠ {name}: κενό email — παράλειψη')
            failed.append((name, '—'))
            continue

        subj     = subj_tpl.replace('{school_year}', SCHOOL_YEAR).replace('{school_name}', name)
        body_txt = body_tpl.replace('{school_year}', SCHOOL_YEAR).replace('{school_name}', name)
        try:
            with open(file_path, 'rb') as f:
                xls_bytes = f.read()
        except Exception as e:
            log(f'  ✗ {name}: δεν ήταν δυνατή η ανάγνωση του αρχείου split: {e}')
            failed.append((name, email_to))
            continue
        fname_att = os.path.basename(file_path)

        msg = MIMEMultipart()
        msg['Subject'] = subj
        msg['From']    = formataddr((str(Header('ΔΙ.Π.Ε. Αν. Θεσ/νίκης', 'utf-8')), ef))
        msg['To']      = formataddr((str(Header(name, 'utf-8')), email_to))
        msg['Date']    = formatdate(localtime=True)
        msg.attach(MIMEText(body_txt, 'plain', 'utf-8'))
        part = MIMEBase('application',
                        'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(xls_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment',
                        filename=Header(fname_att, 'utf-8').encode())
        msg.attach(part)

        try:
            _smtp_send(config, email_to, msg.as_string())
            log(f'  ✓ → {email_to}  [{name}]')
            sent.append((name, email_to))
        except Exception as e:
            log(f'  ✗ {name}: {e}')
            failed.append((name, email_to))

    if sent:
        _send_summary_oloimerou(config, sent, failed, today, log)

    return sent, failed


def _open_oloimerou_template_editor(parent, C):
    """Dialog επεξεργασίας θέματος & κειμένου email."""
    import tkinter as tk
    cur_subj, cur_body = _load_oloimerou_email_template()
    if cur_body == DEFAULT_EMAIL_BODY:
        try:
            import config as _cfg
            sig = getattr(_cfg, 'EMAIL_SIGNATURE', '').strip()
            if sig:
                cur_body = cur_body + '\n\n' + sig
        except Exception:
            pass

    ed = tk.Toplevel(parent)
    ed.title('Πρότυπο Email — Τμήματα Ολοημέρου')
    ed.configure(bg=C['bg'])
    ed.resizable(True, False)
    ed.grab_set()
    ed.transient(parent)

    epad = dict(padx=14, pady=5)
    tk.Label(ed, text='Θέμα:', bg=C['bg'], fg=C['hdr_bg'],
             font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **epad)
    subj_var = tk.StringVar(value=cur_subj)
    tk.Entry(ed, textvariable=subj_var, font=('Arial', 9),
             width=70).pack(fill='x', padx=14, pady=(0, 8))

    tk.Label(ed, text='Κείμενο email:', bg=C['bg'], fg=C['hdr_bg'],
             font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **epad)
    txt = tk.Text(ed, font=('Arial', 9), width=70, height=14,
                  wrap='word', relief='solid', bd=1)
    txt.pack(fill='x', padx=14, pady=(0, 4))
    txt.insert('1.0', cur_body)

    tk.Label(ed,
             text='Χρησιμοποιήστε {school_year} και {school_name} (αντικαθίστανται αυτόματα).',
             bg=C['bg'], fg=C['desc'], font=('Arial', 8),
             anchor='w').pack(fill='x', padx=14, pady=(0, 10))

    def _save():
        _save_oloimerou_email_template(subj_var.get().strip(), txt.get('1.0', 'end-1c'))
        ed.destroy()
        import tkinter.messagebox as _mb
        _mb.showinfo('Αποθήκευση', 'Το πρότυπο αποθηκεύτηκε.', parent=parent)

    def _reset():
        import tkinter.messagebox as _mb
        if _mb.askyesno('Επαναφορά', 'Να επανέλθει το προεπιλεγμένο κείμενο;', parent=ed):
            _save_oloimerou_email_template(DEFAULT_EMAIL_SUBJECT, DEFAULT_EMAIL_BODY)
            ed.destroy()

    br = tk.Frame(ed, bg=C['bg'])
    br.pack(pady=(0, 12))
    tk.Button(br, text='Αποθήκευση', bg=C['btn_bg'], fg=C['btn_fg'],
              font=('Arial', 9, 'bold'), relief='flat', padx=14, pady=5,
              cursor='hand2', command=_save).pack(side='left', padx=4)
    tk.Button(br, text='Επαναφορά προεπιλογής', bg=C['bg2'], fg=C['hdr_bg'],
              font=('Arial', 9), relief='flat', padx=14, pady=5,
              cursor='hand2', command=_reset).pack(side='left', padx=4)
    tk.Button(br, text='Άκυρο', bg=C['bg2'], fg=C['desc'],
              font=('Arial', 9), relief='flat', padx=14, pady=5,
              cursor='hand2', command=ed.destroy).pack(side='left', padx=4)

    ed.update_idletasks()
    x = parent.winfo_x() + (parent.winfo_width()  - ed.winfo_width())  // 2
    y = parent.winfo_y() + (parent.winfo_height() - ed.winfo_height()) // 2
    ed.geometry(f'+{x}+{y}')


def _show_results_dialog(config, df_nip, df_ds, today, out_path, summary_text):
    """
    Εμφανίζει σύνοψη αποτελεσμάτων και κουμπί για άνοιγμα του Excel.
    Τρέχει στο main thread.
    """
    import tkinter as tk
    from tkinter import scrolledtext

    root = tk._default_root
    if root is None:
        return

    win = tk.Toplevel(root)
    win.title(f'Αποτελέσματα — {CHECK_TITLE}')
    win.configure(bg='#FFF8E1')
    win.resizable(True, True)
    win.attributes('-topmost', True)
    # Το CheckRunDialog (γονικό παράθυρο πίσω από αυτό) κάνει ήδη grab_set()
    # στον εαυτό του — παίρνουμε το grab εδώ ώστε τα κουμπιά να δουλεύουν αμέσως.
    win.transient(root)
    win.grab_set()
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f'520x380+{sw//2-260}+{sh//2-190}')

    tk.Frame(win, bg='#E65100', pady=8).pack(fill='x')
    tk.Label(win.children[list(win.children)[-1]],
             text=f'⚠  {CHECK_TITLE}',
             bg='#E65100', fg='white',
             font=('Arial', 11, 'bold')).pack()

    txt = scrolledtext.ScrolledText(win, font=('Arial', 9), wrap='word',
                                     relief='flat', bg='#FFF8E1', height=10)
    txt.pack(fill='both', expand=True, padx=14, pady=8)
    txt.insert('1.0', summary_text)
    txt.config(state='disabled')

    btn_f = tk.Frame(win, bg='#FFF8E1')
    btn_f.pack(pady=(0, 12))

    def _open_excel():
        # os.startfile — subprocess.Popen(['start','',path], shell=True) ΔΕΝ
        # δουλεύει σωστά στα Windows.
        try:
            os.startfile(os.path.normpath(out_path))
        except Exception as e:
            import tkinter.messagebox as _mb
            _mb.showerror('Σφάλμα', f'Δεν ήταν δυνατό το άνοιγμα του αρχείου:\n{e}', parent=win)

    tk.Button(btn_f, text='📄 Άνοιγμα Excel',
              bg='#E65100', fg='white',
              font=('Arial', 9, 'bold'), relief='flat',
              padx=14, pady=5, cursor='hand2',
              command=_open_excel).pack(side='left', padx=4)
    tk.Button(btn_f, text='Κλείσιμο',
              bg='#E8EDF3', fg='#333333',
              font=('Arial', 9), relief='flat',
              padx=14, pady=5, cursor='hand2',
              command=win.destroy).pack(side='left', padx=4)


def _build_email_form(parent, config, df_nip, df_ds, today, out_path, C):
    """
    Χτίζει τη φόρμα αποστολής email απευθείας μέσα στο tab «✉ Αποστολή».
    Αντίγραφο ίδιας λογικής με tmimata_genikis._build_email_form.
    """
    import tkinter as tk
    from tkinter import scrolledtext
    import threading

    pad = dict(padx=14, pady=5)
    split_dir = _split_dir_for(out_path)

    tk.Button(parent, text='✉  Πρότυπο Email (Θέμα & Κείμενο)',
              bg=C['bg2'], fg=C['hdr_bg'], font=('Arial', 9),
              relief='flat', padx=10, pady=4, cursor='hand2',
              command=lambda: _open_oloimerou_template_editor(parent, C)
              ).pack(anchor='w', **pad)

    count_row = tk.Frame(parent, bg=C['bg'])
    count_row.pack(fill='x', **pad)
    count_lbl = tk.Label(count_row, text='', bg=C['bg'], fg=C['desc'],
                          font=('Arial', 8, 'italic'), anchor='w',
                          justify='left', wraplength=480)
    count_lbl.pack(side='left', fill='x', expand=True)

    def _update_count(*_args):
        schools   = _schools_from_split_dir(split_dir)
        with_e    = sum(1 for s in schools if s[2])
        without_e = len(schools) - with_e
        if not schools:
            count_lbl.config(
                text='⚠ Δεν βρέθηκαν αρχεία στο φάκελο «split» — τρέξε πρώτα '
                     'το tab «✂ Διαχωρισμός».', fg='#B00020')
        else:
            txt = f'→ {len(schools)} σχολεία θα λάβουν email (βρέθηκαν στο φάκελο split)'
            if without_e:
                txt += f'  ({without_e} χωρίς email)'
            count_lbl.config(text=txt, fg=C['desc'])

    tk.Button(count_row, text='↻ Ανανέωση', bg=C['bg2'], fg=C['desc'],
              font=('Arial', 8), relief='flat', padx=8, pady=2, cursor='hand2',
              command=_update_count).pack(side='right')

    _update_count()

    try:
        _nb = parent.master.master
        _nb.bind('<<NotebookTabChanged>>', _update_count, add='+')
    except Exception:
        pass

    mode_f = tk.Frame(parent, bg=C['bg'])
    mode_f.pack(fill='x', **pad)
    mode_var = tk.StringVar(value='schools')
    fe = getattr(config, 'FROM_EMAIL', '') or '...'
    tk.Radiobutton(mode_f, text='Αποστολή σε σχολεία (ένα email ανά σχολείο)',
                   variable=mode_var, value='schools',
                   bg=C['bg'], selectcolor=C['sel_bg'],
                   activebackground=C['bg'], font=('Arial', 9)).pack(anchor='w')
    tk.Radiobutton(mode_f, text=f'Test mode — όλα στο: {fe}',
                   variable=mode_var, value='test',
                   bg=C['bg'], selectcolor=C['sel_bg'],
                   activebackground=C['bg'], font=('Arial', 9)).pack(anchor='w')

    log_w = scrolledtext.ScrolledText(parent, height=10, font=('Courier', 8),
                                       wrap='word', relief='solid', bd=1,
                                       state='disabled')
    log_w.pack(fill='both', expand=True, padx=14, pady=8)

    def _log(msg):
        def _append():
            log_w.config(state='normal')
            log_w.insert('end', msg + '\n')
            log_w.see('end')
            log_w.config(state='disabled')
        try:
            log_w.after(0, _append)
        except Exception:
            pass

    btn_f = tk.Frame(parent, bg=C['bg'])
    btn_f.pack(fill='x', padx=14, pady=(0, 12))
    send_btn = tk.Button(btn_f, text='✉  Αποστολή',
                         bg=C['btn_bg'], fg=C['btn_fg'],
                         font=('Arial', 9, 'bold'), relief='flat',
                         padx=14, pady=5, cursor='hand2')
    send_btn.pack(side='left', padx=(0, 8))

    def _start():
        send_btn.config(state='disabled')
        dry_run       = (mode_var.get() == 'test')
        subj, body_tpl = _load_oloimerou_email_template()

        def _worker():
            try:
                sent, failed = _do_send_emails(
                    dry_run, subj, body_tpl, config, today, _log,
                    out_path=out_path, split_dir=split_dir)
                _log(f'\n{"─"*40}')
                _log(f'✓ Εστάλησαν: {len(sent)}   ✗ Αποτυχίες: {len(failed)}')
            except Exception as e:
                _log(f'✗ Σφάλμα: {e}')
            finally:
                try:
                    send_btn.after(0, lambda: send_btn.config(state='normal'))
                except Exception:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

    send_btn.config(command=_start)


_OLOIMEROU_PALETTE = {
    'bg': '#F5F7FA', 'bg2': '#E8EDF3', 'hdr_bg': '#1F4E79',
    'btn_bg': '#1F4E79', 'btn_fg': '#FFFFFF', 'desc': '#666666',
    'sel_bg': '#D6E4F0',
}


def build_send_tab(parent, config):
    """
    Χτίζει απευθείας μέσα στο `parent` (το body του tab «✉ Αποστολή») τη
    φόρμα αποστολής email, με βάση το αποτέλεσμα της τελευταίας επιτυχούς
    Εκτέλεσης (_LAST_RESULT — γεμίζει μέσα στο run()).
    """
    import tkinter as tk

    if not _LAST_RESULT or not _LAST_RESULT.get('out_path'):
        tk.Label(parent, text='Τρέξε πρώτα την «▶ Εκτέλεση» για να ενεργοποιηθεί η αποστολή.',
                 bg=_OLOIMEROU_PALETTE['bg'], fg=_OLOIMEROU_PALETTE['desc'],
                 font=('Arial', 9), anchor='w', justify='left',
                 wraplength=520).pack(fill='x', padx=14, pady=14)
        return

    _build_email_form(parent, config, _LAST_RESULT['df_nip'], _LAST_RESULT['df_ds'],
                       _LAST_RESULT['today'], _LAST_RESULT['out_path'], _OLOIMEROU_PALETTE)


CUSTOM_SEND_TAB = build_send_tab


# ── Tab «✂ Διαχωρισμός» — Excel ανά σχολείο ─────────────────────────────────
def _sanitize_filename(name):
    import re
    return re.sub(r'[^\w\s-]', '_', str(name)).strip() or 'σχολειο'


def _find_latest_output():
    """
    Βρίσκει το πιο πρόσφατο συγκεντρωτικό αρχείο αποτελεσμάτων
    (Documents/MySchoolChecks/results_*/tmimata_oloimerou/*_tmimata_oloimerou.xlsx).
    """
    docs = os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks')
    pattern = os.path.join(docs, 'results_*', RESULTS_FOLDER, f'*_{RESULTS_FOLDER}.xlsx')
    matches = [f for f in _glob.glob(pattern) if os.path.isfile(f)]
    if not matches:
        return None
    matches.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return matches[0]


def split_oloimerou_workbook(file_path, output_dir, log=print):
    """
    Διαβάζει το συγκεντρωτικό αρχείο αποτελεσμάτων (φύλλα «ΝΗΠ-...» / «ΔΣ-...»)
    και γράφει ένα ξεχωριστό, μορφοποιημένο Excel ανά σχολείο μέσα στο
    output_dir. Σε αντίθεση με το tmimata_genikis (που εξαιρεί σχολεία χωρίς
    απόκλιση), εδώ ΔΕΝ εφαρμόζεται κανένα φίλτρο — κάθε γραμμή του
    συγκεντρωτικού αρχείου ήδη πληροί το κριτήριο (χωρίς Αναστολή ΚΑΙ
    < 3 μαθητές), άρα κάθε σχολείο παίρνει ατομικό αρχείο.
    Επιστρέφει το πλήθος αρχείων που γράφτηκαν.
    """
    os.makedirs(output_dir, exist_ok=True)
    today = datetime.today()
    n_written = 0

    for sheet_name in pd.ExcelFile(file_path).sheet_names:
        if sheet_name.startswith('ΝΗΠ-'):
            col_defs, label = columns_nip(), 'Νηπιαγωγεία'
        elif sheet_name.startswith('ΔΣ-'):
            col_defs, label = columns_ds(), 'Δημοτικά'
        else:
            log(f'  ⚠ Άγνωστο φύλλο «{sheet_name}» — παράλειψη.')
            continue

        col_names = [c[0] for c in col_defs]
        # Οι 2 πρώτες γραμμές του φύλλου είναι τίτλος/υπότιτλος (merged
        # cells) — οι στήλες είναι στη γραμμή 3, άρα skiprows=2.
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=2)
        except Exception as e:
            log(f'  ✗ Σφάλμα ανάγνωσης φύλλου «{sheet_name}»: {e}')
            continue

        missing = [c for c in col_names if c not in df.columns]
        if missing:
            log(f'  ✗ Λείπουν στήλες στο «{sheet_name}»: {", ".join(missing)}')
            continue
        df = df[col_names]

        for _, row in df.iterrows():
            name = str(row.get('Ονομασία Σχολ. Μονάδας', '')).strip()
            if not name or name.lower() == 'nan':
                continue

            code = row.get('Κωδ. Υπουργείου Σχολείου', '')
            try:
                code = int(code)
            except Exception:
                pass

            row_df = pd.DataFrame([row])
            xls_bytes = _build_mini_excel(row_df, col_defs, label, today)
            fname = f'{code}_{_sanitize_filename(name)}.xlsx'
            with open(os.path.join(output_dir, fname), 'wb') as f:
                f.write(xls_bytes)
            n_written += 1
            log(f'  ✓ {fname}')

    log(f'\n✓ Διαχωρισμός ολοκληρώθηκε — {n_written} αρχεία → {output_dir}')
    return n_written


def build_split_tab(parent, config):
    """
    Χτίζει το tab «✂ Διαχωρισμός» απευθείας μέσα στο δοσμένο `parent`.
    Ανιχνεύει αυτόματα το πιο πρόσφατο αρχείο αποτελεσμάτων και το χωρίζει
    σε ένα Excel ανά σχολείο μέσα σε υποφάκελο «split».
    """
    import tkinter as tk
    from tkinter import scrolledtext
    import threading

    C = _OLOIMEROU_PALETTE
    pad = dict(padx=14, pady=5)

    tk.Label(parent,
             text='Παίρνει το πιο πρόσφατο συγκεντρωτικό αρχείο αποτελεσμάτων του '
                  'ελέγχου και το χωρίζει σε ξεχωριστά αρχεία Excel — ένα ανά '
                  'σχολείο — μέσα σε φάκελο «split» δίπλα στο αρχείο.\n'
                  'Όλα τα σχολεία που εμφανίζονται στο συγκεντρωτικό αρχείο ήδη '
                  'πληρούν το κριτήριο (χωρίς Αναστολή Ολοημέρου, < 3 μαθητές) — '
                  'δεν εφαρμόζεται επιπλέον φίλτρο εδώ. Στο tab «✉ Αποστολή» θα '
                  'σταλούν emails σε όλα αυτά τα σχολεία.',
             bg=C['bg'], fg=C['desc'], font=('Arial', 8),
             wraplength=560, justify='left', anchor='w').pack(fill='x', **pad)

    info_lbl = tk.Label(parent, text='', bg=C['bg'], fg=C['desc'],
                         font=('Arial', 8, 'italic'), anchor='w',
                         justify='left', wraplength=560)
    info_lbl.pack(fill='x', padx=14, pady=(0, 4))

    state = {'file': None}

    def _detect():
        f = _find_latest_output()
        state['file'] = f
        if f:
            info_lbl.config(text=f'✓ {os.path.basename(f)}', fg='#2E7D32')
        else:
            info_lbl.config(
                text='Δεν βρέθηκε αρχείο αποτελεσμάτων — τρέξε πρώτα την «▶ Εκτέλεση».',
                fg='#B00020')

    _detect()

    log_w = scrolledtext.ScrolledText(parent, height=10, font=('Courier', 8),
                                       wrap='word', relief='solid', bd=1,
                                       state='disabled')
    log_w.pack(fill='both', expand=True, padx=14, pady=8)

    def _log(msg):
        def _append():
            log_w.config(state='normal')
            log_w.insert('end', msg + '\n')
            log_w.see('end')
            log_w.config(state='disabled')
        try:
            log_w.after(0, _append)
        except Exception:
            pass

    btn_f = tk.Frame(parent, bg=C['bg'])
    btn_f.pack(fill='x', padx=14, pady=(0, 12))
    split_btn = tk.Button(btn_f, text='✂  Διαχωρισμός ανά Σχολείο',
                           bg=C['btn_bg'], fg=C['btn_fg'],
                           font=('Arial', 9, 'bold'), relief='flat',
                           padx=14, pady=5, cursor='hand2')
    split_btn.pack(side='left', padx=(0, 8))

    tk.Button(btn_f, text='Ανίχνευση αρχείου', bg=C['bg2'], fg=C['desc'],
              font=('Arial', 9), relief='flat', padx=10, pady=5, cursor='hand2',
              command=_detect).pack(side='left')

    def _start():
        if not state['file']:
            _detect()
        if not state['file']:
            import tkinter.messagebox as _mb
            _mb.showwarning(
                'Προσοχή',
                'Δεν βρέθηκε αρχείο αποτελεσμάτων.\nΤρέξε πρώτα την «▶ Εκτέλεση».',
                parent=parent)
            return

        split_btn.config(state='disabled')
        f = state['file']
        out_dir = os.path.join(os.path.dirname(f), 'split')

        def _worker():
            try:
                split_oloimerou_workbook(f, out_dir, log=_log)
            except Exception as e:
                _log(f'✗ Σφάλμα: {e}')
            finally:
                try:
                    split_btn.after(0, lambda: split_btn.config(state='normal'))
                except Exception:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

    split_btn.config(command=_start)


CUSTOM_SPLIT_TAB = build_split_tab


# ── Λήψη δεδομένων (μόνο μέσω tab «⬇ Λήψη») ──────────────────────────────────
def _download_rids():
    """Τα report-ids που κατεβάζει το _download_inputs."""
    import core.downloader as _dl
    orig_22 = next((r for r in _dl.REPORTS if r[0] == '2.2'), None)
    return ['5.3', '5.4'] + (['2.2'] if orig_22 else [])


def _download_target_dir():
    """Ο φάκελος όπου κατεβαίνουν 5.3/5.4/2.2 — ο κοινός φάκελος λήψεων της
    ημέρας (downloads/{today}). Δεν χρειάζεται υποφάκελος σχολικού έτους
    (σε αντίθεση με το tmimata_genikis) αφού το 5.3/5.4 δεν χρειάζονται
    override έτους — ήδη κατεβαίνουν με το 2026-2027 από προεπιλογή."""
    import core.downloader as _dl
    return _dl.get_downloads_dir(
        os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks'))


def _download_target_info():
    """Επιστρέφει (dest_dir, rids) — καλείται από core/check_dialog.py ΠΡΙΝ
    τη λήψη, ώστε να ελεγχθεί αν υπάρχουν ήδη αρχεία σήμερα."""
    return _download_target_dir(), _download_rids()


CUSTOM_DOWNLOAD_INFO = _download_target_info


def _download_inputs(config, log=print, force=None):
    """
    Κατεβάζει 5.3, 5.4 (+2.2 αν είναι ρυθμισμένο) — καμία αλλαγή σχολικού
    έτους δεν χρειάζεται.

    `force`: λίστα από report-ids προς αναγκαστική επανάληψη. `None`
    (προεπιλογή) σημαίνει «όλα».

    Επιστρέφει (path_53, path_54, path_22) — None όπου δεν κατέβηκε.
    """
    import core.downloader as _dl

    ms_user = getattr(config, 'MYSCHOOL_USER', '').strip()
    ms_pass = getattr(config, 'MYSCHOOL_PASS', '').strip()
    if not ms_user or not ms_pass:
        raise RuntimeError('Συμπλήρωσε username και κωδικό MySchool στις Ρυθμίσεις (⚙).')

    rids = _download_rids()
    force_list = list(rids) if force is None else list(force)

    dest_dir = _download_target_dir()
    os.makedirs(dest_dir, exist_ok=True)

    dl = _dl.MySchoolDownloader(
        username=ms_user, password=ms_pass, dest_dir=dest_dir,
        callback=log, reports=rids,
        browser=getattr(config, 'BROWSER', 'chrome'),
        force=force_list,
    )
    results = dl.run()

    return results.get('5.3'), results.get('5.4'), results.get('2.2')


# Public alias — το tab «⬇ Λήψη» (core/check_dialog.py) καλεί αυτή τη
# συνάρτηση αντί για τον γενικό μηχανισμό λήψης.
download = _download_inputs
CUSTOM_DOWNLOAD = download


def _find_downloaded_inputs():
    """
    Ψάχνει για 5.3/5.4 (+2.2) που έχουν ΗΔΗ κατέβει σήμερα μέσω του tab
    «⬇ Λήψη» (βλ. download() / CUSTOM_DOWNLOAD) — ΔΕΝ κατεβάζει τίποτα.
    Επιστρέφει (path_53, path_54, path_22), None όπου δεν βρέθηκε.
    """
    from core.downloader import FILE_PREFIX_MAP

    dest_dir = _download_target_dir()
    if not os.path.isdir(dest_dir):
        return None, None, None

    def _match(rid):
        prefix  = FILE_PREFIX_MAP.get(rid, rid)
        matches = [f for f in _glob.glob(os.path.join(dest_dir, f'{prefix}*'))
                   if not f.endswith(('.tmp', '.crdownload'))]
        return matches[0] if matches else None

    return _match('5.3'), _match('5.4'), _match('2.2')


# ── CUSTOM RUN ────────────────────────────────────────────────────────────
def run(config):
    import core.framework as _fw
    _fw._current_check_title = CHECK_TITLE

    print('=' * 65)
    print(f'  {CHECK_TITLE}')
    print('=' * 65)

    path_53, path_54, path_22 = _find_downloaded_inputs()

    missing = [rid for rid, p in (('5.3', path_53), ('5.4', path_54)) if not p]
    if missing:
        import tkinter.messagebox as _mb
        _mb.showwarning(
            'Λείπει η λήψη',
            f'Δεν έχουν κατέβει σήμερα: {", ".join(missing)}.\n\n'
            f'Πήγαινε πρώτα στο tab «⬇ Λήψη» και πάτησε το κουμπί λήψης.'
        )
        return

    print(f'\n  ✓ Χρήση ήδη κατεβασμένων αρχείων (σχολικό έτος {SCHOOL_YEAR}).')
    print('-' * 65)

    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    print(f'\n  Ημερομηνία : {today.strftime("%d/%m/%Y")}')
    print(f'  Κριτήριο   : χωρίς Αναστολή Ολοημέρου ΚΑΙ < {THRESHOLD} μαθητές σε Ολοήμερο')
    print('-' * 65)

    if path_22:
        print(f'  ✓ stat2_2: {os.path.basename(path_22)}')
    else:
        print('  ℹ stat2_2 δεν κατέβηκε — Email Σχολείου θα είναι κενό')

    print('\nΕπεξεργασία...')
    try:
        df_nip, df_ds, perif = process(path_53, path_54, path_22=path_22)
    except Exception as e:
        import tkinter.messagebox as _mb
        _mb.showerror('Σφάλμα επεξεργασίας', str(e))
        return

    global _LAST_RESULT
    _LAST_RESULT = {'df_nip': df_nip, 'df_ds': df_ds, 'today': today}

    print(f'  ✓ Νηπιαγωγεία : {len(df_nip)} σχολεία')
    print(f'  ✓ Δημοτικά    : {len(df_ds)} σχολεία')

    if df_nip.empty and df_ds.empty:
        import tkinter as tk
        import tkinter.messagebox as _mb
        _root = tk._default_root
        if _root is not None:
            _root.after(0, lambda: _mb.showinfo(
                CHECK_TITLE, '✓  Δεν βρέθηκαν σχολεία που να πληρούν τα κριτήρια.'))
        return

    _docs   = os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks')
    out_dir = os.path.join(_docs, f'results_{today.strftime("%Y%m%d")}', RESULTS_FOLDER)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_{RESULTS_FOLDER}.xlsx')

    try:
        build_workbook(df_nip, df_ds, perif, today, out_path)
        print(f'\n  ✓ Αποθηκεύτηκε: {os.path.basename(out_path)}')
        _LAST_RESULT['out_path'] = out_path
    except PermissionError:
        import tkinter.messagebox as _mb
        _mb.showwarning(
            'Αρχείο ανοιχτό',
            f'Το αρχείο {os.path.basename(out_path)} είναι ανοιχτό σε άλλο πρόγραμμα.\n'
            f'Κλείστε το και τρέξτε ξανά τον έλεγχο.'
        )
        return

    body = (
        f'Σύνοψη ελέγχου τμημάτων ολοημέρου — {today.strftime("%d/%m/%Y")}\n'
        f'{"─"*50}\n'
        f'Κριτήριο: χωρίς Αναστολή Ολοημέρου ΚΑΙ < {THRESHOLD} μαθητές σε Ολοήμερο\n\n'
        f'Νηπιαγωγεία: {len(df_nip)} σχολεία\n'
        f'Δημοτικά:    {len(df_ds)} σχολεία\n\n'
        f'{"─"*50}\n'
        f'Αποτελέσματα αποθηκεύτηκαν στο φάκελο:\n{out_dir}'
    )
    import tkinter as tk
    _root = tk._default_root
    if _root is not None:
        _summary = body
        _root.after(0, lambda: _show_results_dialog(
            config, df_nip, df_ds, today, out_path, _summary))
