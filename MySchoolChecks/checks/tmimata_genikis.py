"""
checks/tmimata_genikis.py
═════════════════════════
Έλεγχος τμημάτων γενικής παιδείας / δυναμικού — πλήρης αποτύπωση (2026-2027).

Πηγές (κατεβαίνουν ΑΠΕΥΘΕΙΑΣ μέσα στον έλεγχο — δες _download_inputs):
  3.1 (stat3_1) — Κατανομή μαθητών ανά τάξη. 2 γραμμές header (η 2η μόνο για τις
                   στήλες Αγόρια/Κορίτσια/Σύνολο, λόγω merged cell στο excel).
                   Στήλες (θέση): 2=Είδος Σχολείου, 4=Κωδικός Υπ., 10=Τάξη,
                   11=Αριθμός Τμημάτων, 14=Σύνολο μαθητών.
  5.3 (stat5_3) — Αποτύπωση νηπιαγωγείων: Τύπος Σχολείου, Κωδ. Υπουργείου Σχολείου,
                   Ονομασία Σχολ. Μονάδας, Λειτουργικότητα, Οργανικότητα, Μαθητές.
  5.4 (stat5_4) — Αποτύπωση δημοτικών: ίδιες στήλες με το 5.3.

Γιατί η λήψη γίνεται μέσα στον έλεγχο (και όχι μέσω του γενικού «⬇ Λήψη Δεδομένων»):
  Και τα 3 στατιστικά πρέπει να αντληθούν με το σχολικό έτος 2026-2027 ενεργό στο
  MySchool (το 3.1 κανονικά αντλείται με το τρέχον έτος — το χρησιμοποιούν και άλλα
  εργαλεία, π.χ. το «Σχολικές Μονάδες» — γι' αυτό εδώ γίνεται ένα τοπικό override
  μόνο για αυτή τη λήψη, χωρίς να πειράζεται η καθολική ρύθμιση στο downloader.py).
  Η αλλαγή έτους γίνεται αυτόματα στην αρχή (Default.aspx → cmbActiveAcadYear).

Λογική:
  Για κάθε σχολείο (πηγή: 5.3/5.4) υπολογίζεται, ανά τάξη, το άθροισμα Τμημάτων
  και Μαθητών από το 3.1 (Νηπιαγωγεία: μία ενιαία στήλη «ΠΡΟΝΗΠΙΑ-ΝΗΠΙΑ» — αθροίζει
  όλες τις σχετικές τάξεις π.χ. ΝΗΠΙΑ/ΠΡΟΝΗΠΙΑ/ΠΡΟΝΗΠΙΑ-ΝΗΠΙΑ. Δημοτικά: Α-ΣΤ).

  Άθροισμα Τμήματα  = Σ(Τμήματα ανά τάξη)
  Άθροισμα Μαθητών  = Σ(Μαθητές ανά τάξη)
  Διαφορά Τμήματα   = Λειτουργικότητα − Άθροισμα Τμήματα
  Διαφορά Μαθητές   = Ενεργοί Μαθητές (5.3/5.4 «Μαθητές») − Άθροισμα Μαθητών

Αποτέλεσμα: 1 αρχείο Excel με 2 φύλλα (Δημοτικά, Νηπιαγωγεία) — ΟΛΑ τα σχολεία,
με πράσινο/κόκκινο χρωματισμό στις στήλες Διαφορά (0 = πράσινο, ≠0 = κόκκινο).
Καμία αποστολή email.
"""

import glob as _glob
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.framework import _show_results_popup

# ── Μεταδεδομένα ────────────────────────────────────────────────────────────
CHECK_TITLE       = 'Έλεγχος Τμημάτων Γενικής Παιδείας / Δυναμικού'
CHECK_DESCRIPTION = 'Σύγκριση Λειτουργικότητας με πραγματικά τμήματα και μαθητές ανά σχολείο.'
RESULTS_FOLDER    = 'tmimata_genikis'
HAS_EMAIL         = False
CUSTOM_RUN        = True
SCHOOL_YEAR       = '2026-2027'
REQUIRED_REPORTS  = [
    '3.1 — Κατανομή μαθητών ανά τάξη',
    '5.3 — Αποτύπωση νηπιαγωγείων',
    '5.4 — Αποτύπωση δημοτικών',
]

# Αποτέλεσμα της τελευταίας επιτυχούς Εκτέλεσης (df_ds/df_nip/today) — το
# χρησιμοποιεί το open_email_from_last_result() παρακάτω, ώστε το tab
# «✉ Αποστολή» του CheckRunDialog να μπορεί να ανοίξει το email dialog χωρίς
# να χρειάζεται να περάσει από το παράθυρο Αποτελεσμάτων.
_LAST_RESULT = None

DEFAULT_EMAIL_SUBJECT = (
    'Τμήματα Γενικής Παιδείας {school_year} — αποκλίσεις MySchool'
)
DEFAULT_EMAIL_BODY = (
    'Καλημέρα,\n\n'
    'Κατά τον έλεγχο των στοιχείων τμημάτων γενικής παιδείας / δυναμικού '
    'για το σχολικό έτος {school_year} στο MySchool, εντοπίστηκαν αποκλίσεις '
    'μεταξύ Λειτουργικότητας και καταχωρημένων τμημάτων/μαθητών '
    'για το σχολείο σας.\n\n'
    'Παρακαλούμε ελέγξτε το συνημμένο αρχείο και προβείτε στις απαραίτητες '
    'διορθώσεις στο MySchool.\n\n'
    'Παρακαλούμε για τις ενέργειές σας.'
)

GRADES_DS  = ['Α', 'Β', 'Γ', 'Δ', 'Ε', 'ΣΤ']          # Δημοτικά
NIP_GROUP  = 'ΠΡΟΝΗΠΙΑ-ΝΗΠΙΑ'                          # Νηπιαγωγεία — ενιαία στήλη

COLOR_HDR   = '1F4E79'
COLOR_SUB   = 'D6E4F0'
COLOR_ALT   = 'EBF3FB'
COLOR_OK    = '92D050'   # πράσινο — διαφορά 0
COLOR_DEV   = 'FF0000'   # κόκκινο — διαφορά ≠ 0
DEFAULT_PERIFEREIA = 'ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ'

_LEFT_ALIGN_COLS = {'Τύπος', 'Ονομασία', 'Δήμος', 'Email Σχολείου'}


# ── Φόρτωση 3.1 ─────────────────────────────────────────────────────────────
def _find_col(header, *names):
    """Επιστρέφει το index της πρώτης στήλης του header που ταιριάζει με κάποιο name."""
    for i, v in enumerate(header):
        if str(v).strip() in names:
            return i
    return None


def _load_stat31(path):
    """
    Φορτώνει το 3.1 και επιστρέφει DataFrame με στήλες
    _eidos, _code (int), _taxi, _tmimata (int), _mathites (int).

    Το αρχείο έχει 2 γραμμές header (η 2η μόνο για Αγόρια/Κορίτσια/Σύνολο,
    λόγω merged cell) — εντοπίζεται δυναμικά η γραμμή με «Τάξη». Οι στήλες
    εντοπίζονται με βάση το ΟΝΟΜΑ τους (όχι σταθερή θέση), γιατί ο αριθμός/η
    σειρά των στηλών αλλάζει ανάλογα με τις επιλογές ομαδοποίησης που
    τσεκαρίστηκαν κατά τη λήψη (π.χ. αν λείπει η «Τύπος Σχολείου»).
    """
    raw = pd.read_excel(path, header=None)

    hdr_row = 0
    for i in range(min(5, len(raw))):
        if raw.iloc[i].astype(str).str.strip().eq('Τάξη').any():
            hdr_row = i
            break

    header = raw.iloc[hdr_row].astype(str).str.strip().tolist()

    eidos_idx   = _find_col(header, 'Είδος Σχολείου')
    code_idx    = _find_col(header, 'Κωδικός Υπ.', 'Κωδικός Υπουργείου Σχολείου', 'Κωδικός Υπουργείου')
    taxi_idx    = _find_col(header, 'Τάξη')
    tmim_idx    = _find_col(header, 'Αριθμός Τμημάτων')
    plithos_idx = _find_col(header, 'Πλήθος Ενεργών Μαθητών')

    missing = [n for n, v in (
        ('Είδος Σχολείου', eidos_idx), ('Κωδικός Υπ.', code_idx), ('Τάξη', taxi_idx),
        ('Αριθμός Τμημάτων', tmim_idx), ('Πλήθος Ενεργών Μαθητών', plithos_idx),
    ) if v is None]
    if missing:
        raise ValueError(
            f'Το 3.1 δεν έχει τις αναμενόμενες στήλες ({", ".join(missing)}).\n'
            f'Στήλες αρχείου: {header}'
        )

    # «Πλήθος Ενεργών Μαθητών» είναι merged header πάνω από 3 υποστήλες
    # (Αγόρια, Κορίτσια, Σύνολο) — το Σύνολο είναι η 3η από αυτές.
    sum_idx = plithos_idx + 2

    data = raw.iloc[hdr_row + 2:].reset_index(drop=True)

    dimos_idx = _find_col(header, 'Δήμος')
    if dimos_idx is None and len(header) > 7:
        dimos_idx = 7  # fallback: absolute position

    df = pd.DataFrame({
        '_eidos':    data[eidos_idx].astype(str).str.strip(),
        '_code':     pd.to_numeric(data[code_idx], errors='coerce'),
        '_taxi':     data[taxi_idx].astype(str).str.strip(),
        '_tmimata':  pd.to_numeric(data[tmim_idx], errors='coerce').fillna(0).astype(int),
        '_mathites': pd.to_numeric(data[sum_idx], errors='coerce').fillna(0).astype(int),
        '_dimos':    data[dimos_idx].astype(str).str.strip() if dimos_idx is not None else '',
    })
    df = df.dropna(subset=['_code'])
    df['_code'] = df['_code'].astype(int)
    return df


def _find_stat22():
    """Αυτόματη εύρεση stat2_2 από τον φάκελο λήψεων της εφαρμογής ή ~/Downloads."""
    docs = os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks')
    folders = []
    dl_base = os.path.join(docs, 'downloads')
    if os.path.isdir(dl_base):
        folders += sorted(
            [os.path.join(dl_base, d) for d in os.listdir(dl_base)
             if os.path.isdir(os.path.join(dl_base, d))],
            reverse=True
        )
    folders.append(os.path.join(os.path.expanduser('~'), 'Downloads'))
    for folder in folders:
        for pattern in ('stat2_2*.csv', 'stat2_2*.xlsx', 'stat2_2*.xls', 'CSV_*.zip'):
            matches = [f for f in _glob.glob(os.path.join(folder, pattern))
                       if not f.endswith('.tmp') and not f.endswith('.crdownload')]
            if matches:
                return sorted(matches)[-1]
    return None


def _load_email_lookup(path_22):
    """
    Διαβάζει stat2_2 (CSV με 1-column shift) και επιστρέφει {code_int: email}.
    col11 = Κωδ. ΥΠΠΘ, col18 = e-mail σχολείου.
    """
    import re
    lower = path_22.lower()
    if lower.endswith('.xlsx') or lower.endswith('.xls'):
        df = pd.read_excel(path_22, dtype=str)
    else:
        import zipfile, io
        if lower.endswith('.zip'):
            with zipfile.ZipFile(path_22) as z:
                raw = z.read(z.namelist()[0])
        else:
            with open(path_22, 'rb') as f:
                raw = f.read()
        text = raw.decode('cp1253')
        df = pd.read_csv(io.StringIO(text), sep=';', dtype=str, header=0)

    if len(df.columns) <= 18:
        return {}

    c_code  = df.columns[11]
    c_email = df.columns[18]

    def _clean_code(val):
        s = str(val).strip().strip('"').lstrip('=').strip('"').strip()
        s = re.sub(r'\.0$', '', s)
        return s.lstrip('0') or s

    lookup = {}
    for _, row in df.iterrows():
        raw_code = _clean_code(row[c_code])
        email    = str(row[c_email]).strip() if pd.notna(row[c_email]) else ''
        if email.lower() in ('nan', 'none', ''):
            email = ''
        if raw_code.isdigit():
            lookup[int(raw_code)] = email
    return lookup


def _dimos_lookup(df31):
    """dict {code: dimos} — ένας Δήμος ανά κωδικό σχολείου."""
    sub = df31.drop_duplicates(subset=['_code'])[['_code', '_dimos']]
    return {int(r['_code']): str(r['_dimos']).strip() for _, r in sub.iterrows()}


def _grade_lookup_ds(df31):
    """dict {code: {grade: (τμήματα, μαθητές)}} — Δημοτικά (Α-ΣΤ)."""
    sub = df31[df31['_eidos'] == 'Δημοτικά Σχολεία']
    g = sub.groupby(['_code', '_taxi'])[['_tmimata', '_mathites']].sum().reset_index()
    lookup = {}
    for _, r in g.iterrows():
        lookup.setdefault(int(r['_code']), {})[r['_taxi']] = (int(r['_tmimata']), int(r['_mathites']))
    return lookup


def _grade_lookup_nip(df31):
    """dict {code: (τμήματα, μαθητές)} — Νηπιαγωγεία, όλες οι τάξεις αθροισμένες."""
    sub = df31[df31['_eidos'] == 'Νηπιαγωγεία']
    g = sub.groupby('_code')[['_tmimata', '_mathites']].sum().reset_index()
    return {int(r['_code']): (int(r['_tmimata']), int(r['_mathites'])) for _, r in g.iterrows()}


# ── Στήλες εξόδου ─────────────────────────────────────────────────────────
def _base_cols():
    return [
        ('Τύπος',             42),
        ('Κωδ. ΥΠΠΘ',         14),
        ('Ονομασία',          40),
        ('Δήμος',             20),
        ('Email Σχολείου',    32),
        ('Λειτουργικότητα',   14),
        ('Οργανικότητα',      14),
        ('Ενεργοί Μαθητές',   14),
    ]

def _tail_cols():
    return [
        ('Άθροισμα Μαθητών',  16),
        ('Άθροισμα Τμήματα',  16),
        ('Διαφορά Τμήματα',   16),
        ('Διαφορά Μαθητές',   16),
    ]

def columns_ds():
    grade_cols = []
    for g in GRADES_DS:
        grade_cols += [(f'{g} Μαθητές', 10), (f'{g} Τμήματα', 10)]
    return _base_cols() + grade_cols + _tail_cols()

def columns_nip():
    grade_cols = [(f'{NIP_GROUP} Μαθητές', 18), (f'{NIP_GROUP} Τμήματα', 18)]
    return _base_cols() + grade_cols + _tail_cols()


# ── Χτίσιμο DataFrame ανά τύπο σχολείου ─────────────────────────────────────
def _build_records(df_src, grade_lookup, multi_grade, dimos_lookup=None, email_lookup=None):
    """
    df_src        : DataFrame από 5.3 ή 5.4
    grade_lookup  : dict {code: {grade: (tm, ma)}}  (Δημοτικά) ή
                     dict {code: (tm, ma)}           (Νηπιαγωγεία)
    multi_grade   : True → Δημοτικά (πολλαπλές τάξεις Α-ΣΤ)
                     False → Νηπιαγωγεία (μία ενιαία στήλη)
    dimos_lookup  : dict {code: dimos} από το 3.1 (προαιρετικό)
    email_lookup  : dict {code: email} από το 2.2 (προαιρετικό)
    """
    if dimos_lookup is None:
        dimos_lookup = {}
    if email_lookup is None:
        email_lookup = {}
    records = []
    for _, row in df_src.iterrows():
        code  = int(row['Κωδ. Υπουργείου Σχολείου'])
        leit  = int(row['Λειτουργικότητα']) if pd.notna(row['Λειτουργικότητα']) else 0
        org   = int(row['Οργανικότητα'])    if pd.notna(row['Οργανικότητα'])    else 0
        energ = int(row['Μαθητές'])         if pd.notna(row['Μαθητές'])         else 0

        rec = {
            'Τύπος':             str(row['Τύπος Σχολείου']).strip(),
            'Κωδ. ΥΠΠΘ':         code,
            'Ονομασία':          str(row['Ονομασία Σχολ. Μονάδας']).strip(),
            'Δήμος':             dimos_lookup.get(code, ''),
            'Email Σχολείου':    email_lookup.get(code, ''),
            'Λειτουργικότητα':   leit,
            'Οργανικότητα':      org,
            'Ενεργοί Μαθητές':   energ,
        }

        sum_tm = sum_ma = 0
        if multi_grade:
            per_school = grade_lookup.get(code, {})
            for g in GRADES_DS:
                tm, ma = per_school.get(g, (0, 0))
                rec[f'{g} Μαθητές'] = ma
                rec[f'{g} Τμήματα'] = tm
                sum_tm += tm
                sum_ma += ma
        else:
            tm, ma = grade_lookup.get(code, (0, 0))
            rec[f'{NIP_GROUP} Μαθητές'] = ma
            rec[f'{NIP_GROUP} Τμήματα'] = tm
            sum_tm += tm
            sum_ma += ma

        rec['Άθροισμα Μαθητών'] = sum_ma
        rec['Άθροισμα Τμήματα'] = sum_tm
        rec['Διαφορά Τμήματα']  = leit  - sum_tm
        rec['Διαφορά Μαθητές']  = energ - sum_ma
        records.append(rec)

    df_out = pd.DataFrame(records)
    if not df_out.empty:
        df_out = df_out.sort_values('Ονομασία').reset_index(drop=True)
    return df_out


def _periferia_name(df_src):
    try:
        raw = str(df_src['Περιφέρεια'].dropna().iloc[0])
        if 'ΕΚΠ/ΣΗΣ ' in raw:
            return raw.split('ΕΚΠ/ΣΗΣ ')[-1].strip()
    except Exception:
        pass
    return DEFAULT_PERIFEREIA


# ── Λογική ──────────────────────────────────────────────────────────────────
def process(path_31, path_53, path_54, path_22=None):
    """Επιστρέφει (df_ds, df_nip, periфereia_name)."""
    df31 = _load_stat31(path_31)
    df53 = pd.read_excel(path_53)
    df54 = pd.read_excel(path_54)

    dimos = _dimos_lookup(df31)
    email = _load_email_lookup(path_22) if path_22 else {}
    df_ds  = _build_records(df54, _grade_lookup_ds(df31),  multi_grade=True,  dimos_lookup=dimos, email_lookup=email)
    df_nip = _build_records(df53, _grade_lookup_nip(df31), multi_grade=False, dimos_lookup=dimos, email_lookup=email)

    perif = _periferia_name(df54) if not df54.empty else _periferia_name(df53)
    return df_ds, df_nip, perif


# ── Excel ─────────────────────────────────────────────────────────────────
def _brd():
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)


def _write_sheet(ws, df, title, columns, today, subtitle_extra=''):
    brd = _brd()
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    ncols = len(columns)

    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws['A1'] = f'{title}  —  {today.strftime("%d/%m/%Y")}{subtitle_extra}'
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', start_color=COLOR_HDR)
    ws['A1'].alignment = ctr
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
    ws['A2'] = f'Σύνολο σχολείων: {len(df)}'
    ws['A2'].font      = Font(name='Arial', italic=True, size=9)
    ws['A2'].fill      = PatternFill('solid', start_color=COLOR_SUB)
    ws['A2'].alignment = ctr
    ws.row_dimensions[2].height = 16

    for ci, (name, width) in enumerate(columns, 1):
        c = ws.cell(row=3, column=ci, value=name)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill      = PatternFill('solid', start_color=COLOR_HDR)
        c.alignment = ctr
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 28

    diff_cols = {'Διαφορά Τμήματα', 'Διαφορά Μαθητές'}
    col_keys  = [c[0] for c in columns]

    for ri, (_, row) in enumerate(df.iterrows(), start=4):
        base_fill = PatternFill('solid', start_color=COLOR_ALT) if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(col_keys, 1):
            val = row.get(key, '')
            if hasattr(val, 'item'):
                val = val.item()
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = brd
            if key in diff_cols:
                is_dev = val != 0
                c.font = Font(name='Arial', size=9, bold=True,
                               color='FFFFFF' if is_dev else '000000')
                c.fill = PatternFill('solid', start_color=COLOR_DEV if is_dev else COLOR_OK)
            else:
                c.font = Font(name='Arial', size=9)
                c.fill = base_fill
            c.alignment = lft if key in _LEFT_ALIGN_COLS else ctr
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}3'


def build_workbook(df_ds, df_nip, perif, today, out_path):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = f'ΔΣ-{perif}'[:31]
    _write_sheet(ws1, df_ds, CHECK_TITLE, columns_ds(), today, subtitle_extra='  |  Δημοτικά')

    ws2 = wb.create_sheet(f'ΝΗΠ-{perif}'[:31])
    _write_sheet(ws2, df_nip, CHECK_TITLE, columns_nip(), today, subtitle_extra='  |  Νηπιαγωγεία')

    wb.save(out_path)


# ── Email — Βοηθητικά ─────────────────────────────────────────────────────────
def _app_base():
    import sys
    if getattr(sys, 'frozen', False):
        lad = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'MySchoolChecks')
        return lad if os.path.isdir(lad) else os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _load_tmimata_email_template():
    import json
    spath = os.path.join(_app_base(), 'data', 'local_settings.json')
    try:
        with open(spath, encoding='utf-8') as f:
            s = json.load(f)
        t = s.get('tmimata_email', {})
        return t.get('subject', DEFAULT_EMAIL_SUBJECT), t.get('body', DEFAULT_EMAIL_BODY)
    except Exception:
        return DEFAULT_EMAIL_SUBJECT, DEFAULT_EMAIL_BODY


def _save_tmimata_email_template(subject, body):
    import json
    spath = os.path.join(_app_base(), 'data', 'local_settings.json')
    os.makedirs(os.path.dirname(spath), exist_ok=True)
    try:
        with open(spath, encoding='utf-8') as f:
            s = json.load(f)
    except Exception:
        s = {}
    s['tmimata_email'] = {'subject': subject, 'body': body}
    with open(spath, 'w', encoding='utf-8') as f:
        json.dump(s, f, ensure_ascii=False, indent=2)


def _build_mini_excel(row_df, col_defs, label, today):
    """Mini Excel (bytes) για ένα σχολείο."""
    import io as _io
    buf = _io.BytesIO()
    wb  = Workbook()
    ws  = wb.active
    ws.title = label[:31]
    _write_sheet(ws, row_df, CHECK_TITLE, col_defs, today,
                 subtitle_extra=f'  |  {label}')
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _split_dir_for(out_path):
    """Ο φάκελος «split» που αντιστοιχεί σε ένα συγκεντρωτικό αρχείο αποτελεσμάτων."""
    return os.path.join(os.path.dirname(out_path), 'split')


def _schools_from_split_dir(split_dir):
    """
    Διαβάζει κάθε αρχείο Excel μέσα στον φάκελο «split» (βλ.
    split_tmimata_workbook) και επιστρέφει list of (code, name, email,
    file_path) — ένα ανά σχολείο.

    Αυτός είναι ο ΜΟΝΑΔΙΚΟΣ κατάλογος παραληπτών για την κανονική αποστολή:
    στέλνεται email ΜΟΝΟ σε σχολεία που έχουν ήδη ατομικό αρχείο εδώ (δηλ.
    που πέρασαν το φίλτρο αποκλίσεων στο tab «✂ Διαχωρισμός»).
    """
    result = []
    if not split_dir or not os.path.isdir(split_dir):
        return result
    for fn in sorted(os.listdir(split_dir)):
        if not fn.lower().endswith('.xlsx'):
            continue
        fp = os.path.join(split_dir, fn)
        name, email, code = fn, '', ''
        try:
            df = pd.read_excel(fp, skiprows=2)
            if not df.empty:
                row   = df.iloc[0]
                name  = str(row.get('Ονομασία', '')).strip() or fn
                email = str(row.get('Email Σχολείου', '')).strip()
                if email.lower() in ('nan', 'none'):
                    email = ''
                code = row.get('Κωδ. ΥΠΠΘ', '')
                try:
                    code = int(code)
                except Exception:
                    pass
        except Exception:
            pass
        result.append((code, name, email, fp))
    return result


def _smtp_send(config, to_addr, msg_str):
    """
    Αποστολή email — 3 προσπάθειες όπως ακριβώς στο framework.py:
    1) SMTP_SSL 465 με πλήρες cert
    2) SMTP_SSL 465 χωρίς cert check
    3) SMTP STARTTLS 587 χωρίς cert check
    """
    import ssl, smtplib

    def _lenient():
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode    = ssl.CERT_NONE
        return ctx

    ef = config.FROM_EMAIL
    pw = config.FROM_PASSWORD
    sh = config.SMTP_HOST

    sent = False
    try:
        with smtplib.SMTP_SSL(sh, 465, context=ssl.create_default_context()) as s:
            s.login(ef, pw)
            s.sendmail(ef, to_addr, msg_str)
            sent = True
    except Exception:
        pass
    if not sent:
        try:
            with smtplib.SMTP_SSL(sh, 465, context=_lenient()) as s:
                s.login(ef, pw)
                s.sendmail(ef, to_addr, msg_str)
                sent = True
        except Exception:
            pass
    if not sent:
        with smtplib.SMTP(sh, 587) as s:
            s.starttls(context=_lenient())
            s.login(ef, pw)
            s.sendmail(ef, to_addr, msg_str)


def _send_summary_tmimata(config, sent, failed, today, log):
    """Summary email στο FROM_EMAIL μετά την ολοκλήρωση."""
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.utils import formataddr, formatdate
    from email.header import Header

    ef = getattr(config, 'FROM_EMAIL',    '')
    pw = getattr(config, 'FROM_PASSWORD', '')
    if not ef or not pw:
        return

    sent_lines   = '\n'.join(f'  ✓  {n}  ({e})' for n, e in sorted(sent))
    failed_lines = ('\n'.join(f'  ✗  {n}  ({e})' for n, e in sorted(failed))
                    if failed else '  —')
    body = (
        f'Αποστολή email Τμημάτων Γενικής Παιδείας — {today.strftime("%d/%m/%Y")}\n\n'
        f'Εστάλησαν: {len(sent)}\n{sent_lines}\n\n'
        + (f'Αποτυχίες ({len(failed)}):\n{failed_lines}' if failed else '')
    )
    subject = (
        f'[Τμήματα {SCHOOL_YEAR}] Αποστολή ολοκληρώθηκε — {len(sent)} ✓'
        + (f', {len(failed)} ✗' if failed else '')
    )
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From']    = formataddr((str(Header('MySchool Checks', 'utf-8')), ef))
    msg['To']      = ef
    msg['Date']    = formatdate(localtime=True)
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    log(f'\n  ✉ Summary → {ef}')
    try:
        _smtp_send(config, ef, msg.as_string())
        log('  ✓ Summary εστάλη.')
    except Exception as e:
        log(f'  ✗ Summary: {e}')


def _do_send_emails(dry_run, subj_tpl, body_tpl, config, today, log,
                    out_path=None, split_dir=None):
    """
    Αποστολή email. Επιστρέφει (sent, failed).

    - Κανονική αποστολή (dry_run=False): ένα email ανά σχολείο, ΜΟΝΟ για όσα
      έχουν ατομικό αρχείο μέσα στον φάκελο «split» (βλ. tab «✂ Διαχωρισμός» /
      _schools_from_split_dir) — συνημμένο ακριβώς το ίδιο αρχείο.
    - Test mode (dry_run=True): ΕΝΑ συγκεντρωτικό email προς το FROM_EMAIL,
      με ολόκληρο το ΣΥΝΟΛΙΚΟ Excel (out_path) συνημμένο — αντί για ένα
      email ανά σχολείο (θα ήταν πολλά ξεχωριστά test emails).
    """
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email.utils import formataddr, formatdate
    from email.header import Header
    from email import encoders

    ef = getattr(config, 'FROM_EMAIL',    '')
    pw = getattr(config, 'FROM_PASSWORD', '')
    if not ef:
        log('  ✗ FROM_EMAIL δεν έχει οριστεί στις Ρυθμίσεις.')
        return [], []
    if not pw:
        log('  ✗ Κωδικός email δεν έχει οριστεί στις Ρυθμίσεις.')
        return [], []

    schools = _schools_from_split_dir(split_dir)
    if not schools:
        log('  ℹ Δεν βρέθηκαν αρχεία στο φάκελο «split» — τρέξε πρώτα το tab '
            '«✂ Διαχωρισμός» (χωρίζει το συγκεντρωτικό αρχείο σε ένα Excel '
            'ανά σχολείο, μόνο για όσα έχουν απόκλιση).')
        return [], []

    # ── Test mode: ένα συγκεντρωτικό email με το συνολικό Excel ────────────
    if dry_run:
        if not out_path or not os.path.exists(out_path):
            log('  ✗ Δεν βρέθηκε το συνολικό αρχείο Excel για το test email.')
            return [], []

        names = [name for _, name, *_ in schools]
        log(f'  → Test mode: 1 συγκεντρωτικό email ({len(schools)} σχολεία) → {ef}')

        subj = (f'[TEST] {subj_tpl}'.replace('{school_year}', SCHOOL_YEAR)
                .replace('{school_name}', f'Συγκεντρωτικό ({len(schools)} σχολεία)'))
        body_txt = (
            f'Δοκιμαστική αποστολή — συνολικό αρχείο Excel (όλα τα σχολεία).\n\n'
            f'Σχολεία με απόκλιση: {len(schools)}\n'
            + '\n'.join(f'  • {n}' for n in sorted(names))
        )

        with open(out_path, 'rb') as f:
            xls_bytes = f.read()

        msg = MIMEMultipart()
        msg['Subject'] = subj
        msg['From']    = formataddr((str(Header('ΔΙ.Π.Ε. Αν. Θεσ/νίκης', 'utf-8')), ef))
        msg['To']      = ef
        msg['Date']    = formatdate(localtime=True)
        msg.attach(MIMEText(body_txt, 'plain', 'utf-8'))
        part = MIMEBase('application',
                        'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(xls_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment',
                        filename=Header(os.path.basename(out_path), 'utf-8').encode())
        msg.attach(part)

        try:
            _smtp_send(config, ef, msg.as_string())
            log(f'  ✓ → {ef}  (συγκεντρωτικό, {len(schools)} σχολεία)')
            return [(n, ef) for n in names], []
        except Exception as e:
            log(f'  ✗ {e}')
            return [], [(n, ef) for n in names]

    # ── Κανονική αποστολή: ένα email ανά σχολείο (μόνο όσα έχουν αρχείο split) ─
    log(f'  → {len(schools)} σχολεία για αποστολή (βρέθηκαν στο φάκελο split)')
    sent, failed = [], []

    for code, name, email_to, file_path in schools:
        if not email_to:
            log(f'  ⚠ {name}: κενό email — παράλειψη')
            failed.append((name, '—'))
            continue

        subj     = subj_tpl.replace('{school_year}', SCHOOL_YEAR).replace('{school_name}', name)
        body_txt = body_tpl.replace('{school_year}', SCHOOL_YEAR).replace('{school_name}', name)
        try:
            with open(file_path, 'rb') as f:
                xls_bytes = f.read()
        except Exception as e:
            log(f'  ✗ {name}: δεν ήταν δυνατή η ανάγνωση του αρχείου split: {e}')
            failed.append((name, email_to))
            continue
        fname_att = os.path.basename(file_path)

        msg = MIMEMultipart()
        msg['Subject'] = subj
        msg['From']    = formataddr((str(Header('ΔΙ.Π.Ε. Αν. Θεσ/νίκης', 'utf-8')), ef))
        msg['To']      = formataddr((str(Header(name, 'utf-8')), email_to))
        msg['Date']    = formatdate(localtime=True)
        msg.attach(MIMEText(body_txt, 'plain', 'utf-8'))
        part = MIMEBase('application',
                        'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(xls_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment',
                        filename=Header(fname_att, 'utf-8').encode())
        msg.attach(part)

        try:
            _smtp_send(config, email_to, msg.as_string())
            log(f'  ✓ → {email_to}  [{name}]')
            sent.append((name, email_to))
        except Exception as e:
            log(f'  ✗ {name}: {e}')
            failed.append((name, email_to))

    if sent:
        _send_summary_tmimata(config, sent, failed, today, log)

    return sent, failed


def _open_tmimata_template_editor(parent, C):
    """Dialog επεξεργασίας θέματος & κειμένου email."""
    import tkinter as tk
    cur_subj, cur_body = _load_tmimata_email_template()
    if cur_body == DEFAULT_EMAIL_BODY:
        try:
            import config as _cfg
            sig = getattr(_cfg, 'EMAIL_SIGNATURE', '').strip()
            if sig:
                cur_body = cur_body + '\n\n' + sig
        except Exception:
            pass

    ed = tk.Toplevel(parent)
    ed.title('Πρότυπο Email — Τμήματα Γενικής Παιδείας')
    ed.configure(bg=C['bg'])
    ed.resizable(True, False)
    ed.grab_set()
    ed.transient(parent)

    epad = dict(padx=14, pady=5)
    tk.Label(ed, text='Θέμα:', bg=C['bg'], fg=C['hdr_bg'],
             font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **epad)
    subj_var = tk.StringVar(value=cur_subj)
    tk.Entry(ed, textvariable=subj_var, font=('Arial', 9),
             width=70).pack(fill='x', padx=14, pady=(0, 8))

    tk.Label(ed, text='Κείμενο email:', bg=C['bg'], fg=C['hdr_bg'],
             font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **epad)
    txt = tk.Text(ed, font=('Arial', 9), width=70, height=14,
                  wrap='word', relief='solid', bd=1)
    txt.pack(fill='x', padx=14, pady=(0, 4))
    txt.insert('1.0', cur_body)

    tk.Label(ed,
             text='Χρησιμοποιήστε {school_year} και {school_name} (αντικαθίστανται αυτόματα).',
             bg=C['bg'], fg=C['desc'], font=('Arial', 8),
             anchor='w').pack(fill='x', padx=14, pady=(0, 10))

    def _save():
        _save_tmimata_email_template(subj_var.get().strip(), txt.get('1.0', 'end-1c'))
        ed.destroy()
        import tkinter.messagebox as _mb
        _mb.showinfo('Αποθήκευση', 'Το πρότυπο αποθηκεύτηκε.', parent=parent)

    def _reset():
        import tkinter.messagebox as _mb
        if _mb.askyesno('Επαναφορά', 'Να επανέλθει το προεπιλεγμένο κείμενο;', parent=ed):
            _save_tmimata_email_template(DEFAULT_EMAIL_SUBJECT, DEFAULT_EMAIL_BODY)
            ed.destroy()

    br = tk.Frame(ed, bg=C['bg'])
    br.pack(pady=(0, 12))
    tk.Button(br, text='Αποθήκευση', bg=C['btn_bg'], fg=C['btn_fg'],
              font=('Arial', 9, 'bold'), relief='flat', padx=14, pady=5,
              cursor='hand2', command=_save).pack(side='left', padx=4)
    tk.Button(br, text='Επαναφορά προεπιλογής', bg=C['bg2'], fg=C['hdr_bg'],
              font=('Arial', 9), relief='flat', padx=14, pady=5,
              cursor='hand2', command=_reset).pack(side='left', padx=4)
    tk.Button(br, text='Άκυρο', bg=C['bg2'], fg=C['desc'],
              font=('Arial', 9), relief='flat', padx=14, pady=5,
              cursor='hand2', command=ed.destroy).pack(side='left', padx=4)

    ed.update_idletasks()
    x = parent.winfo_x() + (parent.winfo_width()  - ed.winfo_width())  // 2
    y = parent.winfo_y() + (parent.winfo_height() - ed.winfo_height()) // 2
    ed.geometry(f'+{x}+{y}')


def _show_results_dialog(config, df_ds, df_nip, today, out_path, summary_text):
    """
    Αντικαθιστά το _show_results_popup — εμφανίζει σύνοψη αποτελεσμάτων
    και κουμπί για άνοιγμα του email dialog.  Τρέχει στο main thread.
    """
    import tkinter as tk
    from tkinter import scrolledtext

    root = tk._default_root
    if root is None:
        return

    win = tk.Toplevel(root)
    win.title(f'Αποτελέσματα — {CHECK_TITLE}')
    win.configure(bg='#FFF8E1')
    win.resizable(True, True)
    win.attributes('-topmost', True)
    # Το CheckRunDialog (γονικό παράθυρο πίσω από αυτό) κάνει ήδη grab_set()
    # στον εαυτό του — χωρίς δικό της grab, αυτή η νέα Toplevel δεν λαμβάνει
    # κλικ/πληκτρολόγηση (τα events πάνε στο grab-owner) μέχρι να κλείσει το
    # CheckRunDialog. Παίρνουμε το grab εδώ ώστε τα κουμπιά να δουλεύουν αμέσως.
    win.transient(root)
    win.grab_set()
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f'520x380+{sw//2-260}+{sh//2-190}')

    tk.Frame(win, bg='#E65100', pady=8).pack(fill='x')
    tk.Label(win.children[list(win.children)[-1]],
             text=f'⚠  {CHECK_TITLE}',
             bg='#E65100', fg='white',
             font=('Arial', 11, 'bold')).pack()

    txt = scrolledtext.ScrolledText(win, font=('Arial', 9), wrap='word',
                                     relief='flat', bg='#FFF8E1', height=10)
    txt.pack(fill='both', expand=True, padx=14, pady=8)
    txt.insert('1.0', summary_text)
    txt.config(state='disabled')

    btn_f = tk.Frame(win, bg='#FFF8E1')
    btn_f.pack(pady=(0, 12))

    def _open_excel():
        # Σημείωση: subprocess.Popen(['start', '', path], shell=True) ΔΕΝ
        # δουλεύει σωστά στα Windows — με shell=True + λίστα ορισμάτων, το
        # Python χρησιμοποιεί μόνο το πρώτο στοιχείο ('start') και αγνοεί το
        # path, οπότε δεν άνοιγε ποτέ το αρχείο. Το σωστό είναι os.startfile.
        try:
            os.startfile(os.path.normpath(out_path))
        except Exception as e:
            import tkinter.messagebox as _mb
            _mb.showerror('Σφάλμα', f'Δεν ήταν δυνατό το άνοιγμα του αρχείου:\n{e}', parent=win)

    # Σημείωση: το κουμπί «✉ Αποστολή Email» μετακόμισε στο tab «✉ Αποστολή»
    # του CheckRunDialog (βλ. CUSTOM_SEND_TAB / open_email_from_last_result
    # παρακάτω) — δεν εμφανίζεται πια εδώ στο παράθυρο αποτελεσμάτων.

    tk.Button(btn_f, text='📄 Άνοιγμα Excel',
              bg='#E65100', fg='white',
              font=('Arial', 9, 'bold'), relief='flat',
              padx=14, pady=5, cursor='hand2',
              command=_open_excel).pack(side='left', padx=4)
    tk.Button(btn_f, text='Κλείσιμο',
              bg='#E8EDF3', fg='#333333',
              font=('Arial', 9), relief='flat',
              padx=14, pady=5, cursor='hand2',
              command=win.destroy).pack(side='left', padx=4)


def _build_email_form(parent, config, df_ds, df_nip, today, out_path, C):
    """
    Χτίζει τη φόρμα αποστολής email (πρότυπο / όριο απόκλισης / τρόπος
    αποστολής / log / κουμπί Αποστολή) μέσα στο δοσμένο `parent`.

    Κοινή υλοποίηση — χρησιμοποιείται από το build_send_tab() για να χτίσει
    τη φόρμα ΑΠΕΥΘΕΙΑΣ μέσα στο tab «✉ Αποστολή» του CheckRunDialog (χωρίς
    ξεχωριστό popup παράθυρο, όπως ζητήθηκε).
    """
    import tkinter as tk
    from tkinter import scrolledtext
    import threading

    pad = dict(padx=14, pady=5)
    split_dir = _split_dir_for(out_path)

    # Κουμπί πρότυπου
    tk.Button(parent, text='✉  Πρότυπο Email (Θέμα & Κείμενο)',
              bg=C['bg2'], fg=C['hdr_bg'], font=('Arial', 9),
              relief='flat', padx=10, pady=4, cursor='hand2',
              command=lambda: _open_tmimata_template_editor(parent, C)
              ).pack(anchor='w', **pad)

    # Πληροφορία: η «Αποστολή σε σχολεία» στέλνει ΜΟΝΟ σε όσα έχουν ατομικό
    # αρχείο μέσα στο φάκελο «split» — δηλ. πρέπει να έχει τρέξει πρώτα το
    # tab «✂ Διαχωρισμός» (αυτό φιλτράρει και τα σχολεία χωρίς απόκλιση).
    # ΣΗΜΕΙΩΣΗ: αυτό το tab («✉ Αποστολή») χτίζεται ΜΙΑ φορά, μόλις τελειώσει
    # η Εκτέλεση — αν ο χρήστης τρέξει τον Διαχωρισμό ΜΕΤΑ από αυτό, το
    # μήνυμα πρέπει να ανανεωθεί (δεν ξαναχτίζεται μόνο του το tab), γι' αυτό
    # ανανεώνεται (α) κάθε φορά που αλλάζει tab στο Notebook και (β) με το
    # χειροκίνητο κουμπί «↻ Ανανέωση» παρακάτω.
    count_row = tk.Frame(parent, bg=C['bg'])
    count_row.pack(fill='x', **pad)
    count_lbl = tk.Label(count_row, text='', bg=C['bg'], fg=C['desc'],
                          font=('Arial', 8, 'italic'), anchor='w',
                          justify='left', wraplength=480)
    count_lbl.pack(side='left', fill='x', expand=True)

    def _update_count(*_args):
        schools   = _schools_from_split_dir(split_dir)
        with_e    = sum(1 for s in schools if s[2])
        without_e = len(schools) - with_e
        if not schools:
            count_lbl.config(
                text='⚠ Δεν βρέθηκαν αρχεία στο φάκελο «split» — τρέξε πρώτα '
                     'το tab «✂ Διαχωρισμός».', fg='#B00020')
        else:
            txt = f'→ {len(schools)} σχολεία θα λάβουν email (βρέθηκαν στο φάκελο split)'
            if without_e:
                txt += f'  ({without_e} χωρίς email)'
            count_lbl.config(text=txt, fg=C['desc'])

    tk.Button(count_row, text='↻ Ανανέωση', bg=C['bg2'], fg=C['desc'],
              font=('Arial', 8), relief='flat', padx=8, pady=2, cursor='hand2',
              command=_update_count).pack(side='right')

    _update_count()

    # Ανανέωση αυτόματα κάθε φορά που αλλάζει tab στο Notebook (π.χ. ο
    # χρήστης πάει Διαχωρισμός → Αποστολή) — βρίσκουμε το Notebook ανεβαίνοντας
    # στο δέντρο widget· αν κάτι αλλάξει σε αυτή τη δομή, απλά δεν ανανεώνει
    # αυτόματα (παραμένει πάντα το χειροκίνητο κουμπί).
    try:
        _nb = parent.master.master
        _nb.bind('<<NotebookTabChanged>>', _update_count, add='+')
    except Exception:
        pass

    # Τρόπος αποστολής
    mode_f = tk.Frame(parent, bg=C['bg'])
    mode_f.pack(fill='x', **pad)
    mode_var = tk.StringVar(value='schools')
    fe = getattr(config, 'FROM_EMAIL', '') or '...'
    tk.Radiobutton(mode_f, text='Αποστολή σε σχολεία (ένα email ανά σχολείο)',
                   variable=mode_var, value='schools',
                   bg=C['bg'], selectcolor=C['sel_bg'],
                   activebackground=C['bg'], font=('Arial', 9)).pack(anchor='w')
    tk.Radiobutton(mode_f, text=f'Test mode — όλα στο: {fe}',
                   variable=mode_var, value='test',
                   bg=C['bg'], selectcolor=C['sel_bg'],
                   activebackground=C['bg'], font=('Arial', 9)).pack(anchor='w')

    # Log
    log_w = scrolledtext.ScrolledText(parent, height=10, font=('Courier', 8),
                                       wrap='word', relief='solid', bd=1,
                                       state='disabled')
    log_w.pack(fill='both', expand=True, padx=14, pady=8)

    def _log(msg):
        def _append():
            log_w.config(state='normal')
            log_w.insert('end', msg + '\n')
            log_w.see('end')
            log_w.config(state='disabled')
        try:
            # Χρησιμοποιούμε .after() πάνω στο ίδιο το widget (όχι σε
            # dlg/Toplevel) — δουλεύει το ίδιο είτε είμαστε μέσα σε popup
            # είτε embedded σε tab, αφού όλα τα widgets μοιράζονται τον ίδιο
            # Tcl interpreter.
            log_w.after(0, _append)
        except Exception:
            pass

    # Κουμπί αποστολής
    btn_f = tk.Frame(parent, bg=C['bg'])
    btn_f.pack(fill='x', padx=14, pady=(0, 12))
    send_btn = tk.Button(btn_f, text='✉  Αποστολή',
                         bg=C['btn_bg'], fg=C['btn_fg'],
                         font=('Arial', 9, 'bold'), relief='flat',
                         padx=14, pady=5, cursor='hand2')
    send_btn.pack(side='left', padx=(0, 8))

    def _start():
        send_btn.config(state='disabled')
        dry_run       = (mode_var.get() == 'test')
        subj, body_tpl = _load_tmimata_email_template()

        def _worker():
            try:
                sent, failed = _do_send_emails(
                    dry_run, subj, body_tpl, config, today, _log,
                    out_path=out_path, split_dir=split_dir)
                _log(f'\n{"─"*40}')
                _log(f'✓ Εστάλησαν: {len(sent)}   ✗ Αποτυχίες: {len(failed)}')
            except Exception as e:
                _log(f'✗ Σφάλμα: {e}')
            finally:
                try:
                    send_btn.after(0, lambda: send_btn.config(state='normal'))
                except Exception:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

    send_btn.config(command=_start)


_TMIMATA_PALETTE = {
    'bg': '#F5F7FA', 'bg2': '#E8EDF3', 'hdr_bg': '#1F4E79',
    'btn_bg': '#1F4E79', 'btn_fg': '#FFFFFF', 'desc': '#666666',
    'sel_bg': '#D6E4F0',
}


def build_send_tab(parent, config):
    """
    Χτίζει ΑΠΕΥΘΕΙΑΣ μέσα στο δοσμένο `parent` (το body του tab «✉ Αποστολή»
    του CheckRunDialog) τη φόρμα αποστολής email, με βάση το αποτέλεσμα της
    τελευταίας επιτυχούς Εκτέλεσης (_LAST_RESULT — γεμίζει μέσα στο run()).

    Καλείται από core/check_dialog.py (βλ. CUSTOM_SEND_TAB παρακάτω) αφού
    ολοκληρωθεί επιτυχώς η Εκτέλεση — αντικαθιστά το παλιότερο κουμπί
    «✉ Αποστολή Email» που άνοιγε ξεχωριστό popup παράθυρο.
    """
    import tkinter as tk

    if not _LAST_RESULT or not _LAST_RESULT.get('out_path'):
        tk.Label(parent, text='Τρέξε πρώτα την «▶ Εκτέλεση» για να ενεργοποιηθεί η αποστολή.',
                 bg=_TMIMATA_PALETTE['bg'], fg=_TMIMATA_PALETTE['desc'],
                 font=('Arial', 9), anchor='w', justify='left',
                 wraplength=520).pack(fill='x', padx=14, pady=14)
        return

    _build_email_form(parent, config, _LAST_RESULT['df_ds'], _LAST_RESULT['df_nip'],
                       _LAST_RESULT['today'], _LAST_RESULT['out_path'], _TMIMATA_PALETTE)


# Σηματοδοτεί στο core/check_dialog.py ότι το tab «✉ Αποστολή» πρέπει να
# χτίσει τη φόρμα αποστολής ΑΠΕΥΘΕΙΑΣ μέσα στο tab (χωρίς ξεχωριστό popup) —
# βλ. CheckRunDialog._build_send / _start_execute._done.
CUSTOM_SEND_TAB = build_send_tab


# ── Tab «✂ Διαχωρισμός» — Excel ανά σχολείο ─────────────────────────────────
# Ίδια λογική με το tab Διαχωρισμός του ελέγχου Ε.Ε.Α. (smeae/dialog.py +
# smeae/compare.py::split_xlsx): παίρνει το συγκεντρωτικό αρχείο
# αποτελεσμάτων και το χωρίζει σε ένα Excel ανά σχολείο. Εδώ δεν χρειάζεται
# επιλογή σχολικού έτους (SCHOOL_YEAR είναι σταθερό — βλ. πάνω), και κάθε
# γραμμή του συγκεντρωτικού αρχείου αντιστοιχεί ήδη σε ένα σχολείο (όχι
# πολλαπλές εγγραφές ανά σχολείο όπως στο Ε.Ε.Α.) — οπότε ο «διαχωρισμός»
# είναι απλά: μία μορφοποιημένη γραμμή → ένα αρχείο, ξαναχρησιμοποιώντας το
# ίδιο _build_mini_excel που ήδη φτιάχνει τα συνημμένα των email.
def _sanitize_filename(name):
    import re
    return re.sub(r'[^\w\s-]', '_', str(name)).strip() or 'σχολειο'


def _find_latest_output():
    """
    Βρίσκει το πιο πρόσφατο συγκεντρωτικό αρχείο αποτελεσμάτων
    (Documents/MySchoolChecks/results_*/tmimata_genikis/*_tmimata_genikis.xlsx).
    """
    docs = os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks')
    pattern = os.path.join(docs, 'results_*', RESULTS_FOLDER, f'*_{RESULTS_FOLDER}.xlsx')
    matches = [f for f in _glob.glob(pattern) if os.path.isfile(f)]
    if not matches:
        return None
    matches.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return matches[0]


def split_tmimata_workbook(file_path, output_dir, log=print):
    """
    Διαβάζει το συγκεντρωτικό αρχείο αποτελεσμάτων (φύλλα «ΔΣ-...» / «ΝΗΠ-...»)
    και γράφει ένα ξεχωριστό, μορφοποιημένο Excel ανά σχολείο μέσα στο
    output_dir — ίδια μορφοποίηση με το mini Excel που στέλνεται ως συνημμένο
    email (βλ. _build_mini_excel). Επιστρέφει το πλήθος αρχείων που γράφτηκαν.
    """
    os.makedirs(output_dir, exist_ok=True)
    today = datetime.today()
    n_written = 0

    for sheet_name in pd.ExcelFile(file_path).sheet_names:
        if sheet_name.startswith('ΔΣ-'):
            col_defs, label = columns_ds(), 'Δημοτικά'
        elif sheet_name.startswith('ΝΗΠ-'):
            col_defs, label = columns_nip(), 'Νηπιαγωγεία'
        else:
            log(f'  ⚠ Άγνωστο φύλλο «{sheet_name}» — παράλειψη.')
            continue

        col_names = [c[0] for c in col_defs]
        # Οι 2 πρώτες γραμμές του φύλλου είναι τίτλος/υπότιτλος (merged
        # cells) — οι στήλες είναι στη γραμμή 3, άρα skiprows=2.
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=2)
        except Exception as e:
            log(f'  ✗ Σφάλμα ανάγνωσης φύλλου «{sheet_name}»: {e}')
            continue

        missing = [c for c in col_names if c not in df.columns]
        if missing:
            log(f'  ✗ Λείπουν στήλες στο «{sheet_name}»: {", ".join(missing)}')
            continue
        df = df[col_names]

        n_skipped_sheet = 0
        for _, row in df.iterrows():
            name = str(row.get('Ονομασία', '')).strip()
            if not name or name.lower() == 'nan':
                continue

            # Εξαίρεση: σχολεία χωρίς καμία απόκλιση (και οι δύο στήλες
            # Διαφορά Τμήματα / Διαφορά Μαθητές = 0) — δεν χρειάζονται
            # ξεχωριστό αρχείο, γι' αυτό ούτε θα εμφανιστούν στο tab
            # «Αποστολή» (βλ. _schools_from_split_dir παρακάτω).
            diff_tm = row.get('Διαφορά Τμήματα', 0)
            diff_ma = row.get('Διαφορά Μαθητές', 0)
            try:
                diff_tm = int(diff_tm)
            except Exception:
                diff_tm = 0
            try:
                diff_ma = int(diff_ma)
            except Exception:
                diff_ma = 0
            if diff_tm == 0 and diff_ma == 0:
                n_skipped_sheet += 1
                continue

            code = row.get('Κωδ. ΥΠΠΘ', '')
            try:
                code = int(code)
            except Exception:
                pass

            row_df = pd.DataFrame([row])
            xls_bytes = _build_mini_excel(row_df, col_defs, label, today)
            fname = f'{code}_{_sanitize_filename(name)}.xlsx'
            with open(os.path.join(output_dir, fname), 'wb') as f:
                f.write(xls_bytes)
            n_written += 1
            log(f'  ✓ {fname}')

        if n_skipped_sheet:
            log(f'  ℹ «{sheet_name}»: {n_skipped_sheet} σχολεία χωρίς αποκλίσεις — παραλείφθηκαν.')

    log(f'\n✓ Διαχωρισμός ολοκληρώθηκε — {n_written} αρχεία (με αποκλίσεις) → {output_dir}')
    return n_written


def build_split_tab(parent, config):
    """
    Χτίζει το tab «✂ Διαχωρισμός» ΑΠΕΥΘΕΙΑΣ μέσα στο δοσμένο `parent`.
    Ανιχνεύει αυτόματα το πιο πρόσφατο αρχείο αποτελεσμάτων (δεν εξαρτάται
    από το αν έτρεξε η Εκτέλεση μέσα σε αυτή τη σύνοδο) και το χωρίζει σε
    ένα Excel ανά σχολείο μέσα σε υποφάκελο «split».
    """
    import tkinter as tk
    from tkinter import scrolledtext
    import threading

    C = _TMIMATA_PALETTE
    pad = dict(padx=14, pady=5)

    tk.Label(parent,
             text='Παίρνει το πιο πρόσφατο συγκεντρωτικό αρχείο αποτελεσμάτων του '
                  'ελέγχου και το χωρίζει σε ξεχωριστά αρχεία Excel — ένα ανά '
                  'σχολείο — μέσα σε φάκελο «split» δίπλα στο αρχείο.\n'
                  'Εξαιρούνται τα σχολεία χωρίς καμία απόκλιση (Διαφορά Τμήματα = 0 '
                  'ΚΑΙ Διαφορά Μαθητές = 0) — δημιουργείται αρχείο μόνο για τα '
                  'υπόλοιπα. Στο tab «✉ Αποστολή» θα σταλούν emails μόνο σε αυτά '
                  'τα σχολεία (όσα έχουν ατομικό αρχείο εδώ).',
             bg=C['bg'], fg=C['desc'], font=('Arial', 8),
             wraplength=560, justify='left', anchor='w').pack(fill='x', **pad)

    info_lbl = tk.Label(parent, text='', bg=C['bg'], fg=C['desc'],
                         font=('Arial', 8, 'italic'), anchor='w',
                         justify='left', wraplength=560)
    info_lbl.pack(fill='x', padx=14, pady=(0, 4))

    state = {'file': None}

    def _detect():
        f = _find_latest_output()
        state['file'] = f
        if f:
            info_lbl.config(text=f'✓ {os.path.basename(f)}', fg='#2E7D32')
        else:
            info_lbl.config(
                text='Δεν βρέθηκε αρχείο αποτελεσμάτων — τρέξε πρώτα την «▶ Εκτέλεση».',
                fg='#B00020')

    _detect()

    log_w = scrolledtext.ScrolledText(parent, height=10, font=('Courier', 8),
                                       wrap='word', relief='solid', bd=1,
                                       state='disabled')
    log_w.pack(fill='both', expand=True, padx=14, pady=8)

    def _log(msg):
        def _append():
            log_w.config(state='normal')
            log_w.insert('end', msg + '\n')
            log_w.see('end')
            log_w.config(state='disabled')
        try:
            log_w.after(0, _append)
        except Exception:
            pass

    btn_f = tk.Frame(parent, bg=C['bg'])
    btn_f.pack(fill='x', padx=14, pady=(0, 12))
    split_btn = tk.Button(btn_f, text='✂  Διαχωρισμός ανά Σχολείο',
                           bg=C['btn_bg'], fg=C['btn_fg'],
                           font=('Arial', 9, 'bold'), relief='flat',
                           padx=14, pady=5, cursor='hand2')
    split_btn.pack(side='left', padx=(0, 8))

    tk.Button(btn_f, text='Ανίχνευση αρχείου', bg=C['bg2'], fg=C['desc'],
              font=('Arial', 9), relief='flat', padx=10, pady=5, cursor='hand2',
              command=_detect).pack(side='left')

    def _start():
        if not state['file']:
            _detect()
        if not state['file']:
            import tkinter.messagebox as _mb
            _mb.showwarning(
                'Προσοχή',
                'Δεν βρέθηκε αρχείο αποτελεσμάτων.\nΤρέξε πρώτα την «▶ Εκτέλεση».',
                parent=parent)
            return

        split_btn.config(state='disabled')
        f = state['file']
        out_dir = os.path.join(os.path.dirname(f), 'split')

        def _worker():
            try:
                split_tmimata_workbook(f, out_dir, log=_log)
            except Exception as e:
                _log(f'✗ Σφάλμα: {e}')
            finally:
                try:
                    split_btn.after(0, lambda: split_btn.config(state='normal'))
                except Exception:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

    split_btn.config(command=_start)


# Σηματοδοτεί στο core/check_dialog.py ότι υπάρχει επιπλέον tab «✂ Διαχωρισμός»
# — βλ. CheckRunDialog._build.
CUSTOM_SPLIT_TAB = build_split_tab


# ── Λήψη δεδομένων (μέσα στον έλεγχο — απαιτεί αλλαγή σχολικού έτους) ───────
def _with_year(entry, year, idx=11):
    """Επιστρέφει αντίγραφο ενός REPORTS tuple με το report_year ορισμένο στη θέση idx."""
    entry = list(entry)
    while len(entry) <= idx:
        entry.append(None)
    entry[idx] = year
    return tuple(entry)


def _download_inputs(config, log=print):
    """
    Κατεβάζει 3.1, 5.3, 5.4 απευθείας, αφού πρώτα οριστεί το σχολικό έτος
    SCHOOL_YEAR (2026-2027) στο MySchool. Το 3.1 δεν έχει προεπιλεγμένο
    report_year στο core/downloader.py (το χρησιμοποιούν και άλλα εργαλεία
    με το τρέχον έτος) — εδώ γίνεται τοπικό override μόνο για αυτή τη λήψη,
    χωρίς να πειραχτεί η καθολική ρύθμιση.

    Επιστρέφει (path_31, path_53, path_54) — None όπου δεν κατέβηκε.
    """
    import core.downloader as _dl

    ms_user = getattr(config, 'MYSCHOOL_USER', '').strip()
    ms_pass = getattr(config, 'MYSCHOOL_PASS', '').strip()
    if not ms_user or not ms_pass:
        raise RuntimeError('Συμπλήρωσε username και κωδικό MySchool στις Ρυθμίσεις (⚙).')

    orig_31 = next((r for r in _dl.REPORTS if r[0] == '3.1'), None)
    orig_22 = next((r for r in _dl.REPORTS if r[0] == '2.2'), None)
    if orig_31 is None:
        raise RuntimeError('Δεν βρέθηκε ρύθμιση λήψης για το 3.1 στο core/downloader.py.')

    custom_reports = (
        [_with_year(orig_31, SCHOOL_YEAR)] +
        [r for r in _dl.REPORTS if r[0] in ('5.3', '5.4')] +
        ([orig_22] if orig_22 else [])
    )
    rids = ['3.1', '5.3', '5.4'] + (['2.2'] if orig_22 else [])

    today_str = datetime.today().strftime('%Y%m%d')
    dest_dir  = os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks',
                             'downloads', f'{today_str}_{SCHOOL_YEAR}')
    os.makedirs(dest_dir, exist_ok=True)

    orig_reports_backup = _dl.REPORTS
    try:
        _dl.REPORTS = custom_reports
        dl = _dl.MySchoolDownloader(
            username=ms_user, password=ms_pass, dest_dir=dest_dir,
            callback=log, reports=rids,
            browser=getattr(config, 'BROWSER', 'chrome'),
            # force=rids: κάθε πάτημα του «⬇ Λήψη» ξανακατεβάζει (override) τα
            # 3.1/5.3/5.4, ακόμα κι αν υπάρχουν ήδη από νωρίτερα μέσα στην
            # ίδια μέρα — χρήσιμο για επικαιροποίηση εντός της ημέρας.
            force=rids,
        )
        results = dl.run()
    finally:
        _dl.REPORTS = orig_reports_backup

    return results.get('3.1'), results.get('5.3'), results.get('5.4'), results.get('2.2')


# Public alias — το tab «⬇ Λήψη» (core/check_dialog.py) καλεί αυτή τη
# συνάρτηση αντί για τον γενικό μηχανισμό λήψης, ώστε να διατηρηθεί το
# override σχολικού έτους (SCHOOL_YEAR) που χρειάζεται το 3.1/5.3/5.4.
download = _download_inputs
# Σηματοδοτεί στο core/check_dialog.py ότι υπάρχει custom download function
# (βλ. download() παραπάνω) — έτσι το tab «⬇ Λήψη» χρησιμοποιεί αυτήν αντί
# για τον γενικό MySchoolDownloader.
CUSTOM_DOWNLOAD = download


def _find_downloaded_inputs():
    """
    Ψάχνει για 3.1/5.3/5.4 (+2.2) που έχουν ΗΔΗ κατέβει σήμερα μέσω του tab
    «⬇ Λήψη» (βλ. download() / CUSTOM_DOWNLOAD) — ΔΕΝ κατεβάζει τίποτα.
    Επιστρέφει (path_31, path_53, path_54, path_22), None όπου δεν βρέθηκε.
    """
    import glob as _glob2
    from core.downloader import FILE_PREFIX_MAP

    today_str = datetime.today().strftime('%Y%m%d')
    dest_dir  = os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks',
                             'downloads', f'{today_str}_{SCHOOL_YEAR}')
    if not os.path.isdir(dest_dir):
        return None, None, None, None

    def _match(rid):
        prefix  = FILE_PREFIX_MAP.get(rid, rid)
        matches = [f for f in _glob2.glob(os.path.join(dest_dir, f'{prefix}*'))
                   if not f.endswith(('.tmp', '.crdownload'))]
        return matches[0] if matches else None

    return _match('3.1'), _match('5.3'), _match('5.4'), _match('2.2')


# ── CUSTOM RUN ────────────────────────────────────────────────────────────
def run(config):
    import core.framework as _fw
    _fw._current_check_title = CHECK_TITLE

    print('=' * 65)
    print(f'  {CHECK_TITLE}')
    print('=' * 65)

    path_31, path_53, path_54, path_22 = _find_downloaded_inputs()

    missing = [rid for rid, p in (('3.1', path_31), ('5.3', path_53), ('5.4', path_54)) if not p]
    if missing:
        import tkinter.messagebox as _mb
        _mb.showwarning(
            'Λείπει η λήψη',
            f'Δεν έχουν κατέβει σήμερα: {", ".join(missing)}.\n\n'
            f'Πήγαινε πρώτα στο tab «⬇ Λήψη» και πάτησε το κουμπί λήψης.'
        )
        return

    print(f'\n  ✓ Χρήση ήδη κατεβασμένων αρχείων (σχολικό έτος {SCHOOL_YEAR}).')
    print('-' * 65)

    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    print(f'\n  Ημερομηνία : {today.strftime("%d/%m/%Y")}')
    print('-' * 65)

    if path_22:
        print(f'  ✓ stat2_2: {os.path.basename(path_22)}')
    else:
        print('  ℹ stat2_2 δεν κατέβηκε — Email Σχολείου θα είναι κενό')

    print('\nΕπεξεργασία...')
    try:
        df_ds, df_nip, perif = process(path_31, path_53, path_54, path_22=path_22)
    except Exception as e:
        import tkinter.messagebox as _mb
        _mb.showerror('Σφάλμα επεξεργασίας', str(e))
        return

    global _LAST_RESULT
    _LAST_RESULT = {'df_ds': df_ds, 'df_nip': df_nip, 'today': today}

    dev_ds  = int((df_ds['Διαφορά Τμήματα']  != 0).sum() + (df_ds['Διαφορά Μαθητές']  != 0).sum()) if not df_ds.empty  else 0
    dev_nip = int((df_nip['Διαφορά Τμήματα'] != 0).sum() + (df_nip['Διαφορά Μαθητές'] != 0).sum()) if not df_nip.empty else 0
    print(f'  ✓ Δημοτικά    : {len(df_ds)} σχολεία, {dev_ds} αποκλίσεις')
    print(f'  ✓ Νηπιαγωγεία : {len(df_nip)} σχολεία, {dev_nip} αποκλίσεις')

    if df_ds.empty and df_nip.empty:
        # Το run() τρέχει σε background thread — τα Tk widgets (messagebox)
        # πρέπει να ανοίγουν στο main thread μέσω root.after().
        import tkinter as tk
        import tkinter.messagebox as _mb
        _root = tk._default_root
        if _root is not None:
            _root.after(0, lambda: _mb.showinfo(
                CHECK_TITLE, '✓  Δεν βρέθηκαν σχολεία στα αρχεία 5.3/5.4.'))
        return

    _docs   = os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks')
    out_dir = os.path.join(_docs, f'results_{today.strftime("%Y%m%d")}', RESULTS_FOLDER)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_{RESULTS_FOLDER}.xlsx')

    try:
        build_workbook(df_ds, df_nip, perif, today, out_path)
        print(f'\n  ✓ Αποθηκεύτηκε: {os.path.basename(out_path)}')
        _LAST_RESULT['out_path'] = out_path
    except PermissionError:
        import tkinter.messagebox as _mb
        _mb.showwarning(
            'Αρχείο ανοιχτό',
            f'Το αρχείο {os.path.basename(out_path)} είναι ανοιχτό σε άλλο πρόγραμμα.\n'
            f'Κλείστε το και τρέξτε ξανά τον έλεγχο.'
        )
        return

    body = (
        f'Σύνοψη ελέγχου τμημάτων γενικής παιδείας — {today.strftime("%d/%m/%Y")}\n'
        f'{"─"*50}\n'
        f'Δημοτικά:     {len(df_ds)} σχολεία  ·  {dev_ds} αποκλίσεις\n'
        f'Νηπιαγωγεία:  {len(df_nip)} σχολεία  ·  {dev_nip} αποκλίσεις\n\n'
        f'{"─"*50}\n'
        f'Αποτελέσματα αποθηκεύτηκαν στο φάκελο:\n{out_dir}'
    )
    # Το run() τρέχει σε background thread — το dialog πρέπει να ανοίξει
    # στο main thread μέσω root.after(), αλλιώς το Tkinter δεν το επιτρέπει.
    import tkinter as tk
    _root = tk._default_root
    if _root is not None:
        _summary = body
        _root.after(0, lambda: _show_results_dialog(
            config, df_ds, df_nip, today, out_path, _summary))
