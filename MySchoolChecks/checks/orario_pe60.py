"""
checks/orario_pe60.py
══════════════════════
Έλεγχος υποχρεωτικού ωραρίου ΠΕ60 / ΠΕ60.50 σε νηπιαγωγεία, βάσει λειτουργικότητας.

Κανόνας:
  Κατηγορία 1 (Λειτουργικότητα 1, 2 ή 3) → Υποχρεωτικό Ωράριο πρέπει να είναι 25.
  Κατηγορία 2 (Λειτουργικότητα ≥ 4)      → Υποχρεωτικό Ωράριο πρέπει να είναι διάφορο του 25.

Πηγές:
  2.1 (gridResults) — Κατάλογος σχολείων· φιλτράρεται σε Είδος=Νηπιαγωγεία,
                       key = Κωδικός Υπουργείου (7ψήφιος).
  4.1 (stat4_1)     — Οργανικές/προσωρινές τοποθετήσεις εκπαιδευτικών· φιλτράρεται σε
                       Κωδ. Ειδικότητας ΠΕ60/ΠΕ60.50. Στήλη ελέγχου: «Υποχ. Ωράριο»
                       (το πραγματικό Υποχρεωτικό Διδακτικό Ωράριο — ΟΧΙ οι «Ώρες
                       Υποχ. Διδακτικού Ωραρίου Υπηρέτησης στο Φορέα»). Κωδικός σχολείου
                       από «Μονάδα Οργανικής/Προσωρινής Τοποθέτησης».
  4.2 (stat4_2)     — Αποσπασμένοι εκπαιδευτικοί (προαιρετικό αρχείο). Ίδια λογική,
                       κωδικός σχολείου από «Μονάδα Απόσπασης». Αν ένας εκπαιδευτικός
                       εμφανίζεται και στα δύο αρχεία (ίδιο ΑΦΜ) και το 4.2 έχει
                       έγκυρο (μη κενό) κωδικό σχολείου, υπερισχύει η απόσπαση —
                       θεωρείται η πραγματική τρέχουσα τοποθέτηση.

Αποτέλεσμα (1 φύλλο «Αποκλίσεις», χωρίς αποστολή email):
  Εκπαιδευτικοί ΠΕ60/ΠΕ60.50 με ωράριο που παραβιάζει τον κανόνα.
  Τοποθετήσεις σε κωδικούς εκτός της λίστας Νηπιαγωγείων του 2.1 (π.χ. ιδιωτικά
  νηπιαγωγεία, αποκεντρωμένες υπηρεσίες), με άγνωστη/μη αναγνωρίσιμη Λειτουργικότητα,
  ή χωρίς τιμή Υποχρεωτικού Ωραρίου — αγνοούνται εντελώς, δεν εμφανίζονται πουθενά.

CUSTOM_RUN — ένα φύλλο εξόδου, καμία αποστολή email (μόνο λίστα Excel).
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.framework import get_downloaded_file, _missing_file_dialog, _show_results_popup, read_csv_fixed, clean_field

# ── Μεταδεδομένα ────────────────────────────────────────────────────────────
CHECK_TITLE       = 'Υποχρεωτικό Ωράριο ΠΕ60 (Νηπιαγωγεία)'
CHECK_DESCRIPTION = 'Έλεγχος υποχρεωτικού ωραρίου ΠΕ60/ΠΕ60.50 βάσει λειτουργικότητας νηπιαγωγείου τοποθέτησης'
RESULTS_FOLDER    = 'orario_pe60'
HAS_EMAIL         = False
CUSTOM_RUN        = True
NO_SEND_TAB       = True   # Χωρίς email — καθόλου tab «Αποστολή»
REQUIRED_REPORTS  = [
    '2.1 — Κατάλογος σχολείων',
    '4.1 — Οργανικές τοποθετήσεις',
    '4.2 — Αποσπασμένοι εκπαιδευτικοί (προαιρετικό)',
]

SPECIALTIES = {'ΠΕ60', 'ΠΕ60.50'}
CAT1_LEIT   = {1, 2, 3}     # Υποχρεωτικό ωράριο πρέπει να είναι 25
EXPECTED_25 = 25

COLOR_HDR     = '1F4E79'
COLOR_SUB     = 'D6E4F0'
COLOR_DEV     = 'FADBD8'
COLOR_DEV_ALT = 'FDEDEC'
COLOR_ALT     = 'EBF3FB'

_LEFT_ALIGN_COLS = {'Ονομασία Σχολείου', 'Επώνυμο', 'Όνομα'}


# ── Βοηθητικά ────────────────────────────────────────────────────────────────
def _fc(df, *keywords):
    """Επιστρέφει το όνομα της πρώτης στήλης που ταιριάζει με κάποια λέξη-κλειδί."""
    for kw in keywords:
        kw = kw.lower()
        for col in df.columns:
            if kw in str(col).lower():
                return col
    return None


def _norm_code(series):
    """Κανονικοποίηση κωδικού σχολείου: αριθμός ή string → stripped string χωρίς αρχικά μηδενικά."""
    return (series.fillna('').astype(str).str.strip()
                  .str.replace(r'\.0$', '', regex=True)
                  .str.lstrip('0'))


def _to_int(val):
    """Μετατρέπει τιμή σε int, ή None αν δεν είναι έγκυρος αριθμός."""
    try:
        n = pd.to_numeric(val, errors='coerce')
        if pd.isna(n):
            return None
        return int(n)
    except Exception:
        return None


# ── Φόρτωση ──────────────────────────────────────────────────────────────────
def _load_schools(path_21):
    """Επιστρέφει dict {κωδικός: (κατηγορία 1/2/None, λειτουργικότητα, όνομα σχολείου)} — μόνο Νηπιαγωγεία."""
    df = pd.read_excel(path_21, header=0)

    eidos_col = _fc(df, 'είδος', 'ειδος')
    if eidos_col:
        df = df[df[eidos_col].fillna('').astype(str).str.strip() == 'Νηπιαγωγεία'].copy()

    code_col = None
    for col in df.columns:
        if 'κωδικός υπουργείου' in str(col).lower():
            code_col = col
            break
    if code_col is None:
        code_col = _fc(df, 'κωδικός') or df.columns[12]

    name_col = _fc(df, 'ονομασ')                              or df.columns[13]
    leit_col = _fc(df, 'λειτουργικότητα', 'λειτουργικοτητα')  or df.columns[14]

    df['_code'] = _norm_code(df[code_col])
    df = df[df['_code'] != ''].drop_duplicates('_code')

    schools = {}
    for _, row in df.iterrows():
        code = row['_code']
        leit = _to_int(row[leit_col])
        if leit is None:
            cat = None
        elif leit in CAT1_LEIT:
            cat = 1
        elif leit >= 4:
            cat = 2
        else:
            cat = None
        schools[code] = (cat, leit, str(row[name_col]).strip())
    return schools


def _prep_stat4(df, code_col):
    """Φιλτράρει σε ΠΕ60/ΠΕ60.50 και καθαρίζει τις στήλες ενός stat4_1/stat4_2 DataFrame."""
    if df.empty or 'Κωδ. Ειδικότητας' not in df.columns:
        return pd.DataFrame()

    df = df[df['Κωδ. Ειδικότητας'].fillna('').astype(str).str.strip().isin(SPECIALTIES)].copy()
    if df.empty:
        return df

    df['_afm']     = clean_field(df['Α.Φ.Μ.']).str.replace(r'\.0$', '', regex=True)
    df['_am']      = clean_field(df['Α.Μ.'])
    df['_eponymo'] = df['Επώνυμο'].fillna('').astype(str).str.strip()
    df['_onoma']   = df['Όνομα'].fillna('').astype(str).str.strip()
    df['_spec']    = df['Κωδ. Ειδικότητας'].fillna('').astype(str).str.strip()
    df['_code']    = _norm_code(clean_field(df[code_col]))
    df['_school']  = df['Ονομασία'].fillna('').astype(str).str.strip() if 'Ονομασία' in df.columns else ''
    df['_orario']  = df['Υποχ. Ωράριο'].apply(_to_int) if 'Υποχ. Ωράριο' in df.columns else None

    return df[df['_afm'] != ''].copy()


def _load_placements(path_41, path_42):
    """
    Επιστρέφει DataFrame ΠΕ60/ΠΕ60.50 από 4.1 (Οργανικές τοποθετήσεις) + 4.2
    (Αποσπασμένοι, προαιρετικό). Αν ο ίδιος εκπαιδευτικός (ΑΦΜ) εμφανίζεται και
    στα δύο και το 4.2 έχει έγκυρο κωδικό σχολείου, υπερισχύει η απόσπαση.
    """
    df1 = _prep_stat4(read_csv_fixed(path_41), 'Μονάδα Οργανικής/Προσωρινής Τοποθέτησης')

    records = {}
    for _, row in df1.iterrows():
        records[row['_afm']] = row

    if path_42:
        df2 = _prep_stat4(read_csv_fixed(path_42), 'Μονάδα Απόσπασης')
        for _, row in df2.iterrows():
            if row['_code']:
                records[row['_afm']] = row
            elif row['_afm'] not in records:
                records[row['_afm']] = row

    if not records:
        return pd.DataFrame(columns=['_afm', '_am', '_eponymo', '_onoma', '_spec',
                                      '_code', '_school', '_orario'])

    return pd.DataFrame(list(records.values())).reset_index(drop=True)


# ── Λογική ──────────────────────────────────────────────────────────────────
def process(path_21, path_41, path_42=None):
    """Επιστρέφει DataFrame με τις αποκλίσεις."""
    schools = _load_schools(path_21)
    df_t    = _load_placements(path_41, path_42)

    dev_rows = []

    for _, row in df_t.iterrows():
        code = row['_code']
        info = schools.get(code)

        # Σχολείο εκτός λίστας Νηπιαγωγείων 2.1, ή άγνωστη Λειτουργικότητα → αγνοείται εντελώς
        if info is None or info[0] is None:
            continue

        cat, leit, school_name = info
        orario = row['_orario']

        # Χωρίς τιμή Υποχρεωτικού Ωραρίου → δεν μπορεί να ελεγχθεί, αγνοείται
        if orario is None:
            continue

        if cat == 1:
            ok, expected_txt = (orario == EXPECTED_25), '25'
        else:
            ok, expected_txt = (orario != EXPECTED_25), 'διάφορο του 25'

        if ok:
            continue

        dev_rows.append({
            'Κατηγορία':           cat,
            'Κωδικός Σχολείου':    code,
            'Ονομασία Σχολείου':   school_name,
            'Λειτουργικότητα':     leit,
            'ΑΜ':                  row['_am'],
            'ΑΦΜ':                 row['_afm'],
            'Επώνυμο':             row['_eponymo'],
            'Όνομα':               row['_onoma'],
            'Ειδικότητα':          row['_spec'],
            'Υποχρεωτικό Ωράριο':  orario,
            'Αναμενόμενο Ωράριο':  expected_txt,
        })

    df_dev = pd.DataFrame(dev_rows)
    if not df_dev.empty:
        df_dev = df_dev.sort_values(['Κατηγορία', 'Ονομασία Σχολείου', 'Επώνυμο']).reset_index(drop=True)

    return df_dev


# ── Excel ─────────────────────────────────────────────────────────────────
def _brd():
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)


DEV_COLUMNS = [
    ('Κατηγορία',           12),
    ('Κωδικός Σχολείου',    16),
    ('Ονομασία Σχολείου',   40),
    ('Λειτουργικότητα',     14),
    ('ΑΜ',                  11),
    ('ΑΦΜ',                 13),
    ('Επώνυμο',             18),
    ('Όνομα',               16),
    ('Ειδικότητα',          12),
    ('Υποχρεωτικό Ωράριο',  16),
    ('Αναμενόμενο Ωράριο',  18),
]


def _write_sheet(ws, df, title, columns, header_color, today, subtitle_extra='', highlight_col=None):
    brd = _brd()
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    ncols = len(columns)

    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws['A1'] = f'{title}  —  {today.strftime("%d/%m/%Y")}{subtitle_extra}'
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', start_color=header_color)
    ws['A1'].alignment = ctr
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
    ws['A2'] = f'Σύνολο εγγραφών: {len(df)}'
    ws['A2'].font      = Font(name='Arial', italic=True, size=9)
    ws['A2'].fill      = PatternFill('solid', start_color=COLOR_SUB)
    ws['A2'].alignment = ctr
    ws.row_dimensions[2].height = 16

    for ci, (name, width) in enumerate(columns, 1):
        c = ws.cell(row=3, column=ci, value=name)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill      = PatternFill('solid', start_color=header_color)
        c.alignment = ctr
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 28

    col_keys = [c[0] for c in columns]
    for ri, (_, row) in enumerate(df.iterrows(), start=4):
        base_fill = PatternFill('solid', start_color=COLOR_ALT) if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(col_keys, 1):
            val = row.get(key, '')
            if hasattr(val, 'item'):
                val = val.item()
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = brd
            if highlight_col and key == highlight_col:
                c.font = Font(name='Arial', size=9, bold=True, color='8B0000')
                c.fill = PatternFill('solid', start_color=COLOR_DEV if ri % 2 == 0 else COLOR_DEV_ALT)
            else:
                c.font = Font(name='Arial', size=9)
                c.fill = base_fill
            c.alignment = lft if key in _LEFT_ALIGN_COLS else ctr
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}3'


def build_workbook(df_dev, today, out_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Αποκλίσεις'
    _write_sheet(ws1, df_dev, CHECK_TITLE, DEV_COLUMNS, COLOR_HDR, today,
                 subtitle_extra='  |  Αποκλίσεις', highlight_col='Αναμενόμενο Ωράριο')
    wb.save(out_path)


# ── CUSTOM RUN ────────────────────────────────────────────────────────────
def run(config):
    import core.framework as _fw
    _fw._current_check_title = CHECK_TITLE

    print('=' * 65)
    print(f'  {CHECK_TITLE}')
    print('=' * 65)

    path_21 = get_downloaded_file('2.1', 'Αρχείο 2.1 (Κατάλογος σχολείων):',         silent=True)
    path_41 = get_downloaded_file('4.1', 'Αρχείο 4.1 (Οργανικές τοποθετήσεις):',     silent=True)
    path_42 = get_downloaded_file('4.2', 'Αρχείο 4.2 (Αποσπασμένοι εκπαιδευτικοί):', silent=True)

    if path_21 is None or path_41 is None:
        _missing_file_dialog(CHECK_TITLE, REQUIRED_REPORTS)
        return
    if path_42 is None:
        print('  ℹ Αρχείο 4.2 (Αποσπασμένοι) δεν βρέθηκε — θα αγνοηθούν αποσπάσεις.')

    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    print(f'\n  Ημερομηνία : {today.strftime("%d/%m/%Y")}')
    print('-' * 65)

    print('\nΕπεξεργασία...')
    df_dev = process(path_21, path_41, path_42)

    print(f'  ✓ Αποκλίσεις : {len(df_dev)}')

    if df_dev.empty:
        _show_results_popup(
            CHECK_TITLE,
            f'Ημερομηνία ελέγχου: {today.strftime("%d/%m/%Y")}\n\n'
            f'✓  Δεν βρέθηκαν αποκλίσεις υποχρεωτικού ωραρίου ΠΕ60/ΠΕ60.50.\n\n'
            f'Ο έλεγχος ολοκληρώθηκε χωρίς θέματα.',
            result_type='ok'
        )
        return

    _docs   = os.path.join(os.path.expanduser('~'), 'Documents', 'MySchoolChecks')
    out_dir = os.path.join(_docs, f'results_{today.strftime("%Y%m%d")}', RESULTS_FOLDER)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_{RESULTS_FOLDER}.xlsx')

    try:
        build_workbook(df_dev, today, out_path)
        print(f'\n  ✓ Αποθηκεύτηκε: {os.path.basename(out_path)}')
    except PermissionError:
        import tkinter.messagebox as _mb
        _mb.showwarning(
            'Αρχείο ανοιχτό',
            f'Το αρχείο {os.path.basename(out_path)} είναι ανοιχτό σε άλλο πρόγραμμα.\n'
            f'Κλείστε το και τρέξτε ξανά τον έλεγχο.'
        )
        return

    cat_counts = df_dev['Κατηγορία'].value_counts().to_dict()
    body = (
        f'Σύνοψη ελέγχου υποχρεωτικού ωραρίου ΠΕ60 — {today.strftime("%d/%m/%Y")}\n'
        f'{"─"*50}\n'
        f'Αποκλίσεις: {len(df_dev)}\n'
        f'  • Κατηγορία 1 (Λειτ. 1-3, αναμενόμενο 25): {cat_counts.get(1, 0)}\n'
        f'  • Κατηγορία 2 (Λειτ. ≥4, αναμενόμενο ≠25):  {cat_counts.get(2, 0)}\n\n'
        f'{"─"*50}\n'
        f'Αποτελέσματα αποθηκεύτηκαν στο φάκελο:\n{out_dir}'
    )
    _show_results_popup(CHECK_TITLE, body, result_type='warn', excel_path=out_path)
