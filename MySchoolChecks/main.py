"""
main.py
═══════
Κεντρικό σημείο εισόδου — MySchool Αυτοματισμοί.
"""

import sys, os, importlib, threading, queue, time, warnings
warnings.filterwarnings('ignore', message='.*file size.*not 512.*sector size.*')

# DEBUG: Καταγραφή crashes σε αρχείο
def _emergency_log(exc_type, exc_val, exc_tb):
    import traceback
    try:
        _log = os.path.join(os.path.expanduser('~'), 'Desktop', 'crash.log')
        with open(_log, 'w', encoding='utf-8') as _f:
            _f.write(''.join(traceback.format_exception(exc_type, exc_val, exc_tb)))
    except Exception:
        pass
    sys.__excepthook__(exc_type, exc_val, exc_tb)
sys.excepthook = _emergency_log


def _app_base():
    """Επιστρέφει τον βασικό φάκελο για αποθήκευση δεδομένων.
    - Program Files → %LOCALAPPDATA%\MySchoolChecks  (δεν επιτρέπεται εγγραφή στο PF)
    - dist\ (development/portable) → δίπλα στο .exe
    - Development → φάκελος του κώδικα
    """
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        pf  = os.environ.get('PROGRAMFILES',       r'C:\Program Files').lower()
        pf86= os.environ.get('PROGRAMFILES(X86)',  r'C:\Program Files (x86)').lower()
        if exe_dir.lower().startswith(pf) or exe_dir.lower().startswith(pf86):
            data_dir = os.path.join(
                os.environ.get('LOCALAPPDATA', os.path.expanduser('~')),
                'MySchoolChecks')
            os.makedirs(data_dir, exist_ok=True)
            return data_dir
        return exe_dir
    return os.path.dirname(os.path.abspath(__file__))


def _docs_base():
    """Επιστρέφει τον φάκελο Documents\MySchoolChecks — κοινός για downloads και results.
    Εύκολος στην πρόσβαση από τον χρήστη.
    """
    _docs = os.path.join(os.path.expanduser('~'), 'Documents')
    path  = os.path.join(_docs, 'MySchoolChecks')
    os.makedirs(path, exist_ok=True)
    return path
import tkinter as tk
from tkinter import messagebox

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import config
except ImportError:
    import tkinter as tk
    from tkinter import messagebox
    _r = tk.Tk(); _r.withdraw()
    messagebox.showerror('Σφάλμα', 'Δεν βρέθηκε το config.py στον φάκελο του προγράμματος.')
    sys.exit(1)


class GUIStream:
    """Ανακατευθύνει το sys.stdout στο status bar του GUI."""
    def __init__(self):
        self._callback = None
        self._buffer   = []

    def set_callback(self, cb):
        self._callback = cb
        # Εκκρεμή μηνύματα
        for msg in self._buffer:
            cb(msg)
        self._buffer.clear()

    def write(self, text):
        text = text.strip()
        if not text:
            return
        if self._callback:
            self._callback(text)
        else:
            self._buffer.append(text)

    def flush(self):
        pass


_gui_stream = GUIStream()
sys.stdout  = _gui_stream
sys.stderr  = _gui_stream

try:
    import ctypes
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('myschool.checks.1')
except Exception:
    pass

WHITE = '#FFFFFF'
C = {
    'bg'        : '#EEF4F0',
    'bg2'       : '#E2EDEA',
    'border'    : '#C4D9D0',
    'hdr_bg'    : '#1F4E79',
    'hdr_fg'    : '#FFFFFF',
    'hdr_sub'   : '#D6E4F0',
    'btn_bg'    : '#1F4E79',
    'btn_fg'    : '#FFFFFF',
    'btn_act'   : '#2E75B6',
    'btn_dis'   : '#7F9FAF',
    'sel_bg'    : '#D0E8DC',
    'sel_bd'    : '#4CA870',
    'norm_bg'   : '#FFFFFF',
    'norm_bd'   : '#C4D9D0',
    'desc'      : '#4A6860',
    'footer'    : '#7A9A90',
    'status_ok' : '#2E7D32',
    'status_err': '#C62828',
    'status_run': '#E65100',
    'ind_idle'  : '#B0BEC5',
    'ind_run'   : '#FB8C00',
    'ind_ok'    : '#43A047',
    'ind_err'   : '#E53935',
    'ind_out'   : '#ECEFF1',
    'warn'      : '#E65100',
}


CHECK_ORDER = [
    'forma_82',
    'orario_diafora',
    'arnhtika_ypoloipa',
    'adies_aneu',
    'adies',
    'apontes_xwris_adeia',
    'analipsi',
    'dioikitiko_ergo',
    'ypoloipa',
    'tmimata_genikis',
]

# Sentinel για τη γραμμή «Έλεγχος Ε.Ε.Α.» στην κεντρική λίστα — δεν είναι
# module τύπου checks/*.py (δεν έχει ask_inputs/process), ανοίγει το δικό
# του 4-tab SmeaeDialog αντί για το CheckRunDialog. Βλ. LauncherApp._build_ui.
SMEAE_MARKER = '__SMEAE__'

# Checks που εξαιρούνται από το κεντρικό μενού (π.χ. έχουν μεταφερθεί αλλού)
CHECKS_EXCLUDED = {'orario_pe60'}

def load_checks():
    base = os.path.dirname(os.path.abspath(__file__))
    if base not in sys.path:
        sys.path.insert(0, base)

    checks = []

    if getattr(sys, 'frozen', False):
        # ── Frozen exe (PyInstaller) ──────────────────────────────────
        # Τα checks είναι compiled μέσα στο exe. Τα φορτώνουμε με
        # τη σειρά του CHECK_ORDER + όποια άλλα γνωστά.
        # Δεν χρειαζόμαστε os.listdir() — απευθείας import.
        all_known = list(CHECK_ORDER) + [
            'forma_82', 'orario_diafora', 'arnhtika_ypoloipa',
            'adies_aneu', 'adies', 'apontes_xwris_adeia', 'analipsi', 'dioikitiko_ergo', 'ypoloipa',
            'tmimata_genikis',
        ]
        seen = set()
        ordered = []
        for m in all_known:
            if m not in seen:
                seen.add(m)
                ordered.append(m)
    else:
        # ── Development mode ─────────────────────────────────────────
        checks_dir = os.path.join(base, 'checks')
        available = {fname[:-3] for fname in os.listdir(checks_dir)
                     if fname.endswith('.py') and not fname.startswith('_')}
        ordered = CHECK_ORDER + sorted(available - set(CHECK_ORDER) - CHECKS_EXCLUDED)

    _log = os.path.join(os.path.expanduser('~'), 'Desktop', 'checks_errors.log')
    with open(_log, 'w', encoding='utf-8') as _f:
        _f.write(f'frozen={getattr(sys, "frozen", False)}\n')
        _f.write(f'sys.path={sys.path}\n\n')

    for mod_base in ordered:
        mod_name = f'checks.{mod_base}'
        try:
            mod = importlib.import_module(mod_name)
            title = getattr(mod, 'CHECK_TITLE', None)
            if title:
                # Σημείωση: το "(Απαιτούνται: ...)" ΔΕΝ προστίθεται πια εδώ —
                # φαίνεται πλέον μέσα στο tab «⬇ Λήψη» όταν ανοίγει ο
                # έλεγχος (core/check_dialog.py), οπότε η περιγραφή στη
                # λίστα μένει καθαρή.
                desc = getattr(mod, 'CHECK_DESCRIPTION', '')
                checks.append((title, desc, mod))
        except Exception as e:
            import traceback as _tb
            with open(_log, 'a', encoding='utf-8') as _f:
                _f.write(f'\n--- {mod_base} ---\n')
                _f.write(_tb.format_exc())
            print(f'  !! Den fortothike to {mod_base}.py: {e}')
    return checks


def password_is_set():
    """Ελέγχει αν ο κωδικός email έχει οριστεί (keyring ή config)."""
    try:
        import keyring
        val = keyring.get_password('MySchoolChecks', 'FROM_PASSWORD')
        if val:
            return True
    except Exception:
        pass
    return bool(getattr(config, 'FROM_PASSWORD', ''))


def _get_local_settings_path():
    return os.path.join(_app_base(), 'data', 'local_settings.json')


def _load_local_settings():
    """Φορτώνει το data/local_settings.json. Επιστρέφει dict."""
    import json
    path = _get_local_settings_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


_SENSITIVE_KEYS = {'MYSCHOOL_USER', 'MYSCHOOL_PASS', 'FROM_PASSWORD'}
_KEYRING_SERVICE = 'MySchoolChecks'


def _save_config(updates):
    """
    Αποθηκεύει ρυθμίσεις:
      - Ευαίσθητα (MYSCHOOL_USER/PASS, FROM_PASSWORD) → Windows Credential Manager
      - Μη-ευαίσθητα → data/local_settings.json
    Ενημερώνει επίσης το live config object ώστε να ισχύουν άμεσα.
    """
    import json

    sensitive   = {k: v for k, v in updates.items() if k in _SENSITIVE_KEYS}
    nonsensitive = {k: v for k, v in updates.items() if k not in _SENSITIVE_KEYS}

    # ── Αποθήκευση ευαίσθητων στο keyring ───────────────────────────────────
    try:
        import keyring
        for key, val in sensitive.items():
            if val:
                keyring.set_password(_KEYRING_SERVICE, key, val)
    except Exception as e:
        # Αν το keyring αποτύχει, πέσε back στο JSON (δεν σπάμε τη ροή)
        nonsensitive.update(sensitive)
        print(f'[Προσοχή] keyring μη διαθέσιμο, credentials αποθηκεύονται στο JSON: {e}')

    # ── Αποθήκευση μη-ευαίσθητων στο JSON ───────────────────────────────────
    if nonsensitive:
        path = _get_local_settings_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        existing = _load_local_settings()
        # Βεβαιώσου ότι δεν ξαναμπαίνουν sensitive στο JSON
        for sk in _SENSITIVE_KEYS:
            existing.pop(sk, None)
        existing.update(nonsensitive)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)

    # ── Ενημέρωση live config object ────────────────────────────────────────
    for key, val in updates.items():
        setattr(config, key, val)


class Indicator(tk.Canvas):
    SIZE = 14

    def __init__(self, parent, **kw):
        bg = kw.pop('bg', C['norm_bg'])
        super().__init__(parent, width=self.SIZE, height=self.SIZE,
                         bg=bg, highlightthickness=0, **kw)
        self._circle = self.create_oval(1, 1, self.SIZE-1, self.SIZE-1,
                                        fill=C['ind_idle'],
                                        outline=C['ind_out'], width=1)

    def set_state(self, state):
        colors = {
            'idle'   : (C['ind_idle'], C['ind_out']),
            'running': (C['ind_run'],  '#E65100'),
            'ok'     : (C['ind_ok'],   '#2E7D32'),
            'error'  : (C['ind_err'],  '#B71C1C'),
        }
        fill, outline = colors.get(state, colors['idle'])
        self.itemconfig(self._circle, fill=fill, outline=outline)


class SettingsDialog(tk.Toplevel):
    """Παράθυρο ρυθμίσεων — με tabs: Σύνδεση / Email / Αρχεία."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Ρυθμίσεις')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._ady_new_path = tk.StringVar(value='')
        self._build()

        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{px}+{py}')

    # ── helpers ──────────────────────────────────────────────────────────────
    def _cfg(self, key, default=''):
        return getattr(config, key, default)

    def _labeled_entry(self, parent, row, label, var, width=28, show=''):
        """Βοηθητικό: label + entry σε grid."""
        tk.Label(parent, text=label, bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9), anchor='w').grid(
                 row=row, column=0, sticky='w', pady=(6, 2))
        e = tk.Entry(parent, textvariable=var, width=width,
                     font=('Arial', 10), relief='solid', bd=1, show=show)
        e.grid(row=row, column=1, sticky='w', padx=(10, 0), pady=(6, 2))
        return e

    def _pw_row(self, parent, row, label, var):
        """Label + entry κωδικού με κουμπί 👁."""
        tk.Label(parent, text=label, bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9), anchor='w').grid(
                 row=row, column=0, sticky='w', pady=(6, 2))
        frame = tk.Frame(parent, bg=C['bg'])
        frame.grid(row=row, column=1, sticky='w', padx=(10, 0), pady=(6, 2))
        show_var = tk.BooleanVar(value=False)
        entry = tk.Entry(frame, textvariable=var, show='•',
                         width=24, font=('Arial', 10), relief='solid', bd=1)
        entry.pack(side='left')
        tk.Button(frame, text='👁', bg=C['bg'], relief='flat',
                  font=('Arial', 11), cursor='hand2',
                  command=lambda: [show_var.set(not show_var.get()),
                                   entry.configure(show='' if show_var.get() else '•')]
                  ).pack(side='left', padx=(4, 0))
        return entry

    def _section_label(self, parent, row, text):
        tk.Label(parent, text=text, bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').grid(
                 row=row, column=0, columnspan=2, sticky='w', pady=(14, 4))

    # ── build ─────────────────────────────────────────────────────────────────
    def _build(self):
        from tkinter import ttk

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C['hdr_bg'], pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='⚙  Ρυθμίσεις', bg=C['hdr_bg'], fg=C['hdr_fg'],
                 font=('Arial', 12, 'bold')).pack()

        # ── ttk.Notebook ─────────────────────────────────────────────────────
        # Σημ: στα Windows το theme override κάνει το foreground αόρατο σε σκούρο bg.
        # Λύση: επιλεγμένο tab = ανοιχτό μπλε (hdr_sub) με σκούρο κείμενο (hdr_bg).
        style = ttk.Style()
        style.configure('TNotebook',     background=C['bg'])
        style.configure('TNotebook.Tab', background=C['bg2'], foreground=C['desc'],
                        font=('Arial', 9, 'bold'), padding=(12, 5))
        style.map('TNotebook.Tab',
                  background=[('selected', C['hdr_sub']), ('active', C['sel_bg'])],
                  foreground=[('selected', C['hdr_bg']),  ('active', C['hdr_bg'])])

        nb = ttk.Notebook(self)
        nb.pack(fill='both', padx=16, pady=12)

        tab1 = tk.Frame(nb, bg=C['bg'], padx=20, pady=14)
        tab2 = tk.Frame(nb, bg=C['bg'], padx=20, pady=14)
        tab3 = tk.Frame(nb, bg=C['bg'], padx=20, pady=14)
        nb.add(tab1, text='  Σύνδεση  ')
        nb.add(tab2, text='  Email  ')
        nb.add(tab3, text='  Αρχεία  ')

        # ── Tab 1: Σύνδεση (MySchool + email password) ───────────────────────
        self._section_label(tab1, 0, 'MySchool (SSO):')
        self._ms_user_var = tk.StringVar(value=self._cfg('MYSCHOOL_USER'))
        self._labeled_entry(tab1, 1, 'Username:', self._ms_user_var)
        self._ms_pass_var = tk.StringVar(value=self._cfg('MYSCHOOL_PASS'))
        self._pw_row(tab1, 2, 'Κωδικός:', self._ms_pass_var)

        sep = tk.Frame(tab1, bg=C['border'], height=1)
        sep.grid(row=3, column=0, columnspan=2, sticky='ew', pady=(16, 4))

        # ── Επιλογή Browser ───────────────────────────────────────────────────
        self._section_label(tab1, 4, 'Browser για σύνδεση:')
        self._browser_var = tk.StringVar(value=self._cfg('BROWSER', 'chrome'))
        br_frame = tk.Frame(tab1, bg=C['bg'])
        br_frame.grid(row=5, column=0, columnspan=2, sticky='w', pady=(2, 4))
        tk.Radiobutton(br_frame, text='Chrome', variable=self._browser_var,
                       value='chrome', bg=C['bg'], fg=C['hdr_bg'],
                       font=('Arial', 9), activebackground=C['bg'],
                       selectcolor=C['bg2']).pack(side='left', padx=(0, 16))
        tk.Radiobutton(br_frame, text='Firefox', variable=self._browser_var,
                       value='firefox', bg=C['bg'], fg=C['hdr_bg'],
                       font=('Arial', 9), activebackground=C['bg'],
                       selectcolor=C['bg2']).pack(side='left')
        tk.Label(tab1, text='(και οι δύο πρέπει να είναι εγκατεστημένοι για να επιλεγούν)',
                 bg=C['bg'], fg=C['footer'], font=('Arial', 8)).grid(
                 row=6, column=0, columnspan=2, sticky='w')


        # ── Tab 2: Email (ταυτότητα αποστολέα) ───────────────────────────────
        self._section_label(tab2, 0, 'Στοιχεία αποστολέα:')
        self._from_name_var  = tk.StringVar(value=self._cfg('FROM_NAME'))
        self._from_email_var = tk.StringVar(value=self._cfg('FROM_EMAIL'))
        self._smtp_var       = tk.StringVar(value=self._cfg('SMTP_HOST'))

        self._labeled_entry(tab2, 1, 'Εμφανιζόμενο όνομα:', self._from_name_var,  width=30)
        self._labeled_entry(tab2, 2, 'Email αποστολής:',     self._from_email_var, width=30)
        tk.Label(tab2, text='(χρησιμοποιείται και για δοκιμαστική αποστολή)',
                 bg=C['bg'], fg=C['footer'], font=('Arial', 8)).grid(
                 row=3, column=1, sticky='w', padx=(10, 0))

        sep2 = tk.Frame(tab2, bg=C['border'], height=1)
        sep2.grid(row=4, column=0, columnspan=2, sticky='ew', pady=(14, 4))

        self._section_label(tab2, 5, 'Διακομιστής:')
        self._labeled_entry(tab2, 6, 'SMTP Host:', self._smtp_var, width=24)
        tk.Label(tab2, text='(π.χ. mail.sch.gr)', bg=C['bg'], fg=C['footer'],
                 font=('Arial', 8)).grid(row=7, column=1, sticky='w', padx=(10, 0))

        sep3 = tk.Frame(tab2, bg=C['border'], height=1)
        sep3.grid(row=8, column=0, columnspan=2, sticky='ew', pady=(14, 4))

        self._section_label(tab2, 9, 'Υπογραφή email:')
        self._sig_text = tk.Text(tab2, height=6, width=36,
                                  font=('Consolas', 9), relief='solid', bd=1,
                                  bg='white', fg='#1a1a2e',
                                  wrap='word', padx=6, pady=4)
        self._sig_text.grid(row=10, column=0, columnspan=2, sticky='ew', pady=(4, 2))
        sig_val = self._cfg('EMAIL_SIGNATURE')
        if sig_val:
            self._sig_text.insert('1.0', sig_val)
        tk.Label(tab2, text='(εμφανίζεται στο τέλος κάθε email)',
                 bg=C['bg'], fg=C['footer'], font=('Arial', 8)).grid(
                 row=11, column=0, columnspan=2, sticky='w')

        sep4 = tk.Frame(tab2, bg=C['border'], height=1)
        sep4.grid(row=12, column=0, columnspan=2, sticky='ew', pady=(14, 4))

        self._section_label(tab2, 13, 'Λογαριασμός email (προαιρετικό):')
        self._pw_var = tk.StringVar(value=self._cfg('FROM_PASSWORD'))
        self._pw_row(tab2, 14, 'Κωδικός email:', self._pw_var)

        hint = tk.Frame(tab2, bg='#E8F4FD',
                        highlightbackground='#90CAF9', highlightthickness=1)
        hint.grid(row=15, column=0, columnspan=2, sticky='ew', pady=(8, 0))
        tk.Label(hint, text='ℹ  Συμπληρώστε μόνο αν επιθυμείτε αυτόματη αποστολή email.',
                 bg='#E8F4FD', fg='#1565C0',
                 font=('Arial', 8), padx=8, pady=4).pack(anchor='w')

        # ── Tab 3: Αρχεία ─────────────────────────────────────────────────────
        self._section_label(tab3, 0, 'Αρχείο Αδυνατούντων (υπό έγκριση):')

        current_ady = self._cfg('ADY_XORIS_EGKRISI_PATH')
        current_lbl = (os.path.basename(current_ady)
                       if current_ady and os.path.exists(current_ady)
                       else '— δεν έχει οριστεί —')

        ady_frame = tk.Frame(tab3, bg=C['bg'])
        ady_frame.grid(row=1, column=0, columnspan=2, sticky='w', pady=(4, 0))

        self._ady_lbl = tk.Label(ady_frame, text=current_lbl,
                                  bg=C['bg'], fg=C['desc'],
                                  font=('Arial', 8), anchor='w', wraplength=240)
        self._ady_lbl.pack(side='left', padx=(0, 10))

        tk.Button(ady_frame, text='Αλλαγή...',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 8), relief='flat', padx=8, pady=2,
                  cursor='hand2',
                  command=self._browse_ady).pack(side='left')

        tk.Label(tab3,
                 text='Χρησιμοποιείται στον έλεγχο Υπολοίπων.\n'
                      'Ανεβάστε νέο αρχείο μόνο αν άλλαξαν οι αδυνατούντες υπό έγκριση.',
                 bg=C['bg'], fg=C['footer'], font=('Arial', 8),
                 justify='left', anchor='w').grid(
                 row=2, column=0, columnspan=2, sticky='w', pady=(10, 0))

        # ── Κουμπιά ───────────────────────────────────────────────────────────
        btn_frame = tk.Frame(self, bg=C['bg2'], pady=12)
        btn_frame.pack(fill='x')

        tk.Button(btn_frame, text='Αποθήκευση',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 10, 'bold'),
                  relief='flat', padx=18, pady=6,
                  cursor='hand2',
                  command=self._save).pack(side='right', padx=16)

        tk.Button(btn_frame, text='Ακύρωση',
                  bg=C['bg2'], fg=C['desc'],
                  font=('Arial', 10),
                  relief='flat', padx=12, pady=6,
                  cursor='hand2',
                  command=self.destroy).pack(side='right', padx=4)

    def _browse_ady(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title='Αρχείο Αδυνατούντων (υπό έγκριση)',
            filetypes=[('CSV & Excel', '*.csv *.xlsx *.xls'), ('Όλα τα αρχεία', '*.*')],
            parent=self
        )
        if path and os.path.exists(path):
            self._ady_new_path.set(path)
            self._ady_lbl.configure(text=os.path.basename(path), fg=C['status_ok'])

    def _save(self):
        try:
            from_email = self._from_email_var.get().strip()
            updates = {
                'MYSCHOOL_USER'  : self._ms_user_var.get().strip(),
                'MYSCHOOL_PASS'  : self._ms_pass_var.get().strip(),
                'FROM_PASSWORD'  : self._pw_var.get().strip(),
                'FROM_NAME'      : self._from_name_var.get().strip(),
                'FROM_EMAIL'     : from_email,
                'TEST_EMAIL'     : from_email,   # ίδιο με FROM_EMAIL
                'SMTP_HOST'      : self._smtp_var.get().strip(),
                'EMAIL_SIGNATURE': self._sig_text.get('1.0', tk.END).strip(),
                'BROWSER'        : self._browser_var.get(),
            }

            # Αν επιλέχθηκε νέο αρχείο Αδυνατούντων
            new_ady = self._ady_new_path.get().strip()
            if new_ady and os.path.exists(new_ady):
                data_dir = os.path.join(_app_base(), 'data')
                os.makedirs(data_dir, exist_ok=True)
                ext  = os.path.splitext(new_ady)[1]
                dest = os.path.join(data_dir, f'adynatountes_ypo_egkrisi{ext}')
                import shutil as _sh
                _sh.copy2(new_ady, dest)
                updates['ADY_XORIS_EGKRISI_PATH'] = dest

            _save_config(updates)
            messagebox.showinfo('Επιτυχία', 'Οι ρυθμίσεις αποθηκεύτηκαν.', parent=self)
            self.destroy()
        except Exception as e:
            messagebox.showerror('Σφάλμα', f'Δεν ήταν δυνατή η αποθήκευση:\n{e}',
                                 parent=self)


class DownloadDialog(tk.Toplevel):
    """Παράθυρο κατεβάσματος δεδομένων από MySchool."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Λήψη Δεδομένων MySchool')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{px}+{py}')

    def _build(self):
        from core.downloader import downloads_info, REPORTS, get_downloads_dir, FILE_PREFIX_MAP
        import glob as _glob
        base_dir = _docs_base()

        # Βρες αρχεία που υπάρχουν ήδη στον σημερινό φάκελο
        today_dir = get_downloads_dir(base_dir)
        self._already_have = set()
        already_have = self._already_have
        if os.path.exists(today_dir):
            for rid, prefix in FILE_PREFIX_MAP.items():
                matches = [f for f in _glob.glob(os.path.join(today_dir, f'{prefix}*'))
                           if not f.endswith('.tmp') and not f.endswith('.crdownload')]
                if matches:
                    already_have.add(rid)

        # Header
        hdr = tk.Frame(self, bg=C['hdr_bg'], pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='⬇  Λήψη Δεδομένων MySchool',
                 bg=C['hdr_bg'], fg=C['hdr_fg'],
                 font=('Arial', 12, 'bold')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=20, pady=14)
        body.pack(fill='both')

        # Έλεγχος credentials
        ms_user = getattr(config, 'MYSCHOOL_USER', '').strip()
        ms_pass = getattr(config, 'MYSCHOOL_PASS', '').strip()
        if not ms_user or not ms_pass:
            warn = tk.Frame(body, bg='#FFF3E0',
                            highlightbackground='#FFB74D',
                            highlightthickness=1)
            warn.pack(fill='x', pady=(0, 12))
            tk.Label(warn,
                     text='Τα στοιχεία σύνδεσης MySchool δεν εχουν οριστει. '
                          'Πηγαινε στις Ρυθμισεις για να τα συμπληρωσεις.',
                     bg='#FFF3E0', fg=C['warn'],
                     font=('Arial', 9), padx=10, pady=8,
                     justify='left').pack(anchor='w')

        # Σημερινός φάκελος / τελευταία λήψη
        if already_have:
            from datetime import datetime as _dt
            today_str = _dt.now().strftime('%d/%m/%Y')
            tk.Label(body,
                     text=f'Σήμερα ({today_str}): {len(already_have)}/{len(REPORTS)} αρχεία υπάρχουν ήδη.',
                     bg=C['bg'], fg=C['status_ok'], font=('Arial', 8)).pack(anchor='w', pady=(0,4))
        else:
            info = downloads_info(base_dir)
            if info:
                ts_str, found, age_min = info
                age_txt = f'{age_min} λεπτά' if age_min < 60 else f'{age_min//60}ω {age_min%60}λ'
                tk.Label(body,
                         text=f'Τελευταία λήψη: {ts_str}  ({age_txt} πριν)  —  {len(found)}/{len(REPORTS)} αρχεία',
                         bg=C['bg'], fg=C['desc'], font=('Arial', 8)).pack(anchor='w', pady=(0,4))
            else:
                tk.Label(body, text='Δεν υπάρχουν αποθηκευμένα δεδομένα.',
                         bg=C['bg'], fg=C['desc'], font=('Arial', 8)).pack(anchor='w', pady=(0,4))

        if already_have:
            tk.Label(body,
                     text='Τα ✓ αρχεία υπάρχουν ήδη. Επέλεξέ τα αν θέλεις να τα κατεβάσεις ξανά.',
                     bg=C['bg'], fg=C['desc'], font=('Arial', 8, 'italic')).pack(anchor='w', pady=(0,8))

        # Επιλογή αρχείων
        hdr_row = tk.Frame(body, bg=C['bg'])
        hdr_row.pack(fill='x', pady=(0,6))
        tk.Label(hdr_row, text='Επιλέξτε αρχεία για λήψη:',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold')).pack(side='left')
        self._dl_all_selected = False
        self._dl_all_btn = tk.Button(hdr_row, text='Όλα',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 8), relief='flat', padx=6, pady=1,
                  cursor='hand2', command=self._toggle_dl_all)
        self._dl_all_btn.pack(side='left', padx=(10,0))

        self._report_vars = {}
        grid = tk.Frame(body, bg=C['bg'])
        grid.pack(fill='x')

        for i, (rid, label, *_) in enumerate(REPORTS):
            exists = rid in already_have
            var = tk.BooleanVar(value=False)
            self._report_vars[rid] = var
            row, col = divmod(i, 2)
            prefix   = '' if rid in ('topoth',) else (f'4.26/4.27 — ' if rid == '4.26' else f'{rid} — ')
            lbl_text = f'✓ {prefix}{label}' if exists else f'{prefix}{label}'
            lbl_fg   = C['status_ok'] if exists else C['fg'] if 'fg' in C else '#000000'
            tk.Checkbutton(grid, text=lbl_text,
                           variable=var,
                           fg=lbl_fg,
                           bg=C['bg'], font=('Arial', 9),
                           activebackground=C['bg']).grid(
                           row=row, column=col, sticky='w', padx=(0,16), pady=2)

        # Progress label
        self._progress_var = tk.StringVar(value='')
        self._progress_lbl = tk.Label(body, textvariable=self._progress_var,
                                       bg=C['bg'], fg=C['status_run'],
                                       font=('Arial', 8), wraplength=380, justify='left')
        self._progress_lbl.pack(anchor='w', pady=(10,0))

        # Κουμπιά
        btn_frame = tk.Frame(self, bg=C['bg2'], pady=12)
        btn_frame.pack(fill='x')

        self._start_btn = tk.Button(btn_frame, text='⬇  Έναρξη Λήψης',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 10, 'bold'),
                  relief='flat', padx=18, pady=6,
                  cursor='hand2',
                  command=self._start).pack(side='right', padx=16)

        tk.Button(btn_frame, text='Κλείσιμο',
                  bg=C['bg2'], fg=C['desc'],
                  font=('Arial', 10),
                  relief='flat', padx=12, pady=6,
                  cursor='hand2',
                  command=self.destroy).pack(side='right', padx=4)

    def _toggle_dl_all(self):
        self._dl_all_selected = not self._dl_all_selected
        for v in self._report_vars.values():
            v.set(self._dl_all_selected)
        self._dl_all_btn.config(text='Κανένα' if self._dl_all_selected else 'Όλα')

    def _start(self):
        from core.downloader import (MySchoolDownloader,
                                      get_downloads_dir,
                                      cleanup_old_downloads)

        ms_user = getattr(config, 'MYSCHOOL_USER', '').strip()
        ms_pass = getattr(config, 'MYSCHOOL_PASS', '').strip()

        if not ms_user or not ms_pass:
            messagebox.showwarning('Προσοχή',
                'Συμπλήρωσε username και κωδικό MySchool στις Ρυθμίσεις (⚙).',
                parent=self)
            return

        selected = [rid for rid, var in self._report_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning('Προσοχή', 'Επίλεξε τουλάχιστον ένα αρχείο.', parent=self)
            return

        # Αρχεία που ο χρήστης επέλεξε ενώ υπάρχουν ήδη → force re-download
        force_rids = [rid for rid in selected if rid in self._already_have]

        base_dir = _docs_base()
        dest_dir = get_downloads_dir(base_dir)

        self._progress_var.set('Εκκίνηση...')
        self.update()

        def on_progress(msg):
            self.after(0, lambda m=msg: self._progress_var.set(m))

        def task():
            try:
                dl = MySchoolDownloader(
                    username=ms_user,
                    password=ms_pass,
                    dest_dir=dest_dir,
                    callback=on_progress,
                    reports=selected,
                    browser=getattr(config, 'BROWSER', 'chrome'),
                    force=force_rids,
                )
                results = dl.run()
                ok   = sum(1 for v in results.values() if v)
                fail = len(results) - ok

                # Κράτα μόνο τον τελευταίο φάκελο
                cleanup_old_downloads(base_dir, keep=1)

                from core.downloader import REPORTS as _REPS
                _rid_label = {r[0]: (f'4.26/4.27 — {r[1]}' if r[0] == '4.26' else f'{r[0]} — {r[1]}' if r[0] not in ('topoth',) else r[1]) for r in _REPS}
                failed_names = [_rid_label.get(rid, rid) for rid, v in results.items() if not v]

                msg = f'Ολοκληρώθηκε: {ok}/{len(results)} αρχεία κατεβήκαν.'
                if failed_names:
                    msg += '\n\nΔεν κατέβηκε:\n' + '\n'.join(f'• {n}' for n in failed_names)
                self.after(0, lambda m=msg: [
                    self._progress_var.set(m.split('\n')[0]),
                    messagebox.showinfo('Λήψη', m, parent=self)
                ])
            except Exception as e:
                err = str(e)
                self.after(0, lambda m=err: [
                    self._progress_var.set(f'Σφάλμα: {m}'),
                    messagebox.showerror('Σφάλμα Λήψης', m, parent=self)
                ])

        threading.Thread(target=task, daemon=True).start()


class LauncherApp:

    def __init__(self, root, checks):
        self.root         = root
        self.checks       = list(checks) + [(
            'Έλεγχος Ε.Ε.Α. — Στατιστικά',
            'Λήψη, σύγκριση, διαχωρισμός ανά σχολείο και αποστολή στατιστικών ΣΜΕΑΕ/ΕΕΑ',
            SMEAE_MARKER,
        )]
        self.check_frames = []

        root.title('MySchool Checks')
        root.configure(bg=C['bg'])

        self._build_ui()

        # Παγώνουμε το μέγεθος μετά το πλήρες render
        root.update_idletasks()
        root.update()
        root.resizable(False, False)

        # Σύνδεση stdout με status bar
        _gui_stream.set_callback(self._on_print)

        # Έλεγχος για νέα έκδοση (background, αθόρυβος — 3s μετά την εκκίνηση)
        def _on_update(new_ver, dl_url):
            def _show():
                self._set_status(
                    f'⬆  Διαθέσιμη νέα έκδοση v{new_ver} — κλικ για ενημέρωση',
                    C['status_run'])
                self.status_lbl.configure(cursor='hand2')
                self.status_lbl.bind('<Button-1>',
                    lambda e: _do_update(self.root, new_ver, dl_url))
            self.root.after(0, _show)
        root.after(3000, lambda: _check_for_update(_on_update))

        # Αθόρυβο «ping» με τη Διεύθυνση Εκπαίδευσης (βλ. _send_install_ping)
        root.after(3000, _send_install_ping)

        # Αν δεν έχει οριστεί κωδικός, άνοιξε αυτόματα τις ρυθμίσεις
        if not password_is_set():
            root.after(400, self._open_settings)

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg=C['hdr_bg'], pady=8)
        hdr.pack(fill='x')

        # Μόνο ⚙ στο header
        btn_hdr = tk.Frame(hdr, bg=C['hdr_bg'])
        btn_hdr.place(relx=1.0, x=-6, y=2, anchor='ne')

        tk.Button(btn_hdr, text='⚙',
                  bg=C['hdr_bg'], fg=C['hdr_sub'],
                  font=('Arial', 13), relief='flat', cursor='hand2',
                  activebackground=C['hdr_bg'], activeforeground='white',
                  command=self._open_settings).pack(side='right', padx=(2, 0))

        tk.Button(btn_hdr, text='?',
                  bg=C['hdr_bg'], fg=C['hdr_sub'],
                  font=('Arial', 13, 'bold'), relief='flat', cursor='hand2',
                  activebackground=C['hdr_bg'], activeforeground='white',
                  command=self._open_help).pack(side='right', padx=(0, 2))

        tk.Label(hdr, text='MySchool Checks',
                 bg=C['hdr_bg'], fg=C['hdr_fg'],
                 font=('Arial', 15, 'bold')).pack()
        tk.Label(hdr, text='Δ/νση Π.Ε. Ανατολικής Θεσσαλονίκης',
                 bg=C['hdr_bg'], fg=C['hdr_sub'],
                 font=('Arial', 9)).pack()
        tk.Label(hdr, text=f'v{config.APP_VERSION}',
                 bg=C['hdr_bg'], fg=C['hdr_sub'],
                 font=('Arial', 8)).pack()

        # Ένδειξη αν ο κωδικός λείπει
        if not password_is_set():
            warn_bar = tk.Frame(self.root, bg='#FFF3E0',
                                highlightbackground='#FFB74D',
                                highlightthickness=1)
            warn_bar.pack(fill='x')
            tk.Label(warn_bar,
                     text='⚠  Ο κωδικός email δεν έχει οριστεί — κλικ στο ⚙ για να τον ορίσεις',
                     bg='#FFF3E0', fg=C['warn'],
                     font=('Arial', 8), padx=10, pady=4).pack(side='left')

        # Έλεγχος αν υπάρχουν αρχεία σήμερα (χρησιμοποιείται σε toolbar + status bar)
        try:
            import datetime as _dt2, os as _os2
            _today_str = _dt2.date.today().strftime('%Y%m%d')
            _dl_dir = _os2.path.join(_os2.path.expanduser('~'), 'Documents', 'MySchoolChecks', 'downloads', _today_str)
            _has_files = _os2.path.isdir(_dl_dir) and bool(list(_os2.scandir(_dl_dir)))
        except Exception:
            _has_files = False

        # Η παλιά γραμμή εργαλείων (toolbar/toolbar2) καταργήθηκε εντελώς —
        # όλες οι λειτουργίες μοιράστηκαν σε 3 tabs: Έλεγχοι / Αναφορές &
        # Στατιστικά / Ενέργειες MySchool (βλ. _build_checks_tab κ.λπ.).
        from tkinter import ttk
        style = ttk.Style()
        style.configure('TNotebook', background=C['bg'])
        style.configure('TNotebook.Tab', background=C['bg2'], foreground=C['desc'],
                        font=('Arial', 10, 'bold'), padding=(14, 7))
        style.map('TNotebook.Tab',
                  background=[('selected', C['hdr_sub']), ('active', C['sel_bg'])],
                  foreground=[('selected', C['hdr_bg']),  ('active', C['hdr_bg'])])

        nb = ttk.Notebook(self.root)
        nb.pack(fill='both', padx=10, pady=(8, 0))
        self._main_nb = nb

        tab_checks  = tk.Frame(nb, bg=C['bg'])
        tab_reports = tk.Frame(nb, bg=C['bg'])
        tab_actions = tk.Frame(nb, bg=C['bg'])
        tab_users   = tk.Frame(nb, bg=C['bg'])
        nb.add(tab_checks,  text='  🗂  Έλεγχοι  ')
        nb.add(tab_reports, text='  📊  Αναφορές / Στατιστικά  ')
        nb.add(tab_actions, text='  🏛  Ενέργειες MySchool  ')
        nb.add(tab_users,   text='  👥  Χρήστες  ')

        self._build_checks_tab(tab_checks)
        self._build_reports_tab(tab_reports)
        self._build_actions_tab(tab_actions)
        self._build_users_tab(tab_users)

        # Κουμπί ρυθμίσεων ⚙ στο header
        tk.Button(hdr, text='⚙', font=('Arial', 11),
                  bg=C['hdr_bg'], fg=C['hdr_sub'],
                  activebackground=C['btn_act'],
                  activeforeground=WHITE,
                  relief='flat', cursor='hand2',
                  bd=0, padx=6,
                  command=self._open_settings).place(relx=1.0, rely=0.0,
                                                      anchor='ne', x=-8, y=6)

        # Status bar
        _init_status = ('Έτοιμο  •  Δεδομένα σήμερα: ✓' if _has_files
                        else '💡 Άνοιξε έναν έλεγχο — κατεβάζει μόνος του ό,τι χρειάζεται')
        self.status_var = tk.StringVar(value=_init_status)
        status_bar = tk.Frame(self.root, bg=C['bg2'],
                              highlightbackground=C['border'],
                              highlightthickness=1)
        status_bar.pack(fill='x', side='bottom')
        self.status_lbl = tk.Label(status_bar,
                                    textvariable=self.status_var,
                                    bg=C['bg2'], fg=C['footer'],
                                    font=('Arial', 8), anchor='w',
                                    padx=10, pady=4)
        self.status_lbl.pack(side='left')
        tk.Label(status_bar,
                 text=f'{len(self.checks)} έλεγχοι  •  Μιχάλης Κατσιρντάκης  •  2310954145',
                 bg=C['bg2'], fg=C['footer'],
                 font=('Arial', 8), padx=10).pack(side='right')

    # ── Βοηθητικό: scrollable Canvas με mousewheel bound μόνο ενώ ο δείκτης
    #    είναι πάνω από το συγκεκριμένο canvas (χρειάζεται γιατί τώρα υπάρχουν
    #    3 ξεχωριστά scrollable tabs) ─────────────────────────────────────────
    def _make_scrollable(self, parent, n_items):
        _screen_h = self.root.winfo_screenheight()
        # Αρχική εκτίμηση ύψους (χρησιμοποιείται μέχρι να μετρηθεί το πραγματικό
        # ύψος του περιεχομένου — βλ. _finalize_scrollable). Γενναιόδωρη εκτίμηση
        # (78px/γραμμή) ώστε να μην εμφανίζεται περιττή μπάρα κύλισης όσο δεν
        # έχει γίνει ακόμα η τελική προσαρμογή.
        _canvas_h = min(max(n_items, 1) * 78 + 10, _screen_h - 380)
        _max_h = _screen_h - 380

        outer = tk.Frame(parent, bg=C['bg'])
        outer.pack(fill='x', padx=18, pady=(0, 10))

        canvas = tk.Canvas(outer, bg=C['bg'], height=_canvas_h, highlightthickness=0)
        vsb    = tk.Scrollbar(outer, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        canvas.pack(side='left', fill='x', expand=True)

        inner = tk.Frame(canvas, bg=C['bg'])
        win_id = canvas.create_window((0, 0), window=inner, anchor='nw')

        def _on_inner_configure(e):
            canvas.configure(scrollregion=canvas.bbox('all'))

        def _on_canvas_configure(e):
            canvas.itemconfig(win_id, width=e.width)

        inner.bind('<Configure>', _on_inner_configure)
        canvas.bind('<Configure>', _on_canvas_configure)

        def _on_wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')

        def _bind_wheel(_):
            canvas.bind_all('<MouseWheel>', _on_wheel)

        def _unbind_wheel(_):
            canvas.unbind_all('<MouseWheel>')

        canvas.bind('<Enter>', _bind_wheel)
        canvas.bind('<Leave>', _unbind_wheel)

        # Κρατάμε αναφορές ώστε το _finalize_scrollable (καλείται αφού έχουν
        # προστεθεί όλες οι γραμμές) να μπορεί να προσαρμόσει το πραγματικό
        # ύψος του canvas βάσει του πραγματικού (μετρημένου) ύψους περιεχομένου,
        # αντί για εκτίμηση — έτσι αποφεύγεται η μπάρα κύλισης όποτε είναι δυνατό.
        inner._msc_canvas = canvas
        inner._msc_max_h  = _max_h
        return inner

    def _finalize_scrollable(self, inner):
        """Προσαρμόζει το ύψος ενός scrollable canvas (βλ. _make_scrollable)
        στο πραγματικό ύψος του περιεχομένου του, ώστε να μην εμφανίζεται
        μπάρα κύλισης όταν το περιεχόμενο χωράει ήδη στο διαθέσιμο ύψος."""
        canvas = getattr(inner, '_msc_canvas', None)
        if canvas is None:
            return
        max_h = getattr(inner, '_msc_max_h', None)
        try:
            inner.update_idletasks()
            req_h = inner.winfo_reqheight()
            if req_h > 0:
                new_h = min(req_h, max_h) if max_h else req_h
                canvas.configure(height=new_h)
        except Exception:
            pass

    def _tab_label(self, parent, text):
        row = tk.Frame(parent, bg=C['bg'])
        row.pack(fill='x', padx=18, pady=(10, 4))
        tk.Label(row, text=text, bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 10, 'bold'), anchor='w').pack(side='left')

    # ── Tab 1: Έλεγχοι (9 έλεγχοι + Ε.Ε.Α.) — κάθε ένας τρέχει ξεχωριστά,
    #    μέσα στο δικό του 3-tab παράθυρο (core/check_dialog.py::CheckRunDialog).
    #    Δεν υπάρχει πια μαζική εκτέλεση με checkboxes. ─────────────────────────
    def _build_checks_tab(self, parent):
        self._tab_label(parent, 'Διαθέσιμοι έλεγχοι — άνοιξε έναν για Λήψη / Εκτέλεση / Αποστολή:')
        inner = self._make_scrollable(parent, len(self.checks))

        for idx, (title, desc, mod) in enumerate(self.checks, start=1):
            title = f'Α{idx}. {title}'
            f = tk.Frame(inner, bg=C['norm_bg'],
                         highlightbackground=C['norm_bd'],
                         highlightthickness=1,
                         pady=6, padx=10)
            f.pack(fill='x', pady=3)
            self.check_frames.append(f)

            # Δεξιά: το κουμπί «Άνοιγμα» σε δικό του πλαίσιο που καλύπτει
            # όλο το ύψος της κάρτας (fill='y') — έτσι το ίδιο το κουμπί
            # (pack expand=True, χωρίς fill) στοιχίζεται κάθετα στο κέντρο,
            # ό,τι ύψος κι αν έχει η περιγραφή δίπλα του.
            btn_holder = tk.Frame(f, bg=C['norm_bg'])
            btn_holder.pack(side='right', fill='y')

            _is_smeae = (mod == SMEAE_MARKER)
            _open_cmd = self._open_smeae if _is_smeae else (lambda m=mod: self._open_check(m))
            tk.Button(btn_holder, text='▶  Άνοιγμα',
                      bg=C['btn_bg'], fg=C['btn_fg'],
                      font=('Arial', 9, 'bold'), relief='flat',
                      padx=10, pady=3, cursor='hand2',
                      activebackground=C['btn_act'],
                      command=_open_cmd
                      ).pack(expand=True, padx=(4, 0))

            # Σημείωση: το κουμπί ✏ επεξεργασίας προτύπου email μετακόμισε
            # μέσα στο 3ο tab «✉ Αποστολή» κάθε ελέγχου (core/check_dialog.py
            # ::CheckRunDialog._open_email_editor) — δεν εμφανίζεται πια εδώ.

            # Αριστερά: τίτλος + περιγραφή σε δική τους στήλη, ώστε η
            # περιγραφή να αναδιπλώνεται πάντα ΜΕΣΑ στο διαθέσιμο πλάτος
            # (δίπλα από το κουμπί, όχι από κάτω/πίσω του).
            left = tk.Frame(f, bg=C['norm_bg'])
            left.pack(side='left', fill='both', expand=True, padx=(0, 10))

            tk.Label(left, text=title,
                     bg=C['norm_bg'], fg=C['hdr_bg'],
                     font=('Arial', 10, 'bold'), anchor='w'
                     ).pack(fill='x')

            if desc:
                desc_lbl = tk.Label(left, text=desc, bg=C['norm_bg'], fg=C['desc'],
                                     font=('Arial', 8), anchor='w', justify='left')
                desc_lbl.pack(fill='x', padx=4)

                def _on_left_resize(event, _lbl=desc_lbl):
                    new_wrap = max(event.width - 8, 50)
                    if _lbl.cget('wraplength') != new_wrap:
                        _lbl.configure(wraplength=new_wrap)
                left.bind('<Configure>', _on_left_resize)

        self._finalize_scrollable(inner)

    # ── Γενικό helper για τα tabs «Αναφορές/Στατιστικά» και «Ενέργειες
    #    MySchool»: απλή λίστα γραμμών (τίτλος + περιγραφή + «▶ Άνοιγμα»).
    #    items: list of dict {title, desc, cmd} ή {title, desc, menu:[(label,cmd), ...]}
    #    για γραμμές που ανοίγουν popup menu (π.χ. Επιβεβαίωση Δ/νσης). ────────
    def _build_simple_list(self, parent, items, scrollable=True):
        """
        `scrollable=False` χτίζει τις κάρτες απευθείας μέσα σε ένα απλό
        πλαίσιο (χωρίς Canvas/Scrollbar) — χρησιμοποιείται στα tabs
        «Αναφορές/Στατιστικά» και «Ενέργειες MySchool», όπου τα στοιχεία
        είναι λίγα και χωράνε πάντα χωρίς να χρειάζεται πλευρική μπάρα
        κύλισης. Το tab «Έλεγχοι» (_build_checks_tab) συνεχίζει να
        χρησιμοποιεί το scrollable Canvas, καθώς έχει περισσότερα στοιχεία.
        """
        if scrollable:
            inner = self._make_scrollable(parent, len(items))
        else:
            inner = tk.Frame(parent, bg=C['bg'])
            inner.pack(fill='both', expand=True, padx=18, pady=(0, 10))

        for item in items:
            f = tk.Frame(inner, bg=C['norm_bg'],
                         highlightbackground=C['norm_bd'],
                         highlightthickness=1,
                         pady=6, padx=10)
            f.pack(fill='x', pady=3)

            # Δεξιά: το κουμπί σε δικό του πλαίσιο ύψους όσο η κάρτα, ώστε
            # να στοιχίζεται κάθετα στο κέντρο (pack expand=True).
            btn_holder = tk.Frame(f, bg=C['norm_bg'])
            btn_holder.pack(side='right', fill='y')

            menu_items = item.get('menu')
            locked = item.get('locked', False)
            # Κλειδωμένα στοιχεία: εμφανίζονται γκριζαρισμένα σαν ανενεργά,
            # αλλά παραμένουν πατήσιμα — ζητούν κωδικό πρόσβασης πριν ανοίξουν
            # (βλ. _password_gate).
            btn = tk.Button(btn_holder, text='▶  Άνοιγμα',
                      bg=(C['btn_dis'] if locked else C['btn_bg']),
                      fg=C['btn_fg'],
                      font=('Arial', 9, 'bold'), relief='flat',
                      padx=10, pady=3, cursor='hand2',
                      activebackground=(C['btn_dis'] if locked else C['btn_act']))
            btn.pack(expand=True, padx=(4, 0))

            if menu_items:
                popup = tk.Menu(self.root, tearoff=0,
                                 bg='white', fg='#1A1A1A',
                                 activebackground=C['hdr_bg'], activeforeground='white',
                                 font=('Arial', 10), relief='flat', bd=1)
                for label, cmd in menu_items:
                    _cmd = (lambda c=cmd: _password_gate(self.root, c)) if locked else cmd
                    popup.add_command(label=label, command=_cmd)

                def _show_popup(event=None, _b=btn, _m=popup):
                    x = _b.winfo_rootx()
                    y = _b.winfo_rooty() + _b.winfo_height()
                    _m.tk_popup(x, y)
                btn.configure(command=_show_popup)
            elif locked:
                btn.configure(command=lambda c=item['cmd']: _password_gate(self.root, c))
            else:
                btn.configure(command=item['cmd'])

            # Αριστερά: τίτλος + περιγραφή σε δική τους στήλη, ώστε η
            # περιγραφή να αναδιπλώνεται πάντα ΜΕΣΑ στο διαθέσιμο πλάτος
            # (δίπλα από το κουμπί, όχι από κάτω/πίσω του).
            left = tk.Frame(f, bg=C['norm_bg'])
            left.pack(side='left', fill='both', expand=True, padx=(0, 10))

            tk.Label(left, text=item['title'],
                     bg=C['norm_bg'], fg=C['hdr_bg'],
                     font=('Arial', 10, 'bold'), anchor='w'
                     ).pack(fill='x')

            if item.get('desc'):
                desc_lbl = tk.Label(left, text=item['desc'], bg=C['norm_bg'], fg=C['desc'],
                                     font=('Arial', 8), anchor='w', justify='left')
                desc_lbl.pack(fill='x', padx=4)

                def _on_left_resize(event, _lbl=desc_lbl):
                    new_wrap = max(event.width - 8, 50)
                    if _lbl.cget('wraplength') != new_wrap:
                        _lbl.configure(wraplength=new_wrap)
                left.bind('<Configure>', _on_left_resize)

        if scrollable:
            self._finalize_scrollable(inner)

    # ── Tab 2: Αναφορές / Στατιστικά ─────────────────────────────────────────
    def _build_reports_tab(self, parent):
        self._tab_label(parent, 'Αναφορές & εξαγωγές — άνοιξε ένα εργαλείο:')
        items = [
            {'title': 'Ενημερωτικό',
             'desc': 'Αποστολή ενημερωτικού email (με προαιρετικό συνημμένο αρχείο) σε σχολεία.',
             'cmd': self._open_inform_email},
            {'title': 'Σχολικές Μονάδες',
             'desc': 'Εξαγωγή στοιχείων σχολικών μονάδων ανά Δήμο.',
             'cmd': self._open_monada_tool},
            {'title': 'Εκπ/κοί ανά Ειδικότητα',
             'desc': 'Εξαγωγή εκπαιδευτικών ανά ειδικότητα — για αποστολή στοιχείων ενδεικτικά σε '
                     'Συμβούλους Εκπ/σης.',
             'cmd': self._open_eidikotita_tool},
            {'title': 'Εκπαιδευτικοί ανά Ειδικότητα & Θέση Συμβούλου',
             'desc': 'Εξαγωγή εκπαιδευτικών φιλτραρισμένων ανά ειδικότητα και θέση Συμβούλου Εκπ/σης.',
             'cmd': lambda: SymbouloiDialog(self.root),
             'locked': True},
        ]
        for idx, item in enumerate(items, start=1):
            item['title'] = f"Β{idx}. {item['title']}"
        self._build_simple_list(parent, items, scrollable=False)

    # ── Tab 3: Ενέργειες MySchool (καταχωρήσεις) ────────────────────────────
    def _build_actions_tab(self, parent):
        self._tab_label(parent, 'Ενέργειες καταχώρησης στο MySchool — άνοιξε ένα εργαλείο:')
        items = [
            {'title': 'Επεξεργασία αρχείου τοποθετήσεων',
             'desc': 'Μετατροπή αρχικού αρχείου — συμπλήρωση πεδίων και άνοιγμα στο Excel.',
             'cmd': lambda: DipePlacementsDialog(self.root),
             'locked': True},
            {'title': 'Τοποθετήσεις',
             'desc': 'Αυτόματη καταχώρηση τοποθετήσεων εκπαιδευτικών στο MySchool.',
             'cmd': self._open_placements},
            {'title': 'Τερματισμός Τοποθετήσεων',
             'desc': 'Αυτόματος τερματισμός τοποθετήσεων — ορισμός ημερομηνίας λήξης (21/6/2026) '
                     'στο MySchool.',
             'cmd': lambda: TerminationDialog(self.root)},
            {'title': 'Αλλαγή Λειτουργικότητας',
             'desc': 'Αυτόματη ενημέρωση λειτουργικότητας σχολικών μονάδων στο MySchool βάσει '
                     'αρχείου Excel.',
             'cmd': lambda: FunctionalityDialog(self.root)},
            {'title': 'Υποχρεωτικό Ωράριο ΠΕ60',
             'desc': 'Έλεγχος υποχρεωτικού ωραρίου ΠΕ60/ΠΕ60.50 βάσει λειτουργικότητας νηπιαγωγείου '
                     'τοποθέτησης (2.1 + 4.1 + 4.2).',
             'cmd': self._run_orario_pe60},
            {'title': 'Επιβεβαίωση Δ/νσης',
             'desc': 'Καταχώρηση/αναίρεση Γραμματειακής Υποστήριξης (πρώην «PANIC») — Έναρξη ή Λήξη.',
             'menu': [('▶  Έναρξη', self._open_editor), ('⏹  Λήξη', self._open_panic_end)]},
            {'title': 'Καταχώρηση Απουσίας σε Οργανική',
             'desc': 'Αυτόματη καταχώρηση απουσίας Ολικής Διάθεσης στην οργανική τοποθέτηση '
                     'εκπαιδευτικών.',
             'cmd': lambda: AbsencesDialog(self.root)},
        ]
        for idx, item in enumerate(items, start=1):
            item['title'] = f"Ε{idx}. {item['title']}"
        self._build_simple_list(parent, items, scrollable=False)

    # ── Tab 4: Χρήστες — ποιες Διευθύνσεις χρησιμοποιούν το πρόγραμμα και σε
    #    πόσους υπολογιστές η καθεμιά (βλ. _send_install_ping / AppsScript). ──
    def _build_users_tab(self, parent):
        self._tab_label(parent, 'Διευθύνσεις που χρησιμοποιούν το πρόγραμμα:')

        top = tk.Frame(parent, bg=C['bg'])
        top.pack(fill='x', padx=18, pady=(0, 4))
        self._users_status_lbl = tk.Label(top, text='', bg=C['bg'], fg=C['desc'],
                                           font=('Arial', 8, 'italic'), anchor='w')
        self._users_status_lbl.pack(side='left')
        tk.Button(top, text='🔄  Ανανέωση', bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 8, 'bold'), relief='flat', padx=8, pady=2,
                  cursor='hand2', command=self._refresh_users_list
                  ).pack(side='right')

        self._users_list_frame = tk.Frame(parent, bg=C['bg'])
        self._users_list_frame.pack(fill='both', expand=True, padx=18, pady=(4, 10))

        self._refresh_users_list()

    def _refresh_users_list(self):
        self._users_status_lbl.configure(text='Φόρτωση...', fg=C['desc'])
        for w in self._users_list_frame.winfo_children():
            w.destroy()

        def _on_done(data, err):
            self.root.after(0, lambda: self._render_users(data, err))
        _fetch_users_summary(_on_done)

    def _render_users(self, data, err):
        for w in self._users_list_frame.winfo_children():
            w.destroy()

        if err:
            self._users_status_lbl.configure(text=f'⚠ {err}', fg=C['status_err'])
            return
        if not data:
            self._users_status_lbl.configure(
                text='Δεν υπάρχουν ακόμα καταχωρημένες Διευθύνσεις.', fg=C['desc'])
            return

        total_comp = sum(int(row.get('count', 0)) for row in data)
        self._users_status_lbl.configure(
            text=f'✓ {len(data)} Διευθύνσεις  •  {total_comp} υπολογιστές συνολικά',
            fg=C['status_ok'])

        header = tk.Frame(self._users_list_frame, bg=C['bg'])
        header.pack(fill='x', pady=(0, 2))
        tk.Label(header, text='Διεύθυνση', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(side='left', fill='x', expand=True)
        tk.Label(header, text='Υπολογιστές', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='e', width=14).pack(side='right')

        for row in sorted(data, key=lambda r: (r.get('type', ''), r.get('dir', ''))):
            label = f"{row.get('type', '')}  {row.get('dir', '')}".strip()
            count = int(row.get('count', 0))
            f = tk.Frame(self._users_list_frame, bg=C['norm_bg'],
                         highlightbackground=C['norm_bd'], highlightthickness=1,
                         pady=4, padx=10)
            f.pack(fill='x', pady=2)
            tk.Label(f, text=label, bg=C['norm_bg'], fg=C['hdr_bg'],
                     font=('Arial', 9), anchor='w').pack(side='left', fill='x', expand=True)
            tk.Label(f, text=str(count), bg=C['norm_bg'], fg=C['desc'],
                     font=('Arial', 9, 'bold'), anchor='e', width=14).pack(side='right')

    def _run_orario_pe60(self):
        """Υποχρεωτικό Ωράριο ΠΕ60 (πρώην μέσα στο «Νέο Σχ. Έτος»).
        Ανοίγει το ίδιο 2-tab CheckRunDialog (⬇ Λήψη / ▶ Εκτέλεση) με τους
        υπόλοιπους ελέγχους, ώστε να υπάρχει και εδώ λήψη στατιστικών
        (2.1, 4.1, 4.2) πριν την εκτέλεση."""
        import importlib
        try:
            mod = importlib.import_module('checks.orario_pe60')
            from core.check_dialog import CheckRunDialog
            CheckRunDialog(self.root, config, _docs_base(), C, mod)
        except Exception as e:
            messagebox.showerror('Σφάλμα', str(e), parent=self.root)

    def _open_settings(self):
        SettingsDialog(self.root)

    def _open_download(self):
        DownloadDialog(self.root)

    def _open_check(self, mod):
        """Ανοίγει το 3-tab dialog (Λήψη/Εκτέλεση/Αποστολή) για έναν έλεγχο."""
        from core.check_dialog import CheckRunDialog
        CheckRunDialog(self.root, config, _docs_base(), C, mod)

    def _open_eidikotita_tool(self):
        EidikotitaDialog(self.root)

    def _open_monada_tool(self):
        MonadaDialog(self.root)


    def _open_smeae(self):
        import shutil
        docs = _docs_base()
        data_src = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'smeae_column_mappings.json')
        data_dst_dir = os.path.join(docs, 'data')
        os.makedirs(data_dst_dir, exist_ok=True)
        data_dst = os.path.join(data_dst_dir, 'smeae_column_mappings.json')
        if not os.path.exists(data_dst) and os.path.exists(data_src):
            shutil.copy2(data_src, data_dst)
        from smeae.dialog import SmeaeDialog
        SmeaeDialog(self.root, config, docs, C)

    def _open_placements(self):
        PlacementsDialog(self.root)

    def _open_dipe(self):
        DipeDialog(self.root)

    def _open_neo_school_year(self):
        NeoSchoolYearDialog(self.root)

    def _open_editor(self):
        EditorDialog(self.root)

    def _open_panic_end(self):
        PanicEndDialog(self.root)

    def _open_inform_email(self):
        InformEmailDialog(self.root)

    # ── Email template editor ────────────────────────────────────────────────

    def _get_default_email_body(self, module):
        """Επιστρέφει το default body text χωρίς υπογραφή."""
        body_t = getattr(module, 'EMAIL_BODY', '')
        try:
            full = body_t('') if callable(body_t) else body_t
            sig  = config.email_signature()
            if sig and full.endswith(sig):
                return full[:-len(sig)]
            return full
        except Exception:
            return ''

    def _open_email_editor(self, mod_name, module):
        """Dialog επεξεργασίας email template για συγκεκριμένο έλεγχο."""
        import json

        # Φόρτωσε τρέχον template (custom ή default)
        settings = _load_local_settings()
        templates = settings.get('email_templates', {})
        custom = templates.get(mod_name)

        if custom:
            cur_subject = custom.get('subject', '')
            cur_body    = custom.get('body', '')
        else:
            cur_subject = getattr(module, 'EMAIL_SUBJECT', '')
            cur_body    = self._get_default_email_body(module)

        title_str = getattr(module, 'CHECK_TITLE', mod_name)

        # Παράθυρο
        dlg = tk.Toplevel(self.root)
        dlg.title(f'Πρότυπο Email — {title_str}')
        dlg.configure(bg=C['bg'])
        dlg.resizable(True, False)
        dlg.grab_set()
        dlg.transient(self.root)

        pad = dict(padx=14, pady=5)

        tk.Label(dlg, text='Θέμα:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **pad)

        subj_var = tk.StringVar(value=cur_subject)
        tk.Entry(dlg, textvariable=subj_var, font=('Arial', 9),
                 width=60).pack(fill='x', padx=14, pady=(0, 8))

        tk.Label(dlg, text='Κείμενο email:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **pad)

        txt = tk.Text(dlg, font=('Arial', 9), width=60, height=10,
                      wrap='word', relief='solid', bd=1)
        txt.pack(fill='x', padx=14, pady=(0, 4))
        txt.insert('1.0', cur_body)

        tk.Label(dlg, text='Η υπογραφή σας προστίθεται αυτόματα στο τέλος.',
                 bg=C['bg'], fg=C['desc'], font=('Arial', 8),
                 anchor='w').pack(fill='x', padx=14, pady=(0, 10))

        def _save():
            new_subj = subj_var.get().strip()
            new_body = txt.get('1.0', 'end-1c')
            s = _load_local_settings()
            s.setdefault('email_templates', {})[mod_name] = {
                'subject': new_subj,
                'body':    new_body,
            }
            path = _get_local_settings_path()
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(s, f, ensure_ascii=False, indent=2)
            dlg.destroy()
            messagebox.showinfo('Αποθήκευση', 'Το πρότυπο email αποθηκεύτηκε.',
                                parent=self.root)

        def _reset():
            if messagebox.askyesno('Επαναφορά', 'Να επανέλθει το προεπιλεγμένο κείμενο;',
                                   parent=dlg):
                s = _load_local_settings()
                s.get('email_templates', {}).pop(mod_name, None)
                path = _get_local_settings_path()
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(s, f, ensure_ascii=False, indent=2)
                dlg.destroy()

        btn_row = tk.Frame(dlg, bg=C['bg'])
        btn_row.pack(pady=(0, 12))

        tk.Button(btn_row, text='Αποθήκευση',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=14, pady=5, cursor='hand2',
                  command=_save).pack(side='left', padx=4)

        tk.Button(btn_row, text='Επαναφορά προεπιλογής',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 9), relief='flat',
                  padx=14, pady=5, cursor='hand2',
                  command=_reset).pack(side='left', padx=4)

        tk.Button(btn_row, text='Άκυρο',
                  bg=C['bg2'], fg=C['desc'],
                  font=('Arial', 9), relief='flat',
                  padx=14, pady=5, cursor='hand2',
                  command=dlg.destroy).pack(side='left', padx=4)

        dlg.update_idletasks()
        # Κεντράρισμα
        w, h = dlg.winfo_width(), dlg.winfo_height()
        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
        dlg.geometry(f'+{x}+{y}')

    def _open_help(self):
        """Ανοίγει τον οδηγό χρήσης (PDF) με τον προεπιλεγμένο viewer."""
        _show_help(self.root)

    def _open_settings(self):
        """Παράθυρο ρυθμίσεων (email + MySchool credentials)."""
        SettingsDialog(self.root)

    # (Η παλιά μαζική εκτέλεση με checkboxes καταργήθηκε — κάθε έλεγχος
    #  τρέχει πλέον μέσα στο δικό του CheckRunDialog: core/check_dialog.py.
    #  Το _on_print/_set_status παραμένουν για γενικά μηνύματα status bar.)
    def _on_print(self, text):
        """Λαμβάνει μηνύματα από sys.stdout και τα εμφανίζει στο status bar."""
        self.root.after(0, lambda t=text: self._set_status(t))

    def _set_status(self, text, color=None):
        self.status_var.set(text)
        if color:
            self.status_lbl.configure(fg=color)


def _check_for_update(on_update_available):
    """Ελέγχει αν υπάρχει νεότερη έκδοση στο GitHub. Τρέχει σε background thread.
    Αν βρεθεί νεότερη, καλεί on_update_available(new_version, download_url)."""
    def _task():
        try:
            import urllib.request, json as _json
            api_url = 'https://api.github.com/repos/MichalisKat/MySchoolChecks/releases/latest'
            req = urllib.request.Request(api_url, headers={'User-Agent': 'MySchoolChecks'})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = _json.loads(resp.read().decode())
            latest = data.get('tag_name', '').lstrip('v')
            if not latest:
                return
            current = getattr(config, 'APP_VERSION', '0.0.0')
            def _ver(s):
                try: return tuple(int(x) for x in s.split('.'))
                except: return (0,)
            if _ver(latest) > _ver(current):
                # Βρες το .exe asset
                assets = data.get('assets', [])
                dl_url = next(
                    (a['browser_download_url'] for a in assets
                     if a['name'].endswith('.exe')),
                    None)
                if dl_url:
                    on_update_available(latest, dl_url)
        except Exception:
            pass  # Αθόρυβη αποτυχία — δεν επηρεάζει τη λειτουργία
    threading.Thread(target=_task, daemon=True).start()


def _get_directorate():
    """Διαβάζει την επιλεγμένη Διεύθυνση Εκπαίδευσης (γράφτηκε από τον installer
    στο data\\directorate.txt, μορφή "ΠΕ|Α' Αθήνας" ή "ΔΕ|Χαλκιδικής").
    Επιστρέφει (dir_type, dir_name) ή (None, None) αν δεν βρεθεί."""
    path = None
    try:
        if getattr(sys, 'frozen', False):
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(base, 'data', 'directorate.txt')
        with open(path, 'rb') as f:
            raw_bytes = f.read()
        # Ο NSIS installer γράφει το FileWrite σε ANSI (π.χ. Windows-1253 σε
        # ελληνικά Windows), όχι πάντα UTF-8 — δοκιμάζουμε αρκετές πιθανές
        # κωδικοποιήσεις με τη σειρά.
        raw = None
        for enc in ('utf-8-sig', 'utf-8', 'cp1253', 'mbcs', 'latin-1'):
            try:
                raw = raw_bytes.decode(enc).strip()
                break
            except Exception:
                continue
        if raw is None:
            raise ValueError(f'αδύνατη αποκωδικοποίηση ({len(raw_bytes)} bytes)')
        dtype, _, dname = raw.partition('|')
        dtype, dname = dtype.strip(), dname.strip()
        if dname:
            return dtype, dname
        _log_ping_debug(f'_get_directorate: κενό directorate.txt ({path!r}) — raw={raw!r}')
    except Exception as e:
        _log_ping_debug(f'_get_directorate: {type(e).__name__}: {e} (path={path!r})')
    return None, None


def _get_install_id():
    """Επιστρέφει ένα μοναδικό, μόνιμο αναγνωριστικό για ΑΥΤΟΝ τον υπολογιστή/
    εγκατάσταση (δημιουργείται μία φορά, αποθηκεύεται σε εγγράψιμο φάκελο).
    Χρησιμοποιείται ώστε το Apps Script να ξεχωρίζει διαφορετικούς υπολογιστές
    της ΙΔΙΑΣ Διεύθυνσης (και να προσθέτει αύξοντα αριθμό #2, #3, ...)."""
    try:
        import uuid
        path = os.path.join(_app_base(), 'install_id.txt')
        if os.path.isfile(path):
            with open(path, 'r', encoding='utf-8') as f:
                iid = f.read().strip()
            if iid:
                return iid
        iid = uuid.uuid4().hex
        with open(path, 'w', encoding='utf-8') as f:
            f.write(iid)
        return iid
    except Exception:
        return ''


def _send_install_ping():
    """Στέλνει ένα αθόρυβο «ping» με τη Διεύθυνση Εκπαίδευσης στο Apps Script
    (βλ. PING_URL στο config.py) ώστε να ξέρουμε ποιες Διευθύνσεις χρησιμοποιούν
    το πρόγραμμα (και πόσους υπολογιστές η καθεμιά). Τρέχει σε background
    thread — καμία επίπτωση αν αποτύχει (π.χ. χωρίς ίντερνετ) ή αν δεν έχει
    οριστεί PING_URL."""
    ping_url = getattr(config, 'PING_URL', '').strip()
    if not ping_url:
        _log_ping_debug('Δεν στάλθηκε — κενό PING_URL στο config.py.')
        return
    dtype, dname = _get_directorate()
    if not dname:
        _log_ping_debug('Δεν στάλθηκε — δεν βρέθηκε Διεύθυνση (βλ. προηγούμενη γραμμή αν υπάρχει).')
        return

    def _task():
        try:
            import urllib.request, urllib.parse
            version = getattr(config, 'APP_VERSION', '')
            iid = _get_install_id()
            qs = urllib.parse.urlencode({
                'action': 'ping', 'dir': dname, 'type': dtype,
                'version': version, 'iid': iid,
            })
            req = urllib.request.Request(f'{ping_url}?{qs}',
                                          headers={'User-Agent': 'MySchoolChecks'})
            resp_body = urllib.request.urlopen(req, timeout=8).read()
            _log_ping_debug(f'OK — απάντηση: {resp_body[:200]!r}')
        except Exception as e:
            # Δεν επηρεάζει τη λειτουργία του προγράμματος — απλά καταγράφουμε
            # την αιτία σε ένα μικρό log αρχείο για διάγνωση (π.χ. proxy/SSL
            # certificate σχολικού δικτύου), χωρίς να ενοχλούμε τον χρήστη.
            _log_ping_debug(f'ΣΦΑΛΜΑ: {type(e).__name__}: {e}')
    threading.Thread(target=_task, daemon=True).start()


def _log_ping_debug(msg):
    """Καταγράφει μια γραμμή στο ping_debug.log (φάκελος _app_base()) — μόνο
    για διάγνωση προβλημάτων σύνδεσης με το Apps Script, δεν εμφανίζεται
    πουθενά στο UI."""
    try:
        import datetime as _dt
        path = os.path.join(_app_base(), 'ping_debug.log')
        with open(path, 'a', encoding='utf-8') as f:
            f.write(f'[{_dt.datetime.now():%Y-%m-%d %H:%M:%S}] {msg}\n')
    except Exception:
        pass


def _fetch_users_summary(on_done):
    """Φέρνει από το Apps Script τη συγκεντρωτική λίστα Διευθύνσεων/αριθμού
    υπολογιστών (action=list). Τρέχει σε background thread· καλεί
    on_done(list_or_None, error_msg_or_None) στο κύριο thread μέσω `after`."""
    ping_url = getattr(config, 'PING_URL', '').strip()

    def _task():
        if not ping_url:
            on_done(None, 'Δεν έχει ρυθμιστεί PING_URL στο config.py.')
            return
        try:
            import urllib.request, urllib.parse, json as _json
            qs  = urllib.parse.urlencode({'action': 'list'})
            req = urllib.request.Request(f'{ping_url}?{qs}',
                                          headers={'User-Agent': 'MySchoolChecks'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = _json.loads(resp.read().decode('utf-8'))
            on_done(data, None)
        except Exception as e:
            on_done(None, str(e))
    threading.Thread(target=_task, daemon=True).start()


def _do_update(parent, new_ver, dl_url):
    """Κατεβάζει το νέο setup.exe και το εκτελεί. Εμφανίζει progress dialog."""
    import urllib.request, tempfile, subprocess as _sub

    # Ερώτηση επιβεβαίωσης
    answer = messagebox.askyesno(
        'Διαθέσιμη ενημέρωση',
        f'Διαθέσιμη νέα έκδοση v{new_ver}!\n\n'
        f'Θέλεις να κατεβάσεις και να εγκαταστήσεις τώρα;\n\n'
        f'Η εφαρμογή θα κλείσει αυτόματα για την εγκατάσταση.',
        parent=parent)
    if not answer:
        return

    # Progress dialog
    dlg = tk.Toplevel(parent)
    dlg.title('Λήψη ενημέρωσης')
    dlg.configure(bg=C['bg'])
    dlg.resizable(False, False)
    dlg.grab_set()
    dlg.transient(parent)

    tk.Label(dlg, text=f'Λήψη MySchool Checks v{new_ver}...',
             bg=C['bg'], fg=C['hdr_bg'],
             font=('Arial', 10, 'bold')).pack(padx=24, pady=(18, 6))

    from tkinter import ttk as _ttk
    pb = _ttk.Progressbar(dlg, length=320, mode='determinate')
    pb.pack(padx=24, pady=(0, 6))

    status_var = tk.StringVar(value='Σύνδεση...')
    tk.Label(dlg, textvariable=status_var,
             bg=C['bg'], fg=C['footer'],
             font=('Arial', 8)).pack(padx=24, pady=(0, 18))

    dlg.update_idletasks()
    px = parent.winfo_x() + (parent.winfo_width()  - dlg.winfo_width())  // 2
    py = parent.winfo_y() + (parent.winfo_height() - dlg.winfo_height()) // 2
    dlg.geometry(f'+{px}+{py}')

    def _dlg_alive():
        """Ελέγχει αν το progress dialog υπάρχει ακόμα."""
        try:
            return dlg.winfo_exists()
        except Exception:
            return False

    def _safe_after(fn):
        """Εκτελεί fn στο main thread μόνο αν το dialog υπάρχει ακόμα."""
        if _dlg_alive():
            try:
                dlg.after(0, fn)
            except Exception:
                pass

    def _download():
        try:
            tmp_dir  = tempfile.mkdtemp()
            fname    = f'myschool-checksplus-{new_ver}-setup.exe'
            tmp_path = os.path.join(tmp_dir, fname)

            def _reporthook(count, block_size, total_size):
                if total_size > 0:
                    pct = min(100, int(count * block_size * 100 / total_size))
                    mb_done = count * block_size / 1_048_576
                    mb_total = total_size / 1_048_576
                    _safe_after(lambda p=pct, d=mb_done, t=mb_total: [
                        pb.configure(value=p),
                        status_var.set(f'{d:.1f} / {t:.1f} MB  ({p}%)')])

            urllib.request.urlretrieve(dl_url, tmp_path, _reporthook)

            # Κατεβάστηκε — εκτέλεση
            _safe_after(lambda: [
                status_var.set('Εκκίνηση εγκατάστασης...'),
                dlg.update()])
            import time as _t; _t.sleep(0.8)

            import ctypes
            ctypes.windll.shell32.ShellExecuteW(None, 'runas', tmp_path, None, None, 1)
            _safe_after(lambda: [dlg.destroy(), parent.destroy()])

        except Exception as e:
            _safe_after(lambda err=str(e): [
                dlg.destroy(),
                messagebox.showerror('Σφάλμα λήψης',
                    f'Δεν ήταν δυνατή η λήψη:\n{err}\n\n'
                    'Κατέβασέ την χειροκίνητα από:\n'
                    'github.com/MichalisKat/MySchoolChecks/releases',
                    parent=parent)])

    threading.Thread(target=_download, daemon=True).start()


# ── Γενικό tab «⬇ Λήψη» για τα standalone εργαλεία (Σχολικές Μονάδες,
#    Εκπ/κοί ανά Ειδικότητα, Εκπ/κοί ανά Ειδικότητα & Θέση Συμβούλου) —
#    ίδια λογική/εμφάνιση με το tab «⬇ Λήψη» των ελέγχων
#    (core/check_dialog.py::_build_download), προσαρμοσμένη σε αυτά τα
#    εργαλεία που δεν είναι check modules (δεν έχουν REQUIRED_REPORTS). ─────
def _build_report_download_tab(dialog, parent, rids, required, on_done=None):
    """
    Χτίζει το tab «⬇ Λήψη» μέσα στο `parent` (frame ενός tab σε Notebook).
    `rids`     : λίστα από report-ids για το core.downloader.MySchoolDownloader
                 (π.χ. ['3.1', '2.2'] ή ['topoth', '2.1', '4.1', '4.2', '4.16']).
    `required` : λίστα με περιγραφικές γραμμές για εμφάνιση στον χρήστη.
    `on_done`  : callback (χωρίς ορίσματα) που καλείται ΜΕΤΑ από επιτυχή
                 λήψη — ώστε ο caller να ξανα-ανιχνεύσει τα αρχεία και να
                 ξαναχτίσει το tab «Εκτέλεση» με τα φρέσκα αρχεία.
    """
    from core.downloader import FILE_PREFIX_MAP
    from tkinter import scrolledtext as _st
    import glob as _glob, threading as _th
    from datetime import datetime as _dt

    tk.Label(parent, text='Απαιτούμενα στατιστικά:', bg=C['bg'], fg=C['hdr_bg'],
             font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(14, 3))
    txt = ('\n'.join(f'  •  {r}' for r in required) if required
           else '  (δεν έχουν οριστεί)')
    tk.Label(parent, text=txt, bg=C['bg'], fg=C['desc'], font=('Arial', 9),
             justify='left', anchor='w', wraplength=560).pack(fill='x', padx=18, pady=(0, 8))

    today_dir = os.path.join(_docs_base(), 'downloads', _dt.now().strftime('%Y%m%d'))
    status_lbl = tk.Label(parent, bg=C['bg'], font=('Arial', 8, 'italic'), anchor='w')
    status_lbl.pack(fill='x', padx=18, pady=(0, 6))

    def _refresh_status():
        have = set()
        if rids and os.path.isdir(today_dir):
            for rid in rids:
                prefix  = FILE_PREFIX_MAP.get(rid, rid)
                matches = [f for f in _glob.glob(os.path.join(today_dir, f'{prefix}*'))
                           if not f.endswith(('.tmp', '.crdownload'))]
                if matches:
                    have.add(rid)
        if rids:
            complete = (have == set(rids))
            txt2 = (f'✓ Σήμερα υπάρχουν ήδη {len(have)}/{len(rids)} αρχεία.' if have
                    else 'Δεν έχουν κατέβει ακόμα σήμερα.')
            status_lbl.configure(text=txt2, fg=(C['status_ok'] if complete else C['desc']))
    _refresh_status()

    log_w = _st.ScrolledText(parent, height=10, font=('Consolas', 8),
                              relief='solid', bd=1, state='disabled',
                              bg='#F5F5F5', wrap=tk.WORD)
    log_w.pack(fill='both', expand=True, padx=18, pady=(0, 8))

    def on_log(msg):
        def _append():
            try:
                log_w.configure(state='normal')
                log_w.insert(tk.END, msg + '\n')
                log_w.see(tk.END)
                log_w.configure(state='disabled')
            except Exception:
                pass
        try:
            dialog.after(0, _append)
        except Exception:
            pass

    n = len(rids)
    label = f'⬇  Λήψη {n} Αρχεί{"ου" if n == 1 else "ων"}' if n else '⬇  Λήψη'
    br = tk.Frame(parent, bg=C['bg'])
    br.pack(fill='x', padx=18, pady=(0, 14))
    dl_btn = tk.Button(br, text=label, bg=C['btn_bg'], fg=C['btn_fg'],
                        font=('Arial', 10, 'bold'), relief='flat',
                        padx=16, pady=6, cursor='hand2')
    dl_btn.pack(side='right')
    if not rids:
        dl_btn.configure(state='disabled', bg=C['btn_dis'])

    def _start():
        usr = getattr(config, 'MYSCHOOL_USER', '').strip()
        pw  = getattr(config, 'MYSCHOOL_PASS', '').strip()
        if not usr or not pw:
            messagebox.showwarning(
                'Προσοχή',
                'Συμπλήρωσε username και κωδικό MySchool στις Ρυθμίσεις (⚙).',
                parent=dialog)
            return

        # Έλεγχος αν υπάρχουν ήδη αρχεία (πιθανόν από άλλον έλεγχο/εργαλείο
        # που χρησιμοποιεί το ίδιο στατιστικό, π.χ. 3.1) ΠΡΙΝ ξεκινήσει η
        # λήψη — ρωτάμε αν θέλει επαναλήψη αντί να γίνεται σιωπηλή
        # επαναχρησιμοποίηση.
        from core.downloader import get_downloads_dir, find_existing_reports
        dest = get_downloads_dir(_docs_base())
        existing_rids = find_existing_reports(dest, rids)
        force = []
        if existing_rids:
            redo = messagebox.askyesno(
                'Αρχεία υπάρχουν ήδη',
                f'Βρέθηκαν ήδη {len(existing_rids)}/{n} αρχεία σήμερα '
                '(πιθανόν από άλλον έλεγχο).\n\nΝα γίνει λήψη εκ νέου;',
                parent=dialog)
            if redo:
                force = existing_rids
            elif len(existing_rids) == n:
                # Όλα τα απαραίτητα αρχεία υπάρχουν ήδη και δεν θέλουμε νέα
                # λήψη — δεν υπάρχει λόγος να γίνει login στο MySchool.
                messagebox.showinfo('Λήψη',
                    f'Όλα τα {n} αρχεία υπάρχουν ήδη σήμερα — δεν χρειάζεται νέα λήψη.',
                    parent=dialog)
                return

        dl_btn.configure(state='disabled', bg=C['btn_dis'], text='Εκτελείται...')

        def task():
            try:
                from core.downloader import MySchoolDownloader, cleanup_old_downloads
                dl = MySchoolDownloader(
                    usr, pw, dest, callback=on_log, reports=rids, force=force,
                    browser=getattr(config, 'BROWSER', 'chrome'))
                results = dl.run()
                ok = sum(1 for v in results.values() if v)
                cleanup_old_downloads(_docs_base(), keep=1)

                def _finish():
                    dl_btn.configure(state='normal', bg=C['btn_bg'], text=label)
                    _refresh_status()
                    messagebox.showinfo('Λήψη',
                        f'Ολοκληρώθηκε: {ok}/{n} αρχεία κατεβήκαν.', parent=dialog)
                    if callable(on_done):
                        try:
                            on_done()
                        except Exception:
                            import traceback
                            traceback.print_exc()
                dialog.after(0, _finish)
            except Exception as e:
                err = str(e)
                def _err():
                    dl_btn.configure(state='normal', bg=C['btn_bg'], text=label)
                    messagebox.showerror('Σφάλμα Λήψης', err, parent=dialog)
                dialog.after(0, _err)

        _th.Thread(target=task, daemon=True).start()

    dl_btn.configure(command=_start)


def _build_recipients_box(parent, label_text, saved_value, height=3):
    """
    Πεδίο παραληπτών όπως στο «Ενημερωτικό» (InformEmailDialog) — ένα email
    ανά γραμμή, ή περισσότερα στην ίδια γραμμή με διαχωριστικό «;». Γυρνάει
    το ScrolledText widget (διάβασέ το με `_parse_recipients`).
    """
    from tkinter import scrolledtext as _st
    pad = dict(padx=18, pady=2)
    tk.Label(parent, text=label_text, bg=C['bg'], fg=C['hdr_bg'],
             font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **pad)
    tk.Label(parent, text='(ένα ανά γραμμή, ή περισσότερα στην ίδια γραμμή με «;»)',
             bg=C['bg'], fg=C['footer'], font=('Arial', 8),
             anchor='w').pack(fill='x', padx=18)
    txt = _st.ScrolledText(parent, height=height, font=('Consolas', 8),
                            relief='solid', bd=1, wrap=tk.NONE)
    txt.pack(fill='x', padx=18, pady=(0, 6))
    if saved_value:
        txt.insert('1.0', saved_value)
    return txt


def _parse_recipients(text_widget):
    """Εξάγει λίστα emails από ένα recipients ScrolledText (βλ.
    `_build_recipients_box`) — ένα ανά γραμμή, ή περισσότερα στην ίδια
    γραμμή χωρισμένα με «;»."""
    recips = []
    for line in text_widget.get('1.0', tk.END).splitlines():
        for part in line.split(';'):
            part = part.strip()
            if part and '@' in part:
                recips.append(part)
    return recips


# Απαιτούμενα στατιστικά — κοινά για «Εκπ/κοί ανά Ειδικότητα» και
# «Εκπ/κοί ανά Ειδικότητα & Θέση Συμβούλου» (EidikotitaDialog/SymbouloiDialog).
_EIDIK_RIDS = ['topoth', '2.1', '4.1', '4.2', '4.16']
_EIDIK_REQUIRED = [
    'Τοποθετήσεις — Τοποθετήσεις εκπαιδευτικών',
    '2.1 — Κατάλογος σχολείων',
    '4.1 — Οργανικές τοποθετήσεις',
    '4.2 — Αποσπασμένοι εκπαιδευτικοί',
    '4.16 — Αιτιολόγηση απουσίας',
]


class EidikotitaDialog(tk.Toplevel):
    """Εργαλείο εξαγωγής εκπαιδευτικών ανά ειδικότητα."""

    _SETTINGS_KEY   = 'eidikotita_tool'
    _DEFAULT_BODY   = (
        'Αποτύπωση Myschool {date}.\n\n'
        'Καλημέρα σας,\n\n'
        'Επισυνάπτεται πίνακας excel με τους εκπαιδευτικούς ειδικότητας {specialty} '
        'που υπηρετούν στη Δ/νση μας σύμφωνα με τα καταχωρημένα '
        'στοιχεία στο myschool.\n\n\n'
        'Στη διάθεσή σας για οποιαδήποτε πληροφορία'
    )
    _DEFAULT_SUBJECT = 'Στοιχεία τοποθετήσεων εκπ/κών "{specialty}" σε σχολικές μονάδες'

    # Σταθερές στήλες εξόδου
    _OUT_COLS = [
        'ΑΜ',
        'Επώνυμο', 'Όνομα', 'Κύρια Ειδικ.',
        'Email στο ΠΣΔ', 'Email', 'Κινητό',
        'Σχέση εργασίας', 'Σχέση τοποθέτησης',
        'Κατάσταση',
        'Φορέας τοποθέτησης', 'Δήμος',
        'Τηλέφωνο', 'e-mail',
        'ΑΠΟΥΣΙΑ', 'Από', 'Έως',
    ]

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Εκπ/κοί ανά Ειδικότητα')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self._parent = parent

        s = _load_local_settings().get(self._SETTINGS_KEY, {})
        self._saved_subject = s.get('subject',       self._DEFAULT_SUBJECT)
        self._saved_body    = s.get('body',          self._DEFAULT_BODY)
        self._saved_email   = s.get('advisor_email', '')
        self._saved_dir     = s.get('direction',     '')
        # Αν τα αποθηκευμένα δεν έχουν placeholder, reset στα defaults
        if '{specialty}' not in self._saved_subject:
            self._saved_subject = self._DEFAULT_SUBJECT
        if '{specialty}' not in self._saved_body or '{date}' not in self._saved_body:
            self._saved_body = self._DEFAULT_BODY

        # Αυτόματη εύρεση αρχείων από downloads
        self._topoth_path = self._auto_find('Topothetiseis')
        self._grid_path   = self._auto_find('gridResults')
        self._stat_path   = self._auto_find('stat4_16')
        self._stat41_path = self._auto_find('stat4_1')
        self._stat42_path = self._auto_find('stat4_2')

        # 2-tab δομή όπως οι έλεγχοι: «⬇ Λήψη» + «▶ Εκτέλεση» (το ίδιο
        # εργαλείο που υπήρχε πριν, χωρίς αλλαγή στη λειτουργία του).
        from tkinter import ttk as _ttk
        nb = _ttk.Notebook(self)
        nb.pack(fill='both', expand=True)
        tab_dl   = tk.Frame(nb, bg=C['bg'])
        tab_exec = tk.Frame(nb, bg=C['bg'])
        nb.add(tab_dl,   text='  ⬇ Λήψη  ')
        nb.add(tab_exec, text='  ▶ Εκτέλεση  ')
        self._exec_body = tab_exec

        _build_report_download_tab(self, tab_dl, _EIDIK_RIDS, _EIDIK_REQUIRED,
                                    on_done=self._on_download_done)

        self._build_form()
        self.update_idletasks()
        w = 620
        h = self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f'{w}x{h}+{x}+{y}')

    # ── Auto-find ────────────────────────────────────────────────────────────

    @staticmethod
    def _auto_find(prefix):
        """Ψάχνει το αρχείο στους φακέλους downloads (νεότερος πρώτα), μετά ~/Downloads.
        Το αρχείο πρέπει να αρχίζει ΑΚΡΙΒΩΣσε prefix και ο αμέσως επόμενος χαρακτήρας
        δεν πρέπει να είναι αλφαριθμητικός — ώστε π.χ. 'stat4_1' να μην ταιριάζει 'stat4_16'.
        """
        import glob as _glob, re as _re
        _exact = _re.compile(r'(?i)' + _re.escape(prefix) + r'[^a-zA-Z0-9]')

        def _ok(f):
            return (not f.endswith('.tmp') and not f.endswith('.crdownload')
                    and _exact.match(os.path.basename(f)))

        dl_base = os.path.join(_docs_base(), 'downloads')
        if os.path.isdir(dl_base):
            folders = sorted([
                os.path.join(dl_base, d)
                for d in os.listdir(dl_base)
                if os.path.isdir(os.path.join(dl_base, d))
            ], reverse=True)
            for folder in folders:
                matches = [f for f in _glob.glob(os.path.join(folder, f'{prefix}*')) if _ok(f)]
                if matches:
                    return sorted(matches)[-1]
        dl_user = os.path.join(os.path.expanduser('~'), 'Downloads')
        matches = [f for f in _glob.glob(os.path.join(dl_user, f'*{prefix}*')) if _ok(f)]
        return sorted(matches)[-1] if matches else ''

    # ── Κύρια φόρμα ──────────────────────────────────────────────────────────

    def _build_form(self):
        self._clear()

        hdr = tk.Frame(self._exec_body, bg='#0F6E56', pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='📋  Εκπαιδευτικοί ανά Ειδικότητα',
                 bg='#0F6E56', fg='white',
                 font=('Arial', 12, 'bold')).pack()
        tk.Label(hdr, text='για αποστολή στοιχείων ενδεικτικά σε Συμβούλους Εκπ/σης  (Απαιτούνται: Τοποθετήσεις, 2.1, 4.1, 4.2, 4.16)',
                 bg='#0F6E56', fg='#A8D8C8',
                 font=('Arial', 8, 'italic'), wraplength=580, justify='center').pack()

        # Προειδοποίηση αν λείπουν κρίσιμα αρχεία
        missing = []
        if not self._topoth_path: missing.append('Τοποθετήσεις')
        if not self._grid_path:   missing.append('Κατάλογος σχολείων (2.1)')
        if missing:
            warn = tk.Label(self._exec_body,
                text=f'⚠  Δεν βρέθηκαν: {", ".join(missing)}. Κατέβασε τα πρώτα από «Λήψη Δεδομένων».',
                bg='#FFF3E0', fg='#E65100', font=('Arial', 8), anchor='w', padx=10, pady=5,
                wraplength=560, justify='left')
            warn.pack(fill='x', padx=18, pady=(0, 6))
        missing_opt = []
        if not self._stat41_path: missing_opt.append('4.1')
        if not self._stat42_path: missing_opt.append('4.2')
        if missing_opt:
            warn2 = tk.Label(self._exec_body,
                text=f'ℹ  Δεν βρέθηκαν: {", ".join(missing_opt)} — οι στήλες Email ΠΣΔ / Email / Κινητό θα είναι κενές.',
                bg='#E3F2FD', fg='#1565C0', font=('Arial', 8), anchor='w', padx=10, pady=5,
                wraplength=560, justify='left')
            warn2.pack(fill='x', padx=18, pady=(0, 6))

        # ── Διεύθυνση ────────────────────────────────────────────────────────
        tk.Label(self._exec_body, text='Διεύθυνση:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))

        dir_row = tk.Frame(self._exec_body, bg=C['bg'])
        dir_row.pack(fill='x', padx=18, pady=(2, 6))

        self._dir_var = tk.StringVar(value=self._saved_dir)
        from tkinter import ttk as _ttk

        if self._saved_dir:
            # Αποθηκευμένη Διεύθυνση — εμφάνιση ως label με κουμπί αλλαγής
            self._dir_combo = None
            tk.Label(dir_row, textvariable=self._dir_var,
                     bg=C['bg'], fg=C['hdr_bg'],
                     font=('Arial', 9, 'bold')).pack(side='left')
            tk.Button(dir_row, text='Αλλαγή', bg=C['bg'], fg=C['desc'],
                      relief='flat', font=('Arial', 8), cursor='hand2',
                      command=self._reset_direction).pack(side='left', padx=(10, 0))
        else:
            # Πρώτη φορά — εμφάνιση dropdown
            self._dir_combo = _ttk.Combobox(dir_row, textvariable=self._dir_var,
                                             width=46, state='readonly')
            self._dir_combo.pack(side='left')

        self._dir_lbl = tk.Label(dir_row, text='', bg=C['bg'],
                                  fg=C['desc'], font=('Arial', 8))
        self._dir_lbl.pack(side='left', padx=(10, 0))

        # ── Ειδικότητα ───────────────────────────────────────────────────────
        tk.Label(self._exec_body, text='Ειδικότητα:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))

        spec_row = tk.Frame(self._exec_body, bg=C['bg'])
        spec_row.pack(fill='x', padx=18, pady=(2, 6))

        self._spec_var = tk.StringVar()
        self._spec_combo = _ttk.Combobox(spec_row, textvariable=self._spec_var,
                                          width=46, state='readonly')
        self._spec_combo.pack(side='left')

        self._spec_lbl = tk.Label(spec_row, text='Φόρτωση…', bg=C['bg'],
                                   fg=C['desc'], font=('Arial', 8))
        self._spec_lbl.pack(side='left', padx=(10, 0))

        # Όταν αλλάζει η ειδικότητα → ενημέρωσε το θέμα
        self._spec_var.trace_add('write', self._on_spec_change)

        # ── Στήλες εξόδου ────────────────────────────────────────────────────
        tk.Label(self._exec_body, text='Προαιρετικές στήλες εξόδου (επιλέξτε όσες επιθυμείτε):', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(6, 2))

        col_frame = tk.Frame(self._exec_body, bg=C['bg'])
        col_frame.pack(fill='x', padx=18, pady=(0, 6))

        self._col_vars = {}
        for col_name in ('Email στο ΠΣΔ', 'Email', 'Κινητό'):
            var = tk.BooleanVar(value=True)
            self._col_vars[col_name] = var
            tk.Checkbutton(col_frame, text=col_name, variable=var,
                           bg=C['bg'], font=('Arial', 9),
                           activebackground=C['bg']).pack(side='left', padx=(0, 12))

        # ── Email ─────────────────────────────────────────────────────────────
        pad = dict(padx=18, pady=2)

        self._to_txt = _build_recipients_box(
            self._exec_body, 'Προς (email συμβούλου):', self._saved_email)

        tk.Label(self._exec_body, text='Θέμα:',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._subj_var = tk.StringVar(value=self._saved_subject)
        tk.Entry(self._exec_body, textvariable=self._subj_var,
                 font=('Arial', 9)).pack(fill='x', padx=18, pady=(0, 6))

        tk.Label(self._exec_body, text='Κείμενο email:',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._body_txt = tk.Text(self._exec_body, font=('Arial', 9), height=6,
                                  wrap='word', relief='solid', bd=1)
        self._body_txt.pack(fill='x', padx=18, pady=(0, 6))
        from datetime import datetime as _dt
        self._body_txt.insert('1.0',
            self._saved_body.replace('{date}', _dt.today().strftime('%d/%m/%Y'))
                             .replace('{specialty}', self._spec_var.get()))

        btn_row = tk.Frame(self._exec_body, bg=C['bg'])
        btn_row.pack(side='bottom', pady=10)
        tk.Button(btn_row, text='Μόνο Excel (χωρίς email)',
                  bg=C['bg2'], fg=C['hdr_bg'], relief='flat',
                  font=('Arial', 9), padx=10, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=False)).pack(side='left', padx=4)
        tk.Button(btn_row, text='▶  Δημιουργία & Αποστολή',
                  bg=C['btn_bg'], fg=C['btn_fg'], relief='flat',
                  font=('Arial', 9, 'bold'), padx=14, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=True)).pack(side='left', padx=4)

        self.after(100, self._load_specialties)

    def _reset_direction(self):
        """Διαγράφει την αποθηκευμένη Διεύθυνση και επαναφορτώνει τη φόρμα."""
        import json
        s = _load_local_settings()
        if self._SETTINGS_KEY in s:
            s[self._SETTINGS_KEY].pop('direction', None)
            path_s = _get_local_settings_path()
            with open(path_s, 'w', encoding='utf-8') as f:
                json.dump(s, f, ensure_ascii=False, indent=2)
        self._saved_dir = ''
        self._build_form()

    def _on_spec_change(self, *_):
        """Ενημερώνει το θέμα και το body όταν αλλάζει η ειδικότητα."""
        spec = self._spec_var.get()
        self._subj_var.set(self._saved_subject.replace('{specialty}', spec))
        if hasattr(self, '_body_txt'):
            from datetime import datetime as _dt
            self._body_txt.delete('1.0', 'end')
            self._body_txt.insert('1.0',
                self._saved_body
                    .replace('{date}', _dt.today().strftime('%d/%m/%Y'))
                    .replace('{specialty}', spec))

    def _load_specialties(self):
        """Φορτώνει Διευθύνσεις και Ειδικότητες από το Topothetiseis αρχείο."""
        if not self._topoth_path:
            self._spec_lbl.config(text='Δεν βρέθηκε αρχείο Τοποθετήσεων.', fg='#CC0000')
            return
        try:
            import pandas as pd
            df = pd.read_excel(self._topoth_path, header=0)

            # Έλεγχος εγκυρότητας — το Topothetiseis πρέπει να έχει τουλάχιστον 10 γραμμές
            if len(df) < 10:
                self._spec_lbl.config(
                    text=f'⚠ Το αρχείο Τοποθετήσεων φαίνεται ελλιπές ({len(df)} γραμμές). '
                         f'Κατεβάστε το ξανά από το myschool.',
                    fg='#CC0000'
                )
                return

            # Αποθηκεύουμε ολόκληρο το df για φιλτράρισμα αργότερα
            self._topoth_df = df

            spec_col = self._fc(df, 'κλάδ', 'ειδικ') or df.columns[4]
            self._topoth_spec_col = spec_col

            # Στήλη Περιοχή Μετάθεσης Φορέα (col19)
            area_col = self._fc(df, 'περιοχή μετάθεσης φορέα', 'μετάθεσης φορέα') \
                       or df.columns[19]
            self._topoth_area_col = area_col

            # Φόρτωσε Διευθύνσεις (μοναδικές, ταξινομημένες)
            dirs = sorted(df[area_col].dropna().astype(str).str.strip()
                          .replace('', pd.NA).dropna().unique())

            if self._dir_combo is not None:
                # Πρώτη φορά — dropdown
                self._dir_combo.config(values=dirs)
                if dirs and not self._dir_var.get():
                    self._dir_var.set(dirs[0])
                self._dir_lbl.config(text=f'{len(dirs)} διευθύνσεις', fg=C['desc'])
            else:
                # Αποθηκευμένη Διεύθυνση — έλεγχος ότι υπάρχει ακόμα στο αρχείο
                saved = self._dir_var.get()
                if saved not in dirs and dirs:
                    self._dir_var.set(dirs[0])

            # Φόρτωσε ειδικότητες για την επιλεγμένη Διεύθυνση
            self._update_specialties()

            self._spec_lbl.config(
                text=f'{os.path.basename(self._topoth_path)}', fg=C['desc'])

            # Όταν αλλάζει η Διεύθυνση → ανανέωσε ειδικότητες
            self._dir_var.trace_add('write', lambda *_: self._update_specialties())

        except Exception as e:
            self._spec_lbl.config(text=f'Σφάλμα φόρτωσης: {e}', fg='#CC0000')

    def _update_specialties(self):
        """Ανανεώνει τη λίστα ειδικοτήτων βάσει της επιλεγμένης Διεύθυνσης."""
        if not hasattr(self, '_topoth_df'):
            return
        import pandas as pd
        df = self._topoth_df
        selected_dir = self._dir_var.get().strip()
        if selected_dir:
            mask = df[self._topoth_area_col].fillna('').astype(str).str.strip() == selected_dir
            df_filtered = df[mask]
        else:
            df_filtered = df
        specialties = sorted(df_filtered[self._topoth_spec_col].dropna().astype(str).unique())
        self._spec_combo.config(values=specialties)
        if specialties:
            self._spec_var.set(specialties[0])
            if hasattr(self, '_body_txt'):
                from datetime import datetime as _dt
                self._body_txt.delete('1.0', 'end')
                self._body_txt.insert('1.0',
                    self._saved_body
                        .replace('{date}', _dt.today().strftime('%d/%m/%Y'))
                        .replace('{specialty}', specialties[0]))
        self._spec_lbl.config(
            text=f'{len(specialties)} ειδικότητες για "{selected_dir}"', fg=C['desc'])

    # ── Βοηθητικά ───────────────────────────────────────────────────────────

    @staticmethod
    def _fc(df, *keywords):
        """Επιστρέφει το όνομα της πρώτης στήλης που ταιριάζει με κάποια λέξη-κλειδί."""
        for kw in keywords:
            kw = kw.lower()
            for col in df.columns:
                if kw in str(col).lower():
                    return col
        return None

    @staticmethod
    def _norm_code(series):
        """Κανονικοποίηση κωδικού σχολείου: αριθμός ή string → stripped string."""
        return series.fillna('').astype(str).str.strip().str.lstrip('0') \
                     .str.replace(r'\.0$', '', regex=True)

    @staticmethod
    def _clean_afm(val):
        """Καθαρισμός ΑΦΜ από CSV format =\"152159882\" → '152159882'."""
        return str(val).strip().strip('"').lstrip('=').strip('"').strip()

    # ── Εκτέλεση ────────────────────────────────────────────────────────────

    def _execute(self, send=True):
        import json, pandas as pd
        from datetime import datetime, date
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        specialty = self._spec_var.get()
        to_emails = _parse_recipients(self._to_txt)
        subject   = self._subj_var.get().strip()
        body_text = self._body_txt.get('1.0', 'end-1c')
        full_body = body_text + '\n\n' + config.email_signature()

        if send and not to_emails:
            messagebox.showwarning('Email', 'Εισάγετε email παραλήπτη.', parent=self)
            return

        if not specialty:
            messagebox.showwarning('Ειδικότητα', 'Επίλεξε ειδικότητα.', parent=self)
            return
        if not self._topoth_path or not self._grid_path:
            from core.framework import _missing_file_dialog
            _missing_file_dialog('Εκπαιδευτικοί ανά Ειδικότητα', [
                'Τοποθετήσεις εκπαιδευτικών (gridResults / 2.1)',
                'Στατιστικά 4.1 / 4.2 / 4.16',
            ])
            return

        try:
            # ── 1. Τοποθετήσεις ──────────────────────────────────────────────
            df_t = pd.read_excel(self._topoth_path, header=0)

            # Στήλες στο Topothetiseis (βάσει dump: col4=ειδικ, col5=σχέση εργ,
            #   col6=σχέση τοποθ, col7=κωδικός, col8=φορέας, col16=έως)
            spec_col      = self._topoth_spec_col
            code_col      = self._fc(df_t, 'κωδικ')        or df_t.columns[7]
            eos_col       = self._fc(df_t, 'έως', 'εως')   or df_t.columns[16]
            eponym_col    = self._fc(df_t, 'επώνυμ')       or df_t.columns[2]
            org_col       = self._fc(df_t, 'σχέση εργ', 'οργαν') or df_t.columns[5]
            topoth_col    = self._fc(df_t, 'σχέση τοποθ')  or df_t.columns[6]
            school_name_col = self._fc(df_t, 'φορέας τοποθ', 'φορέας') or df_t.columns[8]

            # ΑΦΜ (Α.Φ.Μ.) — join key, υπάρχει σε όλους (μόνιμοι + αναπληρωτές)
            # Α.Φ.Μ. (ΑΦΜ) — join key για stat4_1/4_2 (9 ψηφία)
            afm_col = self._fc(df_t, 'α.φ.μ', 'αφμ') or df_t.columns[1]
            # Α.Μ. — join key για stat4_16 (6 ψηφία) + εμφάνιση στην έξοδο
            am_col = None
            for col in df_t.columns:
                c = str(col).lower().strip()
                if 'α.μ' in c and 'φ' not in c:
                    am_col = col; break
            if am_col is None:
                am_col = df_t.columns[0]

            # Βρες Όνομα εκπ/κού — εξαιρούμε "Ονομασία σχολείου"
            onoma_col = None
            for col in df_t.columns:
                c = str(col).lower()
                if ('όνομ' in c or 'ονομ' in c) and 'ονομασ' not in c and 'σχολ' not in c:
                    onoma_col = col; break

            # ── Φίλτρα καθαρισμού Τοποθετήσεων ──────────────────────────────
            # 1. Αφαίρεση εγγραφών με Κατάσταση = ΠΑΡΗΛΘΕ
            status_col = self._fc(df_t, 'κατάσταση', 'κατασταση') or df_t.columns[17]
            df_t = df_t[df_t[status_col].fillna('').astype(str).str.strip() != 'ΠΑΡΗΛΘΕ'].copy()

            # 2. Αφαίρεση συγκεκριμένων τύπων σχέσης εργασίας
            # Χρήση contains γιατί οι τιμές μπορεί να έχουν παρενθέσεις
            # π.χ. "Ιδιωτικού Δικαίου Αορίστου Χρόνου (Ι.Δ.Α.Χ.)"
            _EXCLUDE_ORG_PAT = (
                r'Με άδεια διδασκαλίας για Ξένο Σχολείο'
                r'|Αναπληρωτής Ιδιωτικής Εκπαίδευσης'
                r'|Ιδιωτικού Δικαίου Αορίστου Χρόνου'
            )
            df_t = df_t[~df_t[org_col].fillna('').astype(str).str.strip()
                        .str.contains(_EXCLUDE_ORG_PAT, regex=True, na=False)].copy()

            # 3. Φιλτράρισμα βάσει επιλεγμένης Διεύθυνσης
            area_mt_col = self._topoth_area_col \
                          if hasattr(self, '_topoth_area_col') \
                          else (self._fc(df_t, 'περιοχή μετάθεσης φορέα', 'μετάθεσης φορέα')
                                or df_t.columns[19])
            selected_dir = self._dir_var.get().strip()
            if selected_dir:
                df_t = df_t[
                    df_t[area_mt_col].fillna('').astype(str).str.strip() == selected_dir
                ].copy()

            # 4. Αφαίρεση συγκεκριμένων τύπων σχέσης τοποθέτησης
            # Χρήση contains γιατί οι τιμές μπορεί να έχουν παρενθέσεις
            # π.χ. "Μερική Διάθεση (αναπληρωτές εκπαιδευτικοί)"
            _EXCLUDE_TOPOTH_PAT = r'Υπερωριακά|Μερική Διάθεση|Τοποθέτηση Διοικητικού'
            df_t = df_t[~df_t[topoth_col].fillna('').astype(str).str.strip()
                        .str.contains(_EXCLUDE_TOPOTH_PAT, regex=True, na=False)].copy()


            # Κανονικοποίηση κωδικού, ΑΦΜ και Α.Μ.
            df_t['_code'] = self._norm_code(df_t[code_col])
            df_t['_afm']  = df_t[afm_col].fillna('').astype(str).str.strip() \
                                         .str.replace(r'\.0$', '', regex=True) \
                                         .str.zfill(9)
            df_t['_am']   = df_t[am_col].fillna('').astype(str).str.strip() \
                                        .str.replace(r'\.0$', '', regex=True)

            # ── 2. gridResults ────────────────────────────────────────────────
            df_g = pd.read_excel(self._grid_path, header=0)

            gc_code  = self._fc(df_g, 'κωδικός', 'κωδ')    or df_g.columns[11]
            gc_name  = self._fc(df_g, 'ονομασ')             or df_g.columns[1]
            gc_phone = self._fc(df_g, 'τηλ')                or df_g.columns[15]
            gc_email = self._fc(df_g, 'e-mail', 'email')    or df_g.columns[17]
            gc_area  = self._fc(df_g, 'περιοχ', 'τοποθεσ') or df_g.columns[18]
            gc_dimos = self._fc(df_g, 'δήμ', 'δημ')

            # Φίλτρα gridResults
            gc_eidos = self._fc(df_g, 'είδος', 'ειδος')
            if gc_eidos:
                df_g = df_g[df_g[gc_eidos].fillna('').astype(str).str.strip() != 'Ιδιωτικά Σχολεία'].copy()

            df_g['_code'] = self._norm_code(df_g[gc_code])
            _g_cols = ['_code', gc_name, gc_phone, gc_email, gc_area]
            _g_names = ['_code', '_school_name', '_phone', '_school_email', '_area']
            if gc_dimos:
                _g_cols.append(gc_dimos)
                _g_names.append('_dimos')
            df_g_lu = df_g[_g_cols].drop_duplicates('_code').copy()
            df_g_lu.columns = _g_names
            df_g_lu['_phone'] = df_g_lu['_phone'].fillna('').astype(str) \
                                    .str.replace(r'\.0$', '', regex=True).str.strip()

            valid_codes = set(df_g_lu['_code'])

            # ── 3. Φιλτράρισμα: ειδικότητα + (προαιρετικά) σχολεία Δ/νσης ──
            # Το φίλτρο valid_codes εφαρμόζεται μόνο αν η Διεύθυνση είναι Π.Ε.
            # (τα σχολεία ΔΔΕ δεν υπάρχουν στο gridResults/2.1)
            selected_dir = self._dir_var.get().strip()
            _is_pe = 'Π.Ε' in selected_dir or not selected_dir
            if _is_pe:
                df_t = df_t[df_t['_code'].isin(valid_codes)].copy()
            df_t = df_t[df_t[spec_col].astype(str) == specialty].copy()

            if df_t.empty:
                messagebox.showwarning('Αποτέλεσμα',
                    f'Δεν βρέθηκαν εκπαιδευτικοί ειδικότητας "{specialty}".', parent=self)
                return

            # ── 4. stat4_16 (απόντες — αιτιολόγηση απουσίας) ────────────────
            def _read_csv_enc(path):
                if not path: return pd.DataFrame()
                import zipfile as _zf, io as _io
                if path.endswith(('.xlsx', '.xls')):
                    try:
                        return pd.read_excel(path, header=0, dtype=str)
                    except Exception:
                        return pd.DataFrame()
                if path.endswith('.zip'):
                    try:
                        with _zf.ZipFile(path) as z:
                            csvname = [n for n in z.namelist() if n.endswith('.csv')][0]
                            data = z.read(csvname)
                    except Exception:
                        return pd.DataFrame()
                    for enc in ['utf-8-sig', 'utf-8', 'iso-8859-7', 'cp1253']:
                        try:
                            return pd.read_csv(_io.BytesIO(data), sep=None, engine='python',
                                               encoding=enc, header=0, dtype=str)
                        except Exception:
                            continue
                    return pd.DataFrame()
                for enc in ['utf-8-sig', 'utf-8', 'iso-8859-7', 'cp1253']:
                    try:
                        return pd.read_csv(path, sep=None, engine='python',
                                           encoding=enc, header=0, dtype=str)
                    except Exception:
                        continue
                return pd.DataFrame()

            df_s16 = _read_csv_enc(self._stat_path)
            if not df_s16.empty:
                # stat4_16: col16 labeled "Α.Μ." αλλά έχει ΑΦΜ δεδομένα (9 ψηφία, ="..." format)
                # → join με ΑΦΜ (ίδιο key με Topothetiseis col1 και stat4_1/4_2 col0)
                # 1-column shift: header[17]='Α.Μ.' → data[16] — χρήση absolute index
                s16_afm_col = df_s16.columns[16]
                # col shift: header[i] περιγράφει data[i-1]
                # header[45]='Αιτιολόγηση Απουσίας' → data[44]
                # header[47]='Από'                   → data[46]
                # header[48]='Έως'                   → data[47]
                s16_abs_col = df_s16.columns[44] if len(df_s16.columns) > 44 else df_s16.columns[45]
                s16_apo_col = df_s16.columns[46] if len(df_s16.columns) > 46 else None
                s16_eos_col = df_s16.columns[47] if len(df_s16.columns) > 47 else None

                df_s16['_afm'] = df_s16[s16_afm_col].apply(self._clean_afm).str.zfill(9)
                keep16 = ['_afm', s16_abs_col]
                if s16_apo_col: keep16.append(s16_apo_col)
                if s16_eos_col: keep16.append(s16_eos_col)
                df_s16_lu = df_s16[keep16].drop_duplicates('_afm').copy()
                rename16 = {'_afm': '_afm', s16_abs_col: '_apoysia'}
                if s16_apo_col: rename16[s16_apo_col] = '_apo'
                if s16_eos_col: rename16[s16_eos_col] = '_eos'
                df_s16_lu = df_s16_lu.rename(columns=rename16)
                if '_apo' not in df_s16_lu: df_s16_lu['_apo'] = ''
                if '_eos' not in df_s16_lu: df_s16_lu['_eos'] = ''
                absent_afms = set(df_s16_lu['_afm'])
            else:
                df_s16_lu = pd.DataFrame(columns=['_afm', '_apoysia', '_apo', '_eos'])
                absent_afms = set()

            # ── 5. stat4_1 & stat4_2 (Email ΠΣΔ, Κινητό) — join με ΑΦΜ ─────
            # stat4_1 col0 labeled "Α.Μ." αλλά έχει ΑΦΜ δεδομένα (9 ψηφία, ="..." format)
            frames_41_42 = []
            for path in [self._stat41_path, self._stat42_path]:
                df_tmp = _read_csv_enc(path)
                if not df_tmp.empty:
                    frames_41_42.append(df_tmp)

            if frames_41_42:
                df_41_42 = pd.concat(frames_41_42, ignore_index=True)
                # stat4_1/4_2 έχουν 1-column shift στα headers.
                # col0  (1-based:  1) → ΑΦΜ data (="..." format)
                # col9  (1-based: 10) → Κινητό data
                # col11 (1-based: 12) → Email προσωπικό data
                # col12 (1-based: 13) → Email ΠΣΔ (sch.gr) data
                s41_afm_col      = df_41_42.columns[0]
                s41_psd_col      = df_41_42.columns[12] if len(df_41_42.columns) > 12 else None
                s41_email_col    = df_41_42.columns[11] if len(df_41_42.columns) > 11 else None
                s41_mobile_col   = df_41_42.columns[9]  if len(df_41_42.columns) > 9  else None

                df_41_42['_afm'] = df_41_42[s41_afm_col].apply(self._clean_afm).str.zfill(9)
                keep = ['_afm']
                if s41_psd_col:    keep.append(s41_psd_col)
                if s41_email_col:  keep.append(s41_email_col)
                if s41_mobile_col: keep.append(s41_mobile_col)
                df_41_lu = df_41_42[keep].drop_duplicates('_afm').copy()
                rename = {}
                if s41_psd_col:    rename[s41_psd_col]    = '_email_psd'
                if s41_email_col:  rename[s41_email_col]  = '_email_personal'
                if s41_mobile_col: rename[s41_mobile_col] = '_kinito'
                df_41_lu = df_41_lu.rename(columns=rename)
                if '_email_psd'      not in df_41_lu: df_41_lu['_email_psd']      = ''
                if '_email_personal' not in df_41_lu: df_41_lu['_email_personal'] = ''
                if '_kinito'         not in df_41_lu: df_41_lu['_kinito']         = ''
            else:
                df_41_lu = pd.DataFrame(columns=['_afm', '_email_psd', '_email_personal', '_kinito'])

            # ── 6. Join ───────────────────────────────────────────────────────
            df_t = df_t.merge(df_g_lu,   on='_code', how='left')
            df_t = df_t.merge(df_s16_lu, on='_afm',  how='left')   # ΑΦΜ join (stat4_16)
            df_t = df_t.merge(df_41_lu,  on='_afm',  how='left')   # ΑΦΜ join (stat4_1/4_2)
            df_t['_absent'] = df_t[status_col].fillna('').astype(str).str.strip() == 'ΑΠΟΥΣΙΑ'

            # ── 7. Χτίσε dataframe εξόδου ─────────────────────────────────────
            def gcol(col):
                if col is not None and col in df_t.columns:
                    return df_t[col].fillna('').astype(str)
                return pd.Series([''] * len(df_t), index=df_t.index)

            out = pd.DataFrame(index=df_t.index)
            out['ΑΜ']                = df_t['_am'].fillna('')
            out['Επώνυμο']            = gcol(eponym_col)
            out['Όνομα']              = gcol(onoma_col)
            out['Κύρια Ειδικ.']      = gcol(spec_col)
            out['Email στο ΠΣΔ']     = df_t['_email_psd'].fillna('')
            out['Email']             = df_t['_email_personal'].fillna('') \
                                        if '_email_personal' in df_t.columns else ''
            out['Κινητό']            = df_t['_kinito'].fillna('')
            out['Σχέση εργασίας']    = gcol(org_col)
            out['Σχέση τοποθέτησης'] = gcol(topoth_col)
            out['Κατάσταση']         = gcol(status_col)
            out['Φορέας τοποθέτησης']= gcol(school_name_col)
            out['Δήμος']             = df_t['_dimos'].fillna('') if '_dimos' in df_t.columns else ''
            out['Τηλέφωνο']          = df_t['_phone'].fillna('')
            out['e-mail']            = df_t['_school_email'].fillna('')
            out['ΑΠΟΥΣΙΑ']           = df_t['_apoysia'].fillna('')
            out['Από']               = df_t['_apo'].fillna('') if '_apo' in df_t.columns else ''
            out['Έως']               = df_t['_eos'].fillna('') if '_eos' in df_t.columns else ''
            out['_absent']           = df_t['_absent']
            # Αιτιολόγηση + ημ/νίες απουσίας μόνο για απόντες
            out.loc[~out['_absent'], 'ΑΠΟΥΣΙΑ'] = ''
            out.loc[~out['_absent'], 'Από']     = ''
            out.loc[~out['_absent'], 'Έως']     = ''

            out = out.sort_values('Επώνυμο', na_position='last').reset_index(drop=True)

            # ── 7. Δημιουργία Excel ──────────────────────────────────────────
            today_str = datetime.today().strftime('%Y%m%d')
            out_dir   = os.path.join(_docs_base(), f'results_{today_str}')
            os.makedirs(out_dir, exist_ok=True)
            spec_safe = specialty.replace('/', '_').replace('\\', '_')
            out_path  = os.path.join(out_dir, f'Εκπαιδευτικοί_{spec_safe}_{today_str}.xlsx')

            # Φιλτράρισμα στηλών βάσει επιλογής χρήστη
            _disabled = {c for c, v in getattr(self, '_col_vars', {}).items() if not v.get()}
            active_cols = [c for c in self._OUT_COLS if c not in _disabled]

            wb = Workbook()
            ws = wb.active
            ws.title = specialty[:31]

            RED = 'FF0000'
            thin   = Side(style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row_align = Alignment(horizontal='left',   vertical='center')

            # Επικεφαλίδες — κόκκινο φόντο, λευκό bold
            for ci, col in enumerate(active_cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                cell.fill      = PatternFill('solid', start_color=RED)
                cell.alignment = hdr_align
                cell.border    = border

            # Δεδομένα
            alt_fill = PatternFill('solid', start_color='FFF0F0')
            for ri, row in out.iterrows():
                is_absent = bool(row.get('_absent', False))
                for ci, col in enumerate(active_cols, 1):
                    val = row.get(col, '')
                    if pd.isna(val):
                        val = ''
                    cell = ws.cell(row=ri + 2, column=ci, value=str(val) if val != '' else '')
                    if is_absent:
                        cell.font = Font(name='Arial', size=9, color=RED, bold=True)
                    else:
                        cell.font = Font(name='Arial', size=9, color='000000')
                        if ri % 2 == 1:
                            cell.fill = alt_fill
                    cell.alignment = row_align
                    cell.border    = border

            # Πλάτη στηλών
            for ci, col in enumerate(active_cols, 1):
                vals = [str(out.iloc[r][col]) for r in range(min(len(out), 50))
                        if col in out.columns and not pd.isna(out.iloc[r][col])]
                w = max([len(col)] + [len(v) for v in vals]) if vals else len(col)
                ws.column_dimensions[get_column_letter(ci)].width = min(w + 3, 42)

            ws.row_dimensions[1].height = 30
            ws.freeze_panes = 'A2'
            wb.save(out_path)

        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror('Σφάλμα', str(e), parent=self)
            return

        # ── 8. Αποθήκευση ρυθμίσεων ──────────────────────────────────────────
        s = _load_local_settings()
        # Αποθήκευση με placeholders {specialty} και {date} ώστε να αντικαθίστανται σωστά την επόμενη φορά
        from datetime import datetime as _dt_save
        _today_save = _dt_save.today().strftime('%d/%m/%Y')
        body_to_save    = body_text.replace(specialty, '{specialty}') if specialty else body_text
        body_to_save    = body_to_save.replace(_today_save, '{date}')
        subject_to_save = subject.replace(specialty, '{specialty}')   if specialty else subject
        subject_to_save = subject_to_save.replace(_today_save, '{date}')
        self._saved_body    = body_to_save
        self._saved_subject = subject_to_save
        s[self._SETTINGS_KEY] = {
            'subject':       subject_to_save,
            'body':          body_to_save,
            'advisor_email': '\n'.join(to_emails),
            'direction':     self._dir_var.get().strip(),
        }
        path_s = _get_local_settings_path()
        os.makedirs(os.path.dirname(path_s), exist_ok=True)
        with open(path_s, 'w', encoding='utf-8') as f:
            json.dump(s, f, ensure_ascii=False, indent=2)

        absent_count = int(out['_absent'].sum())
        total_count  = len(out)

        if not send:
            from core.framework import _show_results_popup
            _show_results_popup('Εκπ/κοί ανά Ειδικότητα',
                f'Αρχείο αποθηκεύτηκε.\n\n'
                f'Σύνολο: {total_count} εκπ/κοί  |  Απόντες (κόκκινο): {absent_count}',
                result_type='warn', excel_path=out_path)
            self.destroy()
            return

        # Αποστολή email
        try:
            from core.framework import send_email
            send_email(config, to_emails, subject, full_body, out_path)
            from core.framework import _show_results_popup
            _show_results_popup('Εκπ/κοί ανά Ειδικότητα',
                f'Email στάλθηκε: {", ".join(to_emails)}\n\n'
                f'Σύνολο: {total_count} εκπ/κοί  |  Απόντες: {absent_count}',
                result_type='ok', excel_path=out_path)
            self.destroy()
        except Exception as e:
            messagebox.showerror('Σφάλμα αποστολής', str(e), parent=self)

    # ── Βοηθητικά ───────────────────────────────────────────────────────────

    def _clear(self):
        for w in self._exec_body.winfo_children():
            w.destroy()

    def _on_download_done(self):
        """Καλείται μετά από επιτυχή λήψη στο tab «⬇ Λήψη» — ξανα-ανιχνεύει
        τα αρχεία και ξαναχτίζει το tab «▶ Εκτέλεση» με τα φρέσκα αρχεία."""
        self._topoth_path = self._auto_find('Topothetiseis')
        self._grid_path   = self._auto_find('gridResults')
        self._stat_path   = self._auto_find('stat4_16')
        self._stat41_path = self._auto_find('stat4_1')
        self._stat42_path = self._auto_find('stat4_2')
        self._build_form()


class SymbouloiDialog(tk.Toplevel):
    """Εκπ/κοί ανά Ειδικότητα & Θέση Συμβούλου — μόνο για ΔΙ.Π.Ε.Αν.Θ."""

    _SETTINGS_KEY    = 'symvouloi_tool'
    _DEFAULT_BODY    = (
        'Αποτύπωση Myschool {date}.\n\n'
        'Καλημέρα σας,\n\n'
        'Επισυνάπτεται πίνακας excel με τους εκπαιδευτικούς ειδικότητας {specialty} '
        'που υπηρετούν στη Δ/νση μας σύμφωνα με τα καταχωρημένα '
        'στοιχεία στο myschool.\n\n\n'
        'Στη διάθεσή σας για οποιαδήποτε πληροφορία'
    )
    _DEFAULT_SUBJECT = 'Στοιχεία τοποθετήσεων εκπ/κών "{specialty}" — {thesi}'

    _OUT_COLS = [
        'ΑΜ',
        'Επώνυμο', 'Όνομα', 'Κύρια Ειδικ.',
        'Email στο ΠΣΔ', 'Email', 'Κινητό',
        'Σχέση εργασίας', 'Σχέση τοποθέτησης',
        'Κατάσταση',
        'Φορέας τοποθέτησης', 'Δήμος',
        'Τηλέφωνο', 'e-mail',
        'ΑΠΟΥΣΙΑ', 'Από', 'Έως',
    ]

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Εκπ/κοί ανά Ειδικότητα & Θέση')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self._parent = parent

        s = _load_local_settings().get(self._SETTINGS_KEY, {})
        self._saved_subject = s.get('subject',       self._DEFAULT_SUBJECT)
        self._saved_body    = s.get('body',          self._DEFAULT_BODY)
        self._saved_email   = s.get('advisor_email',    '')
        self._saved_dir     = s.get('direction',        '')
        self._saved_include_private = s.get('include_private', True)
        if '{specialty}' not in self._saved_subject:
            self._saved_subject = self._DEFAULT_SUBJECT
        if '{specialty}' not in self._saved_body or '{date}' not in self._saved_body:
            self._saved_body = self._DEFAULT_BODY

        self._topoth_path = EidikotitaDialog._auto_find('Topothetiseis')
        self._grid_path   = EidikotitaDialog._auto_find('gridResults')
        self._stat_path   = EidikotitaDialog._auto_find('stat4_16')
        self._stat41_path = EidikotitaDialog._auto_find('stat4_1')
        self._stat42_path = EidikotitaDialog._auto_find('stat4_2')

        self._symv = {}
        self._load_symvouloi_data()

        # 2-tab δομή όπως οι έλεγχοι: «⬇ Λήψη» + «▶ Εκτέλεση» (το ίδιο
        # εργαλείο που υπήρχε πριν, χωρίς αλλαγή στη λειτουργία του).
        from tkinter import ttk as _ttk
        nb = _ttk.Notebook(self)
        nb.pack(fill='both', expand=True)
        tab_dl   = tk.Frame(nb, bg=C['bg'])
        tab_exec = tk.Frame(nb, bg=C['bg'])
        nb.add(tab_dl,   text='  ⬇ Λήψη  ')
        nb.add(tab_exec, text='  ▶ Εκτέλεση  ')
        self._exec_body = tab_exec

        _build_report_download_tab(self, tab_dl, _EIDIK_RIDS, _EIDIK_REQUIRED,
                                    on_done=self._on_download_done)

        self._build_form()
        self.update_idletasks()
        w = 640
        h = self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f'{w}x{h}+{x}+{y}')

    @staticmethod
    def _get_symvouloi_path():
        return os.path.join(_app_base(), 'data', 'symvouloi_ekpaidefsis.xlsx')

    def _load_symvouloi_data(self):
        path = self._get_symvouloi_path()
        if not os.path.exists(path):
            return
        try:
            import openpyxl as _oxl
            wb = _oxl.load_workbook(path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name.upper() == 'ΑΠΟΡΙΑ':
                    continue
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue
                title    = str(rows[0][0] or sheet_name)
                parts    = title.split(' — ', 1)
                spec_pos = parts[0].strip()
                advisor  = parts[1].strip() if len(parts) > 1 else ''
                words    = spec_pos.split()
                spec     = words[0]
                pos_label = ' '.join(words[1:])
                display  = f'{pos_label} — {advisor}' if pos_label else f'(ενιαία) — {advisor}'
                codes = set()
                for row in rows[2:]:
                    val = row[1]
                    if val is not None:
                        codes.add(str(val).strip().lstrip('0')
                                  .replace('.0', '').replace('="', '').replace('"', ''))
                if spec not in self._symv:
                    self._symv[spec] = []
                self._symv[spec].append({'label': display, 'codes': codes})
            wb.close()
        except Exception as e:
            print(f'[SymbouloiDialog] Σφάλμα: {e}')

    def _build_form(self):
        self._clear()
        HDR = '#1F4E79'
        hdr = tk.Frame(self._exec_body, bg=HDR, pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='📋  Εκπ/κοί ανά Ειδικότητα & Θέση Συμβούλου',
                 bg=HDR, fg='white', font=('Arial', 12, 'bold')).pack()
        tk.Label(hdr,
                 text='μόνο για χρήση από Δ/νση Π.Ε. Αν. Θεσσαλονίκης  '
                      '(Απαιτούνται: Τοποθετήσεις, 2.1, 4.1, 4.2, 4.16)',
                 bg=HDR, fg='#A8C4D8',
                 font=('Arial', 8, 'italic'), wraplength=600, justify='center').pack()

        symv_path = self._get_symvouloi_path()
        if not self._symv:
            wf = tk.Frame(self._exec_body, bg='#FFF3E0')
            wf.pack(fill='x', padx=18, pady=(6, 0))
            tk.Label(wf,
                     text=f'⚠  Δεν βρέθηκε το αρχείο Συμβούλων Εκπ/σης. Αναμένεται: {symv_path}',
                     bg='#FFF3E0', fg='#E65100', font=('Arial', 8),
                     anchor='w', padx=10, pady=5, wraplength=500, justify='left').pack(side='left')
            tk.Button(wf, text='📂 Εισαγωγή', bg='#E65100', fg='white', relief='flat',
                      font=('Arial', 8, 'bold'), cursor='hand2', padx=6, pady=4,
                      command=self._import_symvouloi).pack(side='right', padx=6, pady=4)

        missing = []
        if not self._topoth_path: missing.append('Τοποθετήσεις')
        if not self._grid_path:   missing.append('Κατάλογος σχολείων (2.1)')
        if missing:
            tk.Label(self._exec_body,
                text=f'⚠  Δεν βρέθηκαν: {", ".join(missing)}. Κατέβασε τα από «Λήψη Δεδομένων».',
                bg='#FFF3E0', fg='#E65100', font=('Arial', 8),
                anchor='w', padx=28, pady=5, wraplength=580, justify='left').pack(fill='x')
        missing_opt = []
        if not self._stat41_path: missing_opt.append('4.1')
        if not self._stat42_path: missing_opt.append('4.2')
        if missing_opt:
            tk.Label(self._exec_body,
                text=f'ℹ  Δεν βρέθηκαν: {", ".join(missing_opt)} — Email ΠΣΔ / Email / Κινητό θα είναι κενές.',
                bg='#E3F2FD', fg='#1565C0', font=('Arial', 8),
                anchor='w', padx=28, pady=5, wraplength=580, justify='left').pack(fill='x')

        from tkinter import ttk as _ttk

        tk.Label(self._exec_body, text='Διεύθυνση:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(6, 0))
        dir_row = tk.Frame(self._exec_body, bg=C['bg'])
        dir_row.pack(fill='x', padx=18, pady=(2, 4))
        self._dir_var = tk.StringVar(value=self._saved_dir)
        if self._saved_dir:
            self._dir_combo = None
            tk.Label(dir_row, textvariable=self._dir_var,
                     bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold')).pack(side='left')
            tk.Button(dir_row, text='Αλλαγή', bg=C['bg'], fg=C['desc'],
                      relief='flat', font=('Arial', 8), cursor='hand2',
                      command=self._reset_direction).pack(side='left', padx=(10, 0))
        else:
            self._dir_combo = _ttk.Combobox(dir_row, textvariable=self._dir_var,
                                             width=48, state='readonly')
            self._dir_combo.pack(side='left')
        self._dir_lbl = tk.Label(dir_row, text='', bg=C['bg'], fg=C['desc'], font=('Arial', 8))
        self._dir_lbl.pack(side='left', padx=(10, 0))

        tk.Label(self._exec_body, text='Ειδικότητα:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))
        spec_row = tk.Frame(self._exec_body, bg=C['bg'])
        spec_row.pack(fill='x', padx=18, pady=(2, 4))
        self._spec_var = tk.StringVar()
        self._spec_combo = _ttk.Combobox(spec_row, textvariable=self._spec_var,
                                          width=20, state='readonly')
        self._spec_combo.pack(side='left')
        self._spec_lbl = tk.Label(spec_row, text='Φόρτωση…', bg=C['bg'],
                                   fg=C['desc'], font=('Arial', 8))
        self._spec_lbl.pack(side='left', padx=(10, 0))
        self._spec_var.trace_add('write', self._on_spec_change)

        tk.Label(self._exec_body, text='Θέση Συμβούλου:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))
        thesi_row = tk.Frame(self._exec_body, bg=C['bg'])
        thesi_row.pack(fill='x', padx=18, pady=(2, 6))
        self._thesi_var = tk.StringVar()
        self._thesi_combo = _ttk.Combobox(thesi_row, textvariable=self._thesi_var,
                                           width=48, state='readonly')
        self._thesi_combo.pack(side='left')
        self._thesi_lbl = tk.Label(thesi_row, text='', bg=C['bg'],
                                    fg=C['desc'], font=('Arial', 8))
        self._thesi_lbl.pack(side='left', padx=(10, 0))

        priv_frame = tk.Frame(self._exec_body, bg=C['bg'])
        priv_frame.pack(fill='x', padx=18, pady=(4, 2))
        self._include_private_var = tk.BooleanVar(value=self._saved_include_private)
        tk.Checkbutton(priv_frame, text='Συμπερίληψη ιδιωτικών;',
                       variable=self._include_private_var,
                       bg=C['bg'], font=('Arial', 9, 'bold'),
                       fg=C['hdr_bg'], activebackground=C['bg'],
                       selectcolor=C['bg']).pack(side='left')

        tk.Label(self._exec_body, text='Προαιρετικές στήλες εξόδου:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 2))
        col_frame = tk.Frame(self._exec_body, bg=C['bg'])
        col_frame.pack(fill='x', padx=18, pady=(0, 4))
        self._col_vars = {}
        for col_name in ('Email στο ΠΣΔ', 'Email', 'Κινητό'):
            var = tk.BooleanVar(value=True)
            self._col_vars[col_name] = var
            tk.Checkbutton(col_frame, text=col_name, variable=var,
                           bg=C['bg'], font=('Arial', 9),
                           activebackground=C['bg']).pack(side='left', padx=(0, 12))

        pad = dict(padx=18, pady=2)
        self._to_txt = _build_recipients_box(
            self._exec_body, 'Προς (email Συμβούλου):', self._saved_email)

        tk.Label(self._exec_body, text='Θέμα:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **pad)
        self._subj_var = tk.StringVar(value=self._saved_subject)
        tk.Entry(self._exec_body, textvariable=self._subj_var, font=('Arial', 9)).pack(fill='x', padx=18, pady=(0, 4))

        tk.Label(self._exec_body, text='Κείμενο email:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **pad)
        self._body_txt = tk.Text(self._exec_body, font=('Arial', 9), height=6, wrap='word', relief='solid', bd=1)
        self._body_txt.pack(fill='x', padx=18, pady=(0, 4))
        from datetime import datetime as _dt
        self._body_txt.insert('1.0',
            self._saved_body.replace('{date}', _dt.today().strftime('%d/%m/%Y'))
                             .replace('{specialty}', '').replace('{thesi}', ''))

        btn_row = tk.Frame(self._exec_body, bg=C['bg'])
        btn_row.pack(side='bottom', pady=10)
        tk.Button(btn_row, text='Μόνο Excel (χωρίς email)',
                  bg=C['bg2'], fg=C['hdr_bg'], relief='flat',
                  font=('Arial', 9), padx=10, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=False)).pack(side='left', padx=4)
        tk.Button(btn_row, text='▶  Δημιουργία & Αποστολή',
                  bg=C['btn_bg'], fg=C['btn_fg'], relief='flat',
                  font=('Arial', 9, 'bold'), padx=14, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=True)).pack(side='left', padx=4)

        self.after(100, self._load_dir_and_specs)

    def _reset_direction(self):
        import json
        s = _load_local_settings()
        if self._SETTINGS_KEY in s:
            s[self._SETTINGS_KEY].pop('direction', None)
            with open(_get_local_settings_path(), 'w', encoding='utf-8') as f:
                json.dump(s, f, ensure_ascii=False, indent=2)
        self._saved_dir = ''
        self._build_form()

    def _on_spec_change(self, *_):
        self._update_thesi(self._spec_var.get())
        self._refresh_email_fields()

    def _refresh_email_fields(self):
        spec  = self._spec_var.get()
        thesi = self._thesi_var.get()
        if hasattr(self, '_subj_var'):
            self._subj_var.set(
                self._saved_subject.replace('{specialty}', spec).replace('{thesi}', thesi))
        if hasattr(self, '_body_txt'):
            from datetime import datetime as _dt
            self._body_txt.delete('1.0', 'end')
            self._body_txt.insert('1.0',
                self._saved_body
                    .replace('{date}',      _dt.today().strftime('%d/%m/%Y'))
                    .replace('{specialty}', spec)
                    .replace('{thesi}',     thesi))

    def _update_thesi(self, spec):
        if not hasattr(self, '_thesi_combo'):
            return
        entries = self._symv.get(spec, [])
        labels  = [e['label'] for e in entries]
        if not labels:
            labels = ['(δεν βρέθηκε θέση — εμφανίζονται όλοι)']
            self._thesi_lbl.config(text='ειδικότητα δεν βρέθηκε στο αρχείο Συμβούλων', fg='#CC0000')
        else:
            self._thesi_lbl.config(
                text=f'{len(labels)} θέσ{"η" if len(labels)==1 else "εις"}', fg=C['desc'])
        self._thesi_combo.config(values=labels)
        self._thesi_var.set(labels[0] if labels else '')
        self._thesi_combo.bind('<<ComboboxSelected>>', lambda _: self._refresh_email_fields())

    def _load_dir_and_specs(self):
        specs = sorted(self._symv.keys())
        if not specs:
            self._spec_lbl.config(text='Αρχείο Συμβούλων δεν φορτώθηκε.', fg='#CC0000')
            return
        self._spec_combo.config(values=specs)
        self._spec_var.set(specs[0])

        if self._topoth_path:
            try:
                import pandas as pd
                df = pd.read_excel(self._topoth_path, header=0)
                self._topoth_df = df
                spec_col = EidikotitaDialog._fc(df, 'κλάδ', 'ειδικ') or df.columns[4]
                self._topoth_spec_col = spec_col
                area_col = EidikotitaDialog._fc(df, 'περιοχή μετάθεσης φορέα', 'μετάθεσης φορέα') or df.columns[19]
                self._topoth_area_col = area_col
                dirs = sorted(df[area_col].dropna().astype(str).str.strip()
                              .replace('', pd.NA).dropna().unique())
                if self._dir_combo is not None:
                    self._dir_combo.config(values=dirs)
                    if dirs and not self._dir_var.get():
                        self._dir_var.set(dirs[0])
                    self._dir_lbl.config(text=f'{len(dirs)} διευθύνσεις', fg=C['desc'])
                self._dir_var.trace_add('write', lambda *_: None)
            except Exception as e:
                self._dir_lbl.config(text=f'Σφάλμα: {e}', fg='#CC0000')

        self._spec_lbl.config(text=f'{len(specs)} ειδικότητες', fg=C['desc'])
        self._update_thesi(specs[0])

    def _import_symvouloi(self):
        from tkinter import filedialog
        import shutil
        src = filedialog.askopenfilename(parent=self,
            title='Επιλογή αρχείου Συμβούλων Εκπ/σης',
            filetypes=[('Excel', '*.xlsx *.xls'), ('Όλα', '*.*')])
        if not src:
            return
        dest = self._get_symvouloi_path()
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        shutil.copy2(src, dest)
        self._symv = {}
        self._load_symvouloi_data()
        self._build_form()

    def _execute(self, send=True):
        import json, pandas as pd
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        specialty  = self._spec_var.get().strip()
        thesi_lbl  = self._thesi_var.get().strip()
        to_emails  = _parse_recipients(self._to_txt)
        subject    = self._subj_var.get().strip()
        body_text  = self._body_txt.get('1.0', 'end-1c')
        full_body  = body_text + '\n\n' + config.email_signature()

        if send and not to_emails:
            messagebox.showwarning('Email', 'Εισάγετε email παραλήπτη.', parent=self); return
        if not specialty:
            messagebox.showwarning('Ειδικότητα', 'Επίλεξε ειδικότητα.', parent=self); return
        if not self._topoth_path or not self._grid_path:
            from core.framework import _missing_file_dialog
            _missing_file_dialog('Εκπ/κοί ανά Ειδικότητα & Θέση',
                ['Τοποθετήσεις εκπαιδευτικών (gridResults / 2.1)', 'Στατιστικά 4.1 / 4.2 / 4.16'])
            return

        _thesi_codes = None
        for e in self._symv.get(specialty, []):
            if e['label'] == thesi_lbl:
                _thesi_codes = e['codes']; break

        try:
            df_t = pd.read_excel(self._topoth_path, header=0)
            spec_col    = self._topoth_spec_col if hasattr(self, '_topoth_spec_col') \
                          else (EidikotitaDialog._fc(df_t, 'κλάδ', 'ειδικ') or df_t.columns[4])
            code_col    = EidikotitaDialog._fc(df_t, 'κωδικ')        or df_t.columns[7]
            eponym_col  = EidikotitaDialog._fc(df_t, 'επώνυμ')       or df_t.columns[2]
            org_col     = EidikotitaDialog._fc(df_t, 'σχέση εργ', 'οργαν') or df_t.columns[5]
            topoth_col  = EidikotitaDialog._fc(df_t, 'σχέση τοποθ')  or df_t.columns[6]
            school_name_col = EidikotitaDialog._fc(df_t, 'φορέας τοποθ', 'φορέας') or df_t.columns[8]
            afm_col     = EidikotitaDialog._fc(df_t, 'α.φ.μ', 'αφμ') or df_t.columns[1]
            am_col      = next((c for c in df_t.columns if 'α.μ' in str(c).lower() and 'φ' not in str(c).lower()), df_t.columns[0])
            onoma_col   = next((c for c in df_t.columns if ('όνομ' in str(c).lower() or 'ονομ' in str(c).lower()) and 'ονομασ' not in str(c).lower() and 'σχολ' not in str(c).lower()), None)
            status_col  = EidikotitaDialog._fc(df_t, 'κατάσταση', 'κατασταση') or df_t.columns[17]

            # Υπολογισμός _code νωρίς ώστε να χρησιμοποιηθεί σε όλα τα φίλτρα
            df_t['_code'] = EidikotitaDialog._norm_code(df_t[code_col])
            _EXCL_ORG = r'Με άδεια διδασκαλίας για Ξένο Σχολείο|Αναπληρωτής Ιδιωτικής Εκπαίδευσης|Ιδιωτικού Δικαίου Αορίστου Χρόνου'
            # ΠΑΡΗΛΘΕ: φιλτράρει όλους (δημόσια ΚΑΙ ιδιωτικά)
            df_t = df_t[df_t[status_col].fillna('').astype(str).str.strip() != 'ΠΑΡΗΛΘΕ'].copy()
            if self._include_private_var.get():
                # Ιδιωτικά (7xxx): παρακάμπτουν EXCL_ORG (έχουν ΙΔΑΧ / Αναπληρωτής Ιδιωτικής κλπ)
                _priv = df_t['_code'].str.startswith('7')
                df_t = df_t[_priv | ~df_t[org_col].fillna('').astype(str).str.contains(_EXCL_ORG, regex=True, na=False)].copy()
            else:
                df_t = df_t[~df_t[org_col].fillna('').astype(str).str.contains(_EXCL_ORG, regex=True, na=False)].copy()

            area_mt_col = self._topoth_area_col if hasattr(self, '_topoth_area_col') \
                          else (EidikotitaDialog._fc(df_t, 'περιοχή μετάθεσης φορέα') or df_t.columns[19])
            selected_dir = self._dir_var.get().strip()
            if selected_dir:
                _dir_mask = df_t[area_mt_col].fillna('').astype(str).str.strip() == selected_dir
                if self._include_private_var.get():
                    _dir_mask = _dir_mask | df_t['_code'].str.startswith('7')
                df_t = df_t[_dir_mask].copy()

            _EXCL_TOP = r'Υπερωριακά|Μερική Διάθεση|Τοποθέτηση Διοικητικού'
            df_t = df_t[~df_t[topoth_col].fillna('').astype(str).str.contains(_EXCL_TOP, regex=True, na=False)].copy()

            df_t['_code'] = EidikotitaDialog._norm_code(df_t[code_col])
            df_t['_afm']  = df_t[afm_col].fillna('').astype(str).str.strip().str.replace(r'\.0$','',regex=True).str.zfill(9)
            df_t['_am']   = df_t[am_col].fillna('').astype(str).str.strip().str.replace(r'\.0$','',regex=True)

            df_g = pd.read_excel(self._grid_path, header=0)
            gc_code  = EidikotitaDialog._fc(df_g, 'κωδικός', 'κωδ') or df_g.columns[11]
            gc_name  = EidikotitaDialog._fc(df_g, 'ονομασ')          or df_g.columns[1]
            gc_phone = EidikotitaDialog._fc(df_g, 'τηλ')             or df_g.columns[15]
            gc_email = EidikotitaDialog._fc(df_g, 'e-mail', 'email') or df_g.columns[17]
            gc_area  = EidikotitaDialog._fc(df_g, 'περιοχ', 'τοποθεσ') or df_g.columns[18]
            gc_dimos = EidikotitaDialog._fc(df_g, 'δήμ', 'δημ')
            gc_eidos = EidikotitaDialog._fc(df_g, 'είδος', 'ειδος')
            if gc_eidos:
                df_g = df_g[df_g[gc_eidos].fillna('').astype(str).str.strip() != 'Ιδιωτικά Σχολεία'].copy()
            df_g['_code'] = EidikotitaDialog._norm_code(df_g[gc_code])
            _gc = ['_code', gc_name, gc_phone, gc_email, gc_area]
            _gn = ['_code', '_school_name', '_phone', '_school_email', '_area']
            if gc_dimos: _gc.append(gc_dimos); _gn.append('_dimos')
            df_g_lu = df_g[_gc].drop_duplicates('_code').copy()
            df_g_lu.columns = _gn
            df_g_lu['_phone'] = df_g_lu['_phone'].fillna('').astype(str).str.replace(r'\.0$','',regex=True).str.strip()
            valid_codes = set(df_g_lu['_code'])

            _is_pe = 'Π.Ε' in selected_dir or not selected_dir
            if _is_pe:
                _keep = df_t['_code'].isin(valid_codes)
                if self._include_private_var.get():
                    _keep = _keep | df_t['_code'].str.startswith('7')
                df_t = df_t[_keep].copy()
            df_t = df_t[df_t[spec_col].astype(str).str.upper().str.startswith(specialty.upper())].copy()
            if _thesi_codes:
                if self._include_private_var.get():
                    df_t = df_t[df_t['_code'].isin(_thesi_codes) | df_t['_code'].str.startswith('7')].copy()
                else:
                    df_t = df_t[df_t['_code'].isin(_thesi_codes)].copy()
            if not self._include_private_var.get():
                df_t = df_t[~df_t['_code'].str.startswith('7')].copy()

            if df_t.empty:
                messagebox.showwarning('Αποτέλεσμα',
                    f'Δεν βρέθηκαν εκπαιδευτικοί για "{specialty}" / "{thesi_lbl}".', parent=self)
                return

            def _read_csv_enc(path):
                if not path: return pd.DataFrame()
                import zipfile as _zf, io as _io
                if path.endswith(('.xlsx', '.xls')):
                    try: return pd.read_excel(path, header=0, dtype=str)
                    except: return pd.DataFrame()
                if path.endswith('.zip'):
                    try:
                        with _zf.ZipFile(path) as z:
                            data = z.read([n for n in z.namelist() if n.endswith('.csv')][0])
                    except: return pd.DataFrame()
                    for enc in ['utf-8-sig','utf-8','iso-8859-7','cp1253']:
                        try: return pd.read_csv(_io.BytesIO(data), sep=None, engine='python', encoding=enc, header=0, dtype=str)
                        except: continue
                    return pd.DataFrame()
                for enc in ['utf-8-sig','utf-8','iso-8859-7','cp1253']:
                    try: return pd.read_csv(path, sep=None, engine='python', encoding=enc, header=0, dtype=str)
                    except: continue
                return pd.DataFrame()

            df_s16 = _read_csv_enc(self._stat_path)
            if not df_s16.empty:
                s16_afm = df_s16.columns[16]
                s16_abs = df_s16.columns[44] if len(df_s16.columns)>44 else df_s16.columns[45]
                s16_apo = df_s16.columns[46] if len(df_s16.columns)>46 else None
                s16_eos = df_s16.columns[47] if len(df_s16.columns)>47 else None
                df_s16['_afm'] = df_s16[s16_afm].apply(EidikotitaDialog._clean_afm).str.zfill(9)
                keep16 = ['_afm', s16_abs] + ([s16_apo] if s16_apo else []) + ([s16_eos] if s16_eos else [])
                df_s16_lu = df_s16[keep16].drop_duplicates('_afm').rename(columns={s16_abs:'_apoysia',**(({s16_apo:'_apo'} if s16_apo else {})),**(({s16_eos:'_eos'} if s16_eos else {}))}).copy()
                if '_apo' not in df_s16_lu: df_s16_lu['_apo']=''
                if '_eos' not in df_s16_lu: df_s16_lu['_eos']=''
            else:
                df_s16_lu = pd.DataFrame(columns=['_afm','_apoysia','_apo','_eos'])

            frames = [_read_csv_enc(p) for p in [self._stat41_path, self._stat42_path] if p]
            frames = [f for f in frames if not f.empty]
            if frames:
                df_41 = pd.concat(frames, ignore_index=True)
                a0,a9,a11,a12 = df_41.columns[0], (df_41.columns[9] if len(df_41.columns)>9 else None), (df_41.columns[11] if len(df_41.columns)>11 else None), (df_41.columns[12] if len(df_41.columns)>12 else None)
                df_41['_afm'] = df_41[a0].apply(EidikotitaDialog._clean_afm).str.zfill(9)
                keep=[c for c in ['_afm',a12,a11,a9] if c]
                df_41_lu = df_41[keep].drop_duplicates('_afm').rename(columns={k:v for k,v in {a12:'_email_psd',a11:'_email_personal',a9:'_kinito'}.items() if k}).copy()
                for c in ('_email_psd','_email_personal','_kinito'):
                    if c not in df_41_lu: df_41_lu[c]=''
            else:
                df_41_lu = pd.DataFrame(columns=['_afm','_email_psd','_email_personal','_kinito'])

            df_t = df_t.merge(df_g_lu, on='_code', how='left')
            df_t = df_t.merge(df_s16_lu, on='_afm', how='left')
            df_t = df_t.merge(df_41_lu,  on='_afm', how='left')
            df_t['_absent'] = df_t[status_col].fillna('').astype(str).str.strip()=='ΑΠΟΥΣΙΑ'

            def gcol(c):
                return df_t[c].fillna('').astype(str) if c is not None and c in df_t.columns else pd.Series(['']*len(df_t), index=df_t.index)

            out = pd.DataFrame(index=df_t.index)
            out['ΑΜ']                = df_t['_am'].fillna('')
            out['Επώνυμο']            = gcol(eponym_col)
            out['Όνομα']              = gcol(onoma_col)
            out['Κύρια Ειδικ.']      = gcol(spec_col)
            out['Email στο ΠΣΔ']     = df_t['_email_psd'].fillna('')
            out['Email']             = df_t['_email_personal'].fillna('') if '_email_personal' in df_t.columns else ''
            out['Κινητό']            = df_t['_kinito'].fillna('')
            out['Σχέση εργασίας']    = gcol(org_col)
            out['Σχέση τοποθέτησης'] = gcol(topoth_col)
            out['Κατάσταση']         = gcol(status_col)
            out['Φορέας τοποθέτησης']= gcol(school_name_col)
            out['Δήμος']             = df_t['_dimos'].fillna('') if '_dimos' in df_t.columns else ''
            out['Τηλέφωνο']          = df_t['_phone'].fillna('')
            out['e-mail']            = df_t['_school_email'].fillna('')
            out['ΑΠΟΥΣΙΑ']           = df_t['_apoysia'].fillna('')
            out['Από']               = df_t['_apo'].fillna('') if '_apo' in df_t.columns else ''
            out['Έως']               = df_t['_eos'].fillna('') if '_eos' in df_t.columns else ''
            out['_absent']           = df_t['_absent']
            out.loc[~out['_absent'], ['ΑΠΟΥΣΙΑ','Από','Έως']] = ''
            out = out.sort_values('Επώνυμο', na_position='last').reset_index(drop=True)

            today_str = datetime.today().strftime('%Y%m%d')
            out_dir   = os.path.join(_docs_base(), f'results_{today_str}')
            os.makedirs(out_dir, exist_ok=True)
            def _sf(s): return s.replace('/','_').replace('\\','_').replace(' ','_')
            out_path = os.path.join(out_dir, f'Εκπαιδευτικοί_{_sf(specialty)}_{_sf(thesi_lbl[:30])}_{today_str}.xlsx')

            _disabled   = {c for c,v in getattr(self,'_col_vars',{}).items() if not v.get()}
            active_cols = [c for c in self._OUT_COLS if c not in _disabled]

            wb = Workbook(); ws = wb.active
            ws.title = f'{specialty} {thesi_lbl}'[:31]
            RED = 'FF0000'
            thin   = Side(style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ci, col in enumerate(active_cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                cell.fill      = PatternFill('solid', start_color='1F4E79')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border    = border
            alt_fill = PatternFill('solid', start_color='FFF0F0')
            for ri, row in out.iterrows():
                is_abs = bool(row.get('_absent', False))
                for ci, col in enumerate(active_cols, 1):
                    val = row.get(col, '')
                    if pd.isna(val): val = ''
                    cell = ws.cell(row=ri+2, column=ci, value=str(val) if val!='' else '')
                    cell.font      = Font(name='Arial', size=9, color=RED if is_abs else '000000', bold=is_abs)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border    = border
                    if not is_abs and ri%2==1: cell.fill = alt_fill
            for ci, col in enumerate(active_cols, 1):
                vals = [str(out.iloc[r][col]) for r in range(min(len(out),50)) if col in out.columns and not pd.isna(out.iloc[r][col])]
                ws.column_dimensions[get_column_letter(ci)].width = min(max([len(col)]+[len(v) for v in vals] if vals else [len(col)])+3, 42)
            ws.row_dimensions[1].height = 30
            ws.freeze_panes = 'A2'
            wb.save(out_path)

        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror('Σφάλμα', str(e), parent=self); return

        s = _load_local_settings()
        from datetime import datetime as _dts
        td = _dts.today().strftime('%d/%m/%Y')
        bts = body_text.replace(specialty,'{specialty}').replace(td,'{date}').replace(thesi_lbl,'{thesi}')
        sts = subject.replace(specialty,'{specialty}').replace(td,'{date}').replace(thesi_lbl,'{thesi}')
        s[self._SETTINGS_KEY] = {'subject':sts,'body':bts,'advisor_email':'\n'.join(to_emails),'direction':self._dir_var.get().strip(),'include_private':self._include_private_var.get()}
        os.makedirs(os.path.dirname(_get_local_settings_path()), exist_ok=True)
        with open(_get_local_settings_path(),'w',encoding='utf-8') as f:
            json.dump(s, f, ensure_ascii=False, indent=2)

        absent_count = int(out['_absent'].sum())
        total_count  = len(out)

        if not send:
            from core.framework import _show_results_popup
            _show_results_popup('Εκπ/κοί ανά Ειδικότητα & Θέση',
                f'Αρχείο αποθηκεύτηκε.\n\nΕιδικότητα: {specialty}  |  Θέση: {thesi_lbl}\n'
                f'Σύνολο: {total_count} εκπ/κοί  |  Απόντες: {absent_count}',
                result_type='warn', excel_path=out_path)
            self.destroy(); return

        try:
            from core.framework import send_email
            send_email(config, to_emails, subject, full_body, out_path)
            from core.framework import _show_results_popup
            _show_results_popup('Εκπ/κοί ανά Ειδικότητα & Θέση',
                f'Email στάλθηκε: {", ".join(to_emails)}\n\nΕιδικότητα: {specialty}  |  Θέση: {thesi_lbl}\n'
                f'Σύνολο: {total_count} εκπ/κοί  |  Απόντες: {absent_count}',
                result_type='ok', excel_path=out_path)
            self.destroy()
        except Exception as e:
            messagebox.showerror('Σφάλμα αποστολής', str(e), parent=self)

    def _clear(self):
        for w in self._exec_body.winfo_children():
            w.destroy()

    def _on_download_done(self):
        """Καλείται μετά από επιτυχή λήψη στο tab «⬇ Λήψη» — ξανα-ανιχνεύει
        τα αρχεία και ξαναχτίζει το tab «▶ Εκτέλεση» με τα φρέσκα αρχεία."""
        self._topoth_path = EidikotitaDialog._auto_find('Topothetiseis')
        self._grid_path   = EidikotitaDialog._auto_find('gridResults')
        self._stat_path   = EidikotitaDialog._auto_find('stat4_16')
        self._stat41_path = EidikotitaDialog._auto_find('stat4_1')
        self._stat42_path = EidikotitaDialog._auto_find('stat4_2')
        self._build_form()


# Απαιτούμενα στατιστικά για το εργαλείο «Στοιχεία Σχολικών Μονάδων».
_MONADA_RIDS = ['3.1', '2.2']
_MONADA_REQUIRED = [
    '3.1 — Κατανομή μαθητών ανά τάξη',
    '2.2 — Εκτεταμένα Στοιχεία Σχολικών Μονάδων',
]


class MonadaDialog(tk.Toplevel):
    """Εργαλείο εξαγωγής στοιχείων σχολικών μονάδων ανά Δήμο."""

    _SETTINGS_KEY   = 'monada_tool'
    _DEFAULT_BODY   = (
        'Αποτύπωση Myschool {date}.\n\n'
        'Καλημέρα σας,\n\n'
        'Επισυνάπτεται πίνακας excel με τα στοιχεία των σχολικών μονάδων '
        'Δήμου {dimos} σύμφωνα με τα καταχωρημένα στοιχεία στο myschool.\n\n\n'
        'Στη διάθεσή σας για οποιαδήποτε πληροφορία'
    )
    _DEFAULT_SUBJECT = 'Στοιχεία σχολικών μονάδων Δήμου {dimos}'

    # Ταξινόμηση τάξεων (Νηπιαγωγείο + Δημοτικό)
    _CLASS_ORDER = ['ΠΡΟΝΗΠΙΑ', 'ΝΗΠΙΑ', 'ΠΡΟΝΗΠΙΑ-ΝΗΠΙΑ', 'Α', 'Β', 'Γ', 'Δ', 'Ε', 'ΣΤ']

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Στοιχεία Σχολικών Μονάδων')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self._parent = parent

        s = _load_local_settings().get(self._SETTINGS_KEY, {})
        self._saved_subject = s.get('subject',     self._DEFAULT_SUBJECT)
        self._saved_body    = s.get('body',         self._DEFAULT_BODY)
        self._saved_email   = s.get('dimos_email',  '')

        # Αυτόματη εύρεση αρχείων zip
        # CSV_* = χειροκίνητο download, stat2_2* = μέσω app downloader (2.2), gridResults* = 2.1 fallback
        self._csv_path    = self._auto_find_zip('CSV_', 'stat2_2', 'gridResults')
        self._stat31_path = self._auto_find_zip('stat3_1')

        # 2-tab δομή όπως οι έλεγχοι: «⬇ Λήψη» + «▶ Εκτέλεση» (το ίδιο
        # εργαλείο που υπήρχε πριν, χωρίς αλλαγή στη λειτουργία του).
        from tkinter import ttk as _ttk
        nb = _ttk.Notebook(self)
        nb.pack(fill='both', expand=True)
        tab_dl   = tk.Frame(nb, bg=C['bg'])
        tab_exec = tk.Frame(nb, bg=C['bg'])
        nb.add(tab_dl,   text='  ⬇ Λήψη  ')
        nb.add(tab_exec, text='  ▶ Εκτέλεση  ')
        self._exec_body = tab_exec

        _build_report_download_tab(self, tab_dl, _MONADA_RIDS, _MONADA_REQUIRED,
                                    on_done=self._on_download_done)

        self._build_form()
        self.update_idletasks()
        w, h = 630, 620
        x = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f'{w}x{h}+{x}+{y}')

    # ── Auto-find ────────────────────────────────────────────────────────────

    @staticmethod
    def _auto_find_zip(*prefixes):
        """Ψάχνει αρχείο (zip/csv/xlsx) με δοθέν prefix — downloads app πρώτα, μετά ~/Downloads.
        Δέχεται πολλαπλά prefixes (πρώτο εύρημα κερδίζει).
        """
        import glob as _glob
        folders = []
        dl_base = os.path.join(_docs_base(), 'downloads')
        if os.path.isdir(dl_base):
            folders += sorted([
                os.path.join(dl_base, d)
                for d in os.listdir(dl_base)
                if os.path.isdir(os.path.join(dl_base, d))
            ], reverse=True)
        folders.append(os.path.join(os.path.expanduser('~'), 'Downloads'))
        for folder in folders:
            for prefix in prefixes:
                # .zip πρώτα (χειροκίνητο κατέβασμα), μετά .csv/.xlsx/.xls (μέσω app downloader)
                for pattern in (f'{prefix}*.zip', f'{prefix}*.csv', f'{prefix}*.xlsx', f'{prefix}*.xls'):
                    matches = [f for f in _glob.glob(os.path.join(folder, pattern))
                               if not f.endswith('.tmp') and not f.endswith('.crdownload')]
                    if matches:
                        return sorted(matches)[-1]
        return ''

    # ── Βοηθητικά ────────────────────────────────────────────────────────────

    @staticmethod
    def _clean_code(val):
        """Αφαιρεί =\"XXXXX\" format και επιστρέφει τα ψηφία (lstrip 0)."""
        import re
        s = str(val).strip().strip('"').lstrip('=').strip('"').strip()
        s = re.sub(r'\.0$', '', s)
        return s.lstrip('0') or s  # lstrip('0') αλλά όχι αν το αποτέλεσμα είναι κενό

    @staticmethod
    def _s(val):
        """Επιστρέφει string από τιμή — NaN/nan → κενό, .0 stripped."""
        import re
        if val is None: return ''
        s = str(val).strip()
        if s.lower() in ('nan', 'none', ''): return ''
        return re.sub(r'\.0$', '', s)

    @staticmethod
    def _read_zip_csv(path, encoding='cp1253', strip_trailing_sep=False):
        """Διαβάζει CSV/XLSX — από zip, plain CSV, ή xlsx.
        strip_trailing_sep: True για αρχεία με trailing ';' (π.χ. stat3_1).
        """
        import zipfile, io, pandas as pd
        lower = path.lower()
        if lower.endswith('.xlsx') or lower.endswith('.xls'):
            return pd.read_excel(path, dtype=str)
        if lower.endswith('.zip'):
            with zipfile.ZipFile(path) as z:
                raw = z.read(z.namelist()[0])
        else:
            # Plain CSV (κατεβασμένο μέσω app downloader)
            with open(path, 'rb') as f:
                raw = f.read()
        text = raw.decode(encoding)
        if strip_trailing_sep:
            text = '\n'.join(l.rstrip(';') for l in text.splitlines())
        return pd.read_csv(io.StringIO(text), sep=';', dtype=str)

    # ── Κύρια φόρμα ──────────────────────────────────────────────────────────

    def _build_form(self):
        self._clear()

        hdr = tk.Frame(self._exec_body, bg='#0F6E56', pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='🏫  Στοιχεία Σχολικών Μονάδων',
                 bg='#0F6E56', fg='white',
                 font=('Arial', 12, 'bold')).pack()
        tk.Label(hdr, text='για αποστολή στοιχείων ενδεικτικά σε Δήμους  (Απαιτούνται: 3.1, 2.2)',
                 bg='#0F6E56', fg='#A8D8C8',
                 font=('Arial', 8, 'italic')).pack()

        # Προειδοποίηση αν λείπουν αρχεία
        missing = []
        if not self._csv_path:    missing.append('Κατάλογος Μονάδων (CSV_...zip)')
        if not self._stat31_path: missing.append('Στατιστικό 3.1 (stat3_1...)')
        if missing:
            tk.Label(self._exec_body,
                text=f'⚠  Δεν βρέθηκαν: {", ".join(missing)}. Κατέβασέ τα από MySchool.',
                bg='#FFF3E0', fg='#E65100', font=('Arial', 8),
                anchor='w', padx=10, pady=5, wraplength=570, justify='left',
            ).pack(fill='x', padx=18, pady=(0, 6))

        # ── Δήμος ─────────────────────────────────────────────────────────────
        tk.Label(self._exec_body, text='Δήμος:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))

        dimos_row = tk.Frame(self._exec_body, bg=C['bg'])
        dimos_row.pack(fill='x', padx=18, pady=(2, 6))
        self._dimos_var = tk.StringVar()
        from tkinter import ttk as _ttk
        self._dimos_combo = _ttk.Combobox(dimos_row, textvariable=self._dimos_var,
                                           width=40, state='readonly')
        self._dimos_combo.pack(side='left')
        self._dimos_lbl = tk.Label(dimos_row, text='Φόρτωση…',
                                    bg=C['bg'], fg=C['desc'], font=('Arial', 8))
        self._dimos_lbl.pack(side='left', padx=(10, 0))
        self._dimos_var.trace_add('write', self._on_dimos_change)

        # ── Εμφάνιση ──────────────────────────────────────────────────────────
        tk.Label(self._exec_body, text='Εμφάνιση:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))
        mode_row = tk.Frame(self._exec_body, bg=C['bg'])
        mode_row.pack(fill='x', padx=18, pady=(2, 6))
        self._mode_var = tk.StringVar(value='monada')
        tk.Radiobutton(mode_row, text='Ανά Σχολική Μονάδα', variable=self._mode_var,
                       value='monada', bg=C['bg'], font=('Arial', 9),
                       activebackground=C['bg']).pack(side='left', padx=(0, 18))
        tk.Radiobutton(mode_row, text='Ανά Τάξη', variable=self._mode_var,
                       value='taxh', bg=C['bg'], font=('Arial', 9),
                       activebackground=C['bg']).pack(side='left')

        # ── Προαιρετικές στήλες ───────────────────────────────────────────────
        tk.Label(self._exec_body, text='Προαιρετικές στήλες:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))
        col_frame = tk.Frame(self._exec_body, bg=C['bg'])
        col_frame.pack(fill='x', padx=18, pady=(2, 6))
        self._col_vars = {}
        for col_name in ('ΑΦΜ Διευθυντή',):
            var = tk.BooleanVar(value=False)
            self._col_vars[col_name] = var
            tk.Checkbutton(col_frame, text=col_name, variable=var,
                           bg=C['bg'], font=('Arial', 9),
                           activebackground=C['bg']).pack(side='left', padx=(0, 12))

        # ── Email ─────────────────────────────────────────────────────────────
        pad = dict(padx=18, pady=2)
        self._to_txt = _build_recipients_box(
            self._exec_body, 'Προς (email δήμου):', self._saved_email)

        tk.Label(self._exec_body, text='Θέμα:',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._subj_var = tk.StringVar(value=self._saved_subject)
        tk.Entry(self._exec_body, textvariable=self._subj_var,
                 font=('Arial', 9)).pack(fill='x', padx=18, pady=(0, 6))

        tk.Label(self._exec_body, text='Κείμενο email:',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._body_txt = tk.Text(self._exec_body, font=('Arial', 9), height=5,
                                  wrap='word', relief='solid', bd=1)
        self._body_txt.pack(fill='x', padx=18, pady=(0, 6))
        from datetime import datetime as _dt
        _DAYS_GR = ['Δευτέρα','Τρίτη','Τετάρτη','Πέμπτη','Παρασκευή','Σάββατο','Κυριακή']
        self._body_txt.insert('1.0',
            self._saved_body.replace('{date}', _dt.today().strftime('%d/%m/%Y'))
                             .replace('{day}',  _DAYS_GR[_dt.today().weekday()])
                             .replace('{dimos}', self._dimos_var.get()))

        btn_row = tk.Frame(self._exec_body, bg=C['bg'])
        btn_row.pack(side='bottom', pady=10)
        tk.Button(btn_row, text='Μόνο Excel (χωρίς email)',
                  bg=C['bg2'], fg=C['hdr_bg'], relief='flat',
                  font=('Arial', 9), padx=10, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=False)).pack(side='left', padx=4)
        tk.Button(btn_row, text='▶  Δημιουργία & Αποστολή',
                  bg=C['btn_bg'], fg=C['btn_fg'], relief='flat',
                  font=('Arial', 9, 'bold'), padx=14, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=True)).pack(side='left', padx=4)

        self.after(100, self._load_dimos)

    def _on_dimos_change(self, *_):
        dimos = self._dimos_var.get()
        self._subj_var.set(self._saved_subject.replace('{dimos}', dimos))
        # Ενημέρωση body: αντικατάσταση παλιού δήμου με νέο
        from datetime import datetime as _dt
        today = _dt.today().strftime('%d/%m/%Y')
        _DAYS_GR = ['Δευτέρα','Τρίτη','Τετάρτη','Πέμπτη','Παρασκευή','Σάββατο','Κυριακή']
        self._body_txt.delete('1.0', 'end')
        self._body_txt.insert('1.0',
            self._saved_body.replace('{date}', today)
                             .replace('{day}',  _DAYS_GR[_dt.today().weekday()])
                             .replace('{dimos}', dimos))

    def _load_dimos(self):
        """Φορτώνει τους Δήμους — πρώτα stat3_1 col7, fallback CSV col5."""
        src_path = self._stat31_path or self._csv_path
        if not src_path:
            self._dimos_lbl.config(text='Δεν βρέθηκε αρχείο.', fg='#CC0000')
            return
        try:
            import pandas as pd
            use_31 = bool(self._stat31_path)
            df = self._read_zip_csv(src_path, strip_trailing_sep=use_31)
            col_idx   = 7 if use_31 else 5   # stat3_1 col7=Δήμος | CSV col5=Δήμος
            dimos_col = df.columns[col_idx]
            dimos_list = sorted(df[dimos_col].dropna().astype(str).str.strip().unique())
            self._dimos_combo.config(values=dimos_list)
            if dimos_list:
                self._dimos_var.set(dimos_list[0])
            self._dimos_lbl.config(
                text=f'{len(dimos_list)} δήμοι | {os.path.basename(src_path)}',
                fg=C['desc'])
        except Exception as e:
            self._dimos_lbl.config(text=f'Σφάλμα: {e}', fg='#CC0000')

    # ── Εκτέλεση ─────────────────────────────────────────────────────────────

    def _execute(self, send=True):
        import json, pandas as pd
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        dimos     = self._dimos_var.get().strip()
        mode      = self._mode_var.get()
        to_emails = _parse_recipients(self._to_txt)
        subject   = self._subj_var.get().strip()
        body_text = self._body_txt.get('1.0', 'end-1c')
        full_body = body_text + '\n\n' + config.email_signature()

        if send and not to_emails:
            messagebox.showwarning('Email', 'Εισάγετε email παραλήπτη.', parent=self)
            return
        if not dimos:
            messagebox.showwarning('Δήμος', 'Επίλεξε Δήμο.', parent=self)
            return
        if not self._stat31_path or not self._csv_path:
            from core.framework import _missing_file_dialog
            _missing_file_dialog('Σχολικές Μονάδες ανά Δήμο', [
                '3.1 — Κατανομή μαθητών ανά τάξη',
                '2.2 — Εκτεταμένα στοιχεία σχολικών μονάδων (CSV)',
            ])
            return

        try:
            # ── 1. CSV → lookup dict (δευτερεύουσα πηγή, στοιχεία επικοινωνίας) ──
            # Επιβεβαιωμένα offsets λόγω 1-column shift στα headers από col11:
            #   col10 = Είδος (τύπος σχολείου)
            #   col11 = Κωδ. ΥΠΠΘ (αριθμητικός κωδικός)
            #   col12 = Ονομασία
            #   col16 = Τηλέφωνο
            #   col18 = e-mail σχολείου
            #   col20 = Ταχ. Διεύθυνση
            #   col48 = Αναστολή (NAI/OXI)
            #   col55 = Ονομ/μο Διευθυντή
            #   col58 = Κινητό Διευθυντή
            #   col59 = Email Διευθυντή
            #   col60 = Email ΠΣΔ Διευθυντή
            csv_df = self._read_zip_csv(self._csv_path)

            c_eidos    = csv_df.columns[10]
            c_code_csv = csv_df.columns[11]
            c_onoma    = csv_df.columns[12]
            c_phone    = csv_df.columns[16]
            c_email    = csv_df.columns[18]
            c_address  = csv_df.columns[20]
            c_anast    = csv_df.columns[48]   # Αναστολή (NAI = κλειστό)
            c_dir_name = csv_df.columns[55]
            c_dir_mob  = csv_df.columns[58]
            c_dir_mail = csv_df.columns[59]
            c_dir_psd  = csv_df.columns[60]
            # ΑΦΜ Διευθυντή — col54
            # Λόγω 1-column shift στα headers: το header "Α.Φ.Μ. Διευθυντή" βρίσκεται
            # στο col55, αλλά τα πραγματικά ΑΦΜ δεδομένα (="XXXXXXXXX") είναι στο col54.
            # Το col55 περιέχει το Ονοματεπώνυμο (χρησιμοποιείται ήδη ως c_dir_name).
            c_dir_afm = csv_df.columns[54] if len(csv_df.columns) > 54 else None

            # Φίλτρο τύπου: εξαίρεση Ιδιωτικών / Ξένων
            eidos_ser = csv_df[c_eidos].fillna('').astype(str)
            mask_type = (
                ~eidos_ser.str.contains('Ιδιωτικό', na=False) &
                ~eidos_ser.str.contains('Ξένο',     na=False)
            )
            csv_df = csv_df[mask_type].copy()

            # Φίλτρο Αναστολής: εξαίρεση σχολείων με Αναστολή = NAI
            csv_df = csv_df[
                csv_df[c_anast].fillna('').astype(str).str.strip().str.upper() != 'NAI'
            ].copy()

            # Κατασκευή lookup dict: {clean_code → στοιχεία}
            csv_df['_code'] = csv_df[c_code_csv].apply(self._clean_code)
            csv_lookup = {}
            for _, row in csv_df.iterrows():
                code = row['_code']
                if not code:
                    continue
                eidos_val   = self._s(row[c_eidos])
                is_dim      = 'Δημοτικό' in eidos_val
                eidos_short = (eidos_val
                               .replace('Ενιαίου Τύπου Ολοήμερο ', '')
                               .replace('Ολοήμερο ', '')
                               .strip())
                csv_lookup[code] = {
                    'eidos':    eidos_short,
                    'is_dim':   is_dim,
                    'onoma':    self._s(row[c_onoma]),
                    'phone':    self._s(row[c_phone]),
                    'email':    self._s(row[c_email]),
                    'address':  self._s(row[c_address]),
                    'dir_name': self._s(row[c_dir_name]),
                    'dir_afm':  self._s(row[c_dir_afm]) if c_dir_afm is not None else '',
                    'dir_mob':  self._s(row[c_dir_mob]),
                    'dir_mail': self._s(row[c_dir_mail]),
                    'dir_psd':  self._s(row[c_dir_psd]),
                }

            # ── 2. stat3_1 — ΚΥΡΙΑ πηγή (κατανομή ανά τάξη & φύλο) ──────────
            df31 = self._read_zip_csv(self._stat31_path, strip_trailing_sep=True)

            kwd31  = df31.columns[4]   # Κωδικός σχολείου
            dim31  = df31.columns[7]   # Δήμος
            taxh31 = df31.columns[10]  # Τάξη
            tmhm31 = df31.columns[11]  # Αριθμός Τμημάτων
            ag31   = df31.columns[12]  # Αγόρια
            ko31   = df31.columns[13]  # Κορίτσια
            sy31   = df31.columns[14]  # Σύνολο

            df31['_code'] = df31[kwd31].apply(self._clean_code)

            # Φίλτρο Δήμου
            df31 = df31[df31[dim31].fillna('').astype(str).str.strip() == dimos].copy()

            # Κράτα μόνο σχολεία που υπάρχουν στο CSV lookup
            # (αποκλείονται αυτόματα Ιδιωτικά, Ξένα, Αναστολή)
            df31 = df31[df31['_code'].isin(csv_lookup)].copy()

            if df31.empty:
                messagebox.showwarning('Αποτέλεσμα',
                    f'Δεν βρέθηκαν σχολεία στον Δήμο "{dimos}".', parent=self)
                return

            # Αριθμητικές στήλες
            for col in [tmhm31, ag31, ko31, sy31]:
                df31[col] = pd.to_numeric(df31[col], errors='coerce').fillna(0).astype(int)

            # Sorted list κωδικών: Νηπιαγωγεία πρώτα, μετά Δημοτικά, αλφαβητικά
            unique_codes = df31['_code'].unique()
            sorted_codes = sorted(unique_codes,
                key=lambda c: (1 if csv_lookup.get(c, {}).get('is_dim') else 0,
                               csv_lookup.get(c, {}).get('onoma', '')))

            # ── 3. Δημιουργία Excel ───────────────────────────────────────────
            today_str  = datetime.today().strftime('%Y%m%d')
            out_dir    = os.path.join(_docs_base(), f'results_{today_str}')
            os.makedirs(out_dir, exist_ok=True)
            dimos_safe = dimos.replace('/', '_').replace('\\', '_')
            mode_sfx   = 'ανά_τάξη' if mode == 'taxh' else 'ανά_μονάδα'
            out_path   = os.path.join(
                out_dir, f'Σχολικές_Μονάδες_{dimos_safe}_{mode_sfx}_{today_str}.xlsx')

            wb  = Workbook()
            ws  = wb.active
            ws.title = dimos[:31]

            RED        = 'FF0000'
            LIGHT_BLUE = 'DCE6F1'
            LIGHT_RED  = 'FCE4EC'
            thin   = Side(style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ctr_al = Alignment(horizontal='center', vertical='center')
            lft_al = Alignment(horizontal='left',   vertical='center')

            def _hdr_cell(ws, row, col, value):
                c = ws.cell(row=row, column=col, value=value)
                c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                c.fill      = PatternFill('solid', start_color=RED)
                c.alignment = hdr_al
                c.border    = border

            CLASS_RANK = {c: i for i, c in enumerate(self._CLASS_ORDER)}

            _show_afm = self._col_vars.get('ΑΦΜ Διευθυντή', tk.BooleanVar(value=False)).get()

            if mode == 'taxh':
                # ─── Ανά Τάξη ─────────────────────────────────────────────────
                all_cols = [
                    'Είδος', 'Ονομασία', 'Τάξη', 'Τμήματα',
                    'Αγόρια', 'Κορίτσια', 'Σύνολο',
                    'Τηλέφωνο', 'e-mail σχολείου',
                    'Ονομ/μο Διευθυντή',
                ]
                if _show_afm:
                    all_cols.append('ΑΦΜ Διευθυντή')
                all_cols += ['Κινητό Διευθυντή', 'Email Διευθυντή', 'Email ΠΣΔ Διευθυντή']

                for ci, col in enumerate(all_cols, 1):
                    _hdr_cell(ws, 1, ci, col)

                subtot_fill = PatternFill('solid', start_color=LIGHT_BLUE)
                grand_fill  = PatternFill('solid', start_color=LIGHT_RED)
                alt_fill    = PatternFill('solid', start_color='F7F7F7')
                er = 2
                tot_ag_g = tot_ko_g = tot_sy_g = 0

                for code in sorted_codes:
                    info   = csv_lookup.get(code, {})
                    is_dim = info.get('is_dim', False)
                    sc_df  = df31[df31['_code'] == code].copy()
                    if sc_df.empty:
                        continue

                    sc_df['_rank'] = sc_df[taxh31].apply(
                        lambda t: CLASS_RANK.get(str(t).strip(), 99))
                    sc_df = sc_df.sort_values('_rank').reset_index(drop=True)

                    onoma = info.get('onoma', '')
                    sc_ag = sc_ko = sc_sy = sc_tm = 0
                    for row_i, (_, crow) in enumerate(sc_df.iterrows()):
                        taxh = self._s(crow[taxh31])
                        tm   = int(crow[tmhm31] or 0)
                        ag   = int(crow[ag31]   or 0)
                        ko   = int(crow[ko31]   or 0)
                        sy   = int(crow[sy31]   or 0)
                        sc_ag += ag; sc_ko += ko; sc_sy += sy; sc_tm += tm

                        # Στοιχεία επικοινωνίας μόνο στην 1η γραμμή
                        first = row_i == 0
                        vals = [
                            info.get('eidos',    '') if first else '',
                            onoma                    if first else '',
                            taxh, tm, ag, ko, sy,
                            info.get('phone',    '') if first else '',
                            info.get('email',    '') if first else '',
                            info.get('dir_name', '') if first else '',
                        ]
                        if _show_afm:
                            vals.append(info.get('dir_afm', '') if first else '')
                        vals += [
                            info.get('dir_mob',  '') if first else '',
                            info.get('dir_mail', '') if first else '',
                            info.get('dir_psd',  '') if first else '',
                        ]

                        row_fill = alt_fill if row_i % 2 == 1 else None
                        for ci, val in enumerate(vals, 1):
                            cell = ws.cell(row=er, column=ci, value=val)
                            cell.font      = Font(name='Arial', size=9)
                            cell.alignment = ctr_al if 4 <= ci <= 7 else lft_al
                            cell.border    = border
                            if row_fill: cell.fill = row_fill
                        er += 1

                    # Subtotal μόνο για Δημοτικά (τα Νηπιαγωγεία έχουν 1-2 γραμμές, δεν χρειάζεται)
                    tot_ag_g += sc_ag; tot_ko_g += sc_ko; tot_sy_g += sc_sy
                    if is_dim:
                        for ci in range(1, len(all_cols) + 1):
                            cell = ws.cell(row=er, column=ci)
                            cell.font      = Font(name='Arial', size=9, bold=True)
                            cell.fill      = subtot_fill
                            cell.border    = border
                            cell.alignment = ctr_al if 4 <= ci <= 7 else lft_al
                        ws.cell(row=er, column=2, value=f'Σύνολο {onoma}')
                        ws.cell(row=er, column=4, value=sc_tm)
                        ws.cell(row=er, column=5, value=sc_ag)
                        ws.cell(row=er, column=6, value=sc_ko)
                        ws.cell(row=er, column=7, value=sc_sy)
                        er += 1

                # Grand total
                for ci in range(1, len(all_cols) + 1):
                    cell = ws.cell(row=er, column=ci)
                    cell.font      = Font(name='Arial', size=9, bold=True)
                    cell.fill      = grand_fill
                    cell.border    = border
                    cell.alignment = ctr_al if 4 <= ci <= 7 else lft_al
                ws.cell(row=er, column=2, value='ΓΕΝΙΚΟ ΣΥΝΟΛΟ')
                ws.cell(row=er, column=5, value=tot_ag_g)
                ws.cell(row=er, column=6, value=tot_ko_g)
                ws.cell(row=er, column=7, value=tot_sy_g)

                _cw = {
                    'Είδος': 26, 'Ονομασία': 40, 'Τάξη': 20,
                    'Τμήματα': 10, 'Αγόρια': 10, 'Κορίτσια': 12, 'Σύνολο': 10,
                    'Τηλέφωνο': 16, 'e-mail σχολείου': 30,
                    'Ονομ/μο Διευθυντή': 28, 'ΑΦΜ Διευθυντή': 14,
                    'Κινητό Διευθυντή': 18, 'Email Διευθυντή': 32, 'Email ΠΣΔ Διευθυντή': 26,
                }
                for ci, col in enumerate(all_cols, 1):
                    ws.column_dimensions[get_column_letter(ci)].width = _cw.get(col, 15)

            else:
                # ─── Ανά Σχολική Μονάδα ──────────────────────────────────────
                # Ομαδοποίηση stat3_1 ανά σχολείο + επικοινωνία από CSV
                all_cols = [
                    'Είδος', 'Ονομασία',
                    'Τμήματα', 'Αγόρια', 'Κορίτσια', 'Σύνολο',
                    'Τηλέφωνο', 'e-mail σχολείου', 'Ταχ. Διεύθυνση',
                    'Ονομ/μο Διευθυντή',
                ]
                if _show_afm:
                    all_cols.append('ΑΦΜ Διευθυντή')
                all_cols += ['Κινητό Διευθυντή', 'Email Διευθυντή', 'Email ΠΣΔ Διευθυντή']

                for ci, col in enumerate(all_cols, 1):
                    _hdr_cell(ws, 1, ci, col)

                alt_fill   = PatternFill('solid', start_color='FFF0F0')
                tot_ag = tot_ko = tot_sy = tot_tm = 0

                for ri, code in enumerate(sorted_codes):
                    info  = csv_lookup.get(code, {})
                    sc_df = df31[df31['_code'] == code]
                    tm = int(sc_df[tmhm31].sum())
                    ag = int(sc_df[ag31].sum())
                    ko = int(sc_df[ko31].sum())
                    sy = int(sc_df[sy31].sum())
                    tot_tm += tm; tot_ag += ag; tot_ko += ko; tot_sy += sy

                    vals = [
                        info.get('eidos',    ''),
                        info.get('onoma',    ''),
                        tm, ag, ko, sy,
                        info.get('phone',    ''),
                        info.get('email',    ''),
                        info.get('address',  ''),
                        info.get('dir_name', ''),
                    ]
                    if _show_afm:
                        vals.append(info.get('dir_afm', ''))
                    vals += [
                        info.get('dir_mob',  ''),
                        info.get('dir_mail', ''),
                        info.get('dir_psd',  ''),
                    ]

                    fill = alt_fill if ri % 2 == 1 else None
                    er   = ri + 2
                    for ci, val in enumerate(vals, 1):
                        cell = ws.cell(row=er, column=ci, value=val)
                        cell.font      = Font(name='Arial', size=9)
                        cell.alignment = ctr_al if 3 <= ci <= 6 else lft_al
                        cell.border    = border
                        if fill: cell.fill = fill

                # Σειρά ΣΥΝΟΛΟ
                tot_row = len(sorted_codes) + 2
                for ci in range(1, len(all_cols) + 1):
                    cell = ws.cell(row=tot_row, column=ci)
                    cell.font      = Font(name='Arial', size=9, bold=True)
                    cell.fill      = PatternFill('solid', start_color=LIGHT_BLUE)
                    cell.border    = border
                    cell.alignment = ctr_al if 3 <= ci <= 6 else lft_al
                ws.cell(row=tot_row, column=2, value='ΣΥΝΟΛΟ')
                ws.cell(row=tot_row, column=3, value=tot_tm)
                ws.cell(row=tot_row, column=4, value=tot_ag)
                ws.cell(row=tot_row, column=5, value=tot_ko)
                ws.cell(row=tot_row, column=6, value=tot_sy)

                _cw = {
                    'Είδος': 26, 'Ονομασία': 40,
                    'Τμήματα': 10, 'Αγόρια': 10, 'Κορίτσια': 12, 'Σύνολο': 10,
                    'Τηλέφωνο': 16, 'e-mail σχολείου': 30, 'Ταχ. Διεύθυνση': 30,
                    'Ονομ/μο Διευθυντή': 28, 'ΑΦΜ Διευθυντή': 14,
                    'Κινητό Διευθυντή': 18, 'Email Διευθυντή': 32, 'Email ΠΣΔ Διευθυντή': 26,
                }
                for ci, col in enumerate(all_cols, 1):
                    ws.column_dimensions[get_column_letter(ci)].width = _cw.get(col, 15)

            ws.row_dimensions[1].height = 28
            ws.freeze_panes = 'A2'
            wb.save(out_path)
            school_count = len(sorted_codes)

        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror('Σφάλμα', str(e), parent=self)
            return

        # ── Αποθήκευση ρυθμίσεων ─────────────────────────────────────────────
        s = _load_local_settings()
        s[self._SETTINGS_KEY] = {
            'subject':     self._saved_subject,
            'body':        self._saved_body,
            'dimos_email': '\n'.join(to_emails),
        }
        path_s = _get_local_settings_path()
        os.makedirs(os.path.dirname(path_s), exist_ok=True)
        with open(path_s, 'w', encoding='utf-8') as f:
            json.dump(s, f, ensure_ascii=False, indent=2)
        if not send:
            from core.framework import _show_results_popup
            _show_results_popup('Σχολικές Μονάδες',
                f'Αρχείο αποθηκεύτηκε.\n\nΣχολεία: {school_count}',
                result_type='warn', excel_path=out_path)
            self.destroy()
            return

        try:
            from core.framework import send_email
            send_email(config, to_emails, subject, full_body, out_path)
            from core.framework import _show_results_popup
            _show_results_popup('Σχολικές Μονάδες',
                f'Email στάλθηκε: {", ".join(to_emails)}\n\nΣχολεία: {school_count}',
                result_type='ok', excel_path=out_path)
            self.destroy()
        except Exception as e:
            messagebox.showerror('Σφάλμα αποστολής', str(e), parent=self)

    def _clear(self):
        for w in self._exec_body.winfo_children():
            w.destroy()

    def _on_download_done(self):
        """Καλείται μετά από επιτυχή λήψη στο tab «⬇ Λήψη» — ξανα-ανιχνεύει
        τα αρχεία και ξαναχτίζει το tab «▶ Εκτέλεση» με τα φρέσκα αρχεία."""
        self._csv_path    = self._auto_find_zip('CSV_', 'stat2_2', 'gridResults')
        self._stat31_path = self._auto_find_zip('stat3_1')
        self._build_form()


class NeoSchoolYearDialog(tk.Toplevel):
    """Κεντρικό menu Νέου Σχολικού Έτους."""

    _HDR   = '#1B5E20'
    _CARD  = '#F1F8F2'
    _HOVER = '#DCEEDE'

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Νέο Σχ. Έτος')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass
        self._build()
        self.update_idletasks()
        w, h = 480, self.winfo_reqheight()
        pw = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        ph = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f'{w}x{h}+{pw}+{ph}')

    def _build(self):
        hdr = tk.Frame(self, bg=self._HDR, pady=12)
        hdr.pack(fill='x')
        tk.Label(hdr, text='🗓  Νέο Σχολικό Έτος',
                 bg=self._HDR, fg='white', font=('Arial', 13, 'bold')).pack()
        tk.Label(hdr, text='εργαλεία προετοιμασίας νέου σχολικού έτους',
                 bg=self._HDR, fg='#A5D6A7', font=('Arial', 8, 'italic')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=18, pady=14)
        body.pack(fill='both', expand=True)

        self._add_item(body,
            icon='⏹',
            title='Τερματισμός Τοποθετήσεων',
            desc='Αυτόματος τερματισμός τοποθετήσεων — ορισμός ημερομηνίας λήξης (21/6/2026) στο MySchool.',
            cmd=lambda: TerminationDialog(self))

        tk.Frame(body, bg='#C8E6C9', height=1).pack(fill='x', pady=10)

        self._add_item(body,
            icon='🔢',
            title='Αλλαγή Λειτουργικότητας',
            desc='Αυτόματη ενημέρωση λειτουργικότητας σχολικών μονάδων στο MySchool βάσει αρχείου Excel.',
            cmd=lambda: FunctionalityDialog(self))

        tk.Frame(body, bg='#C8E6C9', height=1).pack(fill='x', pady=10)

        self._add_item(body,
            icon='📐',
            title='Έλεγχος Τμημάτων Γενικής Παιδείας / Δυναμικού',
            desc='Πλήρης αποτύπωση Λειτουργικότητας vs Τμήματα/Μαθητές ανά τάξη για νηπιαγωγεία (5.3) και δημοτικά (5.4), με στοιχεία από το 3.1.',
            cmd=self._run_tmimata_genikis)

        tk.Frame(body, bg='#C8E6C9', height=1).pack(fill='x', pady=10)

        self._add_item(body,
            icon='🕐',
            title='Υποχρεωτικό Ωράριο ΠΕ60',
            desc='Έλεγχος υποχρεωτικού ωραρίου ΠΕ60/ΠΕ60.50 βάσει λειτουργικότητας νηπιαγωγείου τοποθέτησης (2.1 + 4.1 + 4.2).',
            cmd=self._run_orario_pe60)

    def _run_tmimata_genikis(self):
        import threading
        import importlib
        try:
            mod = importlib.import_module('checks.tmimata_genikis')
            threading.Thread(target=mod.run, args=(config,), daemon=True).start()
        except Exception as e:
            messagebox.showerror('Σφάλμα', str(e), parent=self)

    def _run_orario_pe60(self):
        import threading
        import importlib
        try:
            mod = importlib.import_module('checks.orario_pe60')
            threading.Thread(target=mod.run, args=(config,), daemon=True).start()
        except Exception as e:
            messagebox.showerror('Σφάλμα', str(e), parent=self)

    def _add_item(self, parent, icon, title, desc, cmd):
        card = tk.Frame(parent, bg=self._CARD, bd=1, relief='solid', cursor='hand2')
        card.pack(fill='x', pady=2)
        card.columnconfigure(1, weight=1)
        tk.Label(card, text=icon, bg=self._CARD,
                 font=('Arial', 20), padx=12, pady=10).grid(row=0, column=0, rowspan=2, sticky='ns')
        tk.Label(card, text=title, bg=self._CARD, fg=self._HDR,
                 font=('Arial', 10, 'bold'), anchor='w').grid(row=0, column=1, sticky='w', pady=(8, 0))
        tk.Label(card, text=desc, bg=self._CARD, fg='#555555',
                 font=('Arial', 8), anchor='w', wraplength=340,
                 justify='left').grid(row=1, column=1, sticky='w', pady=(0, 8))
        tk.Label(card, text='›', bg=self._CARD, fg=self._HDR,
                 font=('Arial', 18, 'bold'), padx=12).grid(row=0, column=2, rowspan=2, sticky='ns')
        all_w = [card] + list(card.winfo_children())
        def _on_enter(_): [w.configure(bg=self._HOVER) for w in all_w]
        def _on_leave(_): [w.configure(bg=self._CARD)  for w in all_w]
        def _on_click(_): cmd()
        for w in all_w:
            w.bind('<Enter>', _on_enter)
            w.bind('<Leave>', _on_leave)
            w.bind('<Button-1>', _on_click)


class DipeDialog(tk.Toplevel):
    """Κεντρικό menu ΔΙ.Π.Ε.Αν.Θ."""

    _HDR   = '#1F4E79'
    _CARD  = '#F0F4F8'
    _HOVER = '#DDE6F0'

    def __init__(self, parent):
        super().__init__(parent)
        self.title('ΔΙ.Π.Ε.Αν.Θ.')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass
        self._build()
        self.update_idletasks()
        w, h = 480, self.winfo_reqheight()
        pw = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        ph = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f'{w}x{h}+{pw}+{ph}')

    def _build(self):
        hdr = tk.Frame(self, bg=self._HDR, pady=12)
        hdr.pack(fill='x')
        tk.Label(hdr, text='🏛  ΔΙ.Π.Ε.Αν.Θ.',
                 bg=self._HDR, fg='white', font=('Arial', 13, 'bold')).pack()
        tk.Label(hdr, text='μόνο για χρήση από Δ/νση Π.Ε. Αν. Θεσσαλονίκης',
                 bg=self._HDR, fg='#FF6B6B', font=('Arial', 8, 'bold')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=18, pady=14)
        body.pack(fill='both', expand=True)

        self._add_item(body,
            icon='📁',
            title='Επεξεργασία αρχείου τοποθετήσεων',
            desc='Μετατροπή αρχικού αρχείου — συμπλήρωση πεδίων και άνοιγμα στο Excel.',
            cmd=lambda: DipePlacementsDialog(self))

        tk.Frame(body, bg='#D0D8E4', height=1).pack(fill='x', pady=10)

        self._add_item(body,
            icon='📋',
            title='Εκπ/κοί ανά Ειδικότητα & Θέση Συμβούλου',
            desc='Εξαγωγή εκπαιδευτικών φιλτραρισμένων ανά ειδικότητα και θέση Συμβούλου Εκπ/σης.',
            cmd=lambda: SymbouloiDialog(self))

        tk.Frame(body, bg='#D0D8E4', height=1).pack(fill='x', pady=10)

        self._add_item(body,
            icon='📝',
            title='Καταχώρηση Απουσίας σε Οργανική',
            desc='Αυτόματη καταχώρηση απουσίας Ολικής Διάθεσης στην οργανική τοποθέτηση εκπαιδευτικών.',
            cmd=lambda: AbsencesDialog(self))

    def _add_item(self, parent, icon, title, desc, cmd):
        card = tk.Frame(parent, bg=self._CARD, bd=1, relief='solid', cursor='hand2')
        card.pack(fill='x', pady=2)
        card.columnconfigure(1, weight=1)
        tk.Label(card, text=icon, bg=self._CARD,
                 font=('Arial', 20), padx=12, pady=10).grid(row=0, column=0, rowspan=2, sticky='ns')
        tk.Label(card, text=title, bg=self._CARD, fg=self._HDR,
                 font=('Arial', 10, 'bold'), anchor='w').grid(row=0, column=1, sticky='w', pady=(8,0))
        tk.Label(card, text=desc, bg=self._CARD, fg='#555555',
                 font=('Arial', 8), anchor='w', wraplength=340,
                 justify='left').grid(row=1, column=1, sticky='w', pady=(0,8))
        tk.Label(card, text='›', bg=self._CARD, fg=self._HDR,
                 font=('Arial', 18, 'bold'), padx=12).grid(row=0, column=2, rowspan=2, sticky='ns')
        all_w = [card] + list(card.winfo_children())
        def _on_enter(_): [w.configure(bg=self._HOVER) for w in all_w]
        def _on_leave(_): [w.configure(bg=self._CARD)  for w in all_w]
        def _on_click(_): cmd()
        for w in all_w:
            w.bind('<Enter>', _on_enter)
            w.bind('<Leave>', _on_leave)
            w.bind('<Button-1>', _on_click)


class DipePlacementsDialog(tk.Toplevel):
    """Επεξεργασία αρχικού αρχείου τοποθετήσεων."""

    _SEC_BG = '#F0F4F8'
    _HDR    = '#1F4E79'

    _EIDOS_OPTIONS = [
        'Οργανικά',
        'Οργανικά από Αμοιβαία Μετάθεση',
        'Οργανικά από Αρση Υπεραριθμίας',
        'Οργανικά σε Τμήμα Ένταξης',
        'Από Διάθεση ΠΥΣΠΕ/ΠΥΣΔΕ',
        'Επί Θητεία',
        'Ειδική Θέση (τ. Σχ. Σύμβουλοι - ν. 1966/1991 άρ.8, παρ.5)',
        'Απόσπαση (με αίτηση - κύριος φορέας)',
        'Ολική Διάθεση (ανάγκες υπηρεσίας - κύριος φορέας)',
        'Μερική Διάθεση (Συμπλήρωση Ωραρίου)',
        'Υπερωριακά',
    ]

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Επεξεργασία αρχείου τοποθετήσεων')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        self._raw_var    = tk.StringVar()
        self._stat41_var = tk.StringVar()
        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass
        self._build()
        self.update_idletasks()
        self.geometry('560x390')
        pw = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

    def _build(self):
        from tkinter import scrolledtext as st2
        hdr = tk.Frame(self, bg=self._HDR, pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='📁  Επεξεργασία αρχείου τοποθετήσεων',
                 bg=self._HDR, fg='white', font=('Arial', 12, 'bold')).pack()
        tk.Label(hdr, text='μόνο για χρήση από Δ/νση Π.Ε. Αν. Θεσσαλονίκης',
                 bg=self._HDR, fg='#FF6B6B', font=('Arial', 8, 'bold')).pack()
        body = tk.Frame(self, bg=C['bg'], padx=16, pady=12)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=1)
        sec = tk.LabelFrame(body,
                            text='  Επιλογή αρχικού αρχείου',
                            bg=self._SEC_BG, fg=self._HDR,
                            font=('Arial', 9, 'bold'), bd=1, relief='groove', padx=10, pady=8)
        sec.grid(row=0, column=0, sticky='ew', pady=(0,6))
        sec.columnconfigure(0, weight=1)
        f1 = tk.Frame(sec, bg=self._SEC_BG)
        f1.grid(row=0, column=0, sticky='ew')
        f1.columnconfigure(0, weight=1)
        tk.Entry(f1, textvariable=self._raw_var, font=('Arial', 9),
                 relief='solid', bd=1).pack(side='left', fill='x', expand=True)
        tk.Button(f1, text='📂', bg=self._SEC_BG, relief='flat',
                  font=('Arial', 11), cursor='hand2',
                  command=self._browse_raw).pack(side='left', padx=(4,0))

        sec2 = tk.LabelFrame(body,
                            text='  Αρχείο 4.1 για ΑΦΜ (προαιρετικό — αλλιώς αυτόματη εύρεση)',
                            bg=self._SEC_BG, fg=self._HDR,
                            font=('Arial', 9, 'bold'), bd=1, relief='groove', padx=10, pady=8)
        sec2.grid(row=1, column=0, sticky='ew', pady=(0,6))
        sec2.columnconfigure(0, weight=1)
        f1b = tk.Frame(sec2, bg=self._SEC_BG)
        f1b.grid(row=0, column=0, sticky='ew')
        f1b.columnconfigure(0, weight=1)
        tk.Entry(f1b, textvariable=self._stat41_var, font=('Arial', 9),
                 relief='solid', bd=1).pack(side='left', fill='x', expand=True)
        tk.Button(f1b, text='📂', bg=self._SEC_BG, relief='flat',
                  font=('Arial', 11), cursor='hand2',
                  command=self._browse_stat41).pack(side='left', padx=(4,0))

        self._conv_btn = tk.Button(sec2,
                  text='Επεξεργασία & άνοιγμα  →',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=10, pady=4, cursor='hand2', command=self._convert)
        self._conv_btn.grid(row=1, column=0, sticky='e', pady=(6,0))
        self._status_var = tk.StringVar(value='')
        tk.Label(body, textvariable=self._status_var, bg=C['bg'], fg=C['status_run'],
                 font=('Arial', 8), anchor='w').grid(row=2, column=0, sticky='w', pady=(0,4))
        tk.Label(body, text='Αρχείο καταγραφής:',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold')).grid(row=3, column=0, sticky='w', pady=(4,2))
        self._log = st2.ScrolledText(body, height=6, font=('Consolas', 8),
                                      relief='solid', bd=1, state='disabled',
                                      bg='#F5F5F5', wrap=tk.WORD)
        self._log.grid(row=4, column=0, sticky='nsew', pady=(0,4))
        body.rowconfigure(4, weight=1)

    def _browse_raw(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(parent=self,
            title='Επιλογή αρχείου τοποθετήσεων',
            filetypes=[('Excel', '*.xlsx *.xls'), ('Όλα', '*.*')])
        if path:
            self._raw_var.set(path)

    def _browse_stat41(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(parent=self,
            title='Επιλογή αρχείου stat4_1 (Οργανικές τοποθετήσεις)',
            filetypes=[('Excel/CSV/ZIP', '*.xlsx *.xls *.csv *.zip'), ('Όλα', '*.*')])
        if path:
            self._stat41_var.set(path)

    def _log_msg(self, msg):
        def _do():
            self._log.configure(state='normal')
            self._log.insert(tk.END, msg + '\n')
            self._log.see(tk.END)
            self._log.configure(state='disabled')
        self.after(0, _do)

    def _ask_batch_fields(self):
        """
        Popup: ζητά ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ + ΑΠΟ/ΕΩΣ μία φορά, για εφαρμογή
        σε ΟΛΕΣ τις γραμμές του αρχείου. Κενά πεδία = δεν συμπληρώνεται
        αυτόματα (μένει όπως πριν, χειροκίνητα στο Excel).
        Επιστρέφει dict {'eidos','apo','eos'} ή None αν πατηθεί Άκυρο.
        """
        win = tk.Toplevel(self)
        win.title('Στοιχεία τοποθέτησης (όλο το αρχείο)')
        win.configure(bg=C['bg'])
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        from datetime import datetime as _dt_bf
        from tkinter import ttk as _ttk_bf
        result = {}

        body = tk.Frame(win, bg=C['bg'], padx=16, pady=14)
        body.pack(fill='both', expand=True)

        tk.Label(body, text='Οι τιμές αυτές θα εφαρμοστούν σε ΟΛΕΣ τις γραμμές.\n'
                             'Άφησε κενό ό,τι θέλεις να συμπληρώσεις χειροκίνητα.',
                 bg=C['bg'], fg=C['desc'], font=('Arial', 8),
                 justify='left').grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 10))

        tk.Label(body, text='Είδος τοποθέτησης:', bg=C['bg'], fg=self._HDR,
                 font=('Arial', 9, 'bold')).grid(row=1, column=0, sticky='w', pady=4)
        eidos_var = tk.StringVar()
        cb = _ttk_bf.Combobox(body, textvariable=eidos_var, values=self._EIDOS_OPTIONS,
                           state='readonly', width=42, font=('Arial', 9))
        cb.grid(row=1, column=1, sticky='ew', pady=4, padx=(8, 0))

        tk.Label(body, text='Από (ΗΗ/ΜΜ/ΕΕΕΕ):', bg=C['bg'], fg=self._HDR,
                 font=('Arial', 9, 'bold')).grid(row=2, column=0, sticky='w', pady=4)
        apo_var = tk.StringVar()
        tk.Entry(body, textvariable=apo_var, font=('Arial', 9),
                  relief='solid', bd=1, width=15).grid(row=2, column=1, sticky='w', pady=4, padx=(8, 0))

        tk.Label(body, text='Έως (ΗΗ/ΜΜ/ΕΕΕΕ):', bg=C['bg'], fg=self._HDR,
                 font=('Arial', 9, 'bold')).grid(row=3, column=0, sticky='w', pady=4)
        eos_var = tk.StringVar()
        tk.Entry(body, textvariable=eos_var, font=('Arial', 9),
                  relief='solid', bd=1, width=15).grid(row=3, column=1, sticky='w', pady=4, padx=(8, 0))

        def _valid_date(s):
            s = s.strip()
            if not s:
                return True
            try:
                _dt_bf.strptime(s, '%d/%m/%Y')
                return True
            except ValueError:
                return False

        def _ok():
            apo = apo_var.get().strip()
            eos = eos_var.get().strip()
            if not _valid_date(apo) or not _valid_date(eos):
                messagebox.showwarning('Προσοχή',
                    'Οι ημερομηνίες πρέπει να είναι σε μορφή ΗΗ/ΜΜ/ΕΕΕΕ (π.χ. 01/09/2025).',
                    parent=win)
                return
            result['eidos'] = eidos_var.get().strip()
            result['apo']   = apo
            result['eos']   = eos
            win.destroy()

        def _cancel():
            result.clear()
            result['_cancelled'] = True
            win.destroy()

        btn_row = tk.Frame(body, bg=C['bg'])
        btn_row.grid(row=4, column=0, columnspan=2, sticky='e', pady=(14, 0))
        tk.Button(btn_row, text='Άκυρο', bg=C['bg2'], fg=C['desc'],
                  font=('Arial', 9), relief='flat', padx=10, pady=4,
                  cursor='hand2', command=_cancel).pack(side='right', padx=(6, 0))
        tk.Button(btn_row, text='Συνέχεια →', bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat', padx=12, pady=4,
                  cursor='hand2', command=_ok).pack(side='right')

        win.protocol('WM_DELETE_WINDOW', _cancel)
        win.update_idletasks()
        pw = self.winfo_x() + (self.winfo_width() - win.winfo_width()) // 2
        ph = self.winfo_y() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry(f'+{pw}+{ph}')
        win.wait_window()

        if result.get('_cancelled'):
            return None
        return result

    def _convert(self):
        import threading as _th
        src = self._raw_var.get().strip()
        if not src:
            messagebox.showwarning('Προσοχή',
                'Επίλεξε πρώτα το αρχικό αρχείο.', parent=self)
            return
        batch = self._ask_batch_fields()
        if batch is None:
            return
        self._conv_btn.configure(state='disabled', text='Μετατροπή...')
        self._log_msg('→ Μετατροπή αρχείου...')
        if batch.get('eidos'):
            self._log_msg(f"  Είδος τοποθέτησης (όλο το αρχείο): {batch['eidos']}")
        if batch.get('apo'):
            self._log_msg(f"  Από (όλο το αρχείο): {batch['apo']}")
        if batch.get('eos'):
            self._log_msg(f"  Έως (όλο το αρχείο): {batch['eos']}")
        stat41 = self._stat41_var.get().strip()
        def _do():
            try:
                import placements
                dest, n, warns = placements.convert_raw_file(
                    src,
                    eidos=batch.get('eidos') or None,
                    apo_override=batch.get('apo') or None,
                    eos_override=batch.get('eos') or None,
                    stat41_path=stat41 or None,
                )
                def _after():
                    self._conv_btn.configure(state='normal', text='Επεξεργασία & άνοιγμα  →')
                    self._log_msg(f'✓ Δημιουργήθηκε: {os.path.basename(dest)} ({n} γραμμές)')
                    # Πρώτα τα info για stat2_2 / stat4_1 (αν βρέθηκαν)
                    info_warns  = [w for w in warns if w.startswith('stat2_2') or w.startswith('stat4_1')]
                    afm_warns   = [w for w in warns if w.startswith('Δεν βρέθηκε ΑΦΜ')]
                    other_warns = [w for w in warns if w not in info_warns and w not in afm_warns]
                    for w in info_warns:
                        self._log_msg(f'  📂 {w}')
                    if not any(w.startswith('stat2_2') for w in info_warns):
                        self._log_msg('  ℹ stat2_2 δεν βρέθηκε — κωδικοί κενοί')
                    if not any(w.startswith('stat4_1') for w in info_warns):
                        self._log_msg('  ℹ stat4_1 δεν βρέθηκε — ΑΦΜ κενά')
                    for w in afm_warns:
                        self._log_msg(f'  ⚠ {w}')
                    for w in other_warns:
                        self._log_msg(f'  ⚠ {w}')
                    self._log_msg('  Άνοιγμα Excel — συμπλήρωσε τα πορτοκαλί κελιά και αποθήκευσε.')
                    self._status_var.set('✓ Μετατροπή ολοκληρώθηκε.')
                    try: os.startfile(dest)
                    except Exception: pass
                self.after(0, _after)
            except Exception as exc:
                def _err():
                    self._conv_btn.configure(state='normal', text='Επεξεργασία & άνοιγμα  →')
                    self._log_msg(f'❌ Σφάλμα: {exc}')
                    self._status_var.set(f'❌ {exc}')
                self.after(0, _err)
        _th.Thread(target=_do, daemon=True).start()


class PlacementsDialog(tk.Toplevel):
    """Παράθυρο αυτόματης καταχώρησης τοποθετήσεων."""

    _SEC_BG   = '#F0F4F8'
    _SEC_BD   = '#C8D4E0'
    _LBL_STEP = '#1F4E79'

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Τοποθετήσεις — Αυτόματη Καταχώρηση')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        self._driver    = None
        self._running   = False
        self._excel_var = tk.StringVar()   # έτοιμο αρχείο τοποθετήσεων

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        self.geometry('600x560')
        pw = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build(self):
        from tkinter import scrolledtext as st2

        hdr = tk.Frame(self, bg='#0F6E56', pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='👥  Τοποθετήσεις — Αυτόματη Καταχώρηση',
                 bg='#0F6E56', fg='white',
                 font=('Arial', 12, 'bold')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=16, pady=12)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=1)

        # ── Εκτέλεση καταχώρησης ─────────────────────────────────────────────
        sec2 = tk.LabelFrame(body,
                             text='  Εκτέλεση καταχώρησης',
                             bg=C['bg'], fg=self._LBL_STEP,
                             font=('Arial', 9, 'bold'),
                             bd=1, relief='groove', padx=10, pady=8)
        sec2.grid(row=0, column=0, sticky='ew', pady=(0, 8))
        sec2.columnconfigure(0, weight=1)

        # Links γραμμή
        lnk_row = tk.Frame(sec2, bg=C['bg'])
        lnk_row.grid(row=0, column=0, sticky='w', pady=(0, 6))

        def _open_template():
            import sys, subprocess
            tpl = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'placements_template.xlsx')
            if os.path.exists(tpl):
                os.startfile(tpl)
            else:
                messagebox.showwarning('Προσοχή', 'Το αρχείο προτύπου δεν βρέθηκε.', parent=self)

        def _open_help():
            help_win = tk.Toplevel(self)
            help_win.title('Οδηγίες αρχείου Excel τοποθετήσεων')
            help_win.configure(bg=C['bg'])
            help_win.resizable(True, True)
            help_win.transient(self)
            help_win.geometry('640x580')
            pw = self.winfo_x() + (self.winfo_width()  - 640) // 2
            ph = self.winfo_y() + (self.winfo_height() - 580) // 2
            help_win.geometry(f'+{pw}+{ph}')
            hdr2 = tk.Frame(help_win, bg='#1F4E79', pady=8)
            hdr2.pack(fill='x')
            tk.Label(hdr2, text='📋  Μορφή αρχείου Excel τοποθετήσεων',
                     bg='#1F4E79', fg='white', font=('Arial', 11, 'bold')).pack()
            import tkinter.scrolledtext as st3
            txt = st3.ScrolledText(help_win, font=('Arial', 9), wrap=tk.WORD,
                                   bg='#F5F5F5', relief='flat', padx=12, pady=10)
            txt.pack(fill='both', expand=True, padx=8, pady=8)
            help_text = (
                "Διαβάζονται δεδομένα από τις παρακάτω στήλες:\n\n"
                "• ΕΙΔΟΣ ΤΟΠΟΘΕΤΗΣΗΣ\n"
                "  Επιτρεπτές τιμές (ακριβώς όπως στο MySchool):\n"
                "    - Οργανικά\n"
                "    - Οργανικά από Αμοιβαία Μετάθεση\n"
                "    - Οργανικά από Αρση Υπεραριθμίας\n"
                "    - Οργανικά σε Τμήμα Ένταξης\n"
                "    - Από Διάθεση ΠΥΣΠΕ/ΠΥΣΔΕ\n"
                "    - Επί Θητεία\n"
                "    - Ειδική Θέση (τ. Σχ. Σύμβουλοι - ν. 1966/1991 άρ.8, παρ.5)\n"
                "    - Απόσπαση (με αίτηση - κύριος φορέας)\n"
                "    - Ολική Διάθεση (ανάγκες υπηρεσίας - κύριος φορέας)\n"
                "    - Μερική Διάθεση (Συμπλήρωση Ωραρίου)\n"
                "    - Υπερωριακά\n\n"
                "• Α.Φ.Μ.\n"
                "• ΕΠΙΘΕΤΟ\n"
                "• ΟΝΟΜΑ\n"
                "• ΚΩΔ. ΣΧΟΛΕΙΟΥ\n"
                "• ΣΧΟΛΕΙΟ\n"
                "• ΩΡΕΣ\n"
                "  Για πλήρη διάθεση → τιμή -1 (δίνει πλήρες ωράριο και επιλέγει όλες τις ημέρες).\n"
                "  Αν βάλεις αριθμό ωρών (π.χ. 25), καταχωρεί τις ώρες αλλά ΔΕΝ επιλέγει ημέρες εβδομάδας.\n\n"
                "• ΑΠΟ  (π.χ. 01/09/2025)\n"
                "• ΕΩΣ  (π.χ. 21/06/2026)\n"
                "• OK      → δεν γράφουμε, συμπληρώνεται από την εφαρμογή\n"
                "• ΣΧΟΛΙΟ → δεν γράφουμε, συμπληρώνεται από την εφαρμογή\n\n"
                "─────────────────────────────────────\n"
                "Παρατηρήσεις:\n\n"
                "• Δεν έχει σημασία με ποια σειρά είναι τοποθετημένες οι στήλες.\n"
                "• ΔΕ ΜΟΡΦΟΠΟΙΟΥΜΕ ΤΟ ΑΡΧΕΙΟ.\n"
                "  ΔΕΝ επιλέγουμε ολόκληρες γραμμές ή στήλες για να εφαρμόσουμε κάποιον κανόνα."
            )
            txt.insert('1.0', help_text)
            txt.configure(state='disabled')
            tk.Button(help_win, text='Κλείσιμο', bg=C['btn_bg'], fg=C['btn_fg'],
                      font=('Arial', 9, 'bold'), relief='flat', padx=12, pady=4,
                      cursor='hand2', command=help_win.destroy).pack(pady=(0, 10))

        lbl_tpl = tk.Label(lnk_row, text='📥 Πρότυπο Excel',
                           bg=C['bg'], fg='#1565C0',
                           font=('Arial', 9, 'underline'), cursor='hand2')
        lbl_tpl.pack(side='left')
        lbl_tpl.bind('<Button-1>', lambda e: _open_template())

        tk.Label(lnk_row, text='  |  ', bg=C['bg'],
                 fg=C['desc'], font=('Arial', 9)).pack(side='left')

        lbl_hlp = tk.Label(lnk_row, text='❓ Οδηγίες αρχείου',
                           bg=C['bg'], fg='#1565C0',
                           font=('Arial', 9, 'underline'), cursor='hand2')
        lbl_hlp.pack(side='left')
        lbl_hlp.bind('<Button-1>', lambda e: _open_help())

        f2 = tk.Frame(sec2, bg=C['bg'])
        f2.grid(row=1, column=0, sticky='ew')
        f2.columnconfigure(1, weight=1)
        tk.Label(f2, text='Αρχείο τοποθετήσεων:',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9)).grid(row=0, column=0, padx=(0, 6))
        tk.Entry(f2, textvariable=self._excel_var, font=('Arial', 9),
                 relief='solid', bd=1).grid(row=0, column=1, sticky='ew')
        tk.Button(f2, text='📂', bg=C['bg'], relief='flat',
                  font=('Arial', 11), cursor='hand2',
                  command=self._browse).grid(row=0, column=2, padx=(4, 0))

        btn_row = tk.Frame(sec2, bg=C['bg'])
        btn_row.grid(row=2, column=0, sticky='w', pady=(8, 0))
        self._run_btn = tk.Button(btn_row,
                  text='▶  Εκτέλεση καταχώρησης',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=12, pady=5, cursor='hand2',
                  command=self._run)
        self._run_btn.pack(side='left')

        # Status
        self._status_var = tk.StringVar(value='Επίλεξε αρχείο Excel και πάτα Εκτέλεση.')
        tk.Label(body, textvariable=self._status_var,
                 bg=C['bg'], fg=C['status_run'],
                 font=('Arial', 8), anchor='w').grid(row=1, column=0, sticky='w', pady=(0, 4))

        # Log
        tk.Label(body, text='Αρχείο καταγραφής:',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold')).grid(row=2, column=0, sticky='w', pady=(4, 2))
        self._log = st2.ScrolledText(body, height=10, font=('Consolas', 8),
                                      relief='solid', bd=1, state='disabled',
                                      bg='#F5F5F5', wrap=tk.WORD)
        self._log.grid(row=3, column=0, sticky='nsew', pady=(0, 4))
        body.rowconfigure(3, weight=1)

        # Footer με κουμπί κλεισίματος
        foot = tk.Frame(self, bg=C['bg2'], pady=8)
        foot.pack(fill='x')
        tk.Button(foot, text='Κλείσιμο',
                  bg=C['bg2'], fg=C['desc'],
                  font=('Arial', 10), relief='flat', padx=12, pady=4,
                  cursor='hand2', command=self._on_close).pack(side='right', padx=16)

        self.protocol('WM_DELETE_WINDOW', self._on_close)

    # ── Σύνδεση / εκτέλεση ───────────────────────────────────────────────────

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=self,
            title='Επιλογή αρχείου Excel τοποθετήσεων',
            filetypes=[('Excel', '*.xlsx *.xls'), ('Όλα', '*.*')])
        if path:
            self._excel_var.set(path)

    def _log_msg(self, msg):
        def _do():
            self._log.configure(state='normal')
            self._log.insert(tk.END, msg + '\n')
            self._log.see(tk.END)
            self._log.configure(state='disabled')
        self.after(0, _do)

    def _ask_decision_mode(self):
        """
        Popup: ρωτά αν ο χρήστης θα ανοίξει ο ίδιος μια υπάρχουσα απόφαση
        τοποθέτησης (χρήσιμο όταν μια καταχώρηση είχε διακοπεί στη μέση),
        ή αν θέλει να δημιουργηθεί νέα (προεπιλεγμένη συμπεριφορά).
        Επιστρέφει 'existing', 'new', ή None αν ακυρωθεί.
        """
        win = tk.Toplevel(self)
        win.title('Απόφαση τοποθέτησης')
        win.configure(bg=C['bg'])
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        result = {}

        body = tk.Frame(win, bg=C['bg'], padx=18, pady=16)
        body.pack(fill='both', expand=True)

        tk.Label(body,
                 text='Θα ανοίξεις εσύ μια υπάρχουσα απόφαση τοποθέτησης\n'
                      'στο παράθυρο Chrome (π.χ. μια που είχε διακοπεί),\n'
                      'ή να δημιουργήσω νέα;',
                 bg=C['bg'], fg=self._LBL_STEP, font=('Arial', 10),
                 justify='left').pack(pady=(0, 14))

        def _pick(mode):
            result['mode'] = mode
            win.destroy()

        btn_row = tk.Frame(body, bg=C['bg'])
        btn_row.pack()
        tk.Button(btn_row, text='📂  Θα ανοίξω υπάρχουσα', bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat', padx=10, pady=6,
                  cursor='hand2', command=lambda: _pick('existing')).pack(side='left', padx=(0, 8))
        tk.Button(btn_row, text='➕  Δημιουργία νέας', bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat', padx=10, pady=6,
                  cursor='hand2', command=lambda: _pick('new')).pack(side='left')

        tk.Button(body, text='Άκυρο', bg=C['bg2'], fg=C['desc'], font=('Arial', 8),
                  relief='flat', padx=8, pady=3, cursor='hand2',
                  command=lambda: _pick(None)).pack(pady=(12, 0))

        win.protocol('WM_DELETE_WINDOW', lambda: _pick(None))
        win.update_idletasks()
        pw = self.winfo_x() + (self.winfo_width() - win.winfo_width()) // 2
        ph = self.winfo_y() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry(f'+{pw}+{ph}')
        win.wait_window()

        return result.get('mode')

    def _run(self):
        import threading as _th
        path = self._excel_var.get().strip()
        if not path:
            messagebox.showwarning('Προσοχή', 'Επίλεξε αρχείο Excel πρώτα.', parent=self)
            return
        mode = 'new'
        if not self._driver:
            mode = self._ask_decision_mode()
            if mode is None:
                return
        self._run_btn.configure(state='disabled', bg=C['btn_dis'])
        self._status_var.set('Σύνδεση στο MySchool...')
        def _do():
            import placements
            if not self._driver:
                self._log_msg('→ Εκκίνηση σύνδεσης...')
                drv = placements.connect(log=self._log_msg, mode=mode)
                if not drv:
                    def _fail():
                        self._run_btn.configure(state='normal', bg=C['btn_bg'])
                        self._status_var.set('❌ Αποτυχία σύνδεσης — έλεγξε credentials στις Ρυθμίσεις.')
                    self.after(0, _fail)
                    return
                self._driver = drv
            self._log_msg('→ Εκτέλεση καταχώρησης...')
            self.after(0, lambda: self._status_var.set('Εκτέλεση καταχώρησης...'))
            placements.run({'excel_path': path}, self._driver, callback=self._log_msg)
            def _after():
                self._run_btn.configure(state='normal', bg=C['btn_bg'])
                self._status_var.set('✓ Ολοκλήρωση.')
                messagebox.showinfo(
                    'Ολοκλήρωση',
                    '✓ Η καταχώρηση τοποθετήσεων ολοκληρώθηκε.\n\n'
                    'Ελέγξτε τις στήλες OK και ΣΧΟΛΙΟ στο Excel για λεπτομέρειες.',
                    parent=self)
                self._on_close()
            self.after(0, _after)
        _th.Thread(target=_do, daemon=True).start()

    def _on_close(self):
        if self._driver:
            try: self._driver.quit()
            except Exception: pass
        self.destroy()


class EditorDialog(tk.Toplevel):
    """Παράθυρο αυτόματης επεξεργασίας καρτέλας εκπαιδευτικού."""

    _HDR_BG  = '#1A5276'   # σκούρο μπλε (διαφορετικό από Τοποθετήσεις)
    _LBL_CLR = '#1A5276'

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Editor — Επεξεργασία Καρτέλας Εκπαιδευτικού')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        self._driver   = None
        self._file_var = tk.StringVar()
        from datetime import date as _date
        self._date_var = tk.StringVar(value=_date.today().strftime('%d/%m/%Y'))

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        self.geometry('580x460')
        pw = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

    def _build(self):
        from tkinter import scrolledtext as st2

        hdr = tk.Frame(self, bg=self._HDR_BG, pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='✏  Editor — Επεξεργασία Καρτέλας Εκπαιδευτικού',
                 bg=self._HDR_BG, fg='white',
                 font=('Arial', 12, 'bold')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=16, pady=12)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=1)

        # Αρχείο
        tk.Label(body, text='Αρχείο εκπαιδευτικών (Excel ή CSV):',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky='w', pady=(0, 3))
        ff = tk.Frame(body, bg=C['bg'])
        ff.grid(row=1, column=0, sticky='ew', pady=(0, 10))
        ff.columnconfigure(0, weight=1)
        tk.Entry(ff, textvariable=self._file_var, font=('Arial', 9),
                 relief='solid', bd=1).pack(side='left', fill='x', expand=True)
        tk.Button(ff, text='📂', bg=C['bg'], relief='flat', font=('Arial', 11),
                  cursor='hand2', command=self._browse).pack(side='left', padx=(4, 0))

        # Απαιτούμενες στήλες
        note = tk.Label(body,
                 text='Απαιτείται στήλη: Α.Φ.Μ.  |  Προαιρετική: Ονομασία Σχολείου (για επιλογή όταν υπάρχουν πολλά αποτελέσματα)',
                 bg=C['bg'], fg='#666666', font=('Arial', 8), anchor='w', wraplength=520, justify='left')
        note.grid(row=2, column=0, sticky='w', pady=(0, 10))

        # Ημερομηνία
        tk.Label(body, text='Ημερομηνία καταχώρησης (ΗΗ/ΜΜ/ΕΕΕΕ):',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=3, column=0, sticky='w', pady=(0, 3))
        date_row = tk.Frame(body, bg=C['bg'])
        date_row.grid(row=4, column=0, sticky='w', pady=(0, 10))
        tk.Entry(date_row, textvariable=self._date_var, font=('Arial', 9),
                 relief='solid', bd=1, width=14).pack(side='left')
        tk.Label(date_row, text='  (προεπιλογή: σήμερα)',
                 bg=C['bg'], fg='#888888', font=('Arial', 8)).pack(side='left')

        # Κουμπί ενοποιημένο
        btn_row = tk.Frame(body, bg=C['bg'])
        btn_row.grid(row=5, column=0, sticky='w', pady=(0, 8))
        self._conn_btn = tk.Button(btn_row,
                  text='▶  Σύνδεση & Εκτέλεση',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=12, pady=5, cursor='hand2',
                  command=self._connect_and_run)
        self._conn_btn.pack(side='left')

        # Status
        self._status_var = tk.StringVar(value='Επίλεξε αρχείο και πάτα Σύνδεση & Εκτέλεση.')
        tk.Label(body, textvariable=self._status_var,
                 bg=C['bg'], fg=C['status_run'],
                 font=('Arial', 8), anchor='w').grid(row=6, column=0, sticky='w', pady=(0, 4))

        # Log
        tk.Label(body, text='Αρχείο καταγραφής:',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=7, column=0, sticky='w', pady=(4, 2))
        self._log = st2.ScrolledText(body, height=12, font=('Consolas', 8),
                                      relief='solid', bd=1, state='disabled',
                                      bg='#F5F5F5', wrap=tk.WORD)
        self._log.grid(row=8, column=0, sticky='nsew', pady=(0, 4))
        body.rowconfigure(8, weight=1)

        self.protocol('WM_DELETE_WINDOW', self._on_close)

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=self,
            title='Επιλογή αρχείου εκπαιδευτικών',
            filetypes=[('Excel/CSV', '*.xlsx *.xls *.csv'), ('Όλα', '*.*')])
        if path:
            self._file_var.set(path)

    def _log_msg(self, msg):
        def _do():
            self._log.configure(state='normal')
            self._log.insert(tk.END, msg + '\n')
            self._log.see(tk.END)
            self._log.configure(state='disabled')
        self.after(0, _do)

    def _connect_and_run(self):
        import threading as _th
        path = self._file_var.get().strip()
        if not path:
            messagebox.showwarning('Προσοχή', 'Επίλεξε αρχείο πρώτα.', parent=self)
            return
        self._conn_btn.configure(state='disabled', text='Εκτελείται...')
        self._status_var.set('Σύνδεση στο MySchool...')
        def _do():
            import editor
            drv = editor.connect(log=self._log_msg)
            if not drv:
                def _fail():
                    self._conn_btn.configure(state='normal', text='▶  Σύνδεση & Εκτέλεση')
                    self._status_var.set('Αποτυχία σύνδεσης — έλεγξε credentials στις Ρυθμίσεις.')
                self.after(0, _fail)
                return
            self._driver = drv
            self.after(0, lambda: self._status_var.set('Εκτέλεση...'))
            editor.run({'file_path': path, 'date': self._date_var.get().strip()}, drv, callback=self._log_msg)
            def _after():
                self._conn_btn.configure(state='normal', text='▶  Σύνδεση & Εκτέλεση')
                self._status_var.set('Ολοκλήρωση.')
            self.after(0, _after)
        _th.Thread(target=_do, daemon=True).start()

    def _on_close(self):
        if self._driver:
            try: self._driver.quit()
            except Exception: pass
        self.destroy()


class _NumberedChoiceDialog(tk.Toplevel):
    """Modal dialog επιλογής από αριθμημένη λίστα (thread-safe — καλείται από main thread)."""

    def __init__(self, parent, title, prompt, options):
        super().__init__(parent)
        self.title(title)
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.result = None   # '1'–'N'  ή  '0' (παράλειψη)

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        body = tk.Frame(self, bg=C['bg'], padx=20, pady=16)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=1)

        # Prompt
        tk.Label(body, text=prompt, bg=C['bg'], font=('Arial', 9),
                 anchor='w', justify='left', wraplength=450).grid(
            row=0, column=0, sticky='w', pady=(0, 10))

        # Listbox + scrollbar
        lf = tk.Frame(body, bg=C['bg'])
        lf.grid(row=1, column=0, sticky='nsew', pady=(0, 10))
        lf.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        sb = tk.Scrollbar(lf)
        sb.pack(side='right', fill='y')

        self._lb = tk.Listbox(lf, yscrollcommand=sb.set,
                              font=('Consolas', 9),
                              height=min(len(options), 8),
                              selectmode='single',
                              activestyle='dotbox',
                              relief='solid', bd=1,
                              bg='#FAFAFA')
        for i, opt in enumerate(options, 1):
            self._lb.insert(tk.END, f'{i}.  {opt}')
        self._lb.pack(side='left', fill='both', expand=True)
        sb.configure(command=self._lb.yview)

        # Buttons
        btn_row = tk.Frame(body, bg=C['bg'])
        btn_row.grid(row=2, column=0, sticky='w')

        tk.Button(btn_row, text='0  Παράλειψη',
                  bg='#E0E0E0', fg='#333333',
                  font=('Arial', 9), relief='flat',
                  padx=10, pady=4, cursor='hand2',
                  command=self._skip).pack(side='left', padx=(0, 8))

        tk.Button(btn_row, text='✓  Επιλογή',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=12, pady=4, cursor='hand2',
                  command=self._confirm).pack(side='left')

        self._lb.bind('<Double-Button-1>', lambda _e: self._confirm())
        self.protocol('WM_DELETE_WINDOW', self._skip)

        # Κεντράρισμα
        self.update_idletasks()
        pw = parent.winfo_rootx() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

        parent.wait_window(self)

    def _confirm(self):
        sel = self._lb.curselection()
        self.result = str(sel[0] + 1) if sel else None
        self.destroy()

    def _skip(self):
        self.result = '0'
        self.destroy()


class TerminationDialog(tk.Toplevel):
    """Τερματισμός Τοποθετήσεων — αυτόματη ενημέρωση ημ. λήξης."""

    _HDR_BG  = '#1F4E79'
    _LBL_CLR = '#1F4E79'

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Τερματισμός Τοποθετήσεων')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        self._driver   = None
        self._file_var = tk.StringVar()
        self._date_var = tk.StringVar(value='21/6/2026')

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        self.geometry('600x480')
        pw = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

    def _build(self):
        from tkinter import scrolledtext as st2

        hdr = tk.Frame(self, bg=self._HDR_BG, pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='⏹  Τερματισμός Τοποθετήσεων',
                 bg=self._HDR_BG, fg='white',
                 font=('Arial', 12, 'bold')).pack()
        tk.Label(hdr, text='Αυτόματη ενημέρωση ημερομηνίας λήξης στο MySchool',
                 bg=self._HDR_BG, fg='#A8C4D8',
                 font=('Arial', 8, 'italic')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=16, pady=12)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=1)

        # Αρχείο
        tk.Label(body, text='Αρχείο εκπαιδευτικών (Excel ή CSV):',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky='w', pady=(0, 3))
        ff = tk.Frame(body, bg=C['bg'])
        ff.grid(row=1, column=0, sticky='ew', pady=(0, 6))
        ff.columnconfigure(0, weight=1)
        tk.Entry(ff, textvariable=self._file_var, font=('Arial', 9),
                 relief='solid', bd=1).pack(side='left', fill='x', expand=True)
        tk.Button(ff, text='📂', bg=C['bg'], relief='flat', font=('Arial', 11),
                  cursor='hand2', command=self._browse).pack(side='left', padx=(4, 0))

        # Σημείωση στήλες
        tk.Label(body,
                 text='Απαιτείται στήλη: Α.Φ.Μ.  |  Προαιρετική: Ονομασία Σχολείου',
                 bg=C['bg'], fg='#666666', font=('Arial', 8),
                 anchor='w').grid(row=2, column=0, sticky='w', pady=(0, 10))

        # Ημερομηνία λήξης
        tk.Label(body, text='Ημερομηνία λήξης (ΗΗ/Μ/ΕΕΕΕ):',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=3, column=0, sticky='w', pady=(0, 3))
        date_row = tk.Frame(body, bg=C['bg'])
        date_row.grid(row=4, column=0, sticky='w', pady=(0, 6))
        tk.Entry(date_row, textvariable=self._date_var, font=('Arial', 9),
                 relief='solid', bd=1, width=14).pack(side='left')
        tk.Label(date_row, text='  (31/8/2026 αντικαθίσταται αυτόματα)',
                 bg=C['bg'], fg='#888888', font=('Arial', 8)).pack(side='left')

        # Κουμπί εκτέλεσης
        btn_row = tk.Frame(body, bg=C['bg'])
        btn_row.grid(row=5, column=0, sticky='w', pady=(4, 8))
        self._conn_btn = tk.Button(btn_row,
                  text='▶  Σύνδεση & Εκτέλεση',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=12, pady=5, cursor='hand2',
                  command=self._connect_and_run)
        self._conn_btn.pack(side='left')

        # Status
        self._status_var = tk.StringVar(value='Επίλεξε αρχείο και πάτα Σύνδεση & Εκτέλεση.')
        tk.Label(body, textvariable=self._status_var,
                 bg=C['bg'], fg=C['status_run'],
                 font=('Arial', 8), anchor='w').grid(row=6, column=0, sticky='w', pady=(0, 4))

        # Log
        tk.Label(body, text='Αρχείο καταγραφής:',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=7, column=0, sticky='w', pady=(4, 2))
        self._log = st2.ScrolledText(body, height=12, font=('Consolas', 8),
                                      relief='solid', bd=1, state='disabled',
                                      bg='#F5F5F5', wrap=tk.WORD)
        self._log.grid(row=8, column=0, sticky='nsew', pady=(0, 4))
        body.rowconfigure(8, weight=1)

        self.protocol('WM_DELETE_WINDOW', self._on_close)

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=self,
            title='Επιλογή αρχείου εκπαιδευτικών',
            filetypes=[('Excel/CSV', '*.xlsx *.xls *.csv'), ('Όλα', '*.*')])
        if path:
            self._file_var.set(path)

    def _log_msg(self, msg):
        def _do():
            self._log.configure(state='normal')
            self._log.insert(tk.END, msg + '\n')
            self._log.see(tk.END)
            self._log.configure(state='disabled')
        self.after(0, _do)

    def _ask_user(self, title, prompt, options=None):
        """
        Thread-safe ask_user callback.
        Εκτελείται από background thread — marshals στο main thread και περιμένει.
        """
        import threading
        result   = [None]
        ev       = threading.Event()

        def _show():
            try:
                if options is not None:
                    dlg = _NumberedChoiceDialog(self, title, prompt, options)
                    result[0] = dlg.result
                else:
                    from tkinter import simpledialog
                    result[0] = simpledialog.askstring(title, prompt, parent=self)
            except Exception:
                result[0] = None
            finally:
                ev.set()

        self.after(0, _show)
        ev.wait(timeout=600)   # 10 λεπτά timeout
        return result[0]

    def _connect_and_run(self):
        import threading as _th
        path = self._file_var.get().strip()
        if not path:
            messagebox.showwarning('Προσοχή', 'Επίλεξε αρχείο πρώτα.', parent=self)
            return
        self._conn_btn.configure(state='disabled', text='Εκτελείται...')
        self._status_var.set('Σύνδεση στο MySchool...')

        def _do():
            import termination
            drv = termination.connect(log=self._log_msg)
            if not drv:
                def _fail():
                    self._conn_btn.configure(state='normal', text='▶  Σύνδεση & Εκτέλεση')
                    self._status_var.set('Αποτυχία σύνδεσης — έλεγξε credentials στις Ρυθμίσεις.')
                self.after(0, _fail)
                return
            self._driver = drv
            self.after(0, lambda: self._status_var.set('Εκτέλεση...'))
            termination.run(
                {'file_path': path, 'date': self._date_var.get().strip()},
                drv,
                callback=self._log_msg,
                ask_user=self._ask_user
            )
            def _after():
                self._conn_btn.configure(state='normal', text='▶  Σύνδεση & Εκτέλεση')
                self._status_var.set('Ολοκλήρωση.')
            self.after(0, _after)

        _th.Thread(target=_do, daemon=True).start()

    def _on_close(self):
        if self._driver:
            try: self._driver.quit()
            except Exception: pass
        self.destroy()


class AbsencesDialog(tk.Toplevel):
    """Καταχώρηση Απουσίας σε Οργανική — Ολική Διάθεση."""

    _HDR_BG  = '#1F4E79'
    _LBL_CLR = '#1F4E79'

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Καταχώρηση Απουσίας σε Οργανική')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        self._driver   = None
        self._file_var = tk.StringVar()

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        self.geometry('600x480')
        pw = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

    def _build(self):
        from tkinter import scrolledtext as st2

        hdr = tk.Frame(self, bg=self._HDR_BG, pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='📝  Καταχώρηση Απουσίας σε Οργανική',
                 bg=self._HDR_BG, fg='white',
                 font=('Arial', 12, 'bold')).pack()
        tk.Label(hdr, text='Ολική Διάθεση — μόνο για χρήση από Δ/νση Π.Ε. Αν. Θεσσαλονίκης',
                 bg=self._HDR_BG, fg='#FF6B6B',
                 font=('Arial', 8, 'italic')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=16, pady=12)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=1)

        # Αρχείο
        tk.Label(body, text='Αρχείο τοποθετήσεων (Excel):',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky='w', pady=(0, 3))
        ff = tk.Frame(body, bg=C['bg'])
        ff.grid(row=1, column=0, sticky='ew', pady=(0, 6))
        ff.columnconfigure(0, weight=1)
        tk.Entry(ff, textvariable=self._file_var, font=('Arial', 9),
                 relief='solid', bd=1).pack(side='left', fill='x', expand=True)
        tk.Button(ff, text='📂', bg=C['bg'], relief='flat', font=('Arial', 11),
                  cursor='hand2', command=self._browse).pack(side='left', padx=(4, 0))

        tk.Label(body,
                 text=('Απαιτούνται στήλες: Α.Μ., Σχέση τοποθέτησης, Έως.\n'
                       'Επεξεργάζονται όσοι έχουν 2η εγγραφή "Ολική Διάθεση '
                       '(ανάγκες υπηρεσίας - κύριος φορέας)" ή "Απόσπαση '
                       '(με αίτηση - κύριος φορέας)".'),
                 bg=C['bg'], fg='#666666', font=('Arial', 8),
                 justify='left', anchor='w').grid(row=2, column=0, sticky='w', pady=(0, 10))

        # Κουμπί εκτέλεσης
        btn_row = tk.Frame(body, bg=C['bg'])
        btn_row.grid(row=3, column=0, sticky='w', pady=(4, 8))
        self._conn_btn = tk.Button(btn_row,
                  text='▶  Σύνδεση & Εκτέλεση',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=12, pady=5, cursor='hand2',
                  command=self._connect_and_run)
        self._conn_btn.pack(side='left')

        # Status
        self._status_var = tk.StringVar(value='Επίλεξε αρχείο και πάτα Σύνδεση & Εκτέλεση.')
        tk.Label(body, textvariable=self._status_var,
                 bg=C['bg'], fg=C['status_run'],
                 font=('Arial', 8), anchor='w').grid(row=4, column=0, sticky='w', pady=(0, 4))

        # Log
        tk.Label(body, text='Αρχείο καταγραφής:',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=5, column=0, sticky='w', pady=(4, 2))
        self._log = st2.ScrolledText(body, height=14, font=('Consolas', 8),
                                      relief='solid', bd=1, state='disabled',
                                      bg='#F5F5F5', wrap=tk.WORD)
        self._log.grid(row=6, column=0, sticky='nsew', pady=(0, 4))
        body.rowconfigure(6, weight=1)

        self.protocol('WM_DELETE_WINDOW', self._on_close)

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=self,
            title='Επιλογή αρχείου τοποθετήσεων',
            filetypes=[('Excel', '*.xlsx *.xls'), ('Όλα', '*.*')])
        if path:
            self._file_var.set(path)

    def _log_msg(self, msg):
        def _do():
            self._log.configure(state='normal')
            self._log.insert(tk.END, msg + '\n')
            self._log.see(tk.END)
            self._log.configure(state='disabled')
        self.after(0, _do)

    def _ask_user(self, title, prompt, options=None):
        """Thread-safe ask_user callback — μαρσαλάρει στο main thread και περιμένει."""
        import threading
        result = [None]
        ev     = threading.Event()

        def _show():
            try:
                if options is not None:
                    dlg = _NumberedChoiceDialog(self, title, prompt, options)
                    result[0] = dlg.result
                else:
                    from tkinter import simpledialog
                    result[0] = simpledialog.askstring(title, prompt, parent=self)
            except Exception:
                result[0] = None
            finally:
                ev.set()

        self.after(0, _show)
        ev.wait(timeout=600)
        return result[0]

    def _connect_and_run(self):
        import threading as _th
        path = self._file_var.get().strip()
        if not path:
            messagebox.showwarning('Προσοχή', 'Επίλεξε αρχείο πρώτα.', parent=self)
            return
        self._conn_btn.configure(state='disabled', text='Εκτελείται...')
        self._status_var.set('Σύνδεση στο MySchool...')

        def _do():
            import absences
            drv = absences.connect(log=self._log_msg)
            if not drv:
                def _fail():
                    self._conn_btn.configure(state='normal', text='▶  Σύνδεση & Εκτέλεση')
                    self._status_var.set('Αποτυχία σύνδεσης — έλεγξε credentials στις Ρυθμίσεις.')
                self.after(0, _fail)
                return
            self._driver = drv
            self.after(0, lambda: self._status_var.set('Εκτέλεση...'))
            absences.run(
                {'file_path': path},
                drv,
                callback=self._log_msg,
                ask_user=self._ask_user
            )
            def _after():
                self._conn_btn.configure(state='normal', text='▶  Σύνδεση & Εκτέλεση')
                self._status_var.set('Ολοκλήρωση.')
            self.after(0, _after)

        _th.Thread(target=_do, daemon=True).start()

    def _on_close(self):
        if self._driver:
            try: self._driver.quit()
            except Exception: pass
        self.destroy()


class FunctionalityDialog(tk.Toplevel):
    """Αλλαγή Λειτουργικότητας σχολικών μονάδων στο MySchool."""

    _HDR_BG  = '#1F4E79'
    _LBL_CLR = '#1F4E79'

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Αλλαγή Λειτουργικότητας')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        self._driver   = None
        self._file_var = tk.StringVar()

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        self.geometry('600x460')
        pw = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

    def _build(self):
        from tkinter import scrolledtext as st2

        hdr = tk.Frame(self, bg=self._HDR_BG, pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='🔢  Αλλαγή Λειτουργικότητας',
                 bg=self._HDR_BG, fg='white',
                 font=('Arial', 12, 'bold')).pack()
        tk.Label(hdr, text='Αυτόματη ενημέρωση λειτουργικότητας σχολικών μονάδων στο MySchool',
                 bg=self._HDR_BG, fg='#A8C4D8',
                 font=('Arial', 8, 'italic')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=16, pady=12)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=1)

        # Αρχείο
        tk.Label(body, text='Αρχείο σχολικών μονάδων (Excel ή CSV):',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky='w', pady=(0, 3))
        ff = tk.Frame(body, bg=C['bg'])
        ff.grid(row=1, column=0, sticky='ew', pady=(0, 4))
        ff.columnconfigure(0, weight=1)
        tk.Entry(ff, textvariable=self._file_var, font=('Arial', 9),
                 relief='solid', bd=1).pack(side='left', fill='x', expand=True)
        tk.Button(ff, text='📂', bg=C['bg'], relief='flat', font=('Arial', 11),
                  cursor='hand2', command=self._browse).pack(side='left', padx=(4, 0))

        # Σημείωση στήλες
        tk.Label(body,
                 text='Απαιτούνται στήλες: Κωδικός  |  Παλιά Λειτουργικότητα  |  Νέα Λειτουργικότητα',
                 bg=C['bg'], fg='#666666', font=('Arial', 8),
                 anchor='w').grid(row=2, column=0, sticky='w', pady=(0, 10))

        # Κουμπί εκτέλεσης
        btn_row = tk.Frame(body, bg=C['bg'])
        btn_row.grid(row=3, column=0, sticky='w', pady=(0, 8))
        self._conn_btn = tk.Button(btn_row,
                  text='▶  Σύνδεση & Εκτέλεση',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=12, pady=5, cursor='hand2',
                  command=self._connect_and_run)
        self._conn_btn.pack(side='left')

        # Status
        self._status_var = tk.StringVar(value='Επίλεξε αρχείο και πάτα Σύνδεση & Εκτέλεση.')
        tk.Label(body, textvariable=self._status_var,
                 bg=C['bg'], fg=C['status_run'],
                 font=('Arial', 8), anchor='w').grid(row=4, column=0, sticky='w', pady=(0, 4))

        # Log
        tk.Label(body, text='Αρχείο καταγραφής:',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=5, column=0, sticky='w', pady=(4, 2))
        self._log = st2.ScrolledText(body, height=12, font=('Consolas', 8),
                                      relief='solid', bd=1, state='disabled',
                                      bg='#F5F5F5', wrap=tk.WORD)
        self._log.grid(row=6, column=0, sticky='nsew', pady=(0, 4))
        body.rowconfigure(6, weight=1)

        self.protocol('WM_DELETE_WINDOW', self._on_close)

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=self,
            title='Επιλογή αρχείου σχολικών μονάδων',
            filetypes=[('Excel/CSV', '*.xlsx *.xls *.csv'), ('Όλα', '*.*')])
        if path:
            self._file_var.set(path)

    def _log_msg(self, msg):
        def _do():
            self._log.configure(state='normal')
            self._log.insert(tk.END, msg + '\n')
            self._log.see(tk.END)
            self._log.configure(state='disabled')
        self.after(0, _do)

    def _connect_and_run(self):
        import threading as _th
        path = self._file_var.get().strip()
        if not path:
            messagebox.showwarning('Προσοχή', 'Επίλεξε αρχείο πρώτα.', parent=self)
            return
        self._conn_btn.configure(state='disabled', text='Εκτελείται...')
        self._status_var.set('Σύνδεση στο MySchool...')

        def _do():
            import functionality
            drv = functionality.connect(log=self._log_msg)
            if not drv:
                def _fail():
                    self._conn_btn.configure(state='normal', text='▶  Σύνδεση & Εκτέλεση')
                    self._status_var.set('Αποτυχία σύνδεσης — έλεγξε credentials στις Ρυθμίσεις.')
                self.after(0, _fail)
                return
            self._driver = drv
            self.after(0, lambda: self._status_var.set('Εκτέλεση...'))
            functionality.run(
                {'file_path': path},
                drv,
                callback=self._log_msg,
            )
            def _after():
                self._conn_btn.configure(state='normal', text='▶  Σύνδεση & Εκτέλεση')
                self._status_var.set('Ολοκλήρωση.')
            self.after(0, _after)

        _th.Thread(target=_do, daemon=True).start()

    def _on_close(self):
        if self._driver:
            try: self._driver.quit()
            except Exception: pass
        self.destroy()


class PanicEndDialog(tk.Toplevel):
    """Λήξη PANIC — Διαγραφή εγγραφών Γραμματειακής Υποστήριξης."""

    _HDR_BG  = '#B71C1C'
    _LBL_CLR = '#B71C1C'

    def __init__(self, parent):
        super().__init__(parent)
        self.title('PANIC — Λήξη')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.transient(parent)
        self._driver   = None
        self._file_var = tk.StringVar()

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        self.geometry('580x460')
        pw = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

    def _build(self):
        from tkinter import scrolledtext as st2

        hdr = tk.Frame(self, bg=self._HDR_BG, pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='⏹  PANIC — Λήξη / Αναίρεση Γραμματειακής Υποστήριξης',
                 bg=self._HDR_BG, fg='white',
                 font=('Arial', 12, 'bold')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=16, pady=12)
        body.pack(fill='both', expand=True)
        body.columnconfigure(0, weight=1)

        # Αρχείο
        tk.Label(body, text='Αρχείο εκπαιδευτικών (Excel ή CSV):',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky='w', pady=(0, 3))
        ff = tk.Frame(body, bg=C['bg'])
        ff.grid(row=1, column=0, sticky='ew', pady=(0, 4))
        ff.columnconfigure(0, weight=1)
        tk.Entry(ff, textvariable=self._file_var, font=('Arial', 9),
                 relief='solid', bd=1).pack(side='left', fill='x', expand=True)
        tk.Button(ff, text='📂', bg=C['bg'], relief='flat', font=('Arial', 11),
                  cursor='hand2', command=self._browse).pack(side='left', padx=(4, 0))

        # Ενημέρωση από τελευταία Έναρξη
        import editor as _ed
        _last = _ed.get_panic_path()
        if _last and os.path.exists(_last):
            self._file_var.set(_last)
            _hint = f'Προφορτώθηκε από την τελευταία Έναρξη: {os.path.basename(_last)}'
            _hint_clr = C.get('status_ok', '#388E3C')
        else:
            _hint = 'Δεν βρέθηκε αρχείο από προηγούμενη Έναρξη — επίλεξε χειροκίνητα.'
            _hint_clr = C.get('warn', '#E65100')
        tk.Label(body, text=_hint, bg=C['bg'], fg=_hint_clr,
                 font=('Arial', 8), anchor='w', wraplength=520).grid(
                 row=2, column=0, sticky='w', pady=(0, 10))

        # Κουμπί
        btn_row = tk.Frame(body, bg=C['bg'])
        btn_row.grid(row=3, column=0, sticky='w', pady=(0, 8))
        self._run_btn = tk.Button(btn_row,
                  text='⏹  Σύνδεση & Διαγραφή',
                  bg='#B71C1C', fg='white',
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=12, pady=5, cursor='hand2',
                  activebackground='#D32F2F', activeforeground='white',
                  command=self._connect_and_run)
        self._run_btn.pack(side='left')

        # Status
        self._status_var = tk.StringVar(value='Επίλεξε αρχείο και πάτα Σύνδεση & Διαγραφή.')
        tk.Label(body, textvariable=self._status_var,
                 bg=C['bg'], fg=C['status_run'],
                 font=('Arial', 8), anchor='w').grid(row=4, column=0, sticky='w', pady=(0, 4))

        # Log
        tk.Label(body, text='Αρχείο καταγραφής:',
                 bg=C['bg'], fg=self._LBL_CLR,
                 font=('Arial', 9, 'bold')).grid(row=5, column=0, sticky='w', pady=(4, 2))
        self._log = st2.ScrolledText(body, height=12, font=('Consolas', 8),
                                      relief='solid', bd=1, state='disabled',
                                      bg='#F5F5F5', wrap=tk.WORD)
        self._log.grid(row=6, column=0, sticky='nsew', pady=(0, 4))
        body.rowconfigure(6, weight=1)

        self.protocol('WM_DELETE_WINDOW', self._on_close)

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=self,
            title='Επιλογή αρχείου εκπαιδευτικών',
            filetypes=[('Excel/CSV', '*.xlsx *.xls *.csv'), ('Όλα', '*.*')])
        if path:
            self._file_var.set(path)

    def _log_msg(self, msg):
        def _do():
            self._log.configure(state='normal')
            self._log.insert(tk.END, msg + '\n')
            self._log.see(tk.END)
            self._log.configure(state='disabled')
        self.after(0, _do)

    def _connect_and_run(self):
        import threading as _th
        path = self._file_var.get().strip()
        if not path:
            messagebox.showwarning('Προσοχή', 'Επίλεξε αρχείο πρώτα.', parent=self)
            return
        if not os.path.exists(path):
            messagebox.showwarning('Προσοχή', 'Το αρχείο δεν βρέθηκε.', parent=self)
            return
        self._run_btn.configure(state='disabled', text='Εκτελείται...')
        self._status_var.set('Σύνδεση στο MySchool...')
        def _do():
            import editor
            drv = editor.connect(log=self._log_msg)
            if not drv:
                def _fail():
                    self._run_btn.configure(state='normal', text='⏹  Σύνδεση & Διαγραφή')
                    self._status_var.set('Αποτυχία σύνδεσης — έλεγξε credentials στις Ρυθμίσεις.')
                self.after(0, _fail)
                return
            self._driver = drv
            self.after(0, lambda: self._status_var.set('Εκτέλεση διαγραφών...'))
            editor.run_delete({'file_path': path}, drv, callback=self._log_msg)
            def _after():
                self._run_btn.configure(state='normal', text='⏹  Σύνδεση & Διαγραφή')
                self._status_var.set('Ολοκλήρωση.')
            self.after(0, _after)
        _th.Thread(target=_do, daemon=True).start()

    def _on_close(self):
        if self._driver:
            try: self._driver.quit()
            except Exception: pass
        self.destroy()


class InformEmailDialog(tk.Toplevel):
    """Αποστολή ενημερωτικού email υπενθύμισης επιβεβαίωσης MySchool."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Ενημερωτικό email')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        pw = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        ph = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{pw}+{ph}')

    @staticmethod
    def _obligation_date(now):
        """Επιστρέφει την πιο πρόσφατη 1η ή 15η του μήνα."""
        return now.replace(day=15) if now.day >= 15 else now.replace(day=1)

    def _build(self):
        from tkinter import scrolledtext as st3
        from datetime import datetime as _dt

        hdr = tk.Frame(self, bg='#0F6E56', pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='✉  Ενημερωτικό email',
                 bg='#0F6E56', fg='white',
                 font=('Arial', 12, 'bold')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=18, pady=14)
        body.pack(fill='both', expand=True)

        # ── Ημερομηνίες ──────────────────────────────────────────────────────
        GREEK_DAYS = ['Δευτέρα', 'Τρίτη', 'Τετάρτη', 'Πέμπτη',
                      'Παρασκευή', 'Σάββατο', 'Κυριακή']
        now          = _dt.now()
        day_name     = GREEK_DAYS[now.weekday()]
        date_send    = f'{now.day}/{now.month}/{now.year}'
        date_send_lg = f'{day_name} {date_send}'
        obl          = self._obligation_date(now)
        date_obl     = f'{obl.day}/{obl.month}/{obl.year}'

        # ── Θέμα ─────────────────────────────────────────────────────────────
        tk.Label(body, text='Θέμα:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky='w', pady=(0, 3))
        self._subj_var = tk.StringVar(
            value=f'Υπενθύμιση επιβεβαίωσης δεδομένων στο myschool για {date_obl}')
        tk.Entry(body, textvariable=self._subj_var, width=58,
                 font=('Arial', 9), relief='solid', bd=1).grid(
                 row=1, column=0, columnspan=2, sticky='ew', pady=(0, 10))

        # ── Σώμα μηνύματος ───────────────────────────────────────────────────
        tk.Label(body, text='Σώμα μηνύματος:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold')).grid(row=2, column=0, sticky='w', pady=(0, 3))
        self._body_txt = st3.ScrolledText(body, width=58, height=9,
                                           font=('Arial', 9), relief='solid', bd=1, wrap=tk.WORD)
        self._body_txt.grid(row=3, column=0, columnspan=2, pady=(0, 6))

        # ── Checkbox αδυναμία αναθέσεων ──────────────────────────────────────
        self._ady_var = tk.BooleanVar(value=False)
        tk.Checkbutton(body,
                       text='Συμπερίληψη πρότασης για αδυναμία αναθέσεων',
                       variable=self._ady_var,
                       bg=C['bg'], fg=C['hdr_bg'],
                       selectcolor=C['sel_bg'], activebackground=C['bg'],
                       font=('Arial', 9),
                       command=self._refresh_body).grid(
                       row=4, column=0, columnspan=2, sticky='w', pady=(0, 10))

        # ── Παραλήπτες ───────────────────────────────────────────────────────
        tk.Label(body, text='Συμπληρώστε παραλήπτες '
                             '(ένα ανά γραμμή, ή περισσότερα στην ίδια γραμμή με «;»):',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold')).grid(row=5, column=0, columnspan=2,
                                                  sticky='w', pady=(0, 3))

        _saved = getattr(config, 'INFORM_RECIPIENTS', '')
        self._recip_txt = st3.ScrolledText(body, width=42, height=4,
                                            font=('Consolas', 8), relief='solid', bd=1, wrap=tk.NONE)
        self._recip_txt.grid(row=6, column=0, pady=(0, 4), sticky='ew')
        self._recip_txt.insert('1.0', _saved)

        tk.Button(body, text='💾 Αποθήκευση',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 8), relief='flat',
                  padx=8, pady=4, cursor='hand2',
                  command=self._save_recipients).grid(
                  row=6, column=1, sticky='nw', padx=(8, 0))

        # ── Επισύναψη αρχείου ─────────────────────────────────────────────────
        self._attach_path = None
        att_row = tk.Frame(body, bg=C['bg'])
        att_row.grid(row=7, column=0, columnspan=2, sticky='ew', pady=(2, 6))
        tk.Button(att_row, text='📎 Επισύναψη αρχείου',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 8), relief='flat',
                  padx=8, pady=4, cursor='hand2',
                  command=self._pick_attachment).pack(side='left')
        self._attach_lbl = tk.Label(att_row, text='Κανένα αρχείο',
                                     bg=C['bg'], fg='#888888',
                                     font=('Arial', 8), anchor='w')
        self._attach_lbl.pack(side='left', padx=(8, 4), fill='x', expand=True)
        self._attach_clear_btn = tk.Button(att_row, text='✕',
                  bg=C['bg'], fg='#cc0000',
                  font=('Arial', 8, 'bold'), relief='flat',
                  padx=4, cursor='hand2',
                  command=self._clear_attachment)
        self._attach_clear_btn.pack(side='left')
        self._attach_clear_btn.pack_forget()  # κρυφό μέχρι να επιλεγεί αρχείο

        # ── Status + κουμπί αποστολής ─────────────────────────────────────────
        self._status_var = tk.StringVar(value='')
        tk.Label(body, textvariable=self._status_var, bg=C['bg'],
                 fg=C['status_run'], font=('Arial', 8), wraplength=420,
                 justify='left').grid(row=8, column=0, columnspan=2, sticky='w', pady=(4, 2))

        self._send_btn = tk.Button(body, text='✉  Αποστολή',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 10, 'bold'), relief='flat',
                  padx=18, pady=6, cursor='hand2',
                  command=self._send)
        self._send_btn.grid(row=9, column=0, columnspan=2, pady=(4, 0), sticky='e')

        body.columnconfigure(0, weight=1)

        self._date_send_lg = date_send_lg
        self._date_send    = date_send
        self._date_obl     = date_obl
        self._refresh_body()

    def _pick_attachment(self):
        from tkinter import filedialog as _fd
        path = _fd.askopenfilename(
            parent=self,
            title='Επιλογή αρχείου επισύναψης',
            filetypes=[('Όλα τα αρχεία', '*.*'),
                       ('Excel', '*.xlsx *.xls'),
                       ('PDF', '*.pdf'),
                       ('Word', '*.docx *.doc')])
        if path:
            self._attach_path = path
            name = os.path.basename(path)
            max_w = 35
            display = name if len(name) <= max_w else name[:max_w - 1] + '…'
            self._attach_lbl.configure(text=display, fg=C['hdr_bg'])
            self._attach_clear_btn.pack(side='left')

    def _clear_attachment(self):
        self._attach_path = None
        self._attach_lbl.configure(text='Κανένα αρχείο', fg='#888888')
        self._attach_clear_btn.pack_forget()

    def _refresh_body(self):
        """Ανανεώνει το σώμα με ή χωρίς την πρόταση αδυναμίας."""
        sig = config.email_signature().strip()
        ady_line = (
            '\nΓια πάγιες περιπτώσεις αδυναμίας αναθέσεων έχουν γίνει ήδη οι απαραίτητες '
            'ενέργειες για να κάνετε επιβεβαίωση σήμερα χωρίς να σας εμφανίσει κάποιο '
            'πρόβλημα αλλιώς επικοινωνείτε με μένα.\n'
        ) if self._ady_var.get() else ''

        text = (f'Καλή σας μέρα,\n\n'
                f'Σήμερα, {self._date_send_lg}, κάνουμε επιβεβαίωση δεδομένων '
                f'στο myschool για {self._date_obl}.\n'
                f'{ady_line}')
        if sig:
            text += f'\n\n{sig}'
        self._body_txt.delete('1.0', tk.END)
        self._body_txt.insert('1.0', text)

    def _save_recipients(self):
        """Αποθηκεύει τους παραλήπτες στο local_settings.json."""
        recips = self._recip_txt.get('1.0', tk.END).strip()
        try:
            _save_config({'INFORM_RECIPIENTS': recips})
            setattr(config, 'INFORM_RECIPIENTS', recips)
            self._status_var.set('✓ Παραλήπτες αποθηκεύτηκαν.')
        except Exception as e:
            self._status_var.set(f'❌ Σφάλμα αποθήκευσης: {e}')

    def _send(self):
        import threading as _th
        from core.framework import send_email as _send_email
        subject = self._subj_var.get().strip()
        body    = self._body_txt.get('1.0', tk.END).strip()
        # Παραλήπτες: ένας ανά γραμμή, ή περισσότεροι στην ίδια γραμμή με
        # διαχωριστικό «;» (π.χ. «a@sch.gr; b@sch.gr»).
        recips = []
        for line in self._recip_txt.get('1.0', tk.END).splitlines():
            for part in line.split(';'):
                part = part.strip()
                if part and '@' in part:
                    recips.append(part)
        if not subject:
            messagebox.showwarning('Προσοχή', 'Συμπλήρωσε το Θέμα.', parent=self)
            return
        if not body:
            messagebox.showwarning('Προσοχή', 'Συμπλήρωσε το σώμα.', parent=self)
            return
        if not recips:
            messagebox.showwarning('Προσοχή',
                'Δεν υπάρχουν παραλήπτες.\nΣυμπλήρωσε email (ένα ανά γραμμή).', parent=self)
            return
        if not getattr(config, 'FROM_PASSWORD', ''):
            messagebox.showwarning('Προσοχή',
                'Ο κωδικός email δεν έχει οριστεί.\nΠήγαινε στις Ρυθμίσεις (⚙).', parent=self)
            return
        n  = len(recips)
        pl = 'η' if n == 1 else 'ες'
        if not messagebox.askyesno('Επιβεβαίωση',
                f'Αποστολή email σε {n} παραλήπτ{pl};\n\n' + '\n'.join(recips), parent=self):
            return
        self._send_btn.configure(state='disabled', bg=C['btn_dis'], text='Αποστολή...')
        self._status_var.set('Αποστολή...')

        _subj        = subject
        _body        = body
        _recips      = recips[:]
        _date        = self._date_send
        _attach_path = self._attach_path

        def _do():
            try:
                _send_email(config, _recips, _subj, _body, _attach_path)

                from_addr = getattr(config, 'FROM_EMAIL', '')
                if from_addr:
                    _confirm_subj = f'Αποστολή ενημερωτικού {_date}'
                    _confirm_body = (
                        f'Εστάλη ενημερωτικό email:\n\n'
                        f'Ημερομηνία: {_date}\n'
                        f'Θέμα: {_subj}\n'
                        f'Παραλήπτες ({n}):\n' + '\n'.join(f'  • {r}' for r in _recips)
                    )
                    _send_email(config, [from_addr], _confirm_subj, _confirm_body, None)

                def _ok():
                    self._send_btn.configure(state='normal', bg=C['btn_bg'], text='✉  Αποστολή')
                    self._status_var.set(f'✓ Εστάλη σε {n} παραλήπτ{pl}.')
                self.after(0, _ok)
            except Exception as err:
                def _err(m=str(err)):
                    self._send_btn.configure(state='normal', bg=C['btn_bg'], text='✉  Αποστολή')
                    self._status_var.set(f'❌ Σφάλμα: {m}')
                self.after(0, _err)
        _th.Thread(target=_do, daemon=True).start()


def _password_gate(parent, action):
    """Ζητάει κωδικό πρόσβασης πριν εκτελέσει το `action` — χρησιμοποιείται
    για στοιχεία που παραμένουν ενεργά αλλά προορίζονται αποκλειστικά για
    χρήση από τη ΔΙΠΕ Αν. Θεσ/κης."""
    from tkinter import simpledialog, messagebox
    pwd = simpledialog.askstring(
        'Απαιτείται κωδικός',
        'Αποκλειστικά για χρήση από ΔΙΠΕ Αν. Θεσ/κης.\n\nΔώσε κωδικό πρόσβασης:',
        show='*', parent=parent)
    if pwd is None:
        return
    if pwd == '1511':
        action()
    else:
        messagebox.showerror('Λάθος κωδικός', 'Ο κωδικός δεν είναι σωστός.', parent=parent)


def _show_help(parent):
    """Ανοίγει τον οδηγό (README) στη σελίδα του GitHub repository — εκεί
    εμφανίζεται σωστά μορφοποιημένος, σε αντίθεση με το ωμό .md αρχείο."""
    HELP_URL = 'https://github.com/MichalisKat/myschool-checks#readme'
    try:
        import webbrowser
        if not webbrowser.open(HELP_URL):
            raise RuntimeError('webbrowser.open() επέστρεψε False')
    except Exception:
        from tkinter import messagebox
        messagebox.showinfo('Βοήθεια',
                            f'Ο οδηγός βρίσκεται στο:\n{HELP_URL}\n\n'
                            'αλλά δεν κατέστη δυνατό να ανοίξει αυτόματα.',
                            parent=parent)


def _show_splash(root):
    splash = tk.Toplevel(root)
    splash.overrideredirect(True)
    splash.configure(bg=C['hdr_bg'])

    w, h = 460, 320
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    x  = (sw - w) // 2
    y  = (sh - h) // 2
    splash.geometry(f'{w}x{h}+{x}+{y}')
    splash.lift()
    splash.attributes('-topmost', True)

    tk.Label(splash, text='MySchool Checks',
             bg=C['hdr_bg'], fg=C['hdr_fg'],
             font=('Arial', 17, 'bold')).pack(pady=(24, 2))
    tk.Label(splash, text='Δ/νση Π.Ε. Ανατολικής Θεσσαλονίκης',
             bg=C['hdr_bg'], fg=C['hdr_sub'],
             font=('Arial', 9)).pack()
    tk.Label(splash, text=f'v{config.APP_VERSION}',
             bg=C['hdr_bg'], fg=C['hdr_sub'],
             font=('Arial', 8)).pack()

    from tkinter import ttk as _ttk
    _style = _ttk.Style(splash)
    _style.theme_use('default')
    _style.configure('Splash.Horizontal.TProgressbar',
                     troughcolor='#163D60', background='#E53935',
                     bordercolor=C['hdr_bg'], lightcolor='#E53935',
                     darkcolor='#C62828')
    _style.configure('SplashDone.Horizontal.TProgressbar',
                     troughcolor='#163D60', background='#4CA870',
                     bordercolor=C['hdr_bg'], lightcolor='#4CA870',
                     darkcolor='#4CA870')
    pb = _ttk.Progressbar(splash, mode='indeterminate', length=360,
                          style='Splash.Horizontal.TProgressbar')
    pb.pack(pady=(14, 8))
    pb.start(10)

    log_frame = tk.Frame(splash, bg='#163D60', padx=16, pady=8)
    log_frame.pack(fill='x', padx=30, pady=(0, 8))
    log_txt = tk.Text(log_frame, height=4,
                      bg='#163D60', fg=C['hdr_sub'],
                      font=('Consolas', 8), relief='flat',
                      cursor='arrow', wrap='word',
                      selectbackground='#163D60',
                      insertbackground='#163D60')
    log_txt.pack(fill='x')

    ready_btn = tk.Button(splash, text='▶   Είσοδος',
                          bg='#2E7D32', fg='white',
                          font=('Arial', 10, 'bold'),
                          relief='flat', padx=24, pady=6,
                          cursor='hand2',
                          activebackground='#43A047',
                          activeforeground='white')

    tk.Label(splash, text='Μιχάλης Κατσιρντάκης  •  2310 954145',
             bg=C['hdr_bg'], fg='#4A7FAF',
             font=('Arial', 8)).pack(side='bottom', pady=(0, 10))

    splash.update()
    return splash, pb, log_txt, ready_btn


def _splash_log(log_txt, msg):
    def _do():
        log_txt.configure(state='normal')
        log_txt.insert(tk.END, msg + '\n')
        log_txt.see(tk.END)
        log_txt.configure(state='disabled')
    try:
        log_txt.after(0, _do)
    except Exception:
        pass


def _launch(root, checks, splash, pb):
    """Κλείνει το splash και ξεκινά την κύρια εφαρμογή."""
    pb.stop()
    try:
        splash.destroy()
    except Exception:
        pass
    root.deiconify()
    root.lift()
    LauncherApp(root, checks)


def main():
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    root = tk.Tk()
    root.withdraw()

    # Ασφάλεια: σφάλματα ΜΕΣΑ σε Tk callback (π.χ. κατά τη δημιουργία του
    # κεντρικού παραθύρου) δεν φτάνουν ποτέ στο sys.excepthook — το Tk τα
    # πιάνει μόνο του και τα τυπώνει στο sys.stderr, το οποίο όμως είναι
    # ανακατευθυνόμενο στο GUIStream (status bar) και χάνονται σιωπηλά.
    # Εδώ τα γράφουμε ΚΑΙ στο crash.log ΚΑΙ σε πραγματικό stderr ΚΑΙ σε
    # popup, ώστε να μη «χάνονται» ποτέ ξανά.
    def _tk_callback_error(exc_type, exc_val, exc_tb):
        import traceback as _tb
        _err = ''.join(_tb.format_exception(exc_type, exc_val, exc_tb))
        try:
            _log = os.path.join(os.path.expanduser('~'), 'Desktop', 'crash.log')
            with open(_log, 'a', encoding='utf-8') as _f:
                _f.write('\n--- Tk callback exception ---\n' + _err)
        except Exception:
            pass
        try:
            if sys.__stderr__:
                sys.__stderr__.write(_err)
        except Exception:
            pass
        try:
            messagebox.showerror('Σφάλμα', _err[-1500:])
        except Exception:
            pass
    root.report_callback_exception = _tk_callback_error

    ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
    if os.path.exists(ico):
        try:
            root.iconbitmap(ico)
        except Exception:
            pass

    splash, pb, log_txt, ready_btn = _show_splash(root)

    checks_result = []
    done_flag     = threading.Event()

    def _startup():
        import subprocess as _sub
        _log_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'crash.log')

        def _log(msg):
            try:
                with open(_log_path, 'a', encoding='utf-8') as _f:
                    _f.write(msg + '\n')
            except Exception:
                pass

        try:
            _splash_log(log_txt, f'✓ Python {sys.version.split()[0]}')
            import time as _time
            _time.sleep(0.3)

            if getattr(sys, 'frozen', False):
                _splash_log(log_txt, '✓ Βιβλιοθήκες εγκατεστημένες')
                _time.sleep(0.2)
            else:
                _base    = os.path.dirname(os.path.abspath(__file__))
                _libs_ok = os.path.join(_base, '.libs_ok')
                _reqs    = [('pandas', 'pandas'), ('openpyxl', 'openpyxl'),
                            ('selenium', 'selenium'), ('xlrd', 'xlrd'),
                            ('html2text', 'html2text')]

                if os.path.exists(_libs_ok):
                    _splash_log(log_txt, '✓ Βιβλιοθήκες εγκατεστημένες')
                    _time.sleep(0.2)
                else:
                    _splash_log(log_txt, 'Έλεγχος βιβλιοθηκών...')
                    for pkg, imp in _reqs:
                        try:
                            __import__(imp)
                            _splash_log(log_txt, f'  ✓ {pkg}')
                        except ImportError:
                            _splash_log(log_txt, f'  ⬇ Εγκατάσταση {pkg}...')
                            try:
                                _sub.run([sys.executable, '-m', 'pip', 'install',
                                          pkg, '--disable-pip-version-check', '-q'],
                                         check=True)
                                _splash_log(log_txt, f'  ✓ {pkg} εγκαταστάθηκε')
                                _time.sleep(0.15)
                            except Exception as _e:
                                _log(f'pip {pkg}: {_e}')
                    try:
                        open(_libs_ok, 'w').close()
                    except Exception:
                        pass

            _splash_log(log_txt, 'Φόρτωση ελέγχων...')
            checks = load_checks()
            checks_result.append(checks)
            _splash_log(log_txt, f'✓ {len(checks)} έλεγχοι φορτώθηκαν')
            _time.sleep(0.4)

        except Exception as _e:
            _log(f'Σφάλμα εκκίνησης: {_e}')
        finally:
            done_flag.set()

    def _poll_checks():
        if not done_flag.is_set():
            root.after(100, _poll_checks)
            return

        checks = checks_result[0] if checks_result else []
        if not checks:
            pb.stop()
            from tkinter import messagebox as _mb
            _log_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'crash.log')
            _mb.showerror('Σφάλμα',
                          f'Δεν φορτώθηκαν έλεγχοι!\n\nΔες το αρχείο:\n{_log_path}')
            try:
                splash.destroy()
            except Exception:
                pass
            sys.exit(1)

        pb.stop()
        pb.configure(style='SplashDone.Horizontal.TProgressbar',
                     mode='determinate', value=100)
        _splash_log(log_txt, '✓ Έτοιμο!')
        root.after(0, lambda: _launch(root, checks, splash, pb))

    threading.Thread(target=_startup, daemon=True).start()
    root.after(100, _poll_checks)
    root.mainloop()


if __name__ == '__main__':
    main()
